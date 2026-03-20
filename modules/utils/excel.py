
# modules/utils/excel.py
from __future__ import annotations
from pathlib import Path
from typing import Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime as dt
import pandas as pd


def ensure_workbook(path: str | Path) -> Tuple[Workbook, bool]:
    """
    Opens an existing workbook if present; otherwise creates a new one.

    Parameters
    ----------
    path : str or Path
        Path to the Excel workbook (.xlsx).

    Returns
    -------
    (openpyxl.Workbook, bool)
        Workbook instance and a flag indicating whether it was newly created.
    """
    p = Path(path)
    if p.exists():
        return load_workbook(p), False
    wb = Workbook()
    # Remove the default 'Sheet' to keep the file tidy for first write.
    if wb.sheetnames == ["Sheet"]:
        wb.remove(wb["Sheet"])
    return wb, True


def ensure_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """
    Returns a worksheet by name, creating it if it does not exist.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Target workbook.
    sheet_name : str
        Worksheet name.

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        Worksheet instance.
    """
    return wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)


def _cell_to_rc(cell_ref: str) -> Tuple[int, int]:
    """
    Converts an A1-style cell reference to numeric (row, column) indices.

    Parameters
    ----------
    cell_ref : str
        A1-style cell reference (e.g., "B3").

    Returns
    -------
    (int, int)
        Row index (1-based) and column index (1-based).
    """
    letters = "".join(c for c in cell_ref if c.isalpha())
    digits = "".join(c for c in cell_ref if c.isdigit())
    return int(digits), column_index_from_string(letters)


def clear_range_keep_format(ws: Worksheet, top_left: str, width: int, height: int) -> None:
    """
    Clears the values in a rectangular region while preserving formatting.

    Parameters
    ----------
    ws : Worksheet
        Target worksheet.
    top_left : str
        A1 reference for the top-left cell of the region.
    width : int
        Number of columns to clear.
    height : int
        Number of rows to clear.

    Returns
    -------
    None
    """
    r0, c0 = _cell_to_rc(top_left)
    for r in range(r0, r0 + height):
        for c in range(c0, c0 + width):
            ws.cell(row=r, column=c).value = None


def _write_headers(ws: Worksheet, df: pd.DataFrame, anchor: str) -> None:
    """
    Writes DataFrame column headers at the anchor row.

    Parameters
    ----------
    ws : Worksheet
        Target worksheet.
    df : pandas.DataFrame
        Source DataFrame.
    anchor : str
        A1 cell where the first header cell will be written.

    Returns
    -------
    None
    """
    r0, c0 = _cell_to_rc(anchor)
    for j, name in enumerate(df.columns):
        ws.cell(row=r0, column=c0 + j, value=str(name))


def _write_values(ws: Worksheet, df: pd.DataFrame, anchor: str, include_header: bool) -> None:
    """
    Writes DataFrame values starting below the header if include_header=True,
    otherwise starting at the anchor row.

    Parameters
    ----------
    ws : Worksheet
    df : pandas.DataFrame
    anchor : str
    include_header : bool

    Returns
    -------
    None
    """
    r0, c0 = _cell_to_rc(anchor)
    start_row = r0 + (1 if include_header else 0)
    for i, (_, row) in enumerate(df.iterrows()):
        for j, val in enumerate(row):
            ws.cell(row=start_row + i, column=c0 + j, value=val)


def _detect_block(ws: Worksheet, anchor: str) -> Tuple[int, int]:
    """
    Detects the existing headers+data block width if the header row contains at least one non-empty value.
     This prevents false detection on empty sheets.

    Parameters
    ----------
    ws : Worksheet
        Worksheet to inspect.
    anchor : str
        A1 reference where the header row is expected.

    Returns
    -------
    (int, int)
        (width, height) of the detected block including header row.
        Returns (0, 0) if no header is found at the anchor row.
    """
    r0, c0 = _cell_to_rc(anchor)

    # Scan header row for *any* non-empty cell
    scan_limit = 256
    header_values = [
        ws.cell(row=r0, column=c0 + i).value for i in range(scan_limit)
    ]

    if not any(v not in (None, "") for v in header_values):
        return(0,0)
    
    #Determine width from contiguous non-empty header cells
    width = 0
    for v in header_values:
        if v in (None, ""):
            break
        width +=1

    # Detect height: count contiguous non-empty rows beneath the header.
    height = 1  # include header row
    r = r0 + 1
    while any(ws.cell(row=r, column=c).value not in (None, "")
              for c in range(c0, c0 + width)):
                height += 1
                r += 1

    return width, height

def _safe_save(wb: Workbook, path: str | Path, make_backup: bool) -> None:
    """
    Saves a workbook with optional timestamped backup.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook to save.
    path : str or Path
        Destination file path.
    make_backup : bool
        If True, creates a timestamped backup before saving.

    Returns
    -------
    None
    """
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    if make_backup and p.exists():
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = p.with_name(f"{p.stem}__bak_{ts}{p.suffix}")
        wb.save(backup_path)

    wb.save(p)


def write_once_then_update(
    path: str | Path,
    sheet: str,
    df: pd.DataFrame,
    anchor: str = "A1",
    include_header: bool = True,
    make_backup: bool = True
) -> None:
    """
    Writes a DataFrame into a worksheet at a fixed anchor location. If the
    block already exists, only refreshes the data region while preserving
    formatting, tables, charts, named ranges, and linked PPT cells.

    Parameters
    ----------
    path : str or Path
        Path to the Excel workbook to write/update.
    sheet : str
        Worksheet name.
    df : pandas.DataFrame
        DataFrame to write.
    anchor : str, default "A1"
        Top-left cell of the written block: header if include_header=True,
        otherwise the first data cell.
    include_header : bool, default True
        Whether to write DataFrame column headers at the anchor row.
    make_backup : bool, default True
        If True, saves a timestamped backup file next to the workbook before
        overwriting.

    Returns
    -------
    None
    """
    wb, _ = ensure_workbook(path)
    ws = ensure_sheet(wb, sheet)

    width, height = _detect_block(ws, anchor)

    if width == 0:
        # First-time write: headers (optional) + values
        if include_header:
            _write_headers(ws, df, anchor)
        _write_values(ws, df, anchor, include_header)
    else:
        # Refresh existing: clear old data region only, then write new values
        r0, c0 = _cell_to_rc(anchor)
        data_row_start = r0 + (1 if include_header else 0)
        existing_rows = max(height - (1 if include_header else 0), 0)

        if existing_rows > 0:
            clear_range_keep_format(
                ws,
                top_left=f"{get_column_letter(c0)}{data_row_start}",
                width=width,
                height=existing_rows
            )

        _write_values(ws, df, anchor, include_header)

    _safe_save(wb, path, make_backup)



