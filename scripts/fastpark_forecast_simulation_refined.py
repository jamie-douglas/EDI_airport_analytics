"""
FASTPARK FORECAST SIMULATION - REFINED / GENERATION 2
=====================================================

Purpose
-------
This script runs an all-date, rolling-origin historical backtest of refined
FastPark daily ENTRY and EXIT forecasts.

Generation 2 changes
--------------------
1. Tests every calendar date in the configured simulation period.
2. Reuses the as-of-safe data preparation and component calculations from
   Generation 1.
3. Uses only the strongest independent forecast components:

   ENTRY:
       - booking visibility curve
       - same-weekday historical demand
       - weekday-month historical demand
       - trend-adjusted weekday-month demand

   EXIT:
       - exit booking visibility curve
       - duration-adjusted exit booking curve
       - same-weekday historical demand
       - trend-adjusted weekday-month demand

4. Selects non-negative component weights that sum to one.
5. Selects weights separately by flow and forecast horizon.
6. Uses rolling-origin train / validation / test folds so the test month is
   not used to select its own weights.
7. Scores demand error only. Pricing and FTE conversion are deliberately not
   included in this version.

IMPORTANT
---------
This file expects the Generation 1 simulation module to be importable.
By default the module name is:

    fastpark_forecast_simulation_v2

If your Generation 1 Python filename is different, change BASE_MODULE_NAME
in get_refined_config(). Do not include the .py extension.
"""

# =============================================================================
# 0. IMPORTS
# =============================================================================

import importlib
import itertools
import pathlib
import sys
import time
from pathlib import Path

import numpy as np
import pandas as pd


PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

from modules.utils.progress import step


# =============================================================================
# 1. CONFIGURATION
# =============================================================================

def get_refined_config():
    """Central configuration for the Generation 2 backtest."""

    return {
        # Generation 1 module containing the functions supplied earlier.
        "base_module_name": "fastpark_forecast_simulation",

        # Historical extraction range.
        "history_start": "2024-01-01",
        "history_end": "2026-07-31",

        # Refined simulation range. Every calendar date is tested.
        "simulation_start": "2025-07-01",
        "simulation_end": "2026-07-31",

        # Same forecast horizons as Generation 1.
        "forecast_horizons_days": [
            0, 1, 2, 3, 4, 5, 6, 7,
            14, 21, 28, 35, 42, 49, 56,
        ],

        # All calendar days are included.
        "test_days_of_month": list(range(1, 32)),

        # Historical components.
        "same_weekday_n": 4,
        "weekday_month_max_n": 8,
        "minimum_weekday_month_observations": 3,

        # Trend adjustment.
        "trend_recent_days": 28,
        "trend_comparison_days": 84,
        "trend_factor_min": 0.70,
        "trend_factor_max": 1.30,

        # Weight grid. 0.05 creates interpretable 5 percentage-point weights.
        "weight_step": 0.025,

        # Rolling-origin design.
        # The first test month must have at least this many earlier months.
        "minimum_training_months": 6,
        # The immediately preceding month is used to choose weights.
        "validation_months": 1,
        # Optional recency limit for weight training. None = all prior history.
        "maximum_weight_training_months": 12,

        # Tie-break penalty. This very small penalty prefers weights near the
        # prior fold's solution when validation WAPE is effectively tied.
        "weight_stability_penalty": 0.0001,

        # Booking-curve shrinkage tested as an additional refined component.
        # Higher values pull booking forecasts more strongly toward the
        # historical baseline when visible bookings are low.
        "booking_shrinkage_strengths": [0, 25, 50, 100],

        # Development controls.
        "smoke_test": False,
        "smoke_test_target_count": 14,
        "smoke_test_horizons": [0, 7, 28, 56],
        "fail_fast": True,

        # Outputs.
        "output_path": (
            PROJECT_ROOT
            / "outputs"
            / "fastpark_forecast_simulation_refined.xlsx"
        ),
    }


# =============================================================================
# 2. GENERATION 1 IMPORT / CONFIGURATION
# =============================================================================

def import_generation_1(config):
    """Import the Generation 1 module by configured module name."""

    module_name = config["base_module_name"]

    try:
        return importlib.import_module(module_name)
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Could not import the Generation 1 simulation module.\n"
            f"Configured module: {module_name}\n"
            "Place this refined script in the same project environment as "
            "Generation 1, or change base_module_name in get_refined_config()."
        ) from exc


def build_generation_1_config(base, refined_config):
    """
    Start with Generation 1 configuration and override only the fields needed
    for the all-date Generation 2 simulation.
    """

    config = base.get_simulation_config()

    config.update(
        {
            "history_start": refined_config["history_start"],
            "history_end": refined_config["history_end"],
            "simulation_start": refined_config["simulation_start"],
            "simulation_end": refined_config["simulation_end"],
            "test_days_of_month": refined_config["test_days_of_month"],
            "forecast_horizons_days": refined_config[
                "forecast_horizons_days"
            ],
            "minimum_weekday_month_observations": refined_config[
                "minimum_weekday_month_observations"
            ],
            "smoke_test": refined_config["smoke_test"],
            "smoke_test_target_count": refined_config[
                "smoke_test_target_count"
            ],
            "smoke_test_horizons": refined_config[
                "smoke_test_horizons"
            ],
            "fail_fast": refined_config["fail_fast"],
            # Hourly and pricing/FTE work is intentionally excluded here.
            "enable_hourly_analysis": False,
        }
    )

    return config


# =============================================================================
# 3. GENERAL HELPERS
# =============================================================================

def clip_forecast(value):
    """Return a non-negative float or NaN."""

    if value is None or pd.isna(value):
        return np.nan

    return max(0.0, float(value))


def safe_mean(values):
    """Mean after removing missing and infinite values."""

    values = (
        pd.Series(values, dtype=float)
        .replace([np.inf, -np.inf], np.nan)
        .dropna()
    )

    if values.empty:
        return np.nan

    return float(values.mean())


def weighted_average(values, weights):
    """
    Weighted average with automatic re-normalisation over available values.
    """

    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)

    valid = np.isfinite(values) & np.isfinite(weights) & (weights >= 0)

    if not valid.any():
        return np.nan

    valid_values = values[valid]
    valid_weights = weights[valid]

    weight_total = valid_weights.sum()

    if weight_total <= 0:
        return np.nan

    return clip_forecast(
        np.average(
            valid_values,
            weights=valid_weights,
        )
    )


def wape(actual, forecast):
    """Weighted absolute percentage error as a decimal."""

    actual = pd.Series(actual, dtype=float)
    forecast = pd.Series(forecast, dtype=float)

    valid = actual.notna() & forecast.notna()
    actual = actual.loc[valid]
    forecast = forecast.loc[valid]

    if actual.empty:
        return np.nan

    denominator = actual.abs().sum()

    if denominator <= 0:
        return np.nan

    return float((forecast - actual).abs().sum() / denominator)


def score_forecast_frame(
    dataframe,
    group_columns,
    forecast_col="forecast_value",
    actual_col="actual_value",
):
    """Calculate demand-error metrics for arbitrary groupings."""

    if dataframe.empty:
        return pd.DataFrame()

    valid = dataframe[
        dataframe[forecast_col].notna()
        & dataframe[actual_col].notna()
    ].copy()

    rows = []

    for group_key, group in valid.groupby(
        group_columns,
        dropna=False,
    ):
        if not isinstance(group_key, tuple):
            group_key = (group_key,)

        identifiers = dict(zip(group_columns, group_key))

        actual = group[actual_col].astype(float)
        forecast = group[forecast_col].astype(float)
        error = forecast - actual
        denominator = actual.abs().sum()

        rows.append(
            {
                **identifiers,
                "observations": len(group),
                "actual_total": actual.sum(),
                "forecast_total": forecast.sum(),
                "actual_mean": actual.mean(),
                "forecast_mean": forecast.mean(),
                "mae": error.abs().mean(),
                "rmse": np.sqrt(error.pow(2).mean()),
                "bias": error.mean(),
                "wape_pct": (
                    error.abs().sum() / denominator * 100
                    if denominator > 0
                    else np.nan
                ),
                "underforecast_days": int((error < 0).sum()),
                "overforecast_days": int((error > 0).sum()),
                "maximum_underforecast": error.min(),
                "maximum_overforecast": error.max(),
            }
        )

    return pd.DataFrame(rows)


# =============================================================================
# 4. ALL-DATE TARGET GENERATION
# =============================================================================

def create_all_test_dates(config):
    """Return every calendar date in the simulation range."""

    dates = pd.date_range(
        start=config["simulation_start"],
        end=config["simulation_end"],
        freq="D",
    )

    if config.get("smoke_test", False):
        dates = dates[: config["smoke_test_target_count"]]

    return pd.DataFrame({"target_date": dates.normalize()})


# =============================================================================
# 5. TREND-ADJUSTED HISTORICAL BASELINES
# =============================================================================

def calculate_level_trend_factor(
    daily_actuals,
    cutoff_timestamp,
    target_col,
    recent_days,
    comparison_days,
    minimum_factor,
    maximum_factor,
):
    """
    Estimate a purely historical level adjustment as of the forecast cutoff.

    recent level:
        mean demand during the most recent `recent_days`

    comparison level:
        mean demand during the immediately preceding `comparison_days`

    trend factor:
        recent level / comparison level

    The factor is clipped to reduce instability from short-lived shocks.
    """

    cutoff_date = pd.Timestamp(cutoff_timestamp).normalize()

    history = daily_actuals.copy()
    history["date"] = pd.to_datetime(
        history["date"],
        errors="coerce",
    ).dt.normalize()

    history = history[
        history["date"].lt(cutoff_date)
    ].sort_values("date")

    recent_start = cutoff_date - pd.Timedelta(days=recent_days)
    comparison_start = recent_start - pd.Timedelta(days=comparison_days)

    recent = history[
        history["date"].ge(recent_start)
    ][target_col]

    comparison = history[
        history["date"].ge(comparison_start)
        & history["date"].lt(recent_start)
    ][target_col]

    if len(recent) < max(7, recent_days // 2):
        return 1.0

    if len(comparison) < max(14, comparison_days // 2):
        return 1.0

    recent_mean = safe_mean(recent)
    comparison_mean = safe_mean(comparison)

    if (
        pd.isna(recent_mean)
        or pd.isna(comparison_mean)
        or comparison_mean <= 0
    ):
        return 1.0

    factor = recent_mean / comparison_mean

    return float(
        np.clip(
            factor,
            minimum_factor,
            maximum_factor,
        )
    )


def forecast_trend_adjusted_seasonal(
    seasonal_forecast,
    trend_factor,
):
    """Apply the historical demand-level trend factor to a seasonal forecast."""

    if pd.isna(seasonal_forecast) or pd.isna(trend_factor):
        return np.nan

    return clip_forecast(
        seasonal_forecast * trend_factor
    )


# =============================================================================
# 6. BOOKING FORECAST SHRINKAGE
# =============================================================================

def shrink_booking_forecast(
    booking_forecast,
    historical_forecast,
    known_bookings,
    shrinkage_strength,
):
    """
    Shrink an unstable booking-curve forecast toward a historical baseline.

    Reliability weight:
        known bookings / (known bookings + shrinkage strength)

    Therefore low-visibility forecasts receive more historical support, while
    high-visibility forecasts remain close to the booking curve.
    """

    if pd.isna(booking_forecast):
        return historical_forecast

    if pd.isna(historical_forecast):
        return booking_forecast

    known_bookings = max(0.0, float(known_bookings))
    shrinkage_strength = max(0.0, float(shrinkage_strength))

    if shrinkage_strength == 0:
        return clip_forecast(booking_forecast)

    reliability = (
        known_bookings
        / (known_bookings + shrinkage_strength)
    )

    return clip_forecast(
        reliability * booking_forecast
        + (1.0 - reliability) * historical_forecast
    )


# =============================================================================
# 7. PREPARE GENERATION 2 MODEL DATA
# =============================================================================

def prepare_refined_model_data(
    base,
    raw_data,
    base_config,
    refined_config,
    test_dates,
    horizons,
):
    """
    Prepare Generation 1 datasets and replace sampled-date caches with
    all-date caches for the refined simulation.
    """

    model_data = base.prepare_forecast_model_data(
        data=raw_data,
        config=base_config,
    )

    # Rebuild caches explicitly from the exact all-date target table used by
    # this script. This makes the refined file independent of Generation 1's
    # sampled target-date logic.
    model_data["known_booking_cache"] = (
        base.precompute_known_booking_counts(
            bookings=model_data["bookings"],
            test_dates=test_dates,
            horizons=horizons,
        )
    )

    model_data["curve_factor_cache"] = (
        base.precompute_booking_curve_factors(
            entry_curve=model_data["entry_booking_curve"],
            exit_curve=model_data["exit_booking_curve"],
            test_dates=test_dates,
            horizons=horizons,
        )
    )

    model_data["weekday_cache"] = (
        base.precompute_same_weekday_history(
            daily_actuals=model_data["daily_actuals"],
            test_dates=test_dates,
            horizons=horizons,
            max_n=max(
                12,
                refined_config["same_weekday_n"],
            ),
        )
    )

    model_data["weekday_month_cache"] = (
        base.precompute_weekday_month_history(
            daily_actuals=model_data["daily_actuals"],
            test_dates=test_dates,
            horizons=horizons,
            max_n=refined_config["weekday_month_max_n"],
        )
    )

    return model_data


# =============================================================================
# 8. BUILD ONE REFINED COMPONENT SNAPSHOT
# =============================================================================

def get_cached_weekday_month_forecast(
    data,
    target_date,
    horizon_days,
    demand_type,
    fallback,
):
    """Retrieve as-of-safe weekday-month forecast with fallback."""

    cached = data.get("weekday_month_cache", {}).get(
        (
            pd.Timestamp(target_date).normalize(),
            int(horizon_days),
        )
    )

    if cached is None:
        return fallback, 0

    count = int(cached.get("count", 0))

    minimum = data["config"][
        "minimum_weekday_month_observations"
    ]

    if count < minimum:
        return fallback, 0

    key = "entries" if demand_type == "entry" else "exits"

    return safe_mean(cached.get(key, [])), count


def build_refined_snapshot(
    base,
    target_date,
    horizon_days,
    data,
    refined_config,
):
    """Calculate all independent Generation 2 components for one snapshot."""

    target_date = pd.Timestamp(target_date).normalize()

    cutoff_timestamp = (
        target_date
        - pd.Timedelta(days=horizon_days)
        + pd.Timedelta(hours=7)
    )

    daily_actuals = data["daily_actuals"]

    entry_weekday = base.forecast_same_weekday(
        daily_actuals=daily_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col="entries",
        n=refined_config["same_weekday_n"],
        data=data,
    )

    exit_weekday = base.forecast_same_weekday(
        daily_actuals=daily_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col="exits",
        n=refined_config["same_weekday_n"],
        data=data,
    )

    entry_month, entry_month_n = get_cached_weekday_month_forecast(
        data=data,
        target_date=target_date,
        horizon_days=horizon_days,
        demand_type="entry",
        fallback=entry_weekday,
    )

    exit_month, exit_month_n = get_cached_weekday_month_forecast(
        data=data,
        target_date=target_date,
        horizon_days=horizon_days,
        demand_type="exit",
        fallback=exit_weekday,
    )

    entry_booking = base.forecast_booking_curve(
        bookings=data["bookings"],
        curve=data["entry_booking_curve"],
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="entry",
        data=data,
    )

    exit_booking = base.forecast_booking_curve(
        bookings=data["bookings"],
        curve=data["exit_booking_curve"],
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="exit",
        data=data,
    )

    known_booking_cache = data["known_booking_cache"]

    entry_known = known_booking_cache.get(
        (target_date, int(horizon_days), "entry"),
        0,
    )

    exit_known = known_booking_cache.get(
        (target_date, int(horizon_days), "exit"),
        0,
    )

    # Duration factors and duration-adjusted exit forecast are kept because
    # Generation 1 showed genuine incremental exit value from this component.
    duration_factors = base.get_duration_exit_factors_as_of(
        duration_curve=data["duration_exit_booking_curve"],
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        lead_days=horizon_days,
        minimum_observations=10,
        maximum_observations=20,
    )

    exit_duration = base.get_duration_adjusted_booking_curve(
        bookings=data["bookings"],
        curve=data["exit_booking_curve"],
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="exit",
        historical_duration_factors=duration_factors,
    )

    entry_trend_factor = calculate_level_trend_factor(
        daily_actuals=daily_actuals,
        cutoff_timestamp=cutoff_timestamp,
        target_col="entries",
        recent_days=refined_config["trend_recent_days"],
        comparison_days=refined_config["trend_comparison_days"],
        minimum_factor=refined_config["trend_factor_min"],
        maximum_factor=refined_config["trend_factor_max"],
    )

    exit_trend_factor = calculate_level_trend_factor(
        daily_actuals=daily_actuals,
        cutoff_timestamp=cutoff_timestamp,
        target_col="exits",
        recent_days=refined_config["trend_recent_days"],
        comparison_days=refined_config["trend_comparison_days"],
        minimum_factor=refined_config["trend_factor_min"],
        maximum_factor=refined_config["trend_factor_max"],
    )

    entry_trend_month = forecast_trend_adjusted_seasonal(
        seasonal_forecast=entry_month,
        trend_factor=entry_trend_factor,
    )

    exit_trend_month = forecast_trend_adjusted_seasonal(
        seasonal_forecast=exit_month,
        trend_factor=exit_trend_factor,
    )

    return {
        "target_date": target_date,
        "cutoff_timestamp": cutoff_timestamp,
        "horizon_days": int(horizon_days),
        "entry_booking": entry_booking,
        "entry_weekday": entry_weekday,
        "entry_month": entry_month,
        "entry_trend_month": entry_trend_month,
        "entry_known_bookings": entry_known,
        "entry_weekday_month_observations": entry_month_n,
        "entry_trend_factor": entry_trend_factor,
        "exit_booking": exit_booking,
        "exit_duration": exit_duration,
        "exit_weekday": exit_weekday,
        "exit_month": exit_month,
        "exit_trend_month": exit_trend_month,
        "exit_known_bookings": exit_known,
        "exit_weekday_month_observations": exit_month_n,
        "exit_trend_factor": exit_trend_factor,
    }


# =============================================================================
# 9. BUILD THE COMPLETE COMPONENT TABLE
# =============================================================================

def build_component_table(
    base,
    data,
    test_dates,
    horizons,
    refined_config,
):
    """Build one row per target date and horizon."""

    rows = []
    total = len(test_dates) * len(horizons)
    counter = 0

    print("\n" + "=" * 80)
    print("BUILDING REFINED COMPONENT TABLE")
    print("=" * 80)

    for target_date in test_dates["target_date"]:
        for horizon_days in horizons:
            counter += 1

            print(
                f"\rSnapshot {counter:,}/{total:,} "
                f"| {pd.Timestamp(target_date).date()} "
                f"| T-{horizon_days}",
                end="",
            )

            try:
                rows.append(
                    build_refined_snapshot(
                        base=base,
                        target_date=target_date,
                        horizon_days=horizon_days,
                        data=data,
                        refined_config=refined_config,
                    )
                )
            except Exception as exc:
                if refined_config["fail_fast"]:
                    raise RuntimeError(
                        "Refined component snapshot failed: "
                        f"target={pd.Timestamp(target_date).date()}, "
                        f"horizon=T-{horizon_days}"
                    ) from exc

                rows.append(
                    {
                        "target_date": pd.Timestamp(target_date).normalize(),
                        "horizon_days": int(horizon_days),
                        "error_message": str(exc),
                    }
                )

    print()

    components = pd.DataFrame(rows)

    actuals = data["daily_actuals"][
        ["date", "entries", "exits"]
    ].copy()

    actuals["date"] = pd.to_datetime(
        actuals["date"],
        errors="coerce",
    ).dt.normalize()

    actuals = actuals.rename(
        columns={
            "date": "target_date",
            "entries": "actual_entries",
            "exits": "actual_exits",
        }
    )

    components = components.merge(
        actuals,
        on="target_date",
        how="left",
    )

    components["weekday"] = components[
        "target_date"
    ].dt.day_name()

    components["weekday_num"] = components[
        "target_date"
    ].dt.weekday

    components["month"] = components[
        "target_date"
    ].dt.month

    components["year"] = components[
        "target_date"
    ].dt.year

    components["month_start"] = components[
        "target_date"
    ].dt.to_period("M").dt.to_timestamp()

    return components


# =============================================================================
# 10. GENERATE SIMPLEX WEIGHT GRID
# =============================================================================

def generate_weight_grid(component_names, step):
    """
    Generate non-negative weights that sum to one.

    Example for step 0.05:
        0.00, 0.05, ..., 1.00
    """

    units = int(round(1.0 / step))
    component_count = len(component_names)
    rows = []

    def compositions(total, parts, prefix=()):
        if parts == 1:
            yield prefix + (total,)
            return

        for value in range(total + 1):
            yield from compositions(
                total - value,
                parts - 1,
                prefix + (value,),
            )

    for allocation in compositions(
        units,
        component_count,
    ):
        weights = np.asarray(allocation, dtype=float) / units
        rows.append(dict(zip(component_names, weights)))

    return pd.DataFrame(rows)


# =============================================================================
# 11. ROLLING-ORIGIN FOLDS
# =============================================================================

def create_rolling_month_folds(
    component_table,
    minimum_training_months,
    validation_months,
    maximum_training_months=None,
):
    """Create expanding or recency-limited monthly rolling-origin folds."""

    months = sorted(
        pd.to_datetime(
            component_table["month_start"]
        ).dropna().unique()
    )

    folds = []

    first_test_index = (
        minimum_training_months
        + validation_months
    )

    for test_index in range(first_test_index, len(months)):
        test_month = pd.Timestamp(months[test_index])

        validation_start_index = (
            test_index - validation_months
        )

        validation_month_values = months[
            validation_start_index:test_index
        ]

        training_end_index = validation_start_index
        training_start_index = 0

        if maximum_training_months is not None:
            training_start_index = max(
                0,
                training_end_index - maximum_training_months,
            )

        training_month_values = months[
            training_start_index:training_end_index
        ]

        if len(training_month_values) < minimum_training_months:
            continue

        folds.append(
            {
                "fold_id": len(folds) + 1,
                "training_months": [
                    pd.Timestamp(x)
                    for x in training_month_values
                ],
                "validation_months": [
                    pd.Timestamp(x)
                    for x in validation_month_values
                ],
                "test_month": test_month,
                "training_start": pd.Timestamp(
                    training_month_values[0]
                ),
                "training_end": pd.Timestamp(
                    training_month_values[-1]
                ),
                "validation_start": pd.Timestamp(
                    validation_month_values[0]
                ),
                "validation_end": pd.Timestamp(
                    validation_month_values[-1]
                ),
            }
        )

    return folds


# =============================================================================
# 12. COMPONENT SPECIFICATIONS
# =============================================================================

def get_component_specifications():
    """Return the independent ensemble components for ENTRY and EXIT."""

    return {
        "entry": {
            "actual_col": "actual_entries",
            "known_col": "entry_known_bookings",
            "booking_col": "entry_booking",
            "historical_col": "entry_trend_month",
            "components": [
                "entry_booking_shrunk",
                "entry_weekday",
                "entry_month",
                "entry_trend_month",
            ],
        },
        "exit": {
            "actual_col": "actual_exits",
            "known_col": "exit_known_bookings",
            "booking_col": "exit_booking",
            "historical_col": "exit_trend_month",
            "components": [
                "exit_booking_shrunk",
                "exit_duration",
                "exit_weekday",
                "exit_trend_month",
            ],
        },
    }


# =============================================================================
# 13. APPLY BOOKING SHRINKAGE CANDIDATE
# =============================================================================

def add_shrunk_booking_component(
    frame,
    demand_type,
    shrinkage_strength,
):
    """Add the configured shrunk booking component to a copy of a frame."""

    output = frame.copy()

    if demand_type == "entry":
        output["entry_booking_shrunk"] = [
            shrink_booking_forecast(
                booking_forecast=booking,
                historical_forecast=historical,
                known_bookings=known,
                shrinkage_strength=shrinkage_strength,
            )
            for booking, historical, known in zip(
                output["entry_booking"],
                output["entry_trend_month"],
                output["entry_known_bookings"],
            )
        ]
    else:
        output["exit_booking_shrunk"] = [
            shrink_booking_forecast(
                booking_forecast=booking,
                historical_forecast=historical,
                known_bookings=known,
                shrinkage_strength=shrinkage_strength,
            )
            for booking, historical, known in zip(
                output["exit_booking"],
                output["exit_trend_month"],
                output["exit_known_bookings"],
            )
        ]

    return output


# =============================================================================
# 14. FORECAST FROM A WEIGHT SET
# =============================================================================

def forecast_from_weights(frame, component_names, weight_row):
    """Calculate ensemble forecasts with missing-component re-normalisation."""

    weight_values = np.asarray(
        [
            float(weight_row[name])
            for name in component_names
        ],
        dtype=float,
    )

    component_matrix = frame[
        component_names
    ].to_numpy(dtype=float)

    forecasts = []

    for component_values in component_matrix:
        forecasts.append(
            weighted_average(
                values=component_values,
                weights=weight_values,
            )
        )

    return pd.Series(
        forecasts,
        index=frame.index,
        dtype=float,
    )


# =============================================================================
# 15. SELECT WEIGHTS FOR ONE FLOW / HORIZON / FOLD
# =============================================================================

def select_best_weights(
    training_frame,
    validation_frame,
    demand_type,
    component_names,
    actual_col,
    weight_grid,
    shrinkage_strengths,
    prior_weights=None,
    stability_penalty=0.0,
):
    """
    Select shrinkage and ensemble weights using validation WAPE.

    Training data is used for diagnostics and availability checks. Weight
    choice is based only on the validation period immediately before the test
    month. The test month is never used here.
    """

    if validation_frame.empty:
        return None, pd.DataFrame()

    results = []

    for shrinkage_strength in shrinkage_strengths:
        validation_candidate = add_shrunk_booking_component(
            frame=validation_frame,
            demand_type=demand_type,
            shrinkage_strength=shrinkage_strength,
        )

        training_candidate = add_shrunk_booking_component(
            frame=training_frame,
            demand_type=demand_type,
            shrinkage_strength=shrinkage_strength,
        )

        for weight_row in weight_grid.to_dict("records"):
            validation_forecast = forecast_from_weights(
                frame=validation_candidate,
                component_names=component_names,
                weight_row=weight_row,
            )

            validation_wape = wape(
                validation_candidate[actual_col],
                validation_forecast,
            )

            training_forecast = forecast_from_weights(
                frame=training_candidate,
                component_names=component_names,
                weight_row=weight_row,
            )

            training_wape = wape(
                training_candidate[actual_col],
                training_forecast,
            )

            distance_from_prior = 0.0

            if prior_weights is not None:
                distance_from_prior = sum(
                    abs(
                        float(weight_row[name])
                        - float(prior_weights.get(name, 0.0))
                    )
                    for name in component_names
                )

            selection_score = (
                validation_wape
                + stability_penalty * distance_from_prior
                if pd.notna(validation_wape)
                else np.nan
            )

            results.append(
                {
                    "demand_type": demand_type,
                    "shrinkage_strength": shrinkage_strength,
                    **weight_row,
                    "training_observations": len(training_candidate),
                    "validation_observations": len(validation_candidate),
                    "training_wape_pct": (
                        training_wape * 100
                        if pd.notna(training_wape)
                        else np.nan
                    ),
                    "validation_wape_pct": (
                        validation_wape * 100
                        if pd.notna(validation_wape)
                        else np.nan
                    ),
                    "distance_from_prior": distance_from_prior,
                    "selection_score": selection_score,
                }
            )

    results_df = pd.DataFrame(results)

    valid_results = results_df[
        results_df["selection_score"].notna()
    ].copy()

    if valid_results.empty:
        return None, results_df

    sort_columns = [
        "selection_score",
        "validation_wape_pct",
        "training_wape_pct",
        "shrinkage_strength",
    ] + component_names

    best = (
        valid_results
        .sort_values(sort_columns)
        .iloc[0]
        .to_dict()
    )

    return best, results_df


# =============================================================================
# 16. RUN THE ROLLING-ORIGIN ENSEMBLE BACKTEST
# =============================================================================

def run_rolling_ensemble_backtest(
    component_table,
    refined_config,
):
    """Select weights on prior data and score each untouched test month."""

    folds = create_rolling_month_folds(
        component_table=component_table,
        minimum_training_months=refined_config[
            "minimum_training_months"
        ],
        validation_months=refined_config[
            "validation_months"
        ],
        maximum_training_months=refined_config[
            "maximum_weight_training_months"
        ],
    )

    specifications = get_component_specifications()
    horizons = sorted(component_table["horizon_days"].unique())

    grids = {
        demand_type: generate_weight_grid(
            component_names=specification["components"],
            step=refined_config["weight_step"],
        )
        for demand_type, specification in specifications.items()
    }

    prediction_frames = []
    selected_weight_rows = []
    all_weight_frames = []

    prior_weight_lookup = {}

    print("\n" + "=" * 80)
    print("RUNNING ROLLING-ORIGIN WEIGHT SELECTION")
    print("=" * 80)

    total = len(folds) * len(horizons) * len(specifications)
    counter = 0

    for fold in folds:
        for demand_type, specification in specifications.items():
            for horizon_days in horizons:
                counter += 1

                print(
                    f"\rCalibration {counter:,}/{total:,} "
                    f"| Test {fold['test_month']:%Y-%m} "
                    f"| {demand_type.upper()} "
                    f"| T-{horizon_days}",
                    end="",
                )

                horizon_frame = component_table[
                    component_table["horizon_days"].eq(horizon_days)
                ].copy()

                training = horizon_frame[
                    horizon_frame["month_start"].isin(
                        fold["training_months"]
                    )
                ].copy()

                validation = horizon_frame[
                    horizon_frame["month_start"].isin(
                        fold["validation_months"]
                    )
                ].copy()

                test = horizon_frame[
                    horizon_frame["month_start"].eq(
                        fold["test_month"]
                    )
                ].copy()

                if training.empty or validation.empty or test.empty:
                    continue

                prior_key = (demand_type, int(horizon_days))
                prior_weights = prior_weight_lookup.get(prior_key)

                best, all_weights = select_best_weights(
                    training_frame=training,
                    validation_frame=validation,
                    demand_type=demand_type,
                    component_names=specification["components"],
                    actual_col=specification["actual_col"],
                    weight_grid=grids[demand_type],
                    shrinkage_strengths=refined_config[
                        "booking_shrinkage_strengths"
                    ],
                    prior_weights=prior_weights,
                    stability_penalty=refined_config[
                        "weight_stability_penalty"
                    ],
                )

                if best is None:
                    continue

                prior_weight_lookup[prior_key] = {
                    name: best[name]
                    for name in specification["components"]
                }

                all_weights["fold_id"] = fold["fold_id"]
                all_weights["test_month"] = fold["test_month"]
                all_weights["horizon_days"] = horizon_days
                all_weight_frames.append(all_weights)

                print(
                    f"\n{demand_type.upper()} "
                    f"T-{horizon_days} "
                    f"| Validation WAPE: "
                    f"{best['validation_wape_pct']:.2f}% "
                    f"| Shrinkage: {best['shrinkage_strength']}"
                )

                for component in specification["components"]:
                    print(
                        f"    {component:<25}"
                        f"{best[component]:>6.1%}"
                    )
                
  

                selected_weight_rows.append(
                    {
                        "fold_id": fold["fold_id"],
                        "demand_type": demand_type,
                        "horizon_days": horizon_days,
                        "horizon_label": f"T-{horizon_days}",
                        "training_start": fold["training_start"],
                        "training_end": fold["training_end"],
                        "validation_start": fold["validation_start"],
                        "validation_end": fold["validation_end"],
                        "test_month": fold["test_month"],
                        "shrinkage_strength": best[
                            "shrinkage_strength"
                        ],
                        **{
                            name: best[name]
                            for name in specification["components"]
                        },
                        "training_wape_pct": best[
                            "training_wape_pct"
                        ],
                        "validation_wape_pct": best[
                            "validation_wape_pct"
                        ],
                    }
                )

                test_candidate = add_shrunk_booking_component(
                    frame=test,
                    demand_type=demand_type,
                    shrinkage_strength=best[
                        "shrinkage_strength"
                    ],
                )

                test_candidate["forecast_value"] = forecast_from_weights(
                    frame=test_candidate,
                    component_names=specification["components"],
                    weight_row=best,
                )

                test_candidate["actual_value"] = test_candidate[
                    specification["actual_col"]
                ]

                test_candidate["demand_type"] = demand_type
                test_candidate["fold_id"] = fold["fold_id"]
                test_candidate["test_month"] = fold["test_month"]
                test_candidate["shrinkage_strength"] = best[
                    "shrinkage_strength"
                ]

                for name in specification["components"]:
                    test_candidate[f"weight_{name}"] = best[name]

                test_candidate["error"] = (
                    test_candidate["forecast_value"]
                    - test_candidate["actual_value"]
                )

                test_candidate["absolute_error"] = (
                    test_candidate["error"].abs()
                )

                test_candidate["squared_error"] = (
                    test_candidate["error"].pow(2)
                )

                prediction_frames.append(test_candidate)

    print()

    predictions = (
        pd.concat(prediction_frames, ignore_index=True)
        if prediction_frames
        else pd.DataFrame()
    )

    selected_weights = pd.DataFrame(selected_weight_rows)

    all_weights = (
        pd.concat(all_weight_frames, ignore_index=True)
        if all_weight_frames
        else pd.DataFrame()
    )

    return predictions, selected_weights, all_weights, pd.DataFrame(folds)


# =============================================================================
# 17. BASELINE COMPARISON
# =============================================================================

def build_baseline_predictions(component_table, predictions):
    """
    Produce comparable raw-component forecasts for exactly the test rows used
    by the rolling ensemble.
    """

    if predictions.empty:
        return pd.DataFrame()

    keys = predictions[
        [
            "target_date",
            "horizon_days",
            "demand_type",
            "fold_id",
            "test_month",
        ]
    ].drop_duplicates()

    merged = keys.merge(
        component_table,
        on=["target_date", "horizon_days"],
        how="left",
    )

    rows = []

    baseline_map = {
        "entry": {
            "actual": "actual_entries",
            "forecasts": {
                "ENTRY_BOOKING": "entry_booking",
                "ENTRY_WEEKDAY": "entry_weekday",
                "ENTRY_MONTH": "entry_month",
                "ENTRY_TREND_MONTH": "entry_trend_month",
            },
        },
        "exit": {
            "actual": "actual_exits",
            "forecasts": {
                "EXIT_BOOKING": "exit_booking",
                "EXIT_DURATION": "exit_duration",
                "EXIT_WEEKDAY": "exit_weekday",
                "EXIT_TREND_MONTH": "exit_trend_month",
            },
        },
    }

    for demand_type, details in baseline_map.items():
        flow = merged[
            merged["demand_type"].eq(demand_type)
        ].copy()

        for model_name, forecast_col in details["forecasts"].items():
            output = flow[
                [
                    "target_date",
                    "horizon_days",
                    "fold_id",
                    "test_month",
                    forecast_col,
                    details["actual"],
                ]
            ].copy()

            output = output.rename(
                columns={
                    forecast_col: "forecast_value",
                    details["actual"]: "actual_value",
                }
            )

            output["demand_type"] = demand_type
            output["model_name"] = model_name
            rows.append(output)

    ensemble = predictions[
        [
            "target_date",
            "horizon_days",
            "fold_id",
            "test_month",
            "demand_type",
            "forecast_value",
            "actual_value",
        ]
    ].copy()

    ensemble["model_name"] = "REFINED_ENSEMBLE"
    rows.append(ensemble)

    return pd.concat(rows, ignore_index=True)


# =============================================================================
# 18. FINAL CURRENT WEIGHTS
# =============================================================================

def create_recommended_weight_table(selected_weights):
    """
    Use the most recent rolling fold's selected weight set as the current
    recommended configuration for each flow and horizon.
    """

    if selected_weights.empty:
        return pd.DataFrame()

    return (
        selected_weights
        .sort_values(
            [
                "demand_type",
                "horizon_days",
                "test_month",
            ]
        )
        .groupby(
            ["demand_type", "horizon_days"],
            as_index=False,
        )
        .tail(1)
        .sort_values(
            ["demand_type", "horizon_days"]
        )
        .reset_index(drop=True)
    )


# =============================================================================
# 19. DIAGNOSTIC OUTPUTS
# =============================================================================

def create_weight_stability_summary(selected_weights):
    """Summarise variation in selected weights across rolling folds."""

    if selected_weights.empty:
        return pd.DataFrame()

    weight_columns = [
        column
        for column in selected_weights.columns
        if column.startswith("entry_")
        or column.startswith("exit_")
    ]

    rows = []

    for (demand_type, horizon_days), group in selected_weights.groupby(
        ["demand_type", "horizon_days"]
    ):
        row = {
            "demand_type": demand_type,
            "horizon_days": horizon_days,
            "folds": len(group),
            "average_validation_wape_pct": group[
                "validation_wape_pct"
            ].mean(),
            "average_shrinkage_strength": group[
                "shrinkage_strength"
            ].mean(),
        }

        for column in weight_columns:
            if column not in group.columns:
                continue

            values = group[column].dropna()

            if values.empty:
                continue

            row[f"average_{column}"] = values.mean()
            row[f"minimum_{column}"] = values.min()
            row[f"maximum_{column}"] = values.max()

        rows.append(row)

    return pd.DataFrame(rows)


def create_worst_days(predictions, top_n=100):
    """Return the largest refined out-of-sample demand errors."""

    if predictions.empty:
        return pd.DataFrame()

    return (
        predictions
        .sort_values(
            "absolute_error",
            ascending=False,
        )
        .head(top_n)
        .reset_index(drop=True)
    )


# =============================================================================
# 20. EXCEL FORMATTING / EXPORT
# =============================================================================

def format_workbook(workbook):
    """Apply functional formatting to all output sheets."""

    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

            column_letter = get_column_letter(
                column_cells[0].column
            )

            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                60,
            )


def export_refined_results(
    output_path,
    component_table,
    predictions,
    selected_weights,
    all_weight_tests,
    folds,
    performance,
    baseline_performance,
    recommended_weights,
    weight_stability,
    worst_days,
):
    """Export all Generation 2 demand-forecast results."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        performance.to_excel(
            writer,
            sheet_name="Refined Performance",
            index=False,
        )

        baseline_performance.to_excel(
            writer,
            sheet_name="Baseline Comparison",
            index=False,
        )

        recommended_weights.to_excel(
            writer,
            sheet_name="Recommended Weights",
            index=False,
        )

        selected_weights.to_excel(
            writer,
            sheet_name="Selected Weights",
            index=False,
        )

        weight_stability.to_excel(
            writer,
            sheet_name="Weight Stability",
            index=False,
        )

        folds.to_excel(
            writer,
            sheet_name="Rolling Folds",
            index=False,
        )

        predictions.to_excel(
            writer,
            sheet_name="Out of Sample Forecasts",
            index=False,
        )

        component_table.to_excel(
            writer,
            sheet_name="Component Table",
            index=False,
        )

        worst_days.to_excel(
            writer,
            sheet_name="Worst Forecast Days",
            index=False,
        )

        # This sheet can be large but is important for auditability.
        all_weight_tests.to_excel(
            writer,
            sheet_name="All Weight Tests",
            index=False,
        )

        format_workbook(writer.book)

    return output_path


# =============================================================================
# 21. CONSOLE SUMMARY
# =============================================================================

def print_final_summary(performance, baseline_performance, output_path):
    """Print the main out-of-sample demand results."""

    print("\n" + "=" * 90)
    print("FASTPARK REFINED FORECAST - OUT-OF-SAMPLE RESULTS")
    print("=" * 90)

    if performance.empty:
        print("No valid refined forecasts were generated.")
        return

    print("\nRefined ensemble performance by flow and horizon:\n")
    print(
        performance[
            [
                "demand_type",
                "horizon_days",
                "observations",
                "wape_pct",
                "mae",
                "rmse",
                "bias",
            ]
        ].to_string(index=False)
    )

    if not baseline_performance.empty:
        print("\nModel comparison, averaged over all tested horizons:\n")

        overall = score_forecast_frame(
            dataframe=baseline_performance.rename(
                columns={
                    "wape_pct": "forecast_value",
                    "mae": "actual_value",
                }
            ),
            group_columns=["model_name"],
        )

        # The detailed comparison is in Excel. Avoid printing a misleading
        # aggregation of already-aggregated metrics to the console.
        print(
            baseline_performance[
                [
                    "demand_type",
                    "model_name",
                    "horizon_days",
                    "wape_pct",
                    "bias",
                ]
            ]
            .sort_values(
                [
                    "demand_type",
                    "horizon_days",
                    "wape_pct",
                ]
            )
            .head(40)
            .to_string(index=False)
        )

    print(f"\nExported refined workbook: {output_path}")
    print("=" * 90)


# =============================================================================
# 22. MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    t0 = time.perf_counter()

    print("\n" + "=" * 90)
    print("FASTPARK FORECAST SIMULATION - REFINED / GENERATION 2")
    print("ALL-DATE ROLLING-ORIGIN DEMAND BACKTEST")
    print("=" * 90)

    refined_config = get_refined_config()
    base = import_generation_1(refined_config)
    base_config = build_generation_1_config(
        base=base,
        refined_config=refined_config,
    )

    test_dates = create_all_test_dates(base_config)

    horizons = list(
        refined_config["forecast_horizons_days"]
    )

    if refined_config["smoke_test"]:
        horizons = list(
            refined_config["smoke_test_horizons"]
        )

    print(f"Target dates: {len(test_dates):,}")
    print(f"Horizons: {len(horizons):,}")
    print(
        "Target range: "
        f"{test_dates['target_date'].min().date()} to "
        f"{test_dates['target_date'].max().date()}"
    )

    print("\nLoading and cleaning source data...")
    raw_data = base.load_simulation_data(base_config)
    t1 = step(t0, "Loaded source data")

    base.validate_loaded_data(
        data=raw_data,
        config=base_config,
    )

    print("\nPreparing all-date model data and caches...")
    model_data = prepare_refined_model_data(
        base=base,
        raw_data=raw_data,
        base_config=base_config,
        refined_config=refined_config,
        test_dates=test_dates,
        horizons=horizons,
    )
    t2 = step(t1, "Prepared model data and caches")

    component_table = build_component_table(
        base=base,
        data=model_data,
        test_dates=test_dates,
        horizons=horizons,
        refined_config=refined_config,
    )

    if component_table.empty:
        raise RuntimeError(
            "The refined component table is empty."
        )

    t3 = step(t2, f"Built component table ({len(component_table):,} rows)")

    predictions, selected_weights, all_weight_tests, folds = (
        run_rolling_ensemble_backtest(
            component_table=component_table,
            refined_config=refined_config,
        )
    )

    if predictions.empty:
        raise RuntimeError(
            "The rolling-origin backtest generated no predictions. "
            "Review the configured date range and minimum training months."
        )

    predictions.to_csv(
        "refined_predictions_backup.csv",
        index=False
    )

    selected_weights.to_csv(
        "refined_selected_weights_backup.csv",
        index=False
    )

    all_weight_tests.to_csv(
        "refined_all_weight_tests_backup.csv",
        index=False
    )

    t4 = step(t3, f"Completed rolling-origin optimisation ({len(predictions):,} forecasts)")

    performance = score_forecast_frame(
        dataframe=predictions,
        group_columns=[
            "demand_type",
            "horizon_days",
        ],
    ).sort_values(
        ["demand_type", "horizon_days"]
    ).reset_index(drop=True)

    baseline_predictions = build_baseline_predictions(
        component_table=component_table,
        predictions=predictions,
    )

    baseline_performance = score_forecast_frame(
        dataframe=baseline_predictions,
        group_columns=[
            "demand_type",
            "model_name",
            "horizon_days",
        ],
    ).sort_values(
        [
            "demand_type",
            "horizon_days",
            "wape_pct",
        ]
    ).reset_index(drop=True)

    recommended_weights = create_recommended_weight_table(
        selected_weights
    )

    weight_stability = create_weight_stability_summary(
        selected_weights
    )

    worst_days = create_worst_days(
        predictions,
        top_n=100,
    )

    output_path = export_refined_results(
        output_path=refined_config["output_path"],
        component_table=component_table,
        predictions=predictions,
        selected_weights=selected_weights,
        all_weight_tests=all_weight_tests,
        folds=folds,
        performance=performance,
        baseline_performance=baseline_performance,
        recommended_weights=recommended_weights,
        weight_stability=weight_stability,
        worst_days=worst_days,
    )

    t5 = step(t4, "Workbook exported")

    print_final_summary(
        performance=performance,
        baseline_performance=baseline_performance,
        output_path=output_path,
    )

    print(
        "Total runtime: "
        f"{time.perf_counter() - t0:.2f} seconds"
    )
