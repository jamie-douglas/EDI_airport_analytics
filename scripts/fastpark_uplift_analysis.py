"""
FastPark booking uplift analysis.

Purpose
-------
This script analyses historical FastPark bookings and actual movements to
calculate recommended booking uplifts at horizons D0 to D28.

The script produces:

1. Overall entry uplift by horizon.
2. Overall direct-exit uplift by horizon.
3. Overall implied-exit uplift by entry horizon.
4. Monthly entry, exit and implied-exit uplifts.
5. Month-by-month comparison tables.
6. Entry and exit lead-time distributions.
7. Duration-specific entry visibility and uplift.
8. Daily underlying calculation detail.
9. Entry and exit booking-versus-actual reconciliation.
10. Copy-ready D1-D14 recommended uplift arrays.

Important population alignment
------------------------------
Entry analysis:
    Bookings are loaded by entryDate.
    Actual movements are loaded by CheckInStarted.

Direct-exit analysis:
    Bookings are loaded separately by exitDate.
    Actual movements are loaded by ActualCheckedOutDate.

Implied-exit analysis:
    The population is selected by exitDate so that it aligns with actual exits.
    Visibility is then assessed relative to each booking's entry date.

Default period
--------------
If no explicit start and end dates are supplied, the script analyses the last
eight completed calendar months. The current partial month is not included.

For example, when run during August 2026, the default period is:

    2025-12-01 to 2026-07-31

Horizon definition
------------------
The default calculation uses calendar-day cutoffs:

    Creation Date day <= target day - horizon days

Therefore:

    D0 = bookings created on or before the target calendar date
    D1 = bookings created on or before the calendar date one day before target
    D14 = bookings created on or before the calendar date 14 days before target

This preserves the horizon definition used in the original diagnostics script.
"""

import sys
import pathlib
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------
sys.path.append(str(pathlib.Path(__file__).resolve().parents[1]))

from modules.utils.query import query


# ---------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------
OUTPUT_DIR = Path("outputs/fastpark_uplift_analysis")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

RUN_DATE_STR = datetime.now().strftime("%Y-%m-%d")
EXCEL_OUTPUT_PATH = (
    OUTPUT_DIR /
    f"FastPark_Uplift_Analysis_{RUN_DATE_STR}.xlsx"
)

SAVE_SEPARATE_CSVS = False

DEFAULT_HORIZONS = range(0, 29)

UPLIFT_ARRAY_HORIZONS = range(1, 15)

DURATION_ORDER = [
    "1 day",
    "2 days",
    "3-6 days",
    "7-9 days",
    "10-13 days",
    "14-20 days",
    "21-29 days",
    "30-59 days",
    "60-89 days",
    "90+ days",
    "Unknown",
]


# ---------------------------------------------------------------------
# Date-window helpers
# ---------------------------------------------------------------------
def get_last_months_window(
    number_of_months=8,
    include_current_partial_month=True,
):
    """
    Return the analysis window for the requested number of calendar months.

    When include_current_partial_month is True:
    - the current month is included up to yesterday
    - the period contains the current partial month plus the preceding
      completed months
    - the current day is excluded because entry and exit data for today
      may not yet be complete

    Example
    -------
    If run on 2026-08-11 with number_of_months=8:

        start = 2026-01-01
        end   = 2026-08-10

    This gives:
        January 2026 to July 2026 as completed months
        August 2026 as a partial month through 10 August

    When include_current_partial_month is False:
    - only completed months are included

    Example
    -------
    If run on 2026-08-11 with number_of_months=8:

        start = 2025-12-01
        end   = 2026-07-31
    """
    today = pd.Timestamp.today().normalize()
    first_day_current_month = today.replace(day=1)

    if include_current_partial_month:
        analysis_end = today - pd.Timedelta(days=1)

        analysis_start = (
            first_day_current_month -
            pd.DateOffset(months=number_of_months - 1)
        )

    else:
        analysis_end = (
            first_day_current_month -
            pd.Timedelta(days=1)
        )

        analysis_start = (
            first_day_current_month -
            pd.DateOffset(months=number_of_months)
        )

    return (
        analysis_start.strftime("%Y-%m-%d"),
        analysis_end.strftime("%Y-%m-%d"),
    )


# ---------------------------------------------------------------------
# General helpers
# ---------------------------------------------------------------------
def safe_divide(numerator, denominator):
    """
    Divide safely, returning NaN where the denominator is zero.
    """
    numerator = np.asarray(numerator, dtype=float)
    denominator = np.asarray(denominator, dtype=float)

    return np.where(
        denominator > 0,
        numerator / denominator,
        np.nan,
    )


def duration_bucket(value):
    """
    Assign a cleaned duration to a reporting bucket.
    """
    if pd.isna(value):
        return "Unknown"
    if value == 1:
        return "1 day"
    if value == 2:
        return "2 days"
    if 3 <= value <= 6:
        return "3-6 days"
    if 7 <= value <= 9:
        return "7-9 days"
    if 10 <= value <= 13:
        return "10-13 days"
    if 14 <= value <= 20:
        return "14-20 days"
    if 21 <= value <= 29:
        return "21-29 days"
    if 30 <= value <= 59:
        return "30-59 days"
    if 60 <= value <= 89:
        return "60-89 days"
    return "90+ days"


def print_section(title):
    """
    Print a clearly separated console section.
    """
    print("\n" + "=" * 100)
    print(title)
    print("=" * 100)


def horizon_number(series):
    """
    Convert D0, D1, D2, etc. to numeric horizon values.
    """
    return (
        series.astype(str)
        .str.replace("D", "", regex=False)
        .astype(int)
    )


def save_csv(df, filename):
    """
    Optionally save a DataFrame as a CSV.
    """
    if not SAVE_SEPARATE_CSVS:
        return

    output_path = OUTPUT_DIR / filename
    df.to_csv(output_path, index=False)

    print(f"Saved CSV: {output_path}")


# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------
def clean_for_excel(df):
    """
    Prepare a DataFrame for Excel output.
    """
    if df is None:
        return pd.DataFrame()

    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame(df)

    df = df.copy()

    for column in df.columns:
        if isinstance(df[column].dtype, pd.PeriodDtype):
            df[column] = df[column].astype(str)

        if isinstance(df[column].dtype, pd.DatetimeTZDtype):
            df[column] = df[column].dt.tz_localize(None)

    return df.replace([np.inf, -np.inf], np.nan)


def make_sheet_name(name, used_names):
    """
    Create a valid and unique Excel worksheet name.
    """
    invalid_characters = [":", "\\", "/", "?", "*", "[", "]"]

    clean_name = str(name)

    for character in invalid_characters:
        clean_name = clean_name.replace(character, "_")

    clean_name = clean_name[:31]

    original_name = clean_name
    counter = 1

    while clean_name in used_names:
        suffix = f"_{counter}"
        clean_name = original_name[:31 - len(suffix)] + suffix
        counter += 1

    used_names.add(clean_name)

    return clean_name


def autosize_excel_columns(writer, sheet_name, df, max_width=45):
    """
    Auto-size worksheet columns, subject to a maximum width.
    """
    worksheet = writer.sheets[sheet_name]

    if df.empty:
        return

    for column_index, column in enumerate(df.columns, start=1):
        column_name = str(column)

        try:
            maximum_length = max(
                df[column].astype(str).map(len).max(),
                len(column_name),
            )
        except Exception:
            maximum_length = len(column_name)

        adjusted_width = min(maximum_length + 2, max_width)

        excel_column = get_column_letter(column_index)
        worksheet.column_dimensions[excel_column].width = adjusted_width


def format_excel_sheet(writer, sheet_name, df):
    """
    Apply standard formatting to an Excel worksheet.
    """
    worksheet = writer.sheets[sheet_name]

    if df.empty:
        return

    header_fill = "1F4E78"
    header_font = "FFFFFF"

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True,
            color=header_font,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=header_fill,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    percentage_keywords = [
        "pct",
        "percentage",
        "visible",
        "uplift",
        "correction",
        "mape",
        "ape",
        "bias",
    ]

    date_keywords = [
        "date",
        "day",
        "service_day",
        "entry_day",
        "exit_day",
        "start",
        "end",
    ]

    for column_index, column_name in enumerate(df.columns, start=1):
        column_name_lower = str(column_name).lower()
        excel_column = get_column_letter(column_index)

        if any(
            keyword in column_name_lower
            for keyword in percentage_keywords
        ):
            for row_number in range(2, worksheet.max_row + 1):
                worksheet[
                    f"{excel_column}{row_number}"
                ].number_format = "0.00"

        elif any(
            keyword in column_name_lower
            for keyword in date_keywords
        ):
            for row_number in range(2, worksheet.max_row + 1):
                worksheet[
                    f"{excel_column}{row_number}"
                ].number_format = "yyyy-mm-dd"

        elif pd.api.types.is_numeric_dtype(df[column_name]):
            for row_number in range(2, worksheet.max_row + 1):
                worksheet[
                    f"{excel_column}{row_number}"
                ].number_format = "#,##0.00"

    autosize_excel_columns(
        writer=writer,
        sheet_name=sheet_name,
        df=df,
    )


def write_workbook(workbook_sheets, output_path):
    """
    Write all workbook sheets to one formatted Excel workbook.
    """
    used_sheet_names = set()

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        for requested_sheet_name, dataframe in workbook_sheets.items():
            sheet_name = make_sheet_name(
                requested_sheet_name,
                used_sheet_names,
            )

            clean_dataframe = clean_for_excel(dataframe)

            clean_dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

            format_excel_sheet(
                writer=writer,
                sheet_name=sheet_name,
                df=clean_dataframe,
            )

    print(f"\nSaved Excel workbook: {output_path}")


# ---------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------
def load_airportx_fastpark_bookings(
    start,
    end,
    anchor_date_column,
    deduplicate=True,
):
    """
    Load FastPark bookings using either entryDate or exitDate as the
    analysis-window anchor.

    Parameters
    ----------
    start : str
        Analysis-window start date.

    end : str
        Analysis-window end date.

    anchor_date_column : str
        Must be either "entryDate" or "exitDate".

    deduplicate : bool
        Whether to retain one record per bookingUuid.
    """
    if anchor_date_column not in {"entryDate", "exitDate"}:
        raise ValueError(
            "anchor_date_column must be 'entryDate' or 'exitDate'."
        )

    df = query(
        table="AirportX.v_Bookings",
        columns=[
            "bookingUuid",
            "bookingId AS [Booking ID]",
            "createdAt AS [Creation Date]",
            "entryDate",
            "exitDate",
            "Duration",
            "status",
            "assetName",
            "productName",
            "productCode",
            "bookingTotal",
            "leadtime",
        ],
        where=[
            "assetName = 'FastPark'",
            "status = 'B'",
        ],
        date_column=anchor_date_column,
        start=start,
        end=end,
    )

    df = pd.DataFrame(df).copy()

    if df.empty:
        return df

    for column in ["Creation Date", "entryDate", "exitDate"]:
        df[column] = pd.to_datetime(
            df[column],
            errors="coerce",
        )

    df["Duration"] = pd.to_numeric(
        df["Duration"],
        errors="coerce",
    )

    df["leadtime"] = pd.to_numeric(
        df["leadtime"],
        errors="coerce",
    )

    if deduplicate:
        df = (
            df.sort_values(
                ["bookingUuid", "Creation Date"]
            )
            .drop_duplicates(
                subset=["bookingUuid"],
                keep="first",
            )
            .reset_index(drop=True)
        )

    return df


def load_entry_bookings(start, end, deduplicate=True):
    """
    Load bookings whose entryDate is within the analysis window.
    """
    return load_airportx_fastpark_bookings(
        start=start,
        end=end,
        anchor_date_column="entryDate",
        deduplicate=deduplicate,
    )


def load_exit_bookings(start, end, deduplicate=True):
    """
    Load bookings whose exitDate is within the analysis window.

    This separate population fixes the exit-boundary problem in the old
    script, which used bookings selected by entryDate for exit analysis.
    """
    return load_airportx_fastpark_bookings(
        start=start,
        end=end,
        anchor_date_column="exitDate",
        deduplicate=deduplicate,
    )


def load_fastpark_actual_entries(start, end):
    """
    Load actual FastPark entries for the analysis window.
    """
    df = query(
        table="FastPark.v_EntryAndExits",
        columns=[
            "BookingReference",
            "CheckInStarted",
            "ActualCheckedOutDate",
        ],
        where=[],
        date_column="CheckInStarted",
        start=start,
        end=end,
    )

    df = pd.DataFrame(df).copy()

    if df.empty:
        return df

    df["CheckInStarted"] = pd.to_datetime(
        df["CheckInStarted"],
        errors="coerce",
    )

    df = df.dropna(
        subset=[
            "BookingReference",
            "CheckInStarted",
        ]
    )

    df = df.drop_duplicates(
        subset=["BookingReference"]
    )

    df["entry_day"] = (
        df["CheckInStarted"]
        .dt.normalize()
    )

    df["entry_month"] = (
        df["entry_day"]
        .dt.to_period("M")
        .astype(str)
    )

    return df.reset_index(drop=True)


def load_fastpark_actual_exits(start, end):
    """
    Load actual FastPark exits for the analysis window.
    """
    df = query(
        table="FastPark.v_EntryAndExits",
        columns=[
            "BookingReference",
            "CheckInStarted",
            "ActualCheckedOutDate",
        ],
        where=[],
        date_column="ActualCheckedOutDate",
        start=start,
        end=end,
    )

    df = pd.DataFrame(df).copy()

    if df.empty:
        return df

    df["ActualCheckedOutDate"] = pd.to_datetime(
        df["ActualCheckedOutDate"],
        errors="coerce",
    )

    df = df.dropna(
        subset=[
            "BookingReference",
            "ActualCheckedOutDate",
        ]
    )

    df = df.drop_duplicates(
        subset=["BookingReference"]
    )

    df["exit_day"] = (
        df["ActualCheckedOutDate"]
        .dt.normalize()
    )

    df["exit_month"] = (
        df["exit_day"]
        .dt.to_period("M")
        .astype(str)
    )

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------
# Booking enrichment
# ---------------------------------------------------------------------
def prepare_bookings(df):
    """
    Prepare booking timestamps, lead times and duration groups.
    """
    df = df.copy()

    if df.empty:
        return df

    df = df.dropna(
        subset=[
            "bookingUuid",
            "Creation Date",
            "entryDate",
            "exitDate",
        ]
    ).copy()

    df["creation_day"] = (
        df["Creation Date"]
        .dt.normalize()
    )

    df["entry_day"] = (
        df["entryDate"]
        .dt.normalize()
    )

    df["exit_day"] = (
        df["exitDate"]
        .dt.normalize()
    )

    df["lead_to_entry_days_exact"] = (
        (
            df["entryDate"] -
            df["Creation Date"]
        ).dt.total_seconds() /
        86400
    )

    df["lead_to_exit_days_exact"] = (
        (
            df["exitDate"] -
            df["Creation Date"]
        ).dt.total_seconds() /
        86400
    )

    df["lead_to_entry_calendar_days"] = (
        df["entry_day"] -
        df["creation_day"]
    ).dt.days

    df["lead_to_exit_calendar_days"] = (
        df["exit_day"] -
        df["creation_day"]
    ).dt.days

    df["invalid_entry_lead"] = (
        df["lead_to_entry_days_exact"] < 0
    )

    df["invalid_exit_lead"] = (
        df["lead_to_exit_days_exact"] < 0
    )

    inferred_duration = (
        df["exit_day"] -
        df["entry_day"]
    ).dt.days

    df["Duration_clean"] = pd.to_numeric(
        df["Duration"],
        errors="coerce",
    )

    missing_duration = df["Duration_clean"].isna()

    df.loc[
        missing_duration,
        "Duration_clean",
    ] = inferred_duration[missing_duration]

    df["duration_bucket"] = (
        df["Duration_clean"]
        .apply(duration_bucket)
    )

    df["entry_month"] = (
        df["entry_day"]
        .dt.to_period("M")
        .astype(str)
    )

    df["exit_month"] = (
        df["exit_day"]
        .dt.to_period("M")
        .astype(str)
    )

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------
# Core calculation helper
# ---------------------------------------------------------------------
def calculate_summary_metrics(
    compare_df,
    actual_column,
    seen_column,
):
    """
    Calculate aggregate, mean and median calibration metrics.

    raw_correction_pct may be negative.

    operational_uplift_pct is floored at zero and is intended for systems
    that only support a positive uplift.
    """
    valid = compare_df.loc[
        (compare_df[actual_column] > 0) &
        (compare_df[seen_column] > 0)
    ].copy()

    if valid.empty:
        return {
            "days_with_actual": int(
                (compare_df[actual_column] > 0).sum()
            ),
            "days_with_seen": int(
                (compare_df[seen_column] > 0).sum()
            ),
            "total_actual": compare_df[actual_column].sum(),
            "total_seen": compare_df[seen_column].sum(),
            "volume_weighted_visible_pct": np.nan,
            "volume_weighted_raw_correction_pct": np.nan,
            "volume_weighted_operational_uplift_pct": np.nan,
            "daily_visible_mean_pct": np.nan,
            "daily_visible_median_pct": np.nan,
            "daily_raw_correction_mean_pct": np.nan,
            "daily_raw_correction_median_pct": np.nan,
            "daily_operational_uplift_mean_pct": np.nan,
            "daily_operational_uplift_median_pct": np.nan,
        }

    total_actual = valid[actual_column].sum()
    total_seen = valid[seen_column].sum()

    volume_weighted_visible_pct = (
        total_seen / total_actual * 100
        if total_actual > 0
        else np.nan
    )

    volume_weighted_raw_correction_pct = (
        (total_actual / total_seen - 1) * 100
        if total_seen > 0
        else np.nan
    )

    volume_weighted_operational_uplift_pct = (
        max(volume_weighted_raw_correction_pct, 0)
        if pd.notna(volume_weighted_raw_correction_pct)
        else np.nan
    )

    return {
        "days_with_actual": int(
            (compare_df[actual_column] > 0).sum()
        ),
        "days_with_seen": int(
            (compare_df[seen_column] > 0).sum()
        ),
        "valid_comparison_days": len(valid),
        "total_actual": total_actual,
        "total_seen": total_seen,
        "volume_weighted_visible_pct":
            volume_weighted_visible_pct,
        "volume_weighted_raw_correction_pct":
            volume_weighted_raw_correction_pct,
        "volume_weighted_operational_uplift_pct":
            volume_weighted_operational_uplift_pct,
        "daily_visible_mean_pct":
            valid["pct_actual_visible"].mean(),
        "daily_visible_median_pct":
            valid["pct_actual_visible"].median(),
        "daily_raw_correction_mean_pct":
            valid["raw_correction_pct"].mean(),
        "daily_raw_correction_median_pct":
            valid["raw_correction_pct"].median(),
        "daily_operational_uplift_mean_pct":
            valid["operational_uplift_pct"].mean(),
        "daily_operational_uplift_median_pct":
            valid["operational_uplift_pct"].median(),
    }


def create_daily_comparison(
    actual_counts,
    final_booking_counts,
    seen_booking_counts,
    actual_column,
    final_column,
    seen_column,
    horizon_label,
    index_name,
):
    """
    Create the daily underlying comparison for one horizon.
    """
    compare = pd.concat(
        [
            actual_counts,
            final_booking_counts,
            seen_booking_counts,
        ],
        axis=1,
    ).fillna(0)

    compare.index.name = index_name

    compare["horizon"] = horizon_label

    compare["pct_actual_visible"] = safe_divide(
        compare[seen_column],
        compare[actual_column],
    ) * 100

    compare["pct_final_booking_visible"] = safe_divide(
        compare[seen_column],
        compare[final_column],
    ) * 100

    compare["raw_correction_pct"] = np.where(
        compare[seen_column] > 0,
        (
            compare[actual_column] /
            compare[seen_column] -
            1
        ) * 100,
        np.nan,
    )

    compare["operational_uplift_pct"] = (
        compare["raw_correction_pct"]
        .clip(lower=0)
    )

    compare = compare.reset_index()

    compare["month"] = (
        pd.to_datetime(compare[index_name])
        .dt.to_period("M")
        .astype(str)
    )

    return compare


# ---------------------------------------------------------------------
# Overall entry, exit and implied-exit uplift analysis
# ---------------------------------------------------------------------
def build_overall_uplifts(
    entry_bookings_df,
    exit_bookings_df,
    actual_entries_df,
    actual_exits_df,
    horizons=DEFAULT_HORIZONS,
):
    """
    Calculate overall entry, direct-exit and implied-exit uplift metrics.
    """
    entry_bookings = entry_bookings_df.loc[
        ~entry_bookings_df["invalid_entry_lead"]
    ].copy()

    exit_bookings = exit_bookings_df.loc[
        (
            ~exit_bookings_df["invalid_entry_lead"]
        ) &
        (
            ~exit_bookings_df["invalid_exit_lead"]
        )
    ].copy()

    actual_entry_counts = (
        actual_entries_df
        .groupby("entry_day")["BookingReference"]
        .nunique()
        .rename("actual_entries")
    )

    actual_exit_counts = (
        actual_exits_df
        .groupby("exit_day")["BookingReference"]
        .nunique()
        .rename("actual_exits")
    )

    final_booked_entries = (
        entry_bookings
        .groupby("entry_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_entries")
    )

    final_booked_exits = (
        exit_bookings
        .groupby("exit_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_exits")
    )

    summary_rows = []

    entry_daily_frames = []
    exit_daily_frames = []
    implied_exit_daily_frames = []

    for horizon in horizons:
        horizon_label = f"D{horizon}"

        # ---------------------------------------------------------
        # Entries visible relative to entry date
        # ---------------------------------------------------------
        entry_seen = entry_bookings.loc[
            entry_bookings["creation_day"] <=
            entry_bookings["entry_day"] -
            pd.Timedelta(days=horizon)
        ].copy()

        entry_seen_counts = (
            entry_seen
            .groupby("entry_day")["bookingUuid"]
            .nunique()
            .rename("seen_entries")
        )

        entry_compare = create_daily_comparison(
            actual_counts=actual_entry_counts,
            final_booking_counts=final_booked_entries,
            seen_booking_counts=entry_seen_counts,
            actual_column="actual_entries",
            final_column="final_booked_entries",
            seen_column="seen_entries",
            horizon_label=horizon_label,
            index_name="entry_day",
        )

        entry_metrics = calculate_summary_metrics(
            compare_df=entry_compare,
            actual_column="actual_entries",
            seen_column="seen_entries",
        )

        entry_daily_frames.append(entry_compare)

        # ---------------------------------------------------------
        # Direct exits visible relative to exit date
        # ---------------------------------------------------------
        direct_exit_seen = exit_bookings.loc[
            exit_bookings["creation_day"] <=
            exit_bookings["exit_day"] -
            pd.Timedelta(days=horizon)
        ].copy()

        direct_exit_seen_counts = (
            direct_exit_seen
            .groupby("exit_day")["bookingUuid"]
            .nunique()
            .rename("seen_exits")
        )

        exit_compare = create_daily_comparison(
            actual_counts=actual_exit_counts,
            final_booking_counts=final_booked_exits,
            seen_booking_counts=direct_exit_seen_counts,
            actual_column="actual_exits",
            final_column="final_booked_exits",
            seen_column="seen_exits",
            horizon_label=horizon_label,
            index_name="exit_day",
        )

        exit_metrics = calculate_summary_metrics(
            compare_df=exit_compare,
            actual_column="actual_exits",
            seen_column="seen_exits",
        )

        exit_daily_frames.append(exit_compare)

        # ---------------------------------------------------------
        # Implied exits
        #
        # Population is selected by exitDate.
        # Visibility is determined relative to entryDate.
        # ---------------------------------------------------------
        implied_exit_seen = exit_bookings.loc[
            exit_bookings["creation_day"] <=
            exit_bookings["entry_day"] -
            pd.Timedelta(days=horizon)
        ].copy()

        implied_exit_seen_counts = (
            implied_exit_seen
            .groupby("exit_day")["bookingUuid"]
            .nunique()
            .rename("implied_seen_exits")
        )

        implied_exit_compare = create_daily_comparison(
            actual_counts=actual_exit_counts,
            final_booking_counts=final_booked_exits,
            seen_booking_counts=implied_exit_seen_counts,
            actual_column="actual_exits",
            final_column="final_booked_exits",
            seen_column="implied_seen_exits",
            horizon_label=horizon_label,
            index_name="exit_day",
        )

        implied_exit_metrics = calculate_summary_metrics(
            compare_df=implied_exit_compare,
            actual_column="actual_exits",
            seen_column="implied_seen_exits",
        )

        implied_exit_daily_frames.append(
            implied_exit_compare
        )

        summary_rows.append({
            "horizon": horizon_label,
            "horizon_num": horizon,

            "entry_days":
                entry_metrics["valid_comparison_days"],
            "entry_total_actual":
                entry_metrics["total_actual"],
            "entry_total_seen":
                entry_metrics["total_seen"],
            "entry_visible_volume_weighted_pct":
                entry_metrics[
                    "volume_weighted_visible_pct"
                ],
            "entry_raw_correction_volume_weighted_pct":
                entry_metrics[
                    "volume_weighted_raw_correction_pct"
                ],
            "entry_recommended_uplift_volume_weighted_pct":
                entry_metrics[
                    "volume_weighted_operational_uplift_pct"
                ],
            "entry_visible_daily_median_pct":
                entry_metrics[
                    "daily_visible_median_pct"
                ],
            "entry_raw_correction_daily_median_pct":
                entry_metrics[
                    "daily_raw_correction_median_pct"
                ],
            "entry_recommended_uplift_daily_median_pct":
                entry_metrics[
                    "daily_operational_uplift_median_pct"
                ],

            "exit_days":
                exit_metrics["valid_comparison_days"],
            "exit_total_actual":
                exit_metrics["total_actual"],
            "exit_total_seen":
                exit_metrics["total_seen"],
            "exit_visible_volume_weighted_pct":
                exit_metrics[
                    "volume_weighted_visible_pct"
                ],
            "exit_raw_correction_volume_weighted_pct":
                exit_metrics[
                    "volume_weighted_raw_correction_pct"
                ],
            "exit_recommended_uplift_volume_weighted_pct":
                exit_metrics[
                    "volume_weighted_operational_uplift_pct"
                ],
            "exit_visible_daily_median_pct":
                exit_metrics[
                    "daily_visible_median_pct"
                ],
            "exit_raw_correction_daily_median_pct":
                exit_metrics[
                    "daily_raw_correction_median_pct"
                ],
            "exit_recommended_uplift_daily_median_pct":
                exit_metrics[
                    "daily_operational_uplift_median_pct"
                ],

            "implied_exit_days":
                implied_exit_metrics[
                    "valid_comparison_days"
                ],
            "implied_exit_total_actual":
                implied_exit_metrics["total_actual"],
            "implied_exit_total_seen":
                implied_exit_metrics["total_seen"],
            "implied_exit_visible_volume_weighted_pct":
                implied_exit_metrics[
                    "volume_weighted_visible_pct"
                ],
            "implied_exit_raw_correction_volume_weighted_pct":
                implied_exit_metrics[
                    "volume_weighted_raw_correction_pct"
                ],
            "implied_exit_recommended_uplift_volume_weighted_pct":
                implied_exit_metrics[
                    "volume_weighted_operational_uplift_pct"
                ],
            "implied_exit_visible_daily_median_pct":
                implied_exit_metrics[
                    "daily_visible_median_pct"
                ],
            "implied_exit_raw_correction_daily_median_pct":
                implied_exit_metrics[
                    "daily_raw_correction_median_pct"
                ],
            "implied_exit_recommended_uplift_daily_median_pct":
                implied_exit_metrics[
                    "daily_operational_uplift_median_pct"
                ],
        })

    overall_summary = (
        pd.DataFrame(summary_rows)
        .sort_values("horizon_num")
        .reset_index(drop=True)
    )

    return {
        "overall_summary": overall_summary,
        "entry_daily_detail": pd.concat(
            entry_daily_frames,
            ignore_index=True,
        ),
        "exit_daily_detail": pd.concat(
            exit_daily_frames,
            ignore_index=True,
        ),
        "implied_exit_daily_detail": pd.concat(
            implied_exit_daily_frames,
            ignore_index=True,
        ),
    }


# ---------------------------------------------------------------------
# Monthly uplift analysis
# ---------------------------------------------------------------------
def summarise_monthly_daily_detail(
    daily_detail,
    actual_column,
    seen_column,
    movement_type,
):
    """
    Summarise daily horizon results by service month.
    """
    rows = []

    grouped = daily_detail.groupby(
        ["month", "horizon"],
        dropna=False,
    )

    for (month, horizon), group in grouped:
        metrics = calculate_summary_metrics(
            compare_df=group,
            actual_column=actual_column,
            seen_column=seen_column,
        )

        rows.append({
            "movement_type": movement_type,
            "month": month,
            "horizon": horizon,
            "horizon_num": int(
                str(horizon).replace("D", "")
            ),
            "valid_comparison_days":
                metrics["valid_comparison_days"],
            "total_actual":
                metrics["total_actual"],
            "total_seen":
                metrics["total_seen"],
            "visible_volume_weighted_pct":
                metrics[
                    "volume_weighted_visible_pct"
                ],
            "raw_correction_volume_weighted_pct":
                metrics[
                    "volume_weighted_raw_correction_pct"
                ],
            "recommended_uplift_volume_weighted_pct":
                metrics[
                    "volume_weighted_operational_uplift_pct"
                ],
            "visible_daily_mean_pct":
                metrics[
                    "daily_visible_mean_pct"
                ],
            "visible_daily_median_pct":
                metrics[
                    "daily_visible_median_pct"
                ],
            "raw_correction_daily_median_pct":
                metrics[
                    "daily_raw_correction_median_pct"
                ],
            "recommended_uplift_daily_median_pct":
                metrics[
                    "daily_operational_uplift_median_pct"
                ],
        })

    return (
        pd.DataFrame(rows)
        .sort_values(
            ["month", "horizon_num"]
        )
        .reset_index(drop=True)
    )


def build_monthly_uplifts(overall_outputs):
    """
    Build detailed monthly entry, exit and implied-exit uplift tables.
    """
    monthly_entry = summarise_monthly_daily_detail(
        daily_detail=overall_outputs[
            "entry_daily_detail"
        ],
        actual_column="actual_entries",
        seen_column="seen_entries",
        movement_type="Entry",
    )

    monthly_exit = summarise_monthly_daily_detail(
        daily_detail=overall_outputs[
            "exit_daily_detail"
        ],
        actual_column="actual_exits",
        seen_column="seen_exits",
        movement_type="Direct Exit",
    )

    monthly_implied_exit = summarise_monthly_daily_detail(
        daily_detail=overall_outputs[
            "implied_exit_daily_detail"
        ],
        actual_column="actual_exits",
        seen_column="implied_seen_exits",
        movement_type="Implied Exit",
    )

    monthly_all = pd.concat(
        [
            monthly_entry,
            monthly_exit,
            monthly_implied_exit,
        ],
        ignore_index=True,
    )

    return {
        "monthly_entry": monthly_entry,
        "monthly_exit": monthly_exit,
        "monthly_implied_exit":
            monthly_implied_exit,
        "monthly_all": monthly_all,
    }


def build_monthly_pivot(
    monthly_df,
    metric_column,
):
    """
    Create a horizon-by-month comparison pivot.
    """
    pivot = monthly_df.pivot(
        index="horizon",
        columns="month",
        values=metric_column,
    )

    pivot["horizon_num"] = horizon_number(
        pivot.index.to_series()
    ).values

    pivot = (
        pivot.sort_values("horizon_num")
        .drop(columns="horizon_num")
        .reset_index()
    )

    return pivot


# ---------------------------------------------------------------------
# Lead-time distributions
# ---------------------------------------------------------------------
def build_leadtime_distribution(
    bookings_df,
    calendar_lead_column,
    exact_lead_column,
    movement_type,
):
    """
    Build a calendar-day lead-time distribution.

    The table includes both:
    - number and percentage of bookings
    - exact lead-time statistics inside each calendar-day group
    """
    df = bookings_df.loc[
        bookings_df[calendar_lead_column] >= 0
    ].copy()

    distribution = (
        df.groupby(calendar_lead_column)
        .agg(
            bookings=("bookingUuid", "nunique"),
            exact_lead_mean_days=(
                exact_lead_column,
                "mean",
            ),
            exact_lead_median_days=(
                exact_lead_column,
                "median",
            ),
            exact_lead_min_days=(
                exact_lead_column,
                "min",
            ),
            exact_lead_max_days=(
                exact_lead_column,
                "max",
            ),
        )
        .reset_index()
        .sort_values(calendar_lead_column)
    )

    distribution["booking_pct"] = (
        distribution["bookings"] /
        distribution["bookings"].sum() *
        100
    )

    distribution["cumulative_booking_pct"] = (
        distribution["booking_pct"]
        .cumsum()
    )

    distribution.insert(
        0,
        "movement_type",
        movement_type,
    )

    return distribution


def build_leadtime_band_summary(
    bookings_df,
    calendar_lead_column,
    movement_type,
):
    """
    Build a compact lead-time-band summary.
    """
    df = bookings_df.loc[
        bookings_df[calendar_lead_column] >= 0
    ].copy()

    bins = [
        -0.001,
        0,
        1,
        2,
        3,
        6,
        13,
        20,
        27,
        59,
        89,
        np.inf,
    ]

    labels = [
        "D0",
        "D1",
        "D2",
        "D3",
        "D4-D6",
        "D7-D13",
        "D14-D20",
        "D21-D27",
        "D28-D59",
        "D60-D89",
        "D90+",
    ]

    df["leadtime_band"] = pd.cut(
        df[calendar_lead_column],
        bins=bins,
        labels=labels,
        right=True,
        include_lowest=True,
    )

    summary = (
        df.groupby(
            "leadtime_band",
            observed=False,
        )
        .agg(
            bookings=("bookingUuid", "nunique")
        )
        .reset_index()
    )

    summary["booking_pct"] = (
        summary["bookings"] /
        summary["bookings"].sum() *
        100
    )

    summary.insert(
        0,
        "movement_type",
        movement_type,
    )

    return summary


# ---------------------------------------------------------------------
# Duration analysis
# ---------------------------------------------------------------------
def build_duration_uplifts(
    entry_bookings_df,
    actual_entries_df,
    horizons=DEFAULT_HORIZONS,
):
    """
    Analyse entry-side booking visibility by duration bucket.

    Actual entries cannot be assigned to a duration group unless the
    actual dataset is matched back to booking-level records. Therefore,
    this table calculates:

        visible bookings / final bookings

    and the corresponding completion uplift:

        final bookings / visible bookings - 1

    This measures the booking-curve uplift within each duration group.
    """
    bookings = entry_bookings_df.loc[
        ~entry_bookings_df["invalid_entry_lead"]
    ].copy()

    rows = []

    for horizon in horizons:
        horizon_label = f"D{horizon}"

        seen = bookings.loc[
            bookings["creation_day"] <=
            bookings["entry_day"] -
            pd.Timedelta(days=horizon)
        ].copy()

        final_counts = (
            bookings
            .groupby(
                [
                    "entry_month",
                    "duration_bucket",
                ]
            )["bookingUuid"]
            .nunique()
            .rename("final_bookings")
            .reset_index()
        )

        seen_counts = (
            seen
            .groupby(
                [
                    "entry_month",
                    "duration_bucket",
                ]
            )["bookingUuid"]
            .nunique()
            .rename("seen_bookings")
            .reset_index()
        )

        compare = final_counts.merge(
            seen_counts,
            on=[
                "entry_month",
                "duration_bucket",
            ],
            how="left",
        )

        compare["seen_bookings"] = (
            compare["seen_bookings"]
            .fillna(0)
        )

        compare["horizon"] = horizon_label
        compare["horizon_num"] = horizon

        compare["visible_pct"] = safe_divide(
            compare["seen_bookings"],
            compare["final_bookings"],
        ) * 100

        compare["raw_completion_correction_pct"] = np.where(
            compare["seen_bookings"] > 0,
            (
                compare["final_bookings"] /
                compare["seen_bookings"] -
                1
            ) * 100,
            np.nan,
        )

        compare["recommended_completion_uplift_pct"] = (
            compare[
                "raw_completion_correction_pct"
            ]
            .clip(lower=0)
        )

        rows.append(compare)

    monthly_detail = pd.concat(
        rows,
        ignore_index=True,
    )

    overall = (
        monthly_detail
        .groupby(
            [
                "horizon",
                "horizon_num",
                "duration_bucket",
            ]
        )
        .agg(
            final_bookings=(
                "final_bookings",
                "sum",
            ),
            seen_bookings=(
                "seen_bookings",
                "sum",
            ),
            months=(
                "entry_month",
                "nunique",
            ),
        )
        .reset_index()
    )

    overall["visible_volume_weighted_pct"] = safe_divide(
        overall["seen_bookings"],
        overall["final_bookings"],
    ) * 100

    overall[
        "raw_completion_correction_volume_weighted_pct"
    ] = np.where(
        overall["seen_bookings"] > 0,
        (
            overall["final_bookings"] /
            overall["seen_bookings"] -
            1
        ) * 100,
        np.nan,
    )

    overall[
        "recommended_completion_uplift_volume_weighted_pct"
    ] = (
        overall[
            "raw_completion_correction_volume_weighted_pct"
        ]
        .clip(lower=0)
    )

    duration_order_mapping = {
        duration: index
        for index, duration in enumerate(DURATION_ORDER)
    }

    overall["duration_sort"] = (
        overall["duration_bucket"]
        .map(duration_order_mapping)
        .fillna(len(DURATION_ORDER))
    )

    overall = (
        overall.sort_values(
            [
                "horizon_num",
                "duration_sort",
            ]
        )
        .drop(columns="duration_sort")
        .reset_index(drop=True)
    )

    monthly_detail["duration_sort"] = (
        monthly_detail["duration_bucket"]
        .map(duration_order_mapping)
        .fillna(len(DURATION_ORDER))
    )

    monthly_detail = (
        monthly_detail.sort_values(
            [
                "entry_month",
                "horizon_num",
                "duration_sort",
            ]
        )
        .drop(columns="duration_sort")
        .reset_index(drop=True)
    )

    return {
        "duration_overall": overall,
        "duration_monthly": monthly_detail,
    }


# ---------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------
def build_entry_reconciliation(
    entry_bookings_df,
    actual_entries_df,
):
    """
    Compare final booked entries with actual entries by service day.
    """
    booked = (
        entry_bookings_df
        .groupby("entry_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_entries")
    )

    actual = (
        actual_entries_df
        .groupby("entry_day")["BookingReference"]
        .nunique()
        .rename("actual_entries")
    )

    result = pd.concat(
        [booked, actual],
        axis=1,
    ).fillna(0)

    result["booking_minus_actual"] = (
        result["final_booked_entries"] -
        result["actual_entries"]
    )

    result["booking_vs_actual_pct"] = safe_divide(
        result["final_booked_entries"],
        result["actual_entries"],
    ) * 100

    result = result.reset_index()

    result["month"] = (
        result["entry_day"]
        .dt.to_period("M")
        .astype(str)
    )

    return result


def build_exit_reconciliation(
    exit_bookings_df,
    actual_exits_df,
):
    """
    Compare final booked exits with actual exits by service day.
    """
    booked = (
        exit_bookings_df
        .groupby("exit_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_exits")
    )

    actual = (
        actual_exits_df
        .groupby("exit_day")["BookingReference"]
        .nunique()
        .rename("actual_exits")
    )

    result = pd.concat(
        [booked, actual],
        axis=1,
    ).fillna(0)

    result["booking_minus_actual"] = (
        result["final_booked_exits"] -
        result["actual_exits"]
    )

    result["booking_vs_actual_pct"] = safe_divide(
        result["final_booked_exits"],
        result["actual_exits"],
    ) * 100

    result = result.reset_index()

    result["month"] = (
        result["exit_day"]
        .dt.to_period("M")
        .astype(str)
    )

    return result

def build_current_month_impact_comparison(
    overall_summary,
    monthly_entry,
    monthly_exit,
    monthly_implied_exit,
):
    """
    Compare the overall eight-month recommendation with the aggregated
    completed-month recommendation excluding the current partial month.

    This shows how much the current partial month changes the recommended
    uplift at each horizon.
    """
    current_month = (
        pd.Timestamp.today()
        .normalize()
        .to_period("M")
        .strftime("%Y-%m")
    )

    comparison = overall_summary[
        [
            "horizon",
            "horizon_num",
            "entry_recommended_uplift_volume_weighted_pct",
            "exit_recommended_uplift_volume_weighted_pct",
            "implied_exit_recommended_uplift_volume_weighted_pct",
        ]
    ].copy()

    comparison = comparison.rename(
        columns={
            "entry_recommended_uplift_volume_weighted_pct":
                "entry_uplift_including_partial_month_pct",
            "exit_recommended_uplift_volume_weighted_pct":
                "exit_uplift_including_partial_month_pct",
            "implied_exit_recommended_uplift_volume_weighted_pct":
                "implied_exit_uplift_including_partial_month_pct",
        }
    )

    def aggregate_completed_months(monthly_df, prefix):
        completed = monthly_df.loc[
            monthly_df["month"] != current_month
        ].copy()

        aggregated = (
            completed
            .groupby(
                ["horizon", "horizon_num"],
                as_index=False,
            )
            .agg(
                total_actual=("total_actual", "sum"),
                total_seen=("total_seen", "sum"),
            )
        )

        aggregated[
            f"{prefix}_raw_correction_excluding_partial_month_pct"
        ] = np.where(
            aggregated["total_seen"] > 0,
            (
                aggregated["total_actual"] /
                aggregated["total_seen"] -
                1
            ) * 100,
            np.nan,
        )

        aggregated[
            f"{prefix}_uplift_excluding_partial_month_pct"
        ] = (
            aggregated[
                f"{prefix}_raw_correction_excluding_partial_month_pct"
            ]
            .clip(lower=0)
        )

        return aggregated[
            [
                "horizon",
                "horizon_num",
                f"{prefix}_uplift_excluding_partial_month_pct",
            ]
        ]

    entry_completed = aggregate_completed_months(
        monthly_entry,
        "entry",
    )

    exit_completed = aggregate_completed_months(
        monthly_exit,
        "exit",
    )

    implied_completed = aggregate_completed_months(
        monthly_implied_exit,
        "implied_exit",
    )

    comparison = comparison.merge(
        entry_completed,
        on=["horizon", "horizon_num"],
        how="left",
    )

    comparison = comparison.merge(
        exit_completed,
        on=["horizon", "horizon_num"],
        how="left",
    )

    comparison = comparison.merge(
        implied_completed,
        on=["horizon", "horizon_num"],
        how="left",
    )

    comparison["entry_partial_month_effect_pct_points"] = (
        comparison[
            "entry_uplift_including_partial_month_pct"
        ] -
        comparison[
            "entry_uplift_excluding_partial_month_pct"
        ]
    )

    comparison["exit_partial_month_effect_pct_points"] = (
        comparison[
            "exit_uplift_including_partial_month_pct"
        ] -
        comparison[
            "exit_uplift_excluding_partial_month_pct"
        ]
    )

    comparison[
        "implied_exit_partial_month_effect_pct_points"
    ] = (
        comparison[
            "implied_exit_uplift_including_partial_month_pct"
        ] -
        comparison[
            "implied_exit_uplift_excluding_partial_month_pct"
        ]
    )

    return (
        comparison
        .sort_values("horizon_num")
        .reset_index(drop=True)
    )

# ---------------------------------------------------------------------
# Recommended arrays
# ---------------------------------------------------------------------
def build_recommended_arrays(
    overall_summary,
    monthly_outputs,
):
    """
    Build copy-ready overall and monthly D1-D14 arrays.

    Primary recommendation:
        volume-weighted operational uplift

    Supporting comparison:
        daily median operational uplift
    """
    selected_horizons = [
        f"D{horizon}"
        for horizon in UPLIFT_ARRAY_HORIZONS
    ]

    overall = overall_summary.loc[
        overall_summary["horizon"].isin(
            selected_horizons
        )
    ].copy()

    overall = overall.sort_values("horizon_num")

    array_rows = [
        {
            "scope": "Overall",
            "movement_type": "Entry",
            "method":
                "Volume-weighted operational uplift",
            "month": "All 8 months",
            "python_array": str(
                overall[
                    "entry_recommended_uplift_volume_weighted_pct"
                ]
                .round(2)
                .tolist()
            ),
        },
        {
            "scope": "Overall",
            "movement_type": "Entry",
            "method":
                "Daily median operational uplift",
            "month": "All 8 months",
            "python_array": str(
                overall[
                    "entry_recommended_uplift_daily_median_pct"
                ]
                .round(2)
                .tolist()
            ),
        },
        {
            "scope": "Overall",
            "movement_type": "Direct Exit",
            "method":
                "Volume-weighted operational uplift",
            "month": "All 8 months",
            "python_array": str(
                overall[
                    "exit_recommended_uplift_volume_weighted_pct"
                ]
                .round(2)
                .tolist()
            ),
        },
        {
            "scope": "Overall",
            "movement_type": "Direct Exit",
            "method":
                "Daily median operational uplift",
            "month": "All 8 months",
            "python_array": str(
                overall[
                    "exit_recommended_uplift_daily_median_pct"
                ]
                .round(2)
                .tolist()
            ),
        },
        {
            "scope": "Overall",
            "movement_type": "Implied Exit",
            "method":
                "Volume-weighted operational uplift",
            "month": "All 8 months",
            "python_array": str(
                overall[
                    "implied_exit_recommended_uplift_volume_weighted_pct"
                ]
                .round(2)
                .tolist()
            ),
        },
        {
            "scope": "Overall",
            "movement_type": "Implied Exit",
            "method":
                "Daily median operational uplift",
            "month": "All 8 months",
            "python_array": str(
                overall[
                    "implied_exit_recommended_uplift_daily_median_pct"
                ]
                .round(2)
                .tolist()
            ),
        },
    ]

    monthly_mapping = [
        (
            "Entry",
            monthly_outputs["monthly_entry"],
        ),
        (
            "Direct Exit",
            monthly_outputs["monthly_exit"],
        ),
        (
            "Implied Exit",
            monthly_outputs[
                "monthly_implied_exit"
            ],
        ),
    ]

    for movement_type, monthly_df in monthly_mapping:
        for month in sorted(
            monthly_df["month"].dropna().unique()
        ):
            month_data = monthly_df.loc[
                (
                    monthly_df["month"] == month
                ) &
                (
                    monthly_df["horizon"].isin(
                        selected_horizons
                    )
                )
            ].copy()

            month_data = month_data.sort_values(
                "horizon_num"
            )

            array_rows.append({
                "scope": "Monthly",
                "movement_type": movement_type,
                "method":
                    "Volume-weighted operational uplift",
                "month": month,
                "python_array": str(
                    month_data[
                        "recommended_uplift_volume_weighted_pct"
                    ]
                    .round(2)
                    .tolist()
                ),
            })

    return pd.DataFrame(array_rows)


# ---------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------
def build_summary_sheet(
    start,
    end,
    entry_bookings,
    exit_bookings,
    actual_entries,
    actual_exits,
    overall_summary,
):
    """
    Create a compact run summary.
    """
    rows = [
        {
            "section": "Run Information",
            "metric": "analysis_start",
            "value": start,
        },
        {
            "section": "Run Information",
            "metric": "analysis_end",
            "value": end,
        },
        {
            "section": "Run Information",
            "metric": "run_datetime",
            "value": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        },
        {
            "section": "Data Volumes",
            "metric": "entry_date_bookings_loaded",
            "value": len(entry_bookings),
        },
        {
            "section": "Data Volumes",
            "metric": "exit_date_bookings_loaded",
            "value": len(exit_bookings),
        },
        {
            "section": "Data Volumes",
            "metric": "actual_entries_loaded",
            "value": len(actual_entries),
        },
        {
            "section": "Data Volumes",
            "metric": "actual_exits_loaded",
            "value": len(actual_exits),
        },
        {
            "section": "Method",
            "metric": "primary_recommendation",
            "value":
                "Volume-weighted operational uplift",
        },
        {
            "section": "Method",
            "metric": "horizon_definition",
            "value":
                "Calendar-day creation cutoff",
        },
        {
            "section": "Method",
            "metric": "uplift_formula",
            "value":
                "(total actual / total visible - 1) * 100",
        },
        {
            "section": "Method",
            "metric": "operational_uplift_floor",
            "value":
                "Negative raw corrections are floored at 0 only in operational uplift columns",
        },
    ]

    for horizon in ["D0", "D1", "D7", "D14", "D21", "D28"]:
        horizon_row = overall_summary.loc[
            overall_summary["horizon"] == horizon
        ]

        if horizon_row.empty:
            continue

        horizon_row = horizon_row.iloc[0]

        rows.extend([
            {
                "section": f"{horizon} Headline",
                "metric":
                    "entry_recommended_uplift_volume_weighted_pct",
                "value":
                    horizon_row[
                        "entry_recommended_uplift_volume_weighted_pct"
                    ],
            },
            {
                "section": f"{horizon} Headline",
                "metric":
                    "exit_recommended_uplift_volume_weighted_pct",
                "value":
                    horizon_row[
                        "exit_recommended_uplift_volume_weighted_pct"
                    ],
            },
            {
                "section": f"{horizon} Headline",
                "metric":
                    "implied_exit_recommended_uplift_volume_weighted_pct",
                "value":
                    horizon_row[
                        "implied_exit_recommended_uplift_volume_weighted_pct"
                    ],
            },
        ])

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------
def run_fastpark_uplift_analysis(
    start=None,
    end=None,
    number_of_completed_months=8,
    include_partial_window_month=True,
    deduplicate=True,
):
    """
    Run the complete FastPark uplift analysis.

    If start and end are both omitted, the last eight completed calendar
    months are used.

    If start and end are supplied, those exact dates are used.
    """
    if start is None and end is None:
        start, end = get_last_months_window(
            number_of_months=number_of_completed_months,
            include_current_partial_month=include_partial_window_month,
        )

    elif start is None or end is None:
        raise ValueError(
            "Supply both start and end, or leave both as None."
        )

    print_section("FASTPARK UPLIFT ANALYSIS")

    print(f"Analysis window: {start} to {end}")
    print(
        "Population basis: entry bookings by entryDate; "
        "exit bookings by exitDate"
    )
    print(f"Output folder: {OUTPUT_DIR}")

    # ---------------------------------------------------------
    # Load data
    # ---------------------------------------------------------
    print_section("LOADING DATA")

    entry_bookings_raw = load_entry_bookings(
        start=start,
        end=end,
        deduplicate=deduplicate,
    )

    exit_bookings_raw = load_exit_bookings(
        start=start,
        end=end,
        deduplicate=deduplicate,
    )

    entry_bookings = prepare_bookings(
        entry_bookings_raw
    )

    exit_bookings = prepare_bookings(
        exit_bookings_raw
    )

    actual_entries = load_fastpark_actual_entries(
        start=start,
        end=end,
    )

    actual_exits = load_fastpark_actual_exits(
        start=start,
        end=end,
    )

    print(
        f"Entry-date bookings loaded: "
        f"{len(entry_bookings):,.0f}"
    )

    print(
        f"Exit-date bookings loaded: "
        f"{len(exit_bookings):,.0f}"
    )

    print(
        f"Actual entries loaded: "
        f"{len(actual_entries):,.0f}"
    )

    print(
        f"Actual exits loaded: "
        f"{len(actual_exits):,.0f}"
    )

    if entry_bookings.empty:
        raise ValueError(
            "No entry-date bookings returned."
        )

    if exit_bookings.empty:
        raise ValueError(
            "No exit-date bookings returned."
        )

    if actual_entries.empty:
        raise ValueError(
            "No actual entries returned."
        )

    if actual_exits.empty:
        raise ValueError(
            "No actual exits returned."
        )

    # ---------------------------------------------------------
    # Lead-time distributions
    # ---------------------------------------------------------
    print_section("LEAD-TIME DISTRIBUTIONS")

    entry_leadtime_distribution = (
        build_leadtime_distribution(
            bookings_df=entry_bookings,
            calendar_lead_column=
                "lead_to_entry_calendar_days",
            exact_lead_column=
                "lead_to_entry_days_exact",
            movement_type="Entry",
        )
    )

    exit_leadtime_distribution = (
        build_leadtime_distribution(
            bookings_df=exit_bookings,
            calendar_lead_column=
                "lead_to_exit_calendar_days",
            exact_lead_column=
                "lead_to_exit_days_exact",
            movement_type="Direct Exit",
        )
    )

    entry_leadtime_bands = (
        build_leadtime_band_summary(
            bookings_df=entry_bookings,
            calendar_lead_column=
                "lead_to_entry_calendar_days",
            movement_type="Entry",
        )
    )

    exit_leadtime_bands = (
        build_leadtime_band_summary(
            bookings_df=exit_bookings,
            calendar_lead_column=
                "lead_to_exit_calendar_days",
            movement_type="Direct Exit",
        )
    )

    leadtime_bands = pd.concat(
        [
            entry_leadtime_bands,
            exit_leadtime_bands,
        ],
        ignore_index=True,
    )

    print("\nEntry lead-time bands:")
    print(
        entry_leadtime_bands
        .round(2)
        .to_string(index=False)
    )

    print("\nExit lead-time bands:")
    print(
        exit_leadtime_bands
        .round(2)
        .to_string(index=False)
    )

    # ---------------------------------------------------------
    # Overall uplifts
    # ---------------------------------------------------------
    print_section("OVERALL RECOMMENDED UPLIFTS")

    overall_outputs = build_overall_uplifts(
        entry_bookings_df=entry_bookings,
        exit_bookings_df=exit_bookings,
        actual_entries_df=actual_entries,
        actual_exits_df=actual_exits,
        horizons=DEFAULT_HORIZONS,
    )

    overall_summary = (
        overall_outputs["overall_summary"]
        .round(3)
    )

    print("\nOverall uplift summary:")
    print(
        overall_summary.to_string(index=False)
    )

    # ---------------------------------------------------------
    # Monthly uplifts
    # ---------------------------------------------------------
    print_section("MONTHLY RECOMMENDED UPLIFTS")

    monthly_outputs = build_monthly_uplifts(
        overall_outputs=overall_outputs
    )

    monthly_entry = (
        monthly_outputs["monthly_entry"]
        .round(3)
    )

    monthly_exit = (
        monthly_outputs["monthly_exit"]
        .round(3)
    )

    monthly_implied_exit = (
        monthly_outputs[
            "monthly_implied_exit"
        ]
        .round(3)
    )

    entry_monthly_uplift_pivot = (
        build_monthly_pivot(
            monthly_df=monthly_entry,
            metric_column=
                "recommended_uplift_volume_weighted_pct",
        )
        .round(2)
    )

    exit_monthly_uplift_pivot = (
        build_monthly_pivot(
            monthly_df=monthly_exit,
            metric_column=
                "recommended_uplift_volume_weighted_pct",
        )
        .round(2)
    )

    implied_exit_monthly_uplift_pivot = (
        build_monthly_pivot(
            monthly_df=monthly_implied_exit,
            metric_column=
                "recommended_uplift_volume_weighted_pct",
        )
        .round(2)
    )

    entry_monthly_visibility_pivot = (
        build_monthly_pivot(
            monthly_df=monthly_entry,
            metric_column=
                "visible_volume_weighted_pct",
        )
        .round(2)
    )

    exit_monthly_visibility_pivot = (
        build_monthly_pivot(
            monthly_df=monthly_exit,
            metric_column=
                "visible_volume_weighted_pct",
        )
        .round(2)
    )

    implied_exit_monthly_visibility_pivot = (
        build_monthly_pivot(
            monthly_df=monthly_implied_exit,
            metric_column=
                "visible_volume_weighted_pct",
        )
        .round(2)
    )

    partial_month_impact = (
        build_current_month_impact_comparison(
            overall_summary=overall_summary,
            monthly_entry=monthly_entry,
            monthly_exit=monthly_exit,
            monthly_implied_exit=monthly_implied_exit,
        )
        .round(3)
    )

    print("\nEntry recommended uplift by month:")
    print(
        entry_monthly_uplift_pivot
        .to_string(index=False)
    )

    print("\nDirect-exit recommended uplift by month:")
    print(
        exit_monthly_uplift_pivot
        .to_string(index=False)
    )

    print("\nImplied-exit recommended uplift by month:")
    print(
        implied_exit_monthly_uplift_pivot
        .to_string(index=False)
    )

    # ---------------------------------------------------------
    # Duration analysis
    # ---------------------------------------------------------
    print_section("DURATION UPLIFT ANALYSIS")

    duration_outputs = build_duration_uplifts(
        entry_bookings_df=entry_bookings,
        actual_entries_df=actual_entries,
        horizons=DEFAULT_HORIZONS,
    )

    duration_overall = (
        duration_outputs["duration_overall"]
        .round(3)
    )

    duration_monthly = (
        duration_outputs["duration_monthly"]
        .round(3)
    )

    print("\nOverall duration uplift:")
    print(
        duration_overall.to_string(index=False)
    )

    # ---------------------------------------------------------
    # Reconciliation
    # ---------------------------------------------------------
    print_section("BOOKING AND ACTUAL RECONCILIATION")

    entry_reconciliation = (
        build_entry_reconciliation(
            entry_bookings_df=entry_bookings,
            actual_entries_df=actual_entries,
        )
        .round(3)
    )

    exit_reconciliation = (
        build_exit_reconciliation(
            exit_bookings_df=exit_bookings,
            actual_exits_df=actual_exits,
        )
        .round(3)
    )

    print("\nEntry reconciliation sample:")
    print(
        entry_reconciliation
        .tail(20)
        .to_string(index=False)
    )

    print("\nExit reconciliation sample:")
    print(
        exit_reconciliation
        .tail(20)
        .to_string(index=False)
    )

    # ---------------------------------------------------------
    # Copy-ready arrays
    # ---------------------------------------------------------
    print_section("RECOMMENDED D1-D14 ARRAYS")

    recommended_arrays = build_recommended_arrays(
        overall_summary=overall_summary,
        monthly_outputs=monthly_outputs,
    )

    overall_array_rows = recommended_arrays.loc[
        recommended_arrays["scope"] == "Overall"
    ]

    print("\nOverall recommended arrays:")
    print(
        overall_array_rows.to_string(index=False)
    )

    # ---------------------------------------------------------
    # Summary sheet
    # ---------------------------------------------------------
    summary = build_summary_sheet(
        start=start,
        end=end,
        entry_bookings=entry_bookings,
        exit_bookings=exit_bookings,
        actual_entries=actual_entries,
        actual_exits=actual_exits,
        overall_summary=overall_summary,
    )

    # ---------------------------------------------------------
    # Workbook
    # ---------------------------------------------------------
    print_section("WRITING EXCEL WORKBOOK")

    workbook_sheets = {
        "01_Summary":
            summary,

        "02_Recommended_Arrays":
            recommended_arrays,

        "03_Overall_Uplifts":
            overall_summary,

        "04_Partial_Month_Impact":
            partial_month_impact,

        "05_Entry_Monthly_Uplifts":
            entry_monthly_uplift_pivot,

        "06_Exit_Monthly_Uplifts":
            exit_monthly_uplift_pivot,

        "07_Implied_Exit_Monthly":
            implied_exit_monthly_uplift_pivot,

        "08_Entry_Monthly_Visibility":
            entry_monthly_visibility_pivot,

        "09_Exit_Monthly_Visibility":
            exit_monthly_visibility_pivot,

        "10_Implied_Exit_Visibility":
            implied_exit_monthly_visibility_pivot,

        "11_Monthly_Entry_Detail":
            monthly_entry,

        "12_Monthly_Exit_Detail":
            monthly_exit,

        "13_Monthly_Implied_Detail":
            monthly_implied_exit,

        "14_Leadtime_Bands":
            leadtime_bands,

        "15_Entry_Leadtime_Detail":
            entry_leadtime_distribution,

        "16_Exit_Leadtime_Detail":
            exit_leadtime_distribution,

        "17_Duration_Overall":
            duration_overall,

        "18_Duration_Monthly":
            duration_monthly,

        "19_Entry_Reconciliation":
            entry_reconciliation,

        "20_Exit_Reconciliation":
            exit_reconciliation,

        "21_Entry_Daily_Detail":
            overall_outputs[
                "entry_daily_detail"
            ],

        "22_Exit_Daily_Detail":
            overall_outputs[
                "exit_daily_detail"
            ],

        "23_Implied_Exit_Detail":
            overall_outputs[
                "implied_exit_daily_detail"
            ],
    }

    write_workbook(
        workbook_sheets=workbook_sheets,
        output_path=EXCEL_OUTPUT_PATH,
    )

    save_csv(
        overall_summary,
        "overall_uplifts_D0_D28.csv",
    )

    save_csv(
        monthly_entry,
        "monthly_entry_uplifts_D0_D28.csv",
    )

    save_csv(
        monthly_exit,
        "monthly_exit_uplifts_D0_D28.csv",
    )

    save_csv(
        monthly_implied_exit,
        "monthly_implied_exit_uplifts_D0_D28.csv",
    )

    print(f"\nSaved to: {EXCEL_OUTPUT_PATH}")

    return {
        "analysis_start": start,
        "analysis_end": end,
        "entry_bookings": entry_bookings,
        "exit_bookings": exit_bookings,
        "actual_entries": actual_entries,
        "actual_exits": actual_exits,
        "overall_summary": overall_summary,
        "monthly_entry": monthly_entry,
        "monthly_exit": monthly_exit,
        "monthly_implied_exit":
            monthly_implied_exit,
        "recommended_arrays":
            recommended_arrays,
        "duration_overall": duration_overall,
        "duration_monthly": duration_monthly,
        "entry_leadtime_distribution":
            entry_leadtime_distribution,
        "exit_leadtime_distribution":
            exit_leadtime_distribution,
        "entry_reconciliation":
            entry_reconciliation,
        "exit_reconciliation":
            exit_reconciliation,
        "entry_daily_detail":
            overall_outputs[
                "entry_daily_detail"
            ],
        "exit_daily_detail":
            overall_outputs[
                "exit_daily_detail"
            ],
        "implied_exit_daily_detail":
            overall_outputs[
                "implied_exit_daily_detail"
            ],
    }


# ---------------------------------------------------------------------
# Script entry point
# ---------------------------------------------------------------------
if __name__ == "__main__":
    outputs = run_fastpark_uplift_analysis(
        start=None,
        end=None,
        number_of_completed_months=13,
        include_partial_window_month=True,
        deduplicate=True,
    )