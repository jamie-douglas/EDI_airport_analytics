import pandas as pd
from pathlib import Path
from dateutil.relativedelta import relativedelta
from datetime import datetime
import re


# ============================================================
# CONFIG
# ============================================================

# Put your two input Excel files here.
# The script treats the first file as Day 1 and the second as Day 2.
#
# Example:
#   071326.xlsx
#   071426.xlsx
#
INPUT_FILES = [
    "scripts/071326.xlsx",
    "scripts/071426.xlsx",
]

OUTPUT_FILE = "combined_anonymised_parking_data.xlsx"

REQUIRED_SHEETS = ["Onsite", "Due In"]

BOOKING_REF_COLUMN = "Booking Ref"

# UK-style dates such as 13/07/2026
DAYFIRST = True


# ============================================================
# DATE ANONYMISATION CONFIG
# ============================================================

# Move all dates by the same amount.
#
# Example:
#   SHIFT_MONTHS = -9
#   SHIFT_DAYS = -4
#
# This would move:
#   13/07/2026 03:30
# to approximately:
#   09/10/2025 03:30
#
# The timestamp/time of day is preserved.
SHIFT_MONTHS = -9
SHIFT_DAYS = -4


# Date columns to shift in each sheet
DATE_COLUMNS_BY_SHEET = {
    "Onsite": [
        "Check-in Date",
        "Expected Return",
        "Last Vehicle Movement",
    ],
    "Due In": [
        "Expected Arrival",
        "Expected Return",
    ],
}


# ============================================================
# BOOKING REF ANONYMISATION CONFIG
# ============================================================

# This is a fixed substitution code.
#
# Letters:
#   A becomes J
#   B becomes K
#   C becomes L
#   ...
#
# Numbers:
#   0 becomes 6
#   1 becomes 7
#   2 becomes 8
#   ...
#
# Because this is character-based, the same original booking reference
# always becomes the same anonymised booking reference across all files,
# all sheets and all days.

LETTER_MAP = {
    "A": "J",
    "B": "K",
    "C": "L",
    "D": "M",
    "E": "N",
    "F": "O",
    "G": "P",
    "H": "Q",
    "I": "R",
    "J": "S",
    "K": "T",
    "L": "U",
    "M": "V",
    "N": "W",
    "O": "X",
    "P": "Y",
    "Q": "Z",
    "R": "A",
    "S": "B",
    "T": "C",
    "U": "D",
    "V": "E",
    "W": "F",
    "X": "G",
    "Y": "H",
    "Z": "I",
}

NUMBER_MAP = {
    "0": "6",
    "1": "7",
    "2": "8",
    "3": "9",
    "4": "0",
    "5": "1",
    "6": "2",
    "7": "3",
    "8": "4",
    "9": "5",
}

# If True, each output sheet keeps:
#   Booking Ref Original
#   Booking Ref New
#
# If False, each output sheet only keeps:
#   Booking Ref New
#
# The separate Booking Ref Mapping sheet is always exported.
KEEP_ORIGINAL_BOOKING_REF_IN_OUTPUT = True


# ============================================================
# OPTIONAL VEHICLE REG ANONYMISATION
# ============================================================

# Your Due In sheet contains Vehicle Reg.
# If you want to keep vehicle registrations unchanged, set this to False.
ANONYMISE_VEHICLE_REG = True

VEHICLE_REG_COLUMN = "Vehicle Reg"

KEEP_ORIGINAL_VEHICLE_REG_IN_OUTPUT = True


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_column_names(df):
    """
    Strip leading/trailing spaces from all column names.
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def make_safe_sheet_name(name):
    """
    Excel sheet names cannot exceed 31 characters
    """
    invalid_chars = r'[:\\/?*\[\]]'
    safe = re.sub(invalid_chars, "-", name)
    return safe[:31]


def anonymise_code(value):
    """
    Anonymise a booking reference or vehicle reg using fixed character substitution.

    This keeps the structure of the value:
    - letters remain letters
    - numbers remain numbers
    - other characters are kept as-is

    Example:
        AXJLGYYHR becomes JGSUPHHQA
        1428262503 becomes 7084828169
    """
    if pd.isna(value):
        return value

    value_str = str(value).strip()

    output_chars = []

    for char in value_str:
        upper_char = char.upper()

        if upper_char in LETTER_MAP:
            new_char = LETTER_MAP[upper_char]

            if char.islower():
                new_char = new_char.lower()

            output_chars.append(new_char)

        elif char in NUMBER_MAP:
            output_chars.append(NUMBER_MAP[char])

        else:
            output_chars.append(char)

    return "".join(output_chars)


def parse_uk_datetime_value(value, column_name, file_name, sheet_name):
    """
    Parses a datetime value from the input sheets using UK date logic.

    This handles two cases:

    1. Normal UK text dates:
       13/07/2026 11:00
       8/7/2026 13:40

    2. Excel/pandas already-converted datetime values:
       2026-08-07 13:40:00

       These can be wrong if Excel interpreted 8/7/2026 as 7 August
       instead of 8 July. For ambiguous dates where both the month and day
       are <= 12, this function swaps them back to UK day/month meaning.
    """

    if pd.isna(value):
        return pd.NaT

    value_str = str(value).strip()

    if value_str == "":
        return pd.NaT

    # ------------------------------------------------------------
    # Case 1: Excel/pandas datetime object
    # ------------------------------------------------------------

    if isinstance(value, (pd.Timestamp, datetime)):
        dt = pd.Timestamp(value)

        year = dt.year
        month = dt.month
        day = dt.day

        # If both month and day are <= 12, assume Excel may have read
        # a UK dd/mm date as mm/dd and swap back.
        if month <= 12 and day <= 12:
            dt = pd.Timestamp(
                year=year,
                month=day,
                day=month,
                hour=dt.hour,
                minute=dt.minute,
                second=dt.second
            )

        return dt

    # ------------------------------------------------------------
    # Case 2: ISO-style string created from an Excel datetime
    # Example: 2026-08-07 13:40:00
    # ------------------------------------------------------------

    if "-" in value_str and value_str[0:4].isdigit():
        try:
            dt = pd.to_datetime(
                value_str,
                errors="raise"
            )

            year = dt.year
            month = dt.month
            day = dt.day

            # Swap ambiguous Excel-parsed dates back to UK meaning.
            # Example:
            #   2026-08-07 13:40:00
            # becomes:
            #   2026-07-08 13:40:00
            if month <= 12 and day <= 12:
                dt = pd.Timestamp(
                    year=year,
                    month=day,
                    day=month,
                    hour=dt.hour,
                    minute=dt.minute,
                    second=dt.second
                )

            return dt

        except Exception as exc:
            raise ValueError(
                f"Could not parse ISO-style datetime.\n"
                f"File: {file_name}\n"
                f"Sheet: {sheet_name}\n"
                f"Column: {column_name}\n"
                f"Value: {value_str}\n"
                f"Error: {exc}"
            )

    # ------------------------------------------------------------
    # Case 3: UK slash-format strings
    # Examples:
    #   13/07/2026 11:00
    #   8/7/2026 13:40
    #   8/7/2026 13:40:00
    # ------------------------------------------------------------

    possible_formats = [
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%y %H:%M",
        "%d/%m/%y %H:%M:%S",
    ]

    for date_format in possible_formats:
        try:
            return pd.Timestamp(
                datetime.strptime(value_str, date_format)
            )
        except ValueError:
            continue

    # Final fallback using pandas with dayfirst=True
    # This is only used if the explicit formats above fail.
    try:
        return pd.to_datetime(
            value_str,
            dayfirst=True,
            errors="raise"
        )
    except Exception as exc:
        raise ValueError(
            f"Invalid datetime value found.\n"
            f"File: {file_name}\n"
            f"Sheet: {sheet_name}\n"
            f"Column: {column_name}\n"
            f"Value: {value_str}\n"
            f"Error: {exc}"
        )


def validate_datetime_column(df, column_name, file_name, sheet_name):
    """
    Validates that all non-blank values in a datetime column can be parsed.

    Returns a parsed datetime Series.

    This version fixes the issue where Excel/pandas has already converted
    ambiguous UK dates into ISO-style strings incorrectly.

    Example:

        Input shown as:
            8/7/2026 13:40

        May be read as:
            2026-08-07 13:40:00

        This function corrects it back to:
            2026-07-08 13:40:00
    """

    if column_name not in df.columns:
        raise ValueError(
            f"Missing required date column '{column_name}' "
            f"in sheet '{sheet_name}' of file '{file_name}'"
        )

    converted_values = []

    for value in df[column_name]:
        converted_value = parse_uk_datetime_value(
            value=value,
            column_name=column_name,
            file_name=file_name,
            sheet_name=sheet_name
        )

        converted_values.append(converted_value)

    return pd.Series(
        converted_values,
        index=df.index,
        dtype="datetime64[ns]"
    )


def validate_timestamp_preservation(original_dates, shifted_dates, column_name, file_name, sheet_name):
    """
    Validates that the time part did not change after shifting.

    For example:
        13/07/2026 03:30
    may become:
        09/10/2025 03:30

    The date changes, but 03:30 must remain 03:30.
    """
    valid_mask = original_dates.notna() & shifted_dates.notna()

    original_valid = original_dates.loc[valid_mask]
    shifted_valid = shifted_dates.loc[valid_mask]

    original_time = (
        original_valid.dt.hour * 3600
        + original_valid.dt.minute * 60
        + original_valid.dt.second
    )

    shifted_time = (
        shifted_valid.dt.hour * 3600
        + shifted_valid.dt.minute * 60
        + shifted_valid.dt.second
    )

    mismatch_mask = original_time != shifted_time
    mismatch_count = int(mismatch_mask.sum())

    if mismatch_count > 0:
        mismatch_examples = pd.DataFrame({
            "Original Date": original_valid.loc[mismatch_mask].head(10),
            "Shifted Date": shifted_valid.loc[mismatch_mask].head(10),
        })

        raise ValueError(
            f"Timestamp changed during date shifting.\n"
            f"File: {file_name}\n"
            f"Sheet: {sheet_name}\n"
            f"Column: {column_name}\n"
            f"Mismatch count: {mismatch_count}\n"
            f"First mismatch examples:\n{mismatch_examples}"
        )


def shift_datetime_value(value):
    """
    Shift a single datetime by configured months and days.
    Preserves the time of day.
    """
    if pd.isna(value):
        return value

    return value + relativedelta(
        months=SHIFT_MONTHS,
        days=SHIFT_DAYS
    )


def shift_date_columns(df, sheet_name, file_name):
    """
    Applies the configured month/day date shift to the relevant columns.

    Also validates:
    - the dates are parseable before shifting
    - the timestamp is preserved after shifting
    """
    df = df.copy()

    date_columns = DATE_COLUMNS_BY_SHEET.get(sheet_name, [])

    for col in date_columns:
        original_dates = validate_datetime_column(
            df=df,
            column_name=col,
            file_name=file_name,
            sheet_name=sheet_name
        )

        print("\n")
        print("COLUMN:", col)
        print(df[[col]].head(10))
        print(original_dates.head(10))

        shifted_dates = original_dates.apply(shift_datetime_value)

        validate_timestamp_preservation(
            original_dates=original_dates,
            shifted_dates=shifted_dates,
            column_name=col,
            file_name=file_name,
            sheet_name=sheet_name
        )

        df[col] = shifted_dates

    return df


def anonymise_booking_ref_column(df, sheet_name, file_name):
    """
    Replaces Booking Ref with anonymised code.
    Optionally keeps the original Booking Ref for validation.
    """
    df = df.copy()

    if BOOKING_REF_COLUMN not in df.columns:
        raise ValueError(
            f"Missing required column '{BOOKING_REF_COLUMN}' "
            f"in sheet '{sheet_name}' of file '{file_name}'"
        )

    original_refs = df[BOOKING_REF_COLUMN].astype(str).str.strip()
    new_refs = original_refs.apply(anonymise_code)

    original_col_position = df.columns.get_loc(BOOKING_REF_COLUMN)

    df.insert(
        loc=original_col_position,
        column="Booking Ref Original",
        value=original_refs,
    )

    df.insert(
        loc=original_col_position + 1,
        column="Booking Ref New",
        value=new_refs,
    )

    df = df.drop(columns=[BOOKING_REF_COLUMN])

    if not KEEP_ORIGINAL_BOOKING_REF_IN_OUTPUT:
        df = df.drop(columns=["Booking Ref Original"])

    return df


def anonymise_vehicle_reg_column(df):
    """
    Anonymises Vehicle Reg in the Due In sheet if configured.
    Uses the same fixed character substitution approach.
    """
    df = df.copy()

    if not ANONYMISE_VEHICLE_REG:
        return df

    if VEHICLE_REG_COLUMN not in df.columns:
        return df

    original_regs = df[VEHICLE_REG_COLUMN].astype(str).str.strip()
    new_regs = original_regs.apply(anonymise_code)

    original_col_position = df.columns.get_loc(VEHICLE_REG_COLUMN)

    df.insert(
        loc=original_col_position,
        column="Vehicle Reg Original",
        value=original_regs,
    )

    df.insert(
        loc=original_col_position + 1,
        column="Vehicle Reg New",
        value=new_regs,
    )

    df = df.drop(columns=[VEHICLE_REG_COLUMN])

    if not KEEP_ORIGINAL_VEHICLE_REG_IN_OUTPUT:
        df = df.drop(columns=["Vehicle Reg Original"])

    return df


def process_sheet(df, sheet_name, file_name):
    """
    Full processing for one sheet:
    - clean columns
    - anonymise Booking Ref
    - anonymise Vehicle Reg if applicable
    - validate datetime columns
    - shift date columns
    - validate timestamp preservation
    """
    df = clean_column_names(df)

    df = anonymise_booking_ref_column(
        df=df,
        sheet_name=sheet_name,
        file_name=file_name
    )

    if sheet_name == "Due In":
        df = anonymise_vehicle_reg_column(df)

    df = shift_date_columns(
        df=df,
        sheet_name=sheet_name,
        file_name=file_name
    )

    return df


def collect_mapping_values(input_files):
    """
    Collects original and anonymised booking references and vehicle registrations
    for the mapping sheets.

    Because anonymisation is code-based, the mapping is not needed to generate
    the same new value again, but it is useful for validation.
    """
    booking_refs = []
    vehicle_regs = []

    for file_path in input_files:
        file_path = Path(file_path)

        for sheet_name in REQUIRED_SHEETS:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str
            )

            df = clean_column_names(df)

            if BOOKING_REF_COLUMN in df.columns:
                refs = (
                    df[BOOKING_REF_COLUMN]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                booking_refs.extend(refs)

            if sheet_name == "Due In" and VEHICLE_REG_COLUMN in df.columns:
                regs = (
                    df[VEHICLE_REG_COLUMN]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .tolist()
                )
                vehicle_regs.extend(regs)

    unique_booking_refs = sorted(set([x for x in booking_refs if x != ""]))
    unique_vehicle_regs = sorted(set([x for x in vehicle_regs if x != ""]))

    booking_mapping_df = pd.DataFrame({
        "Booking Ref Original": unique_booking_refs,
        "Booking Ref New": [anonymise_code(x) for x in unique_booking_refs],
    })

    vehicle_mapping_df = pd.DataFrame({
        "Vehicle Reg Original": unique_vehicle_regs,
        "Vehicle Reg New": [anonymise_code(x) for x in unique_vehicle_regs],
    })

    return booking_mapping_df, vehicle_mapping_df


def autosize_excel_columns(writer, sheet_name, df):
    """
    Autosize columns in the Excel output.
    """
    worksheet = writer.sheets[sheet_name]

    for idx, col in enumerate(df.columns, start=1):
        max_length = len(str(col))

        for value in df[col].head(5000):
            if pd.notna(value):
                max_length = max(max_length, len(str(value)))

        adjusted_width = min(max_length + 2, 45)

        column_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_excel_formats(writer, sheet_name, df):
    """
    Applies date/time formatting and simple header styling.
    """
    from openpyxl.styles import Font, PatternFill

    worksheet = writer.sheets[sheet_name]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    date_format = "dd/mm/yyyy hh:mm"

    date_columns = [
        "Check-in Date",
        "Expected Return",
        "Last Vehicle Movement",
        "Expected Arrival",
    ]

    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in date_columns:
            for row_idx in range(2, len(df) + 2):
                worksheet.cell(
                    row=row_idx,
                    column=col_idx
                ).number_format = date_format

    worksheet.freeze_panes = "A2"


def validate_input_files(input_files):
    """
    Validates that:
    - exactly two files have been provided
    - both files exist
    - both required sheets exist in each file
    """
    if len(input_files) != 2:
        raise ValueError("This script expects exactly two input Excel files.")

    for file_path in input_files:
        if not file_path.exists():
            raise FileNotFoundError(f"Input file not found: {file_path}")

        workbook = pd.ExcelFile(file_path, engine="openpyxl")
        available_sheets = workbook.sheet_names

        for required_sheet in REQUIRED_SHEETS:
            if required_sheet not in available_sheets:
                raise ValueError(
                    f"Missing required sheet '{required_sheet}' "
                    f"in file '{file_path.name}'. "
                    f"Available sheets are: {available_sheets}"
                )


# ============================================================
# MAIN SCRIPT
# ============================================================

def main():
    input_files = [Path(f) for f in INPUT_FILES]

    import os

    validate_input_files(input_files)

    processed_sheets = {}

    for day_number, file_path in enumerate(input_files, start=1):
        for sheet_name in REQUIRED_SHEETS:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str
            )

            processed_df = process_sheet(
                df=df,
                sheet_name=sheet_name,
                file_name=file_path.name
            )

            output_sheet_name = make_safe_sheet_name(
                f"Day {day_number} - {sheet_name}"
            )

            processed_sheets[output_sheet_name] = processed_df

    booking_mapping_df, vehicle_mapping_df = collect_mapping_values(input_files)

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
        datetime_format="dd/mm/yyyy hh:mm"
    ) as writer:

        for sheet_name, df in processed_sheets.items():
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            autosize_excel_columns(
                writer=writer,
                sheet_name=sheet_name,
                df=df
            )

            apply_excel_formats(
                writer=writer,
                sheet_name=sheet_name,
                df=df
            )

        booking_mapping_sheet_name = "Booking Ref Mapping"

        booking_mapping_df.to_excel(
            writer,
            sheet_name=booking_mapping_sheet_name,
            index=False
        )

        autosize_excel_columns(
            writer=writer,
            sheet_name=booking_mapping_sheet_name,
            df=booking_mapping_df
        )

        apply_excel_formats(
            writer=writer,
            sheet_name=booking_mapping_sheet_name,
            df=booking_mapping_df
        )

        if ANONYMISE_VEHICLE_REG:
            vehicle_mapping_sheet_name = "Vehicle Reg Mapping"

            vehicle_mapping_df.to_excel(
                writer,
                sheet_name=vehicle_mapping_sheet_name,
                index=False
            )

            autosize_excel_columns(
                writer=writer,
                sheet_name=vehicle_mapping_sheet_name,
                df=vehicle_mapping_df
            )

            apply_excel_formats(
                writer=writer,
                sheet_name=vehicle_mapping_sheet_name,
                df=vehicle_mapping_df
            )

    print("Complete.")
    print(f"Output written to: {OUTPUT_FILE}")
    print(f"Date shift applied: {SHIFT_MONTHS} months and {SHIFT_DAYS} days")
    print(f"Unique booking refs anonymised: {len(booking_mapping_df)}")

    if ANONYMISE_VEHICLE_REG:
        print(f"Unique vehicle regs anonymised: {len(vehicle_mapping_df)}")

    print("Datetime validation passed.")
    print("Timestamp preservation validation passed.")



if __name__ == "__main__":

    main()