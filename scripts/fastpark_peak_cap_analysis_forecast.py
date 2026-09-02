"""
fastpark_peak_cap_analysis_forecast.py
======================================

Standalone FastPark peak-cap analysis and six-week forecast scenario.

HISTORICAL ANALYSIS
For vehicles physically arriving in 03:00-05:59 or 10:00-13:59, the script
classifies the booked entry slot as EARLIER than the peak, INSIDE the peak, or
LATER than the peak. It reports the result by actual hour, booked hour,
weekday, duration band and relative price-per-day quintile. Price quintiles
are calculated within duration band so that a cheap long stay is not compared
naively with an expensive short stay.

The analysis also derives lower-price behaviour inputs from Q1, the cheapest
20 percent within each duration band. These inputs show:
  * the lower-price share booking outside the peak but arriving in it; and
  * where lower-price earlier-slot bookings were placed, particularly the
    one, two and three hours preceding each peak.

FORECAST
The forecast reproduces the existing operational baseline logic:
  * previous two completed weeks of actual entries, exits and passengers;
  * weekday entry and exit penetration rates;
  * six weeks of future passenger flights;
  * current FastPark on-books totals for the first 14 forecast dates;
  * the existing first-14-day uplifts and September multiplier; and
  * weekday/hour entry and exit profiles from those same two completed weeks.

For each day, each complete peak is capped at 350 booked entries. Volume above
the cap is divided into:
  * lost FastPark bookings, using an explicit scenario loss assumption;
  * customers booking a cheaper earlier slot but still arriving in the peak,
    informed by lower-price historical behaviour; and
  * retained demand physically shifted outside the peak, allocated across the
    first, second and third preceding hours using lower-price historical
    destination weights.

Lost entries remove future exits using the observed peak-specific booking
duration distribution. The script writes baseline versus changed outputs at
both hourly and daily level for six weeks. It writes only to its own output
folder and never updates the operational DWH, Workforce or accuracy files.

Run modes
---------
Analysis only:
    RUN_HISTORICAL_ANALYSIS = True
    RUN_FORECAST_SCENARIOS = False

Analysis and forecast:
    RUN_HISTORICAL_ANALYSIS = True
    RUN_FORECAST_SCENARIOS = True

Forecast from previously saved analysis outputs:
    RUN_HISTORICAL_ANALYSIS = False
    RUN_FORECAST_SCENARIOS = True
"""

from __future__ import annotations

from pathlib import Path
import numpy as np
import pandas as pd
import pyodbc

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURATION
# =============================================================================

RUN_HISTORICAL_ANALYSIS = True
RUN_FORECAST_SCENARIOS = True

DSN = "AzureConnection"
ACTUAL_ARRIVAL_COLUMN = "CheckInStarted"  # CheckInEnded may be used instead.

ANALYSIS_START_DATE = "2024-01-01"
ANALYSIS_END_DATE = "2026-08-27"  # Exclusive. Update when rerunning later.

FORECAST_WEEKS = 6
PROFILE_HISTORY_WEEKS = 2
PEAK_CAP = 350.0
PEAKS = {
    "peak_03_06": (3, 6),
    "peak_10_14": (10, 14),
}

PRICE_QUANTILES = 5
DURATION_BINS = [-np.inf, 1, 3, 7, 10, 14, 21, np.inf]
DURATION_LABELS = [
    "0-1 days",
    "2-3 days",
    "4-7 days",
    "8-10 days",
    "11-14 days",
    "15-21 days",
    "22+ days",
]
MAX_DURATION_DAYS = 60

# Scenario assumptions apply only to forecast entry volume above 350.
# lost_rate is explicit because a cap response cannot be observed historically.
# circumvention_multiplier scales the Q1 lower-price historical share that
# booked outside the peak but still arrived inside it.
SCENARIOS = {
    "LOW_LOSS": {
        "lost_rate": 0.20,
        "circumvention_multiplier": 1.0,
    },
    "CENTRAL": {
        "lost_rate": 0.40,
        "circumvention_multiplier": 1.5,
    },
    "HIGH_LOSS": {
        "lost_rate": 0.60,
        "circumvention_multiplier": 2.0,
    },
}

# Upper bound prevents an implausibly large historical multiplier from consuming
# all retained excess. The remainder is physically shifted earlier.
MAX_CIRCUMVENTION_RATE_OF_EXCESS = 0.60

# Preserve the current operational forecast adjustments.
APPLY_CURRENT_MANUAL_UPLIFTS = True
AUG_ENTRY = [5, 10, 17.9, 24.9, 31.4, 38, 44.6, 51.3, 57.3, 62.9, 68.5, 74.3, 79.9, 85.49]
AUG_EXIT = [1, 10.2, 17.9, 24.7, 31.2, 37.8, 43.9, 50, 55.3, 60.2, 64.01, 68.7, 72.9, 77.32]
ENTRY_INC = [5, 10, 15, 20, 25, 30, 35, 38, 41, 44, 47, 50, 53, 56]
EXIT_INC = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
SEPTEMBER_MULTIPLIER = 1.28

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "fastpark_peak_cap_outputs"
ANALYSIS_DIR = OUTPUT_DIR / "analysis"
FORECAST_DIR = OUTPUT_DIR / "forecast"


# =============================================================================
# GENERAL HELPERS
# =============================================================================

def stamp(value) -> str:
    return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")


def connect_db():
    return pyodbc.connect(f"DSN={DSN};")


def relation_to_window(booked, actual, start_hour, end_hour) -> pd.Series:
    start = actual.dt.normalize() + pd.to_timedelta(start_hour, unit="h")
    end = actual.dt.normalize() + pd.to_timedelta(end_hour, unit="h")
    labels = np.select(
        [booked < start, (booked >= start) & (booked < end), booked >= end],
        ["booked_earlier_slot", "booked_inside_window", "booked_later_slot"],
        default="unknown",
    )
    return pd.Series(labels, index=booked.index)


def rounded_profile(values: pd.Series, total: float) -> pd.Series:
    values = values.fillna(0).clip(lower=0).astype(float)
    integer_total = int(round(float(total)))
    if values.sum() <= 0:
        values[:] = 1 / len(values)
    raw = values / values.sum() * integer_total
    result = np.floor(raw).astype(int)
    remainder = integer_total - int(result.sum())
    if remainder > 0:
        order = (raw - result).sort_values(ascending=False).index[:remainder]
        result.loc[order] += 1
    return result


def add_volume(df: pd.DataFrame, timestamp, column: str, amount: float) -> None:
    timestamp = pd.Timestamp(timestamp)
    if timestamp not in df.index:
        df.loc[timestamp, column] = 0.0
    df.loc[timestamp, column] += float(amount)


# =============================================================================
# DATA EXTRACTION
# =============================================================================

def historical_bookings(connection, start, end) -> pd.DataFrame:
    return pd.read_sql(
        f"""
        SELECT bookingId, createdAt, entryDate, exitDate, status, assetName,
               bookingTotal, productPrice, productQuantity, duration
        FROM AirportX.v_Bookings
        WHERE assetName = 'FastPark'
          AND entryDate >= '{stamp(start)}'
          AND entryDate <  '{stamp(end)}'
        """,
        connection,
    )


def operations(connection, start, end) -> pd.DataFrame:
    return pd.read_sql(
        f"""
        SELECT BookingReference, CheckInStarted, CheckInEnded,
               ActualCheckedOutDate, RecordUpdatedDateTime
        FROM FastPark.v_EntryAndExits
        WHERE ({ACTUAL_ARRIVAL_COLUMN} >= '{stamp(start)}'
           AND {ACTUAL_ARRIVAL_COLUMN} <  '{stamp(end)}')
           OR (ActualCheckedOutDate >= '{stamp(start)}'
           AND ActualCheckedOutDate <  '{stamp(end)}')
        """,
        connection,
    )


def historical_flights(connection, start, end) -> pd.DataFrame:
    return pd.read_sql(
        f"""
        SELECT ActualDateTime_Local, ArrDeptureCode,
               IsPassengerFlight, Pax_MostConfident
        FROM EAL.FlightPerformance
        WHERE ActualDateTime_Local >= '{stamp(start)}'
          AND ActualDateTime_Local <  '{stamp(end)}'
        """,
        connection,
    )


def future_flights(connection, start, end) -> pd.DataFrame:
    return pd.read_sql(
        f"""
        SELECT ScheduledDateTime_Local, ArrDeptureCode,
               IsPassengerFlight, Pax_MostConfident
        FROM EAL.FlightPerformance_FutureFlights
        WHERE ScheduledDateTime_Local >= '{stamp(start)}'
          AND ScheduledDateTime_Local <  '{stamp(end)}'
        """,
        connection,
    )


def future_bookings(connection, start, end) -> pd.DataFrame:
    return pd.read_sql(
        f"""
        SELECT bookingId, entryDate, exitDate
        FROM AirportX.v_Bookings
        WHERE assetName = 'FastPark'
          AND status = 'B'
          AND (
                (entryDate >= '{stamp(start)}' AND entryDate < '{stamp(end)}')
             OR (exitDate  >= '{stamp(start)}' AND exitDate  < '{stamp(end)}')
          )
        """,
        connection,
    )


# =============================================================================
# HISTORICAL ANALYSIS
# =============================================================================

def build_movement_fact(connection) -> pd.DataFrame:
    bookings = historical_bookings(connection, ANALYSIS_START_DATE, ANALYSIS_END_DATE)
    actuals = operations(connection, ANALYSIS_START_DATE, ANALYSIS_END_DATE)

    for column in ["createdAt", "entryDate", "exitDate"]:
        bookings[column] = pd.to_datetime(bookings[column], errors="coerce")
    for column in ["CheckInStarted", "CheckInEnded", "ActualCheckedOutDate", "RecordUpdatedDateTime"]:
        actuals[column] = pd.to_datetime(actuals[column], errors="coerce")

    actuals["has_arrival"] = actuals[ACTUAL_ARRIVAL_COLUMN].notna().astype(int)
    actuals = (
        actuals.sort_values(["BookingReference", "has_arrival", "RecordUpdatedDateTime"])
        .drop_duplicates("BookingReference", keep="last")
    )

    fact = bookings.merge(actuals, left_on="bookingId", right_on="BookingReference", how="left")
    fact = fact.loc[
        fact["status"].eq("B")
        & fact[ACTUAL_ARRIVAL_COLUMN].notna()
        & fact["entryDate"].notna()
        & fact["exitDate"].notna()
    ].copy()

    fact["booked_entry_ts"] = fact["entryDate"]
    fact["actual_arrival_ts"] = fact[ACTUAL_ARRIVAL_COLUMN]
    fact["booked_entry_hour"] = fact["booked_entry_ts"].dt.hour
    fact["actual_arrival_hour"] = fact["actual_arrival_ts"].dt.hour
    fact["weekday_num"] = fact["actual_arrival_ts"].dt.weekday
    fact["weekday"] = fact["actual_arrival_ts"].dt.day_name()
    fact["duration_days"] = (fact["exitDate"] - fact["entryDate"]).dt.total_seconds() / 86400
    fact["duration_rounded"] = fact["duration_days"].round().clip(0, MAX_DURATION_DAYS)
    fact["duration_band"] = pd.cut(fact["duration_days"], DURATION_BINS, labels=DURATION_LABELS)
    fact["price_per_day"] = fact["bookingTotal"] / fact["duration_days"].replace(0, np.nan)
    fact["price_quintile"] = pd.NA

    eligible = fact["price_per_day"].notna() & fact["duration_band"].notna()
    for _, indexes in fact.loc[eligible].groupby("duration_band", observed=True).groups.items():
        if fact.loc[indexes, "price_per_day"].nunique() >= PRICE_QUANTILES:
            fact.loc[indexes, "price_quintile"] = pd.qcut(
                fact.loc[indexes, "price_per_day"].rank(method="first"),
                PRICE_QUANTILES,
                labels=[f"Q{i}_low_to_high" for i in range(1, PRICE_QUANTILES + 1)],
            )

    for peak_name, (start_hour, end_hour) in PEAKS.items():
        fact[f"{peak_name}_actual"] = fact["actual_arrival_hour"].between(start_hour, end_hour - 1)
        fact[f"{peak_name}_relation"] = relation_to_window(
            fact["booked_entry_ts"], fact["actual_arrival_ts"], start_hour, end_hour
        )
    return fact


def derive_lower_price_parameters(fact: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    parameter_rows = []
    weight_rows = []

    for peak_name, (start_hour, _) in PEAKS.items():
        peak = fact.loc[
            fact[f"{peak_name}_actual"] & fact["price_quintile"].eq("Q1_low_to_high")
        ].copy()
        total = peak["bookingId"].nunique()
        outside = peak[f"{peak_name}_relation"].isin(
            ["booked_earlier_slot", "booked_later_slot"]
        )
        outside_count = peak.loc[outside, "bookingId"].nunique()
        lower_price_outside_share = outside_count / total if total else 0.0

        parameter_rows.append(
            {
                "peak": peak_name,
                "lower_price_peak_bookings": total,
                "lower_price_booked_outside_peak": outside_count,
                "lower_price_outside_peak_share": lower_price_outside_share,
            }
        )

        earlier = peak.loc[peak[f"{peak_name}_relation"].eq("booked_earlier_slot")].copy()
        earlier["hours_before_peak"] = (
            start_hour - earlier["booked_entry_hour"]
        ) % 24
        earlier = earlier.loc[earlier["hours_before_peak"].between(1, 3)]
        counts = earlier.groupby("hours_before_peak")["bookingId"].nunique()

        if counts.sum() == 0:
            counts = pd.Series({1: 1.0, 2: 0.0, 3: 0.0})
            source = "fallback_previous_hour"
        else:
            source = "lower_price_history"

        for hours_before in [1, 2, 3]:
            count = float(counts.get(hours_before, 0))
            weight_rows.append(
                {
                    "peak": peak_name,
                    "hours_before_peak": hours_before,
                    "bookings": count,
                    "weight": count / counts.sum(),
                    "source": source,
                }
            )

    return pd.DataFrame(parameter_rows), pd.DataFrame(weight_rows)


def run_analysis(connection) -> dict[str, pd.DataFrame]:
    fact = build_movement_fact(connection)
    summary_rows = []
    matrices = []
    price_tables = []
    weekday_tables = []

    for peak_name in PEAKS:
        peak = fact.loc[fact[f"{peak_name}_actual"]].copy()
        total = peak["bookingId"].nunique()

        for slot_relation in ["booked_earlier_slot", "booked_inside_window", "booked_later_slot"]:
            count = peak.loc[peak[f"{peak_name}_relation"].eq(slot_relation), "bookingId"].nunique()
            summary_rows.append(
                {
                    "actual_window": peak_name,
                    "slot_relation": slot_relation,
                    "bookings": count,
                    "share": count / total if total else np.nan,
                }
            )

        matrix = (
            peak.groupby(["actual_arrival_hour", "booked_entry_hour"])["bookingId"]
            .nunique().rename("bookings").reset_index()
        )
        matrix["share_within_actual_hour"] = (
            matrix["bookings"] / matrix.groupby("actual_arrival_hour")["bookings"].transform("sum")
        )
        matrix.insert(0, "actual_window", peak_name)
        matrices.append(matrix)

        price = (
            peak.groupby(
                ["duration_band", "price_quintile", f"{peak_name}_relation"],
                observed=True,
                dropna=False,
            )
            .agg(
                bookings=("bookingId", "nunique"),
                avg_price_per_day=("price_per_day", "mean"),
                median_price_per_day=("price_per_day", "median"),
            )
            .reset_index()
            .rename(columns={f"{peak_name}_relation": "slot_relation"})
        )
        price["share_within_duration_price"] = (
            price["bookings"]
            / price.groupby(["duration_band", "price_quintile"], observed=True)["bookings"].transform("sum")
        )
        price.insert(0, "actual_window", peak_name)
        price_tables.append(price)

        weekday = (
            peak.groupby(["weekday_num", "weekday", f"{peak_name}_relation"])["bookingId"]
            .nunique().rename("bookings").reset_index()
            .rename(columns={f"{peak_name}_relation": "slot_relation"})
        )
        weekday["share_within_weekday"] = (
            weekday["bookings"]
            / weekday.groupby(["weekday_num", "weekday"])["bookings"].transform("sum")
        )
        weekday.insert(0, "actual_window", peak_name)
        weekday_tables.append(weekday)

    duration_rows = []
    for peak_name in PEAKS:
        subset = fact.loc[fact[f"{peak_name}_actual"] & fact["duration_rounded"].notna()]
        counts = subset.groupby(subset["duration_rounded"].astype(int))["bookingId"].nunique()
        for days, count in counts.items():
            duration_rows.append(
                {
                    "peak": peak_name,
                    "duration_days": int(days),
                    "bookings": int(count),
                    "weight": count / counts.sum(),
                }
            )

    low_price_parameters, low_price_shift_weights = derive_lower_price_parameters(fact)

    outputs = {
        "peak_slot_summary": pd.DataFrame(summary_rows),
        "actual_vs_booked_hour_matrix": pd.concat(matrices, ignore_index=True),
        "peak_slot_by_price_duration": pd.concat(price_tables, ignore_index=True),
        "peak_slot_by_weekday": pd.concat(weekday_tables, ignore_index=True),
        "lower_price_behaviour_parameters": low_price_parameters,
        "lower_price_shift_weights": low_price_shift_weights,
        "peak_duration_distribution": pd.DataFrame(duration_rows),
    }

    for name, table in outputs.items():
        table.to_csv(ANALYSIS_DIR / f"{name}.csv", index=False)
    return outputs


def load_analysis() -> dict[str, pd.DataFrame]:
    names = [
        "lower_price_behaviour_parameters",
        "lower_price_shift_weights",
        "peak_duration_distribution",
    ]
    return {name: pd.read_csv(ANALYSIS_DIR / f"{name}.csv") for name in names}


# =============================================================================
# CURRENT FORECAST LOGIC
# =============================================================================

def weekday_rate(actuals, flights, direction: str, timestamp_column: str) -> pd.Series:
    actual = actuals.dropna(subset=[timestamp_column]).drop_duplicates("BookingReference").copy()
    actual["weekday_num"] = actual[timestamp_column].dt.weekday
    relevant = flights.loc[
        flights["ArrDeptureCode"].eq(direction) & flights["IsPassengerFlight"].eq(1)
    ].copy()
    relevant["weekday_num"] = relevant["ActualDateTime_Local"].dt.weekday
    rates = (
        actual.groupby("weekday_num")["BookingReference"].nunique()
        / relevant.groupby("weekday_num")["Pax_MostConfident"].sum().replace(0, np.nan)
    )
    fallback = actual["BookingReference"].nunique() / relevant["Pax_MostConfident"].sum()
    return rates.reindex(range(7)).fillna(fallback)


def hour_profile(actuals, timestamp_column: str) -> pd.DataFrame:
    actual = actuals.dropna(subset=[timestamp_column]).drop_duplicates("BookingReference").copy()
    actual["weekday_num"] = actual[timestamp_column].dt.weekday
    actual["hour"] = actual[timestamp_column].dt.hour
    counts = actual.groupby(["weekday_num", "hour"])["BookingReference"].nunique()
    index = pd.MultiIndex.from_product([range(7), range(24)], names=["weekday_num", "hour"])
    counts = counts.reindex(index, fill_value=0)
    shares = counts / counts.groupby(level=0).transform("sum").replace(0, np.nan)
    return shares.fillna(1 / 24).rename("share").reset_index()


def expand_daily_to_hourly(daily: pd.Series, profile: pd.DataFrame, output_column: str) -> pd.DataFrame:
    lookup = profile.set_index(["weekday_num", "hour"])["share"]
    rows = []
    for date, total in daily.items():
        date = pd.Timestamp(date).normalize()
        shares = pd.Series({hour: lookup.get((date.weekday(), hour), 0) for hour in range(24)})
        allocation = rounded_profile(shares, total)
        rows.extend(
            {
                "datetime": date + pd.Timedelta(hours=int(hour)),
                output_column: int(value),
            }
            for hour, value in allocation.items()
        )
    return pd.DataFrame(rows).set_index("datetime").sort_index()


def baseline_forecast(connection):
    today = pd.Timestamp.today().normalize()
    start_of_this_week = today - pd.Timedelta(days=today.weekday())
    history_start = start_of_this_week - pd.Timedelta(weeks=PROFILE_HISTORY_WEEKS)
    future_end = today + pd.Timedelta(weeks=FORECAST_WEEKS)

    history_flights = historical_flights(connection, history_start, start_of_this_week)
    forecast_flights = future_flights(connection, today, future_end)
    bookings = future_bookings(connection, today, future_end)
    actuals = operations(connection, history_start, start_of_this_week)

    history_flights["ActualDateTime_Local"] = pd.to_datetime(
        history_flights["ActualDateTime_Local"], errors="coerce"
    )
    forecast_flights["ScheduledDateTime_Local"] = pd.to_datetime(
        forecast_flights["ScheduledDateTime_Local"], errors="coerce"
    )
    for column in ["CheckInStarted", "CheckInEnded", "ActualCheckedOutDate"]:
        actuals[column] = pd.to_datetime(actuals[column], errors="coerce")
    bookings["entryDate"] = pd.to_datetime(bookings["entryDate"], errors="coerce")
    bookings["exitDate"] = pd.to_datetime(bookings["exitDate"], errors="coerce")

    entry_rate = weekday_rate(actuals, history_flights, "D", ACTUAL_ARRIVAL_COLUMN)
    exit_rate = weekday_rate(actuals, history_flights, "A", "ActualCheckedOutDate")

    departures = forecast_flights.loc[
        forecast_flights["ArrDeptureCode"].eq("D")
        & forecast_flights["IsPassengerFlight"].eq(1)
    ].copy()
    arrivals = forecast_flights.loc[
        forecast_flights["ArrDeptureCode"].eq("A")
        & forecast_flights["IsPassengerFlight"].eq(1)
    ].copy()
    departures["date"] = departures["ScheduledDateTime_Local"].dt.normalize()
    arrivals["date"] = arrivals["ScheduledDateTime_Local"].dt.normalize()
    departing_pax = departures.groupby("date")["Pax_MostConfident"].sum()
    arriving_pax = arrivals.groupby("date")["Pax_MostConfident"].sum()

    dates = pd.date_range(today, future_end - pd.Timedelta(days=1), freq="D")
    daily = pd.DataFrame(index=dates)
    daily["entry_forecast"] = (
        departing_pax
        * pd.Series(departing_pax.index.weekday, index=departing_pax.index).map(entry_rate)
    ).reindex(dates)
    daily["exit_forecast"] = (
        arriving_pax
        * pd.Series(arriving_pax.index.weekday, index=arriving_pax.index).map(exit_rate)
    ).reindex(dates)

    daily["booking_entries"] = (
        bookings.groupby(bookings["entryDate"].dt.normalize())["bookingId"]
        .nunique().reindex(dates).fillna(0)
    )
    daily["booking_exits"] = (
        bookings.groupby(bookings["exitDate"].dt.normalize())["bookingId"]
        .nunique().reindex(dates).fillna(0)
    )

    near_term = daily.index < today + pd.Timedelta(days=14)
    daily.loc[near_term, "entry_forecast"] = daily.loc[near_term, "booking_entries"]
    daily.loc[near_term, "exit_forecast"] = daily.loc[near_term, "booking_exits"]

    if APPLY_CURRENT_MANUAL_UPLIFTS:
        for position in range(min(14, len(daily))):
            row_date = daily.index[position]
            if row_date.year == 2026 and row_date.month == 8:
                entry_pct = AUG_ENTRY[position]
                exit_pct = AUG_EXIT[position]
            else:
                entry_pct = ENTRY_INC[position]
                exit_pct = EXIT_INC[position]
            daily.iloc[position, daily.columns.get_loc("entry_forecast")] *= 1 + entry_pct / 100
            daily.iloc[position, daily.columns.get_loc("exit_forecast")] *= 1 + exit_pct / 100

        daily.loc[daily.index.month == 9, ["entry_forecast", "exit_forecast"]] *= (
            SEPTEMBER_MULTIPLIER
        )

    daily[["entry_forecast", "exit_forecast"]] = (
        daily[["entry_forecast", "exit_forecast"]].fillna(0).round().astype(int)
    )

    entry_profile = hour_profile(actuals, ACTUAL_ARRIVAL_COLUMN)
    exit_profile = hour_profile(actuals, "ActualCheckedOutDate")
    hourly = expand_daily_to_hourly(
        daily["entry_forecast"], entry_profile, "entry_forecast"
    ).join(
        expand_daily_to_hourly(daily["exit_forecast"], exit_profile, "exit_forecast")
    )
    return daily, hourly, entry_profile, exit_profile


# =============================================================================
# CAP FORECAST
# =============================================================================

def lost_exit_detail(cap_summary, duration_distribution) -> pd.DataFrame:
    rows = []
    for _, cap_row in cap_summary.loc[cap_summary["lost_fastpark_entries"] > 0].iterrows():
        weights = duration_distribution.loc[
            duration_distribution["peak"].eq(cap_row["peak"])
        ].copy()
        weights["weight"] = weights["weight"] / weights["weight"].sum()
        for _, duration_row in weights.iterrows():
            days = int(duration_row["duration_days"])
            rows.append(
                {
                    "scenario": cap_row["scenario"],
                    "entry_date": pd.Timestamp(cap_row["date"]),
                    "peak": cap_row["peak"],
                    "duration_days": days,
                    "duration_weight": duration_row["weight"],
                    "exit_date": pd.Timestamp(cap_row["date"]) + pd.Timedelta(days=days),
                    "lost_exits": cap_row["lost_fastpark_entries"] * duration_row["weight"],
                }
            )
    return pd.DataFrame(rows)

def create_hourly_pivot(
    hourly_df: pd.DataFrame,
    value_column: str,
) -> pd.DataFrame:
    """
    Convert a long hourly forecast into the standard reporting format:

        Date | 00:00 | 01:00 | ... | 23:00 | Peak 03-06 |
        Peak 10-14 | Total

    Parameters
    ----------
    hourly_df:
        DataFrame indexed by hourly datetime.

    value_column:
        Column containing the hourly forecast values.

    Returns
    -------
    pandas.DataFrame
        One row per date, with one column per hour, two peak totals
        and a daily total.
    """

    df = hourly_df.copy()

    if not isinstance(df.index, pd.DatetimeIndex):
        raise TypeError(
            "hourly_df must be indexed by a pandas DatetimeIndex."
        )

    if value_column not in df.columns:
        raise KeyError(
            f"Column '{value_column}' was not found in hourly_df."
        )

    df["Date"] = df.index.normalize()
    df["Hour"] = df.index.hour

    pivot = df.pivot_table(
        index="Date",
        columns="Hour",
        values=value_column,
        aggfunc="sum",
        fill_value=0,
    )

    # Ensure all 24 hours are present.
    pivot = pivot.reindex(
        columns=range(24),
        fill_value=0,
    )

    pivot.columns = [
        f"{hour:02d}:00"
        for hour in range(24)
    ]

    # Peak totals use the exact capped periods.
    pivot["Peak 03-06"] = pivot[
        ["03:00", "04:00", "05:00"]
    ].sum(axis=1)

    pivot["Peak 10-14"] = pivot[
        ["10:00", "11:00", "12:00", "13:00"]
    ].sum(axis=1)

    # Total must include only hourly columns, not the peak subtotal columns.
    hour_columns = [
        f"{hour:02d}:00"
        for hour in range(24)
    ]

    pivot["Total"] = pivot[
        hour_columns
    ].sum(axis=1)

    return pivot.reset_index()


def format_hourly_forecast_sheet(
    writer,
    sheet_name: str,
) -> None:
    """
    Apply consistent formatting and traffic-light bands.

    Hourly colour rules
    -------------------
    0 to 20:
        Green.

    Greater than 20 to 50:
        Yellow.

    Greater than 50 to less than 100:
        Orange.

    100 or more:
        Red.

    Conditional formatting is applied to the 24 hourly cells only.
    Peak totals and the daily total have separate formatting.
    """

    worksheet = writer.sheets[sheet_name]

    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24

    # -----------------------------------------------------------------
    # Header formatting
    # -----------------------------------------------------------------
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # -----------------------------------------------------------------
    # Column layout
    #
    # A       = Date
    # B:Y     = 00:00 to 23:00
    # Z       = Peak 03-06
    # AA      = Peak 10-14
    # AB      = Total
    # -----------------------------------------------------------------
    worksheet.column_dimensions["A"].width = 13

    for column_number in range(2, 26):
        column_letter = get_column_letter(
            column_number
        )

        worksheet.column_dimensions[
            column_letter
        ].width = 9

    worksheet.column_dimensions["Z"].width = 13
    worksheet.column_dimensions["AA"].width = 13
    worksheet.column_dimensions["AB"].width = 12

    # -----------------------------------------------------------------
    # Peak and total formatting
    # -----------------------------------------------------------------
    morning_peak_fill = PatternFill(
        fill_type="solid",
        fgColor="DDEBF7",
    )

    daytime_peak_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    total_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E1F2",
    )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        date_cell = worksheet.cell(
            row=row_number,
            column=1,
        )

        date_cell.number_format = "dd/mm/yyyy"
        date_cell.alignment = Alignment(
            horizontal="center"
        )

        # B:AB contains all numeric forecast cells.
        for column_number in range(2, 29):
            forecast_cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            forecast_cell.number_format = "0.0"
            forecast_cell.alignment = Alignment(
                horizontal="center"
            )

        morning_peak_cell = worksheet.cell(
            row=row_number,
            column=26,
        )

        morning_peak_cell.fill = morning_peak_fill
        morning_peak_cell.font = Font(
            bold=True
        )

        daytime_peak_cell = worksheet.cell(
            row=row_number,
            column=27,
        )

        daytime_peak_cell.fill = daytime_peak_fill
        daytime_peak_cell.font = Font(
            bold=True
        )

        total_cell = worksheet.cell(
            row=row_number,
            column=28,
        )

        total_cell.fill = total_fill
        total_cell.font = Font(
            bold=True
        )

    # -----------------------------------------------------------------
    # Hourly conditional formatting
    # -----------------------------------------------------------------
    hourly_range = (
        f"B2:Y{worksheet.max_row}"
    )

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    orange_fill = PatternFill(
        fill_type="solid",
        fgColor="F4B183",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="F4CCCC",
    )

    worksheet.conditional_formatting.add(
        hourly_range,
        CellIsRule(
            operator="between",
            formula=["0", "20"],
            fill=green_fill,
        ),
    )

    worksheet.conditional_formatting.add(
        hourly_range,
        CellIsRule(
            operator="between",
            formula=["20.0000001", "50"],
            fill=yellow_fill,
        ),
    )

    worksheet.conditional_formatting.add(
        hourly_range,
        CellIsRule(
            operator="between",
            formula=["50.0000001", "99.9999999"],
            fill=orange_fill,
        ),
    )

    worksheet.conditional_formatting.add(
        hourly_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["100"],
            fill=red_fill,
        ),
    )


def write_all_scenarios_hourly_workbook(
    baseline_hourly: pd.DataFrame,
    scenario_outputs: dict,
) -> Path:
    """
    Write baseline and scenario hourly forecasts to one workbook.

    Sheets produced
    ---------------
    Baseline:
        Baseline Entries
        Baseline Exits
        Baseline Movements

    Each scenario:
        Scenario Entries
        Scenario Exits
        Scenario Movements

    Scenario Entries uses booking_compliant_entries because this is the
    booked-slot profile after applying the 350 cap and redistribution.

    The detailed hourly CSV still contains adjusted_physical_entries,
    which shows the expected physical-arrival profile after circumvention.
    """

    workbook_path = (
        FORECAST_DIR
        / "all_scenarios_hourly_comparison.xlsx"
    )

    baseline = baseline_hourly.copy()

    baseline["total_movements"] = (
        baseline["entry_forecast"]
        + baseline["exit_forecast"]
    )

    workbook_sheets = {
        "Baseline Entries": create_hourly_pivot(
            hourly_df=baseline,
            value_column="entry_forecast",
        ),
        "Baseline Exits": create_hourly_pivot(
            hourly_df=baseline,
            value_column="exit_forecast",
        ),
        "Baseline Movements": create_hourly_pivot(
            hourly_df=baseline,
            value_column="total_movements",
        ),
    }

    for scenario_name in [
        "LOW_LOSS",
        "CENTRAL",
        "HIGH_LOSS",
    ]:
        if scenario_name not in scenario_outputs:
            raise KeyError(
                f"Scenario '{scenario_name}' is missing from "
                "scenario_outputs."
            )

        scenario_hourly = (
            scenario_outputs[
                scenario_name
            ]["hourly_comparison"]
            .copy()
        )

        scenario_hourly["datetime"] = pd.to_datetime(
            scenario_hourly["datetime"],
            errors="coerce",
        )

        scenario_hourly = (
            scenario_hourly
            .dropna(subset=["datetime"])
            .set_index("datetime")
            .sort_index()
        )

        # The scenario workbook shows the booking profile that obeys
        # the cap and the adjusted exits after lost demand.
        scenario_hourly[
            "total_movements"
        ] = (
            scenario_hourly[
                "booking_compliant_entries"
            ]
            + scenario_hourly[
                "adjusted_exits"
            ]
        )

        display_name = (
            scenario_name
            .replace("_", " ")
            .title()
        )

        workbook_sheets[
            f"{display_name} Entries"
        ] = create_hourly_pivot(
            hourly_df=scenario_hourly,
            value_column=(
                "booking_compliant_entries"
            ),
        )

        workbook_sheets[
            f"{display_name} Exits"
        ] = create_hourly_pivot(
            hourly_df=scenario_hourly,
            value_column="adjusted_exits",
        )

        workbook_sheets[
            f"{display_name} Movements"
        ] = create_hourly_pivot(
            hourly_df=scenario_hourly,
            value_column="total_movements",
        )

    with pd.ExcelWriter(
        workbook_path,
        engine="openpyxl",
    ) as writer:
        for (
            sheet_name,
            dataframe,
        ) in workbook_sheets.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

            format_hourly_forecast_sheet(
                writer=writer,
                sheet_name=sheet_name[:31],
            )

    return workbook_path


def run_forecast(
    connection,
    analysis_outputs,
) -> dict:
    """
    Create the six-week baseline and all cap scenarios.

    The forecast comparison is performed once:

        current daily forecast
        -> current hourly forecast
        -> apply cap to forecast demand above 350
        -> redistribute retained demand using lower-price historical weights
        -> remove lost entries
        -> remove future exits using historical duration weights
        -> compare baseline with each scenario

    The adjusted profile is the result of the scenario transformation.
    The forecast is not passed through the adjusted profile a second time.
    """

    (
        baseline_daily,
        baseline_hourly,
        _,
        exit_profile,
    ) = baseline_forecast(connection)

    behaviour = analysis_outputs[
        "lower_price_behaviour_parameters"
    ].copy()

    shift_weights = analysis_outputs[
        "lower_price_shift_weights"
    ].copy()

    durations = analysis_outputs[
        "peak_duration_distribution"
    ].copy()

    scenario_outputs = {}

    for (
        scenario_name,
        assumptions,
    ) in SCENARIOS.items():
        print(
            f"Running scenario: {scenario_name}"
        )

        booking_profile = (
            baseline_hourly[
                ["entry_forecast"]
            ]
            .astype(float)
            .copy()
        )

        physical_profile = (
            baseline_hourly[
                ["entry_forecast"]
            ]
            .astype(float)
            .copy()
        )

        audit_rows = []

        forecast_dates = pd.DatetimeIndex(
            baseline_hourly
            .index
            .normalize()
            .unique()
        ).sort_values()

        for date in forecast_dates:
            date = pd.Timestamp(
                date
            ).normalize()

            for (
                peak_name,
                (
                    start_hour,
                    end_hour,
                ),
            ) in PEAKS.items():
                peak_index = pd.date_range(
                    start=(
                        date
                        + pd.Timedelta(
                            hours=start_hour
                        )
                    ),
                    end=(
                        date
                        + pd.Timedelta(
                            hours=end_hour - 1
                        )
                    ),
                    freq="h",
                ).intersection(
                    baseline_hourly.index
                )

                if len(peak_index) == 0:
                    continue

                baseline_peak = float(
                    baseline_hourly.loc[
                        peak_index,
                        "entry_forecast",
                    ].sum()
                )

                excess = max(
                    0.0,
                    baseline_peak - PEAK_CAP,
                )

                lost_rate = float(
                    assumptions["lost_rate"]
                )

                retained_rate = (
                    1.0 - lost_rate
                )

                behaviour_row = behaviour.loc[
                    behaviour["peak"].eq(
                        peak_name
                    )
                ]

                if behaviour_row.empty:
                    raise ValueError(
                        "No lower-price historical "
                        f"behaviour found for {peak_name}."
                    )

                historical_share = float(
                    behaviour_row[
                        "lower_price_outside_peak_share"
                    ].iloc[0]
                )

                circumvention_rate = min(
                    retained_rate,
                    (
                        historical_share
                        * float(
                            assumptions[
                                "circumvention_multiplier"
                            ]
                        )
                    ),
                    MAX_CIRCUMVENTION_RATE_OF_EXCESS,
                )

                physical_shift_rate = (
                    retained_rate
                    - circumvention_rate
                )

                lost = (
                    excess
                    * lost_rate
                )

                circumvent = (
                    excess
                    * circumvention_rate
                )

                physically_shifted = (
                    excess
                    * physical_shift_rate
                )

                if (
                    excess > 0
                    and baseline_peak > 0
                ):
                    peak_shape = (
                        baseline_hourly.loc[
                            peak_index,
                            "entry_forecast",
                        ]
                        / baseline_peak
                    )

                    # Only 350 booked entries remain inside the peak.
                    booking_profile.loc[
                        peak_index,
                        "entry_forecast",
                    ] = (
                        peak_shape
                        * PEAK_CAP
                    )

                    # Physical peak also includes customers booking an
                    # earlier lower-price slot but still arriving in the peak.
                    physical_profile.loc[
                        peak_index,
                        "entry_forecast",
                    ] = (
                        peak_shape
                        * (
                            PEAK_CAP
                            + circumvent
                        )
                    )

                    weights = shift_weights.loc[
                        shift_weights[
                            "peak"
                        ].eq(peak_name)
                    ].copy()

                    if (
                        weights.empty
                        or weights["weight"].sum() <= 0
                    ):
                        raise ValueError(
                            "No valid historical redistribution "
                            f"weights found for {peak_name}."
                        )

                    weights["weight"] = (
                        weights["weight"]
                        / weights["weight"].sum()
                    )

                    # Circumvention changes the booked-entry slot only.
                    # Physical arrival remains inside the original peak.
                    for _, weight_row in weights.iterrows():
                        target = (
                            date
                            + pd.Timedelta(
                                hours=(
                                    start_hour
                                    - int(
                                        weight_row[
                                            "hours_before_peak"
                                        ]
                                    )
                                )
                            )
                        )

                        add_volume(
                            df=booking_profile,
                            timestamp=target,
                            column="entry_forecast",
                            amount=(
                                circumvent
                                * float(
                                    weight_row[
                                        "weight"
                                    ]
                                )
                            ),
                        )

                    # Physically shifted demand moves in both the booking
                    # profile and the expected physical-arrival profile.
                    for _, weight_row in weights.iterrows():
                        target = (
                            date
                            + pd.Timedelta(
                                hours=(
                                    start_hour
                                    - int(
                                        weight_row[
                                            "hours_before_peak"
                                        ]
                                    )
                                )
                            )
                        )

                        shifted_volume = (
                            physically_shifted
                            * float(
                                weight_row["weight"]
                            )
                        )

                        add_volume(
                            df=booking_profile,
                            timestamp=target,
                            column="entry_forecast",
                            amount=shifted_volume,
                        )

                        add_volume(
                            df=physical_profile,
                            timestamp=target,
                            column="entry_forecast",
                            amount=shifted_volume,
                        )

                audit_rows.append(
                    {
                        "scenario": scenario_name,
                        "date": date,
                        "peak": peak_name,
                        "peak_start_hour": start_hour,
                        "peak_end_hour": end_hour,
                        "peak_cap": PEAK_CAP,
                        "baseline_peak_entries":
                            baseline_peak,
                        "entries_above_cap":
                            excess,
                        "lost_rate":
                            lost_rate,
                        "lower_price_historical_outside_share":
                            historical_share,
                        "circumvention_multiplier":
                            float(
                                assumptions[
                                    "circumvention_multiplier"
                                ]
                            ),
                        "circumvention_rate":
                            circumvention_rate,
                        "physical_shift_rate":
                            physical_shift_rate,
                        "circumvent_book_earlier_arrive_peak":
                            circumvent,
                        "physically_shifted_earlier":
                            physically_shifted,
                        "lost_fastpark_entries":
                            lost,
                        "booking_compliant_peak_entries":
                            min(
                                baseline_peak,
                                PEAK_CAP,
                            ),
                        "expected_physical_peak_entries":
                            (
                                min(
                                    baseline_peak,
                                    PEAK_CAP,
                                )
                                + circumvent
                            ),
                    }
                )

        cap_summary = pd.DataFrame(
            audit_rows
        )

        lost_detail = lost_exit_detail(
            cap_summary,
            durations,
        )

        if lost_detail.empty:
            lost_exits = pd.Series(
                dtype=float,
                name="lost_exits",
            )
        else:
            lost_exits = (
                lost_detail
                .groupby("exit_date")[
                    "lost_exits"
                ]
                .sum()
            )

        booking_daily = (
            booking_profile[
                "entry_forecast"
            ]
            .groupby(
                booking_profile
                .index
                .normalize()
            )
            .sum()
        )

        physical_daily = (
            physical_profile[
                "entry_forecast"
            ]
            .groupby(
                physical_profile
                .index
                .normalize()
            )
            .sum()
        )

        daily = baseline_daily.copy()

        daily["baseline_entries"] = (
            daily["entry_forecast"]
        )

        daily[
            "booking_compliant_entries"
        ] = (
            booking_daily
            .reindex(daily.index)
            .fillna(
                daily["entry_forecast"]
            )
        )

        daily[
            "adjusted_physical_entries"
        ] = (
            physical_daily
            .reindex(daily.index)
            .fillna(
                daily["entry_forecast"]
            )
        )

        daily["lost_entries"] = (
            daily["baseline_entries"]
            - daily[
                "adjusted_physical_entries"
            ]
        ).clip(lower=0)

        daily["baseline_exits"] = (
            daily["exit_forecast"]
        )

        daily["lost_exits"] = (
            lost_exits
            .reindex(daily.index)
            .fillna(0)
        )

        daily["adjusted_exits"] = (
            daily["baseline_exits"]
            - daily["lost_exits"]
        ).clip(lower=0)

        daily["baseline_movements"] = (
            daily["baseline_entries"]
            + daily["baseline_exits"]
        )

        daily["adjusted_movements"] = (
            daily[
                "adjusted_physical_entries"
            ]
            + daily["adjusted_exits"]
        )

        daily["entry_change"] = (
            daily[
                "adjusted_physical_entries"
            ]
            - daily["baseline_entries"]
        )

        daily["exit_change"] = (
            daily["adjusted_exits"]
            - daily["baseline_exits"]
        )

        daily["movement_change"] = (
            daily["adjusted_movements"]
            - daily["baseline_movements"]
        )

        adjusted_hourly_exits = (
            expand_daily_to_hourly(
                daily=daily[
                    "adjusted_exits"
                ],
                profile=exit_profile,
                output_column="adjusted_exits",
            )
        )

        hourly = (
            baseline_hourly
            .rename(
                columns={
                    "entry_forecast":
                        "baseline_entries",
                    "exit_forecast":
                        "baseline_exits",
                }
            )
            .join(
                booking_profile.rename(
                    columns={
                        "entry_forecast":
                            "booking_compliant_entries"
                    }
                ),
                how="outer",
            )
            .join(
                physical_profile.rename(
                    columns={
                        "entry_forecast":
                            "adjusted_physical_entries"
                    }
                ),
                how="outer",
            )
            .join(
                adjusted_hourly_exits,
                how="outer",
            )
            .fillna(0)
            .sort_index()
        )

        hourly["baseline_movements"] = (
            hourly["baseline_entries"]
            + hourly["baseline_exits"]
        )

        hourly[
            "booking_compliant_movements"
        ] = (
            hourly[
                "booking_compliant_entries"
            ]
            + hourly["adjusted_exits"]
        )

        hourly[
            "adjusted_physical_movements"
        ] = (
            hourly[
                "adjusted_physical_entries"
            ]
            + hourly["adjusted_exits"]
        )

        hourly["entry_change"] = (
            hourly[
                "adjusted_physical_entries"
            ]
            - hourly["baseline_entries"]
        )

        hourly["exit_change"] = (
            hourly["adjusted_exits"]
            - hourly["baseline_exits"]
        )

        hourly["movement_change"] = (
            hourly[
                "adjusted_physical_movements"
            ]
            - hourly["baseline_movements"]
        )

        hourly["date"] = (
            hourly.index.date
        )

        hourly["time"] = (
            hourly.index.time
        )

        outputs = {
            "daily_comparison":
                daily.reset_index(
                    names="date"
                ),
            "hourly_comparison":
                hourly.reset_index(
                    names="datetime"
                ),
            "cap_summary":
                cap_summary,
            "lost_entry_exit_detail":
                lost_detail,
        }

        for (
            output_name,
            output_table,
        ) in outputs.items():
            output_table.to_csv(
                FORECAST_DIR
                / (
                    f"{scenario_name.lower()}_"
                    f"{output_name}.csv"
                ),
                index=False,
            )

        scenario_outputs[
            scenario_name
        ] = outputs

    all_daily_comparisons = pd.concat(
        [
            tables[
                "daily_comparison"
            ].assign(
                scenario=scenario_name
            )
            for (
                scenario_name,
                tables,
            ) in scenario_outputs.items()
        ],
        ignore_index=True,
    )

    all_hourly_comparisons = pd.concat(
        [
            tables[
                "hourly_comparison"
            ].assign(
                scenario=scenario_name
            )
            for (
                scenario_name,
                tables,
            ) in scenario_outputs.items()
        ],
        ignore_index=True,
    )

    all_cap_summaries = pd.concat(
        [
            tables["cap_summary"]
            for tables
            in scenario_outputs.values()
        ],
        ignore_index=True,
    )

    all_daily_comparisons.to_csv(
        FORECAST_DIR
        / "all_scenarios_daily_comparison.csv",
        index=False,
    )

    all_hourly_comparisons.to_csv(
        FORECAST_DIR
        / "all_scenarios_hourly_comparison.csv",
        index=False,
    )

    all_cap_summaries.to_csv(
        FORECAST_DIR
        / "all_scenarios_cap_summary.csv",
        index=False,
    )

    hourly_workbook_path = (
        write_all_scenarios_hourly_workbook(
            baseline_hourly=
                baseline_hourly,
            scenario_outputs=
                scenario_outputs,
        )
    )

    print(
        "Hourly comparison workbook written to: "
        f"{hourly_workbook_path}"
    )

    return {
        "baseline_daily":
            baseline_daily,
        "baseline_hourly":
            baseline_hourly,
        "scenarios":
            scenario_outputs,
        "hourly_comparison_workbook":
            hourly_workbook_path,
    }
    


# =============================================================================
# VALIDATION AND ENTRY POINT
# =============================================================================

def validate_outputs(analysis_outputs) -> None:
    for peak_name in PEAKS:
        weights = analysis_outputs["lower_price_shift_weights"].loc[
            analysis_outputs["lower_price_shift_weights"]["peak"].eq(peak_name), "weight"
        ]
        if not np.isclose(weights.sum(), 1.0):
            raise AssertionError(f"Lower-price shift weights do not total 1 for {peak_name}")

        durations = analysis_outputs["peak_duration_distribution"].loc[
            analysis_outputs["peak_duration_distribution"]["peak"].eq(peak_name), "weight"
        ]
        if not np.isclose(durations.sum(), 1.0):
            raise AssertionError(f"Duration weights do not total 1 for {peak_name}")


def main() -> None:
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
    FORECAST_DIR.mkdir(parents=True, exist_ok=True)

    if not RUN_HISTORICAL_ANALYSIS and not RUN_FORECAST_SCENARIOS:
        print("Nothing to run. Enable an analysis or forecast flag.")
        return

    connection = connect_db()
    try:
        analysis_outputs = (
            run_analysis(connection) if RUN_HISTORICAL_ANALYSIS else load_analysis()
        )
        validate_outputs(analysis_outputs)
        if RUN_FORECAST_SCENARIOS:
            run_forecast(connection, analysis_outputs)
    finally:
        connection.close()

    print(f"Complete. Outputs written to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
