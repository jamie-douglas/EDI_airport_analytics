# %%
import pyodbc
import pandas as pd
import datetime
from datetime import datetime

dsn = 'AzureConnection'
user = 'jamie_douglas'
username = '{user}@edinburghairport.com'

connect = pyodbc.connect('DSN=' + dsn + ';')

for row in connect.cursor().tables():
    if row.table_schem != 'sys':
        print(row.table_schem+'.' + row.table_name)





# %%
# Calculate the start and end dates for the last 2 weeks, not including this week
today = datetime.now()
current_year = datetime.now().year
start_of_this_week = today - pd.Timedelta(days=today.weekday())
end_of_last_week = start_of_this_week - pd.Timedelta(days=1)
start_of_two_weeks_ago = start_of_this_week - pd.Timedelta(weeks=2)

# Add 1 more day to end_of_last_week
end_of_last_week += pd.Timedelta(days=1)



sql = f"""
SELECT *
FROM EAL.FlightPerformance
WHERE [ActualDateTime_Local] >= '{start_of_two_weeks_ago.strftime('%Y-%m-%d')}'
    AND [ActualDateTime_Local] <= '{end_of_last_week.strftime('%Y-%m-%d')}'
"""

flight_performance_df = pd.read_sql(sql, connect)

# Pull EAL.FlightPerformance_FutureFlights for the next two months
start_of_next_period = today
end_of_next_two_months = today + pd.DateOffset(months=2)
print(start_of_next_period.strftime('%Y-%m-%d'), end_of_next_two_months.strftime('%Y-%m-%d'))

sql_future = f"""
SELECT *
FROM EAL.FlightPerformance_FutureFlights
WHERE [ScheduledDateTime_Local] >= '{start_of_next_period.strftime('%Y-%m-%d')}'
    AND [ScheduledDateTime_Local] <= '{end_of_next_two_months.strftime('%Y-%m-%d')}'
"""

flight_performance_future_df = pd.read_sql(sql_future, connect)

# Select FastPark entries for the last two weeks, not including this week
sql = f"""
SELECT *
FROM FastPark.v_EntryAndExits
WHERE [CheckInStarted] >= '{start_of_two_weeks_ago.strftime('%Y-%m-%d')}'
    AND [CheckInStarted] <= '{end_of_last_week.strftime('%Y-%m-%d')}'
"""
fastpark_actual_entries_df = pd.read_sql(sql, connect)

# Select FastPark exits for the last two weeks, not including this week
sql = f"""
SELECT *
FROM FastPark.v_EntryAndExits
WHERE [ActualCheckedOutDate] >= '{start_of_two_weeks_ago.strftime('%Y-%m-%d')}'
    AND [ActualCheckedOutDate] <= '{end_of_last_week.strftime('%Y-%m-%d')}'
"""
fastpark_actual_exits_df = pd.read_sql(sql, connect)





# %%
# Pull the entire AirportX.v_Bookings table for the next 2 months based on entryDate
start_of_next_period_str = start_of_next_period.strftime('%Y-%m-%d')
end_of_next_two_months_str = end_of_next_two_months.strftime('%Y-%m-%d')

sql_airportx_entry = f"""
SELECT *
FROM AirportX.v_Bookings
WHERE [entryDate] > '{start_of_next_period_str}'
    AND [entryDate] <= '{end_of_next_two_months_str}'
"""
airportx_bookings_entry_df = pd.read_sql(sql_airportx_entry, connect)

sql_airportx_exit = f"""
SELECT *
FROM AirportX.v_Bookings
WHERE [exitDate] > '{start_of_next_period_str}'
    AND [exitDate] <= '{end_of_next_two_months_str}'
"""
airportx_bookings_exit_df = pd.read_sql(sql_airportx_exit, connect)

# Index(['bookingUuid', 'bookingId', 'transactionId', 'createdAt', 'updatedAt',
#        'cancelledAt', 'channel', 'productCode', 'productGroup', 'productName',
#        'productPrice', 'productQuantity', 'productTotal', 'entryDate',
#        'exitDate', 'status', 'assetCode', 'assetName', 'bookingTotal',
#        'leadtime', 'duration', 'nationality', 'inboundAirline',
#        'inboundFlight', 'inboundRoute', 'outboundAirline', 'outboundFlight',
#        'outboundRoute', 'promoCode'],
#       dtype='object')


# %%
booking_entry_fastpark_df = airportx_bookings_entry_df[
    (airportx_bookings_entry_df['assetName'] == 'FastPark') &
    (airportx_bookings_entry_df['status'] == 'B')
]

booking_exit_fastpark_df = airportx_bookings_exit_df[
    (airportx_bookings_exit_df['assetName'] == 'FastPark') &
    (airportx_bookings_exit_df['status'] == 'B')
]

# Convert entryDate and exitDate to datetime if not already
booking_entry_fastpark_df['entryDate'] = pd.to_datetime(booking_entry_fastpark_df['entryDate'])
booking_exit_fastpark_df['exitDate'] = pd.to_datetime(booking_exit_fastpark_df['exitDate'])

# Group by entryDate and exitDate to get daily counts
daily_booking_entries = booking_entry_fastpark_df.groupby(booking_entry_fastpark_df['entryDate'].dt.date).size()
daily_booking_exits = booking_exit_fastpark_df.groupby(booking_exit_fastpark_df['exitDate'].dt.date).size()

# Combine into a DataFrame for clarity
booking_daily_counts_df = pd.DataFrame({
    'booking_entries': daily_booking_entries,
    'booking_exits': daily_booking_exits
}).fillna(0).astype(int)

print(booking_daily_counts_df)


# %%
df_entries = fastpark_actual_entries_df.drop_duplicates(subset='BookingReference')

daily_entry_counts = df_entries['CheckInStarted'].dt.date.value_counts().sort_index()

# Group entries by weekday (0=Monday, 6=Sunday)
df_entries['weekday'] = df_entries['CheckInStarted'].dt.weekday
weekday_entry_counts = df_entries.groupby('weekday').size()

# Map weekday numbers to names for clarity
weekday_entry_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
weekday_entry_summary = pd.Series(weekday_entry_counts.values, index=[weekday_entry_names[i] for i in weekday_entry_counts.index])
print(weekday_entry_summary)

mask = (flight_performance_df['ArrDeptureCode'] == 'D') & (flight_performance_df['IsPassengerFlight'] == 1)
passengers_per_day_departure = flight_performance_df.loc[mask].groupby(flight_performance_df['ActualDateTime_Local'].dt.date)['Pax_MostConfident'].sum()

# Combine total passengers per weekday (Monday to Sunday) for departures
passengers_by_weekday = flight_performance_df.loc[mask].groupby(flight_performance_df['ActualDateTime_Local'].dt.weekday)['Pax_MostConfident'].sum()
passengers_by_weekday_named = pd.Series(passengers_by_weekday.values, index=[weekday_entry_names[i] for i in passengers_by_weekday.index])
print(passengers_by_weekday_named)

penetration_rate_by_weekday_entry = (weekday_entry_summary / passengers_by_weekday_named) * 100
print(penetration_rate_by_weekday_entry.round(2).astype(str) + '%')


# %%
# Use df_exits for exit analysis

df_exits = fastpark_actual_exits_df.drop_duplicates(subset='BookingReference')

daily_exit_counts = df_exits['ActualCheckedOutDate'].dt.date.value_counts().sort_index()

# Group exits by weekday (0=Monday, 6=Sunday)
df_exits['weekday'] = df_exits['ActualCheckedOutDate'].dt.weekday
weekday_exit_counts = df_exits.groupby('weekday').size()
weekday_exit_counts.index = weekday_exit_counts.index.astype(int)
weekday_exit_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
weekday_exit_summary = pd.Series(weekday_exit_counts.values, index=[weekday_exit_names[i] for i in weekday_exit_counts.index])
print(weekday_exit_summary)

mask_arrival = (flight_performance_df['ArrDeptureCode'] == 'A') & (flight_performance_df['IsPassengerFlight'] == 1)
passengers_per_day_arrival = flight_performance_df.loc[mask_arrival].groupby(flight_performance_df['ActualDateTime_Local'].dt.date)['Pax_MostConfident'].sum()

passengers_by_weekday_arrival = flight_performance_df.loc[mask_arrival].groupby(flight_performance_df['ActualDateTime_Local'].dt.weekday)['Pax_MostConfident'].sum()
passengers_by_weekday_arrival_named = pd.Series(passengers_by_weekday_arrival.values, index=[weekday_exit_names[i] for i in passengers_by_weekday_arrival.index])
print(passengers_by_weekday_arrival_named)

penetration_rate_by_weekday_exit = (weekday_exit_summary / passengers_by_weekday_arrival_named) * 100
print(penetration_rate_by_weekday_exit.round(2).astype(str) + '%')


# %%

future_departures = flight_performance_future_df[
    (flight_performance_future_df['ArrDeptureCode'] == 'D') &
    (flight_performance_future_df['IsPassengerFlight'] == 1)
]

future_departures['date'] = future_departures['ScheduledDateTime_Local'].dt.date
daily_future_departures_pax = future_departures.groupby('date')['Pax_MostConfident'].sum()
# print(daily_future_departures_pax)

# Map weekday names to penetration rates for entry
penetration_rate_map = penetration_rate_by_weekday_entry

# Get weekday name for each date in daily_future_departures_pax
future_departures_weekdays = pd.to_datetime(daily_future_departures_pax.index).weekday
future_departures_weekday_names = [weekday_entry_names[i] for i in future_departures_weekdays]

# Calculate forecasted entries per day using penetration rate
forecasted_entries = daily_future_departures_pax.values * penetration_rate_map[future_departures_weekday_names].values / 100

# Create a DataFrame for clarity
forecasted_entries_df = pd.DataFrame({
    'date': daily_future_departures_pax.index,
    'forecasted_entries': forecasted_entries
}).set_index('date')

# print(forecasted_entries_df)

# Add columns for passengers and penetration rate for each row
forecasted_entries_df['passengers'] = daily_future_departures_pax.values
forecasted_entries_df['penetration_rate'] = [
    penetration_rate_map[weekday] for weekday in future_departures_weekday_names
]
print(forecasted_entries_df)

# forecasted_entries_df['date'] = pd.to_datetime(forecasted_entries_df.index)

# # Increase forecasted_entries by 10% for every date in March
# march_mask = forecasted_entries_df['date'].dt.month == 3
# forecasted_entries_df.loc[march_mask, 'forecasted_entries'] *= 1.10

# forecasted_entries_df.drop(columns=['date'], inplace=True)

# print(forecasted_entries_df.head(30))


# %%

future_arrivals = flight_performance_future_df[
    (flight_performance_future_df['ArrDeptureCode'] == 'A') &
    (flight_performance_future_df['IsPassengerFlight'] == 1)
]

future_arrivals['date'] = future_arrivals['ScheduledDateTime_Local'].dt.date
daily_future_arrivals_pax = future_arrivals.groupby('date')['Pax_MostConfident'].sum()
# print(daily_future_arrivals_pax)


# Map weekday names to penetration rates for exits
penetration_rate_exit_map = penetration_rate_by_weekday_exit

# Get weekday name for each date in daily_future_arrivals_pax
future_arrivals_weekdays = pd.to_datetime(daily_future_arrivals_pax.index).weekday
future_arrivals_weekday_names = [weekday_exit_names[i] for i in future_arrivals_weekdays]

# Calculate forecasted exits per day using penetration rate for exits
forecasted_exits = daily_future_arrivals_pax.values * penetration_rate_exit_map[future_arrivals_weekday_names].values / 100

# Create a DataFrame for clarity
forecasted_exits_df = pd.DataFrame({
    'date': daily_future_arrivals_pax.index,
    'forecasted_exits': forecasted_exits
}).set_index('date')

# Add columns for passengers and penetration rate for each row
forecasted_exits_df['passengers'] = daily_future_arrivals_pax.values
forecasted_exits_df['penetration_rate'] = [
    penetration_rate_exit_map[weekday] for weekday in future_arrivals_weekday_names
]

print(forecasted_exits_df)

# forecasted_exits_df['date'] = pd.to_datetime(forecasted_exits_df.index)

# # Increase forecasted_exits by 10% for every date in March
# march_mask = forecasted_exits_df['date'].dt.month == 3
# forecasted_exits_df.loc[march_mask, 'forecasted_exits'] *= 1.10

# forecasted_exits_df.drop(columns=['date'], inplace=True)

# print(forecasted_exits_df.head(30))


# %%

# Combine entry and exit forecasts into a single DataFrame
# Ensure both DataFrames have the same dates and length
if len(forecasted_entries_df) != len(forecasted_exits_df):
    # Align on index (date), keep only common dates
    common_dates = forecasted_entries_df.index.intersection(forecasted_exits_df.index)
    forecasted_entries_df = forecasted_entries_df.loc[common_dates]
    forecasted_exits_df = forecasted_exits_df.loc[common_dates]

combined_forecast_df = pd.DataFrame({
    'date': forecasted_entries_df.index,
    'entry_forecast': forecasted_entries_df['forecasted_entries'].values,
    'exit_forecast': forecasted_exits_df['forecasted_exits'].values
}).set_index('date')
combined_forecast_df['entry_forecast'] = combined_forecast_df['entry_forecast'].apply(lambda x: int(round(x)))
combined_forecast_df['exit_forecast'] = combined_forecast_df['exit_forecast'].apply(lambda x: int(round(x)))

print(combined_forecast_df.head(30))

# combined_forecast_df.to_csv('combined_forecast.csv')



# %%
from datetime import datetime

# Merge booking_daily_counts_df with combined_forecast_df on date
# Ensure index is datetime for merging
booking_daily_counts_df.index = pd.to_datetime(booking_daily_counts_df.index)
combined_forecast_df.index = pd.to_datetime(combined_forecast_df.index)

merged_df = booking_daily_counts_df.merge(
    combined_forecast_df,
    left_index=True,
    right_index=True,
    how='outer'
)
# Remove today's row if present
merged_df = merged_df[merged_df.index.date != today.date()]

# Set the first 14 forecasts to match booking_entries
merged_df.iloc[:14, merged_df.columns.get_loc('entry_forecast')] = merged_df.iloc[:14]['booking_entries']
merged_df.iloc[:14, merged_df.columns.get_loc('exit_forecast')] = merged_df.iloc[:14]['booking_exits']



entry_increase = [5, 10, 15, 20, 25, 30, 35, 38, 41, 44, 47, 50, 53, 56]
exit_increase = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

# Increase entry_forecast by 5% for the first day, 10% for the second day, up to 7 days
for i, percentage in enumerate(entry_increase):
    if i < len(merged_df):
        merged_df.iloc[i, merged_df.columns.get_loc('entry_forecast')] *= (1 + percentage / 100)

# Increase exit_forecast by 0.5% for the first day, 1% for the second, ..., up to 3.5% for the seventh day
for i, percentage in enumerate(exit_increase):
    if i < len(merged_df):
        merged_df.iloc[i, merged_df.columns.get_loc('exit_forecast')] *= (1 + percentage / 100)

print(merged_df.head(30))


merged_df['date'] = pd.to_datetime(merged_df.index)

# Increase forecasted_exits by 15% for M-W and 7% for Th-S
march_mask = merged_df['date'].dt.month == 3
weekday = merged_df['date'].dt.dayofweek
mtw_mask = weekday.isin([0,1,2]) #Mon, Tue, Wed
ths_mask = weekday.isin([3,4,5,6]) #Thu-Sun
merged_df.loc[march_mask & mtw_mask, ['entry_forecast', 'exit_forecast'] *= 1.15
merged_df.loc[march_mask & ths_mask, ['entry_forecast', 'exit_forecast'] *= 1.07

merged_df.drop(columns=['date'], inplace=True)

print(merged_df.head(30))

# Round entry_forecast and exit_forecast to nearest whole number
merged_df['entry_forecast'] = merged_df['entry_forecast'].round().astype('Int64')
merged_df['exit_forecast'] = merged_df['exit_forecast'].round().astype('Int64')


# merged_df.to_csv('merged_booking_forecast.csv')
# Replace forecasted_entries_df with merged_df['entry_forecast'] for future use
forecasted_entries_df['forecasted_entries'] = merged_df['entry_forecast']

# Replace forecasted_exits_df with merged_df['exit_forecast'] for future use
forecasted_exits_df['forecasted_exits'] = merged_df['exit_forecast']
# Save merged_df to CSV with today's date in the filename
today_str = today.strftime('%Y-%m-%d')
merged_df.to_csv(r'C:\Users\jamie_douglas\Edinburgh Airport Limited\Shared Files - Business Planning\Haoran\FastPark\FastPark Model Rebuild 2025\FastPark Short-term Model Rebuild\Daily Forecast Output\{}_daily_forecast.csv'.format(today_str))

# %%
# Calculate percentage of entries in each hour for each weekday (Monday to Sunday)
hourly_distribution = {}

for i, weekday in enumerate(weekday_entry_names):
    # Filter entries for the current weekday
    entries_weekday = df_entries[df_entries['weekday'] == i]
    # Extract hour from CheckInStarted
    entries_weekday['hour'] = entries_weekday['CheckInStarted'].dt.hour
    # Count entries per hour
    hourly_counts = entries_weekday['hour'].value_counts().sort_index()
    # Calculate percentage
    hourly_percent = (hourly_counts / hourly_counts.sum()) * 100
    # Store in dictionary
    hourly_distribution[weekday] = hourly_percent

# Convert to DataFrame for easier viewing
hourly_distribution_df = pd.DataFrame(hourly_distribution).fillna(0).round(2)
# print(hourly_distribution_df)

# Expand forecasted entries into 24 hours for each day using hourly_distribution_df

# Prepare a list to collect expanded rows
expanded_entries = []


for date, row in forecasted_entries_df.iterrows():
    # Get the weekday name for this date
    weekday = pd.to_datetime(date).day_name()
    # Get the hourly distribution for this weekday (as percentages)
    hourly_percentages = hourly_distribution_df[weekday] / 100  # convert to fraction
    # For each hour, calculate forecasted entries
    for hour in range(24):
        percent = hourly_percentages.get(hour, 0)
        forecasted_hourly_entry = row['forecasted_entries'] * percent
        expanded_entries.append({
            'date': date,
            'hour': hour,
            'forecasted_entries': forecasted_hourly_entry
        })

# Create DataFrame
forecasted_entries_hourly_df = pd.DataFrame(expanded_entries)
forecasted_entries_hourly_df = forecasted_entries_hourly_df.set_index(['date', 'hour'])
print(forecasted_entries_hourly_df.head(30))

# Combine 'date' and 'hour' into a single datetime column
forecasted_entries_hourly_df = forecasted_entries_hourly_df.reset_index()
forecasted_entries_hourly_df['datetime'] = pd.to_datetime(forecasted_entries_hourly_df['date']) + pd.to_timedelta(forecasted_entries_hourly_df['hour'], unit='h')
forecasted_entries_hourly_df = forecasted_entries_hourly_df.set_index('datetime')
forecasted_entries_hourly_df = forecasted_entries_hourly_df.drop(['date', 'hour'], axis=1)
# Save to CSV
# forecasted_entries_hourly_df.to_csv('forecasted_entries_hourly.csv')

# %%
# Calculate percentage of exits in each hour for each weekday (Monday to Sunday)
hourly_exit_distribution = {}

for i, weekday in enumerate(weekday_exit_names):
    # Filter exits for the current weekday
    exits_weekday = df_exits[df_exits['weekday'] == i]
    # Extract hour from ActualCheckedOutDate
    exits_weekday['hour'] = exits_weekday['ActualCheckedOutDate'].dt.hour
    # Count exits per hour
    hourly_exit_counts = exits_weekday['hour'].value_counts().sort_index()
    # Calculate percentage
    hourly_exit_percent = (hourly_exit_counts / hourly_exit_counts.sum()) * 100
    # Store in dictionary
    hourly_exit_distribution[weekday] = hourly_exit_percent

# Convert to DataFrame for easier viewing
hourly_exit_distribution_df = pd.DataFrame(hourly_exit_distribution).fillna(0).round(2)
# print(hourly_exit_distribution_df)

# Expand forecasted exits into 24 hours for each day using hourly_exit_distribution_df

# Prepare a list to collect expanded rows
expanded_exits = []

for date, row in forecasted_exits_df.iterrows():
    # Get the weekday name for this date
    weekday = pd.to_datetime(date).day_name()
    # Get the hourly distribution for this weekday (as percentages)
    hourly_percentages = hourly_exit_distribution_df[weekday] / 100  # convert to fraction
    # For each hour, calculate forecasted exits
    for hour in range(24):
        percent = hourly_percentages.get(hour, 0)
        forecasted_hourly_exit = row['forecasted_exits'] * percent
        expanded_exits.append({
            'date': date,
            'hour': hour,
            'forecasted_exits': forecasted_hourly_exit
        })

# Create DataFrame
forecasted_exits_hourly_df = pd.DataFrame(expanded_exits)
forecasted_exits_hourly_df = forecasted_exits_hourly_df.set_index(['date', 'hour'])


# Combine 'date' and 'hour' into a single datetime column
forecasted_exits_hourly_df = forecasted_exits_hourly_df.reset_index()
forecasted_exits_hourly_df['datetime'] = pd.to_datetime(forecasted_exits_hourly_df['date']) + pd.to_timedelta(forecasted_exits_hourly_df['hour'], unit='h')
forecasted_exits_hourly_df = forecasted_exits_hourly_df.set_index('datetime')
forecasted_exits_hourly_df = forecasted_exits_hourly_df.drop(['date', 'hour'], axis=1)

print(forecasted_exits_hourly_df.head(30))
# Save to CSV
# forecasted_exits_hourly_df.to_csv('forecasted_exits_hourly.csv')



# %%
import openpyxl
from openpyxl import load_workbook
import csv
from datetime import datetime, timedelta
import numpy as np

# Combine forecasted_entries_hourly_df and forecasted_exits_hourly_df into a single DataFrame


combined_hourly_forecast_df = forecasted_entries_hourly_df.join(
    forecasted_exits_hourly_df, how='outer', lsuffix='_entry', rsuffix='_exit'
)


# Rename columns for clarity
combined_hourly_forecast_df = combined_hourly_forecast_df.rename(
    columns={
        'forecasted_entries': 'entry_forecast',
        'forecasted_exits': 'exit_forecast'
    }
)

# Insert 'date' and 'time' as the second and third columns
combined_hourly_forecast_df['date'] = combined_hourly_forecast_df.index.date
combined_hourly_forecast_df['time'] = combined_hourly_forecast_df.index.time

# Overwrite only the dates present in combined_hourly_forecast_df in 'FastPark Hourly Forecast Output.csv'

csv_path = 'FastPark Hourly Forecast Output.csv'

exisiting_df = pd.read_csv(csv_path, index_col=0)

exisiting_df['date'] = pd.to_datetime(exisiting_df['date'], dayfirst=True).dt.date

# Convert exisiting_df['date'] from str to datetime.date
exisiting_df['date'] = pd.to_datetime(exisiting_df['date'], errors='coerce').dt.date

tomorrow = datetime.now().date() + timedelta(days=1)
# print(tomorrow)
# Remove all rows where exisiting_df['date'] is greater than today
exisiting_df = exisiting_df[exisiting_df['date'] < tomorrow]



# Append combined_hourly_forecast_df to exisiting_df and output to CSV
exisiting_df = pd.concat([exisiting_df, combined_hourly_forecast_df])


exisiting_df.to_csv('FastPark Hourly Forecast Output.csv')
# print(exisiting_df)



exisiting_df_copy = exisiting_df.copy()

# Move 'date' and 'time' columns to the first and second positions
cols = ['date', 'time'] + [col for col in exisiting_df_copy.columns if col not in ['date', 'time']]
exisiting_df_copy = exisiting_df_copy[cols]

# Rename columns as requested
exisiting_df_copy['entry_forecast'] = np.ceil(exisiting_df_copy['entry_forecast']).astype(int)
exisiting_df_copy['exit_forecast'] = np.ceil(exisiting_df_copy['exit_forecast']).astype(int)

exisiting_df_copy = exisiting_df_copy.rename(columns={
    'date': 'Date',
    'time': 'intervalStartTime_Local',
    'entry_forecast': 'Entries',
    'exit_forecast': 'Exits'
})


# Check for duplicate rows based on 'Date' and 'intervalStartTime_Local', keep the first occurrence
exisiting_df_copy = exisiting_df_copy.drop_duplicates(subset=['Date', 'intervalStartTime_Local'], keep='first')

print(exisiting_df_copy)




DWH_path = rf"\\BIPOEVSAPP01\FilesForUpload\BusinessAnalytics\FastParkForecast\FP_Forecast_output_{datetime.now().strftime('%Y%m%d')}.csv"

exisiting_df_copy.to_csv(r'C:\Users\jamie_douglas\Edinburgh Airport Limited\Shared Files - Business Planning\Haoran\FastPark\FastPark Model Rebuild 2025\FastPark Short-term Model Rebuild\Daily Forecast Output\{}_daily_forecast_DWH.csv'.format(today_str), index=False)
exisiting_df_copy.to_csv(DWH_path, index=False)




# %%



# Save combined_hourly_forecast_df to the additional directory
additional_dir = r'C:\Users\{}\Edinburgh Airport Limited\Workforce Management - FastPark Forecast'.format(user)
exisiting_df.to_csv(f'{additional_dir}\\FastPark Hourly Forecast Output.csv')

if today.weekday() == 0:  # Monday is 0
    exisiting_df.to_csv('FastPark Hourly Forecast Output Weekly.csv')
    exisiting_df.to_csv(f'{additional_dir}\\FastPark Hourly Forecast Output Weekly.csv')

# print(combined_hourly_forecast_df.head(24))
# Split to entry_forecast only
entry_hourly_forecast_df = combined_hourly_forecast_df[['entry_forecast', 'date', 'time']].copy()
entry_hourly_forecast_df.index.name = None  # remove the index name (datetime)
entry_hourly_forecast_df = entry_hourly_forecast_df.reset_index(drop=True)

# Pivot the DataFrame so that each hour is a column and each row is a date
entry_hourly_pivot_df = entry_hourly_forecast_df.pivot(index='date', columns='time', values='entry_forecast')
entry_hourly_pivot_df.columns = [str(col) for col in entry_hourly_pivot_df.columns]  # convert time objects to string for clarity
entry_hourly_pivot_df = entry_hourly_pivot_df.sort_index(axis=1)  # ensure columns are in hour order
entry_hourly_pivot_df['Total'] = entry_hourly_pivot_df.sum(axis=1)
# print(entry_hourly_pivot_df.head())

entry_hourly_pivot_df.to_csv('FastPark Entry Hourly Forecast Output.csv')

# Split to exit_forecast only
exit_hourly_forecast_df = combined_hourly_forecast_df[['exit_forecast', 'date', 'time']].copy()
exit_hourly_forecast_df.index.name = None  # remove the index name (datetime)
exit_hourly_forecast_df = exit_hourly_forecast_df.reset_index(drop=True)

# Pivot the DataFrame so that each hour is a column and each row is a date
exit_hourly_pivot_df = exit_hourly_forecast_df.pivot(index='date', columns='time', values='exit_forecast')
exit_hourly_pivot_df.columns = [str(col) for col in exit_hourly_pivot_df.columns]  # convert time objects to string for clarity
exit_hourly_pivot_df = exit_hourly_pivot_df.sort_index(axis=1)  # ensure columns are in hour order)
exit_hourly_pivot_df['Total'] = exit_hourly_pivot_df.sum(axis=1)
# print(exit_hourly_pivot_df.head())

exit_hourly_pivot_df.to_csv('FastPark Exit Hourly Forecast Output.csv')
# Append entry_hourly_pivot_df to the Excel sheet "FastPark Forecast Accuracy Tracker.xlsx" in the tab "Entries Forecast" from the correct date


excel_path = "FastPark Forecast Accuracy Tracker NEW.xlsx"
sheet_name = "Entries Forecast"
# Load the workbook and worksheet
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# Find the first date in entry_hourly_pivot_df
first_date = entry_hourly_pivot_df.index.min()
last_date = entry_hourly_pivot_df.index.max()


# Find the column headers in the sheet (assume headers are in the first row)
header_row = 1
headers = [cell.value for cell in ws[header_row]]

# Find the date column (should be the first column)
date_col_idx = 1  # openpyxl is 1-based

# Find the starting row to write (match first_date)
startrow = None
for row in range(2, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=date_col_idx).value
    if isinstance(cell_value, datetime):
        cell_date = cell_value.date()
    elif isinstance(cell_value, str):
        try:
            cell_date = pd.to_datetime(cell_value).date()
        except Exception:
            continue
    else:
        continue
    if cell_date == first_date:
        startrow = row
        break

if startrow is None:
    # If not found, append at the end
    startrow = ws.max_row + 1

# Skip overwriting for dates between 19/12/2025 and 03/01/2026 (inclusive)
skip_dates = pd.date_range('2024-12-19', '2024-01-03').date

for i, (date, row) in enumerate(entry_hourly_pivot_df.iterrows()):
    if date in skip_dates:
        continue  # skip these dates
    excel_row = startrow + i
    ws.cell(row=excel_row, column=1, value=date)
    for j, col in enumerate(entry_hourly_pivot_df.columns, start=2):
        ws.cell(row=excel_row, column=j, value=row[col])

if today.weekday() == 0:  # Monday is 0
    # Save the workbook
    wb.save(excel_path)

# Repeat for exit_hourly_pivot_df, sheet name is "Exit Forecast"
exit_sheet_name = "Exits Forecast"
wb_exit = load_workbook(excel_path)
ws_exit = wb_exit[exit_sheet_name]

first_exit_date = exit_hourly_pivot_df.index.min()
last_exit_date = exit_hourly_pivot_df.index.max()

# Find the column headers in the sheet (assume headers are in the first row)
exit_header_row = 1
exit_headers = [cell.value for cell in ws_exit[exit_header_row]]

exit_date_col_idx = 1  # openpyxl is 1-based

exit_startrow = None
for row in range(2, ws_exit.max_row + 1):
    cell_value = ws_exit.cell(row=row, column=exit_date_col_idx).value
    if isinstance(cell_value, datetime):
        cell_date = cell_value.date()
    elif isinstance(cell_value, str):
        try:
            cell_date = pd.to_datetime(cell_value).date()
        except Exception:
            continue
    else:
        continue
    if cell_date == first_exit_date:
        exit_startrow = row
        break

if exit_startrow is None:
    exit_startrow = ws_exit.max_row + 1

# Skip overwriting for dates between 25/12/2025 and 05/01/2026 (inclusive)
skip_exit_dates = pd.date_range('2024-12-25', '2025-01-05').date

for i, (date, row) in enumerate(exit_hourly_pivot_df.iterrows()):
    if date in skip_exit_dates:
        continue  # skip these dates
    excel_row = exit_startrow + i
    ws_exit.cell(row=excel_row, column=1, value=date)
    for j, col in enumerate(exit_hourly_pivot_df.columns, start=2):
        ws_exit.cell(row=excel_row, column=j, value=row[col])


if today.weekday() == 0:  # Monday is 0
    wb_exit.save(excel_path)






# %%
# Create a copy to avoid SettingWithCopyWarning
df_entries_hourly = df_entries.copy()
df_entries_hourly['date'] = df_entries_hourly['CheckInStarted'].dt.date
df_entries_hourly['hour'] = df_entries_hourly['CheckInStarted'].dt.strftime('%H:00')

# Pivot to get dates as rows and hours as columns
entries_pivot = df_entries_hourly.pivot_table(
    index='date',
    columns='hour',
    values='BookingReference',
    aggfunc='count',
    fill_value=0
)

# Ensure columns are ordered from 00:00 to 23:00
hour_columns = [f"{str(h).zfill(2)}:00" for h in range(24)]
entries_pivot = entries_pivot.reindex(columns=hour_columns, fill_value=0)

# Add a 'Total' column summing across all hour columns
entries_pivot['Total'] = entries_pivot[hour_columns].sum(axis=1)


print(entries_pivot.head())

entries_pivot.to_csv('FastPark Entries Actuals Hourly.csv')

df_exits_hourly = df_exits.copy()
df_exits_hourly['date'] = df_exits_hourly['ActualCheckedOutDate'].dt.date
df_exits_hourly['hour'] = df_exits_hourly['ActualCheckedOutDate'].dt.strftime('%H:00')

# Pivot to get dates as rows and hours as columns
exits_pivot = df_exits_hourly.pivot_table(
    index='date',
    columns='hour',
    values='BookingReference',
    aggfunc='count',
    fill_value=0
)

# Ensure columns are ordered from 00:00 to 23:00
exits_pivot = exits_pivot.reindex(columns=hour_columns, fill_value=0)

# Add a 'Total' column summing across all hour columns
exits_pivot['Total'] = exits_pivot[hour_columns].sum(axis=1)


exits_pivot.to_csv('FastPark Exits Actuals Hourly.csv')

# print(exits_pivot.head())

excel_path = "FastPark Forecast Accuracy Tracker NEW.xlsx"
sheet_name = "Entries Actuals"

# Load the workbook and worksheet
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# Find the first date in entries_pivot
first_date = entries_pivot.index.min()
last_date = entries_pivot.index.max()


# Find the column headers in the sheet (assume headers are in the first row)
header_row = 1
headers = [cell.value for cell in ws[header_row]]

# Find the date column (should be the first column)
date_col_idx = 1  # openpyxl is 1-based

# Find the starting row to write (match first_date)
startrow = None
for row in range(2, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=date_col_idx).value
    if isinstance(cell_value, datetime):
        cell_date = cell_value.date()
    elif isinstance(cell_value, str):
        try:
            cell_date = pd.to_datetime(cell_value).date()
        except Exception:
            continue
    else:
        continue
    if cell_date == first_date:
        startrow = row
        break

if startrow is None:
    # If not found, append at the end
    startrow = ws.max_row + 1

# Write the data row by row, overwriting only the matching dates
for i, (date, row) in enumerate(entries_pivot.iterrows()):
    excel_row = startrow + i
    # Write date in the first column
    ws.cell(row=excel_row, column=1, value=date)
    # Write each hour column (assume order matches headers)
    for j, col in enumerate(entries_pivot.columns, start=2):
        ws.cell(row=excel_row, column=j, value=row[col])

# Save the workbook
wb.save(excel_path)

# Repeat for exits_pivot, sheet name is "Exit Forecast"
exit_sheet_name = "Exits Actuals"
wb_exit = load_workbook(excel_path)
ws_exit = wb_exit[exit_sheet_name]

first_exit_date = exits_pivot.index.min()
last_exit_date = exits_pivot.index.max()

# Find the column headers in the sheet (assume headers are in the first row)
exit_header_row = 1
exit_headers = [cell.value for cell in ws_exit[exit_header_row]]

exit_date_col_idx = 1  # openpyxl is 1-based

exit_startrow = None
for row in range(2, ws_exit.max_row + 1):
    cell_value = ws_exit.cell(row=row, column=exit_date_col_idx).value
    if isinstance(cell_value, datetime):
        cell_date = cell_value.date()
    elif isinstance(cell_value, str):
        try:
            cell_date = pd.to_datetime(cell_value).date()
        except Exception:
            continue
    else:
        continue
    if cell_date == first_exit_date:
        exit_startrow = row
        break

if exit_startrow is None:
    exit_startrow = ws_exit.max_row + 1

for i, (date, row) in enumerate(exits_pivot.iterrows()):
    excel_row = exit_startrow + i
    ws_exit.cell(row=excel_row, column=1, value=date)
    for j, col in enumerate(exits_pivot.columns, start=2):
        ws_exit.cell(row=excel_row, column=j, value=row[col])

wb_exit.save(excel_path)
