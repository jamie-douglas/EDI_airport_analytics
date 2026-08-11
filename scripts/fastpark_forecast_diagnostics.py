import sys
import pathlib
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
import numpy as np

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------
sys.path.append(str(pathlib.Path(__file__).resolve().parents[1]))

from modules.utils.query import query


# ---------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------
OUTPUT_DIR = Path("outputs/fastpark_forecast_diagnostics")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

RUN_DATE_STR = datetime.now().strftime("%Y-%m-%d")
EXCEL_OUTPUT_PATH = OUTPUT_DIR / f"FastPark_Forecast_Diagnostics_V2{RUN_DATE_STR}.xlsx"

# Set to True if you still want the separate CSVs as well as the Excel workbook
SAVE_SEPARATE_CSVS = False

WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

DURATION_ORDER = [
    "1 day",
    "2 days",
    "3-6 days",
    "7-9 days",
    "10-13 days",
    "14-20 days",
    "21-29 days",
    "30-59 days",
    "60-89 days",
    "90+ days",
    "Unknown",
]

DEFAULT_HORIZONS = range(0, 29)


# ---------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------
def safe_divide(numerator, denominator):
    return np.where(
        denominator > 0,
        numerator / denominator,
        np.nan
    )


def duration_bucket(x):
    if pd.isna(x):
        return "Unknown"
    elif x == 1:
        return "1 day"
    elif x == 2:
        return "2 days"
    elif 3 <= x <= 6:
        return "3-6 days"
    elif 7 <= x <= 9:
        return "7-9 days"
    elif 10 <= x <= 13:
        return "10-13 days"
    elif 14 <= x <= 20:
        return "14-20 days"
    elif 21 <= x <= 29:
        return "21-29 days"
    elif 30 <= x <= 59:
        return "30-59 days"
    elif 60 <= x <= 89:
        return "60-89 days"
    else:
        return "90+ days"


def print_section(title):
    print("\n" + "=" * 100)
    print(title)
    print("=" * 100)


def save_csv(df, filename):
    """
    Optional CSV saver.
    Controlled by SAVE_SEPARATE_CSVS.
    """
    if not SAVE_SEPARATE_CSVS:
        return

    path = OUTPUT_DIR / filename
    df.to_csv(path, index=True)
    print(f"Saved CSV: {path}")

def clean_for_excel(df):
    """
    Clean a DataFrame before writing to Excel:
    - copy it
    - convert Period columns to string
    - remove timezone info from datetime columns
    - replace inf with NaN
    """
    if df is None:
        return pd.DataFrame()

    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame(df)

    df = df.copy()

    for col in df.columns:
        if isinstance(df[col].dtype, pd.PeriodDtype):
            df[col] = df[col].astype(str)

        if isinstance(df[col].dtype, pd.DatetimeTZDtype):
            df[col] = df[col].dt.tz_localize(None)

    df = df.replace([np.inf, -np.inf], np.nan)

    return df


def make_sheet_name(name, used_names):
    """
    Excel sheet names must be <= 31 characters and cannot contain:
    colon, backslash, slash, question mark, star, square brackets
    """

    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]

    clean_name = str(name)
    for char in invalid_chars:
        clean_name = clean_name.replace(char, "_")

    clean_name = clean_name[:31]

    original_name = clean_name
    counter = 1

    while clean_name in used_names:
        suffix = f"_{counter}"
        clean_name = original_name[:31 - len(suffix)] + suffix
        counter += 1

    used_names.add(clean_name)
    return clean_name


def autosize_excel_columns(writer, sheet_name, df, max_width=45):
    """
    Auto-size columns in an Excel sheet using openpyxl.
    """
    worksheet = writer.sheets[sheet_name]

    if df.empty:
        return

    for idx, col in enumerate(df.columns, start=1):
        col_as_str = str(col)

        try:
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col_as_str)
            )
        except Exception:
            max_len = len(col_as_str)

        adjusted_width = min(max_len + 2, max_width)
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = adjusted_width


def format_excel_sheet(writer, sheet_name, df):
    """
    Apply basic formatting:
    - bold headers
    - freeze top row
    - autofilter
    - sensible number formats
    """
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    if df.empty:
        return

    header_fill = "1F4E78"
    header_font = "FFFFFF"

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color=header_font)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    if max_row >= 1 and max_col >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    # Number formats
    percentage_keywords = ["pct", "percentage", "penetration", "visible", "uplift", "mape", "ape"]
    date_keywords = ["date", "day", "week_start", "service_day", "entry_day", "exit_day"]

    for col_idx, column_name in enumerate(df.columns, start=1):
        column_name_lower = str(column_name).lower()
        excel_col = get_column_letter(col_idx)

        if any(keyword in column_name_lower for keyword in percentage_keywords):
            for row in range(2, max_row + 1):
                worksheet[f"{excel_col}{row}"].number_format = "0.00"

        elif any(keyword in column_name_lower for keyword in date_keywords):
            for row in range(2, max_row + 1):
                worksheet[f"{excel_col}{row}"].number_format = "yyyy-mm-dd"

        else:
            if pd.api.types.is_numeric_dtype(df[column_name]):
                for row in range(2, max_row + 1):
                    worksheet[f"{excel_col}{row}"].number_format = "#,##0.00"

    autosize_excel_columns(writer, sheet_name, df)


def add_summary_sheet(
    workbook_sheets,
    start,
    end,
    bookings,
    actual_entries,
    actual_exits,
    flight_df,
    visibility_summary,
    volatility,
    uplift_backtest
):
    """
    Create a compact executive summary sheet.
    This gives you one front-page view of the diagnostics.
    """
    summary_rows = []

    summary_rows.extend([
        {"section": "Run Info", "metric": "analysis_start", "value": start},
        {"section": "Run Info", "metric": "analysis_end", "value": end},
        {"section": "Run Info", "metric": "run_datetime", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"section": "Data Volumes", "metric": "bookings_loaded", "value": len(bookings)},
        {"section": "Data Volumes", "metric": "actual_entries_loaded", "value": len(actual_entries)},
        {"section": "Data Volumes", "metric": "actual_exits_loaded", "value": len(actual_exits)},
        {"section": "Data Volumes", "metric": "flight_rows_loaded", "value": len(flight_df)},
    ])

    # Current D14 hard-coded uplift from your script
    current_entry_d14 = 56
    current_exit_d14 = 14

    # Recommended D14 from analysis
    d14_row = visibility_summary.loc[visibility_summary["horizon"] == "D14"].copy()

    if not d14_row.empty:
        summary_rows.extend([
            {
                "section": "D14 Recommended Uplift",
                "metric": "current_entry_d14_uplift_pct",
                "value": current_entry_d14,
            },
            {
                "section": "D14 Recommended Uplift",
                "metric": "recommended_entry_d14_uplift_median_pct",
                "value": d14_row["entry_uplift_median"].iloc[0],
            },
            {
                "section": "D14 Recommended Uplift",
                "metric": "current_exit_d14_uplift_pct",
                "value": current_exit_d14,
            },
            {
                "section": "D14 Recommended Uplift",
                "metric": "recommended_exit_d14_uplift_median_pct",
                "value": d14_row["exit_uplift_median"].iloc[0],
            },
            {
                "section": "D14 Recommended Uplift",
                "metric": "recommended_implied_exit_d14_uplift_median_pct",
                "value": d14_row["implied_exit_uplift_median"].iloc[0],
            },
        ])

    # Best penetration stability helper
    if volatility is not None and not volatility.empty:
        vol = volatility.copy()

        entry_vol = vol[vol["metric"].str.contains("entry", case=False, na=False)].copy()
        exit_vol = vol[vol["metric"].str.contains("exit", case=False, na=False)].copy()

        if not entry_vol.empty:
            best_entry = entry_vol.sort_values("coefficient_of_variation").iloc[0]
            summary_rows.append({
                "section": "Penetration Stability",
                "metric": "lowest_entry_penetration_cv_metric",
                "value": best_entry["metric"],
            })
            summary_rows.append({
                "section": "Penetration Stability",
                "metric": "lowest_entry_penetration_cv",
                "value": best_entry["coefficient_of_variation"],
            })

        if not exit_vol.empty:
            best_exit = exit_vol.sort_values("coefficient_of_variation").iloc[0]
            summary_rows.append({
                "section": "Penetration Stability",
                "metric": "lowest_exit_penetration_cv_metric",
                "value": best_exit["metric"],
            })
            summary_rows.append({
                "section": "Penetration Stability",
                "metric": "lowest_exit_penetration_cv",
                "value": best_exit["coefficient_of_variation"],
            })

    # Current uplift backtest headline
    if uplift_backtest is not None and not uplift_backtest.empty:
        d14_bt = uplift_backtest.loc[uplift_backtest["horizon"] == "D14"].copy()

        if not d14_bt.empty:
            summary_rows.extend([
                {
                    "section": "Current Uplift Backtest",
                    "metric": "entry_d14_mape",
                    "value": d14_bt["entry_mape"].iloc[0],
                },
                {
                    "section": "Current Uplift Backtest",
                    "metric": "exit_d14_mape",
                    "value": d14_bt["exit_mape"].iloc[0],
                },
                {
                    "section": "Current Uplift Backtest",
                    "metric": "entry_d14_mean_error",
                    "value": d14_bt["entry_mean_error"].iloc[0],
                },
                {
                    "section": "Current Uplift Backtest",
                    "metric": "exit_d14_mean_error",
                    "value": d14_bt["exit_mean_error"].iloc[0],
                },
            ])

    summary_df = pd.DataFrame(summary_rows)

    workbook_sheets["01_Summary"] = summary_df

    return workbook_sheets


def add_recommended_arrays_sheet(workbook_sheets, visibility_summary):
    """
    Creates a sheet with D1-D14 uplift arrays ready to copy into the forecast script.
    """
    d1_d14 = visibility_summary[
        visibility_summary["horizon"].isin([f"D{i}" for i in range(1, 15)])
    ].copy()

    d1_d14["horizon_num"] = d1_d14["horizon"].str.replace("D", "", regex=False).astype(int)
    d1_d14 = d1_d14.sort_values("horizon_num")

    entry_array = d1_d14["entry_uplift_median"].round(2).tolist()
    exit_array = d1_d14["exit_uplift_median"].round(2).tolist()
    implied_exit_array = d1_d14["implied_exit_uplift_median"].round(2).tolist()

    arrays_df = pd.DataFrame({
        "item": [
            "entry_increase_median_D1_D14",
            "exit_increase_median_D1_D14",
            "implied_exit_increase_median_D1_D14",
        ],
        "python_array": [
            str(entry_array),
            str(exit_array),
            str(implied_exit_array),
        ],
    })

    uplift_table = d1_d14[
        [
            "horizon",
            "entry_uplift_mean",
            "entry_uplift_median",
            "exit_uplift_mean",
            "exit_uplift_median",
            "implied_exit_uplift_mean",
            "implied_exit_uplift_median",
            "entry_pct_actual_visible_median",
            "exit_pct_actual_visible_median",
            "implied_exit_pct_actual_visible_median",
        ]
    ].copy()

    # Put arrays first, then detailed table underneath
    blank = pd.DataFrame([{}])
    combined = pd.concat(
        [
            arrays_df,
            blank,
            pd.DataFrame({"item": ["Detailed horizon table below"], "python_array": [""]}),
            uplift_table.rename(columns={"horizon": "item"}),
        ],
        ignore_index=True,
        sort=False
    )

    workbook_sheets["02_Recommended_Uplifts"] = combined

    return workbook_sheets


def write_diagnostics_workbook(workbook_sheets, output_path):
    """
    Write all diagnostic outputs to one Excel workbook, one sheet per output.
    """
    used_sheet_names = set()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for requested_sheet_name, df in workbook_sheets.items():
            sheet_name = make_sheet_name(requested_sheet_name, used_sheet_names)
            clean_df = clean_for_excel(df)

            clean_df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_excel_sheet(writer, sheet_name, clean_df)

    print(f"\nSaved Excel workbook: {output_path}")

# ---------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------
def load_airportx_fastpark_bookings(start, end, deduplicate=True):
    """
    Load FastPark bookings from AirportX.v_Bookings.
    This pulls by entryDate over the historical analysis window.

    Important:
    - This is entry-date anchored.
    - We use exitDate later for exit-side analysis.
    """
    df = query(
        table="AirportX.v_Bookings",
        columns=[
            "bookingUuid",
            "bookingId AS [Booking ID]",
            "createdAt AS [Creation Date]",
            "entryDate",
            "exitDate",
            "Duration",
            "status",
            "assetName",
            "productName",
            "productCode",
            "bookingTotal",
            "leadtime",
        ],
        where=[
            "assetName = 'FastPark'",
            "status = 'B'"
        ],
        date_column="entryDate",
        start=start,
        end=end,
    )

    df = pd.DataFrame(df).copy()

    if df.empty:
        return df

    for col in ["Creation Date", "entryDate", "exitDate"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    df["Duration"] = pd.to_numeric(df["Duration"], errors="coerce")

    if deduplicate and "bookingUuid" in df.columns:
        df = (
            df.sort_values(["bookingUuid", "Creation Date"])
              .drop_duplicates(subset=["bookingUuid"], keep="first")
              .reset_index(drop=True)
        )

    return df


def load_fastpark_actuals(start, end):
    """
    Load actual FastPark entries and exits from FastPark.v_EntryAndExits.
    """
    cols = [
        "BookingReference",
        "CheckInStarted",
        "ActualCheckedOutDate",
    ]

    entry_df = query(
        table="FastPark.v_EntryAndExits",
        columns=cols,
        where=[],
        date_column="CheckInStarted",
        start=start,
        end=end,
    )

    exit_df = query(
        table="FastPark.v_EntryAndExits",
        columns=cols,
        where=[],
        date_column="ActualCheckedOutDate",
        start=start,
        end=end,
    )

    entry_df = pd.DataFrame(entry_df).copy()
    exit_df = pd.DataFrame(exit_df).copy()

    if not entry_df.empty:
        entry_df["CheckInStarted"] = pd.to_datetime(entry_df["CheckInStarted"], errors="coerce")
        entry_df = entry_df.dropna(subset=["BookingReference", "CheckInStarted"])
        entry_df = entry_df.drop_duplicates(subset=["BookingReference"])
        entry_df["entry_day"] = entry_df["CheckInStarted"].dt.normalize()
        entry_df["weekday"] = entry_df["entry_day"].dt.weekday
        entry_df["weekday_name"] = entry_df["weekday"].map(lambda x: WEEKDAY_NAMES[x])

    if not exit_df.empty:
        exit_df["ActualCheckedOutDate"] = pd.to_datetime(exit_df["ActualCheckedOutDate"], errors="coerce")
        exit_df = exit_df.dropna(subset=["BookingReference", "ActualCheckedOutDate"])
        exit_df = exit_df.drop_duplicates(subset=["BookingReference"])
        exit_df["exit_day"] = exit_df["ActualCheckedOutDate"].dt.normalize()
        exit_df["weekday"] = exit_df["exit_day"].dt.weekday
        exit_df["weekday_name"] = exit_df["weekday"].map(lambda x: WEEKDAY_NAMES[x])

    return entry_df, exit_df


def load_flight_performance(start, end):
    """
    Load EAL.FlightPerformance for passenger denominator analysis.
    """
    df = query(
        table="EAL.FlightPerformance",
        columns=[
            "ActualDateTime_Local",
            "ArrDeptureCode",
            "IsPassengerFlight",
            "Pax_MostConfident",
        ],
        where=[
            "IsPassengerFlight = 1"
        ],
        date_column="ActualDateTime_Local",
        start=start,
        end=end,
    )

    df = pd.DataFrame(df).copy()

    if df.empty:
        return df

    df["ActualDateTime_Local"] = pd.to_datetime(df["ActualDateTime_Local"], errors="coerce")
    df["Pax_MostConfident"] = pd.to_numeric(df["Pax_MostConfident"], errors="coerce").fillna(0)
    df["service_day"] = df["ActualDateTime_Local"].dt.normalize()
    df["weekday"] = df["service_day"].dt.weekday
    df["weekday_name"] = df["weekday"].map(lambda x: WEEKDAY_NAMES[x])

    return df


# ---------------------------------------------------------------------
# Enrichment
# ---------------------------------------------------------------------
def prepare_bookings(df):
    """
    Prepare booking-level lead-time and duration variables.
    """
    df = df.copy()

    if df.empty:
        return df

    df = df.dropna(subset=["bookingUuid", "Creation Date", "entryDate", "exitDate"])

    df["creation_day"] = df["Creation Date"].dt.normalize()
    df["entry_day"] = df["entryDate"].dt.normalize()
    df["exit_day"] = df["exitDate"].dt.normalize()

    df["lead_to_entry_days_exact"] = (
        (df["entryDate"] - df["Creation Date"]).dt.total_seconds() / 86400
    )

    df["lead_to_exit_days_exact"] = (
        (df["exitDate"] - df["Creation Date"]).dt.total_seconds() / 86400
    )

    df["lead_to_entry_calendar_days"] = (
        df["entry_day"] - df["creation_day"]
    ).dt.days

    df["lead_to_exit_calendar_days"] = (
        df["exit_day"] - df["creation_day"]
    ).dt.days

    df["invalid_entry_lead"] = df["lead_to_entry_days_exact"] < 0
    df["invalid_exit_lead"] = df["lead_to_exit_days_exact"] < 0

    inferred_duration = (df["exit_day"] - df["entry_day"]).dt.days

    df["Duration_clean"] = pd.to_numeric(df["Duration"], errors="coerce")
    df.loc[df["Duration_clean"].isna(), "Duration_clean"] = inferred_duration[df["Duration_clean"].isna()]
    df["Duration_clean"] = pd.to_numeric(df["Duration_clean"], errors="coerce")

    df["duration_bucket"] = df["Duration_clean"].apply(duration_bucket)

    df["entry_weekday"] = df["entry_day"].dt.weekday
    df["entry_weekday_name"] = df["entry_weekday"].map(lambda x: WEEKDAY_NAMES[x])

    df["exit_weekday"] = df["exit_day"].dt.weekday
    df["exit_weekday_name"] = df["exit_weekday"].map(lambda x: WEEKDAY_NAMES[x])

    df["entry_month"] = df["entry_day"].dt.to_period("M").astype(str)
    df["entry_week"] = df["entry_day"].dt.to_period("W-MON").astype(str)

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------
# Lead-time visibility analysis
# ---------------------------------------------------------------------
def build_booking_visibility_curves(bookings_df, actual_entries_df, actual_exits_df, horizons=DEFAULT_HORIZONS):
    """
    For each service day and horizon D1-D28, calculate:
    - entries already booked by Dn before entry
    - exits already booked by Dn before exit
    - implied exits from entry-side seen bookings
    - % coverage of actuals
    - recommended uplift
    """
    bookings = bookings_df.copy()

    bookings = bookings.loc[
        (~bookings["invalid_entry_lead"]) &
        (~bookings["invalid_exit_lead"])
    ].copy()

    actual_entry_counts = (
        actual_entries_df
        .groupby("entry_day")["BookingReference"]
        .nunique()
        .rename("actual_entries")
    )

    actual_exit_counts = (
        actual_exits_df
        .groupby("exit_day")["BookingReference"]
        .nunique()
        .rename("actual_exits")
    )

    final_booked_entries = (
        bookings
        .groupby("entry_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_entries")
    )

    final_booked_exits = (
        bookings
        .groupby("exit_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_exits")
    )

    entry_daily_frames = []
    exit_daily_frames = []
    implied_exit_daily_frames = []

    summary_rows = []

    for h in horizons:
        horizon_label = f"D{h}"

        # ---------------------------------------------------------
        # Entry-visible bookings by Dn before entry
        # ---------------------------------------------------------
        entry_seen = bookings.loc[
            bookings["creation_day"] <= bookings["entry_day"] - pd.Timedelta(days=h)
        ].copy()

        entry_seen_counts = (
            entry_seen
            .groupby("entry_day")["bookingUuid"]
            .nunique()
            .rename("seen_entries")
        )

        entry_compare = pd.concat(
            [
                actual_entry_counts,
                final_booked_entries,
                entry_seen_counts,
            ],
            axis=1
        ).fillna(0)

        entry_compare["horizon"] = horizon_label
        entry_compare["pct_actual_visible"] = safe_divide(
            entry_compare["seen_entries"],
            entry_compare["actual_entries"]
        ) * 100

        entry_compare["pct_final_booking_visible"] = safe_divide(
            entry_compare["seen_entries"],
            entry_compare["final_booked_entries"]
        ) * 100

        entry_compare["recommended_uplift_pct"] = np.where(
            entry_compare["seen_entries"] > 0,
            ((entry_compare["actual_entries"] / entry_compare["seen_entries"]) - 1) * 100,
            np.nan
        )

        entry_compare["recommended_uplift_pct"] = entry_compare["recommended_uplift_pct"].clip(lower=0)

        entry_compare = (
            entry_compare
            .reset_index()
            .rename(columns={"entry_day": "service_day"})
        )

        entry_daily_frames.append(entry_compare)

        # ---------------------------------------------------------
        # Exit-visible bookings by Dn before exit
        # ---------------------------------------------------------
        exit_seen = bookings.loc[
            bookings["creation_day"] <= bookings["exit_day"] - pd.Timedelta(days=h)
        ].copy()

        exit_seen_counts = (
            exit_seen
            .groupby("exit_day")["bookingUuid"]
            .nunique()
            .rename("seen_exits")
        )

        exit_compare = pd.concat(
            [
                actual_exit_counts,
                final_booked_exits,
                exit_seen_counts,
            ],
            axis=1
        ).fillna(0)

        exit_compare["horizon"] = horizon_label
        exit_compare["pct_actual_visible"] = safe_divide(
            exit_compare["seen_exits"],
            exit_compare["actual_exits"]
        ) * 100

        exit_compare["pct_final_booking_visible"] = safe_divide(
            exit_compare["seen_exits"],
            exit_compare["final_booked_exits"]
        ) * 100

        exit_compare["recommended_uplift_pct"] = np.where(
            exit_compare["seen_exits"] > 0,
            ((exit_compare["actual_exits"] / exit_compare["seen_exits"]) - 1) * 100,
            np.nan
        )

        exit_compare["recommended_uplift_pct"] = exit_compare["recommended_uplift_pct"].clip(lower=0)

        exit_compare = (
            exit_compare
            .reset_index()
            .rename(columns={"exit_day": "service_day"})
        )

        exit_daily_frames.append(exit_compare)

        # ---------------------------------------------------------
        # Implied exits from bookings seen by Dn before entry
        # This mirrors the model logic where entries are booked first
        # but exits are implied by duration/exitDate.
        # ---------------------------------------------------------
        implied_exit_seen_counts = (
            entry_seen
            .groupby("exit_day")["bookingUuid"]
            .nunique()
            .rename("implied_seen_exits_from_entry_seen_bookings")
        )

        implied_exit_compare = pd.concat(
            [
                actual_exit_counts,
                final_booked_exits,
                implied_exit_seen_counts,
            ],
            axis=1
        ).fillna(0)

        implied_exit_compare["horizon"] = horizon_label

        implied_exit_compare["pct_actual_visible"] = safe_divide(
            implied_exit_compare["implied_seen_exits_from_entry_seen_bookings"],
            implied_exit_compare["actual_exits"]
        ) * 100

        implied_exit_compare["pct_final_booking_visible"] = safe_divide(
            implied_exit_compare["implied_seen_exits_from_entry_seen_bookings"],
            implied_exit_compare["final_booked_exits"]
        ) * 100

        implied_exit_compare["recommended_uplift_pct"] = np.where(
            implied_exit_compare["implied_seen_exits_from_entry_seen_bookings"] > 0,
            (
                implied_exit_compare["actual_exits"] /
                implied_exit_compare["implied_seen_exits_from_entry_seen_bookings"]
                - 1
            ) * 100,
            np.nan
        )

        implied_exit_compare["recommended_uplift_pct"] = implied_exit_compare["recommended_uplift_pct"].clip(lower=0)

        implied_exit_compare = (
            implied_exit_compare
            .reset_index()
            .rename(columns={"exit_day": "service_day"})
        )

        implied_exit_daily_frames.append(implied_exit_compare)

        # ---------------------------------------------------------
        # Summary row
        # ---------------------------------------------------------
        summary_rows.append({
            "horizon": horizon_label,

            "entry_pct_actual_visible_mean": entry_compare["pct_actual_visible"].mean(),
            "entry_pct_actual_visible_median": entry_compare["pct_actual_visible"].median(),
            "entry_uplift_mean": entry_compare["recommended_uplift_pct"].mean(),
            "entry_uplift_median": entry_compare["recommended_uplift_pct"].median(),

            "exit_pct_actual_visible_mean": exit_compare["pct_actual_visible"].mean(),
            "exit_pct_actual_visible_median": exit_compare["pct_actual_visible"].median(),
            "exit_uplift_mean": exit_compare["recommended_uplift_pct"].mean(),
            "exit_uplift_median": exit_compare["recommended_uplift_pct"].median(),

            "implied_exit_pct_actual_visible_mean": implied_exit_compare["pct_actual_visible"].mean(),
            "implied_exit_pct_actual_visible_median": implied_exit_compare["pct_actual_visible"].median(),
            "implied_exit_uplift_mean": implied_exit_compare["recommended_uplift_pct"].mean(),
            "implied_exit_uplift_median": implied_exit_compare["recommended_uplift_pct"].median(),
        })

    entry_daily_detail = pd.concat(entry_daily_frames, ignore_index=True)
    exit_daily_detail = pd.concat(exit_daily_frames, ignore_index=True)
    implied_exit_daily_detail = pd.concat(implied_exit_daily_frames, ignore_index=True)

    summary = pd.DataFrame(summary_rows)

    return {
        "visibility_summary": summary,
        "entry_daily_detail": entry_daily_detail,
        "exit_daily_detail": exit_daily_detail,
        "implied_exit_daily_detail": implied_exit_daily_detail,
    }


# ---------------------------------------------------------------------
# Duration and weekday visibility
# ---------------------------------------------------------------------
def build_segmented_visibility(bookings_df, horizons=DEFAULT_HORIZONS):
    """
    Calculate booking visibility by:
    - duration bucket
    - weekday
    - month
    """
    bookings = bookings_df.copy()

    bookings = bookings.loc[
        (~bookings["invalid_entry_lead"]) &
        (~bookings["invalid_exit_lead"])
    ].copy()

    rows = []

    for h in horizons:
        horizon_label = f"D{h}"

        seen = bookings.loc[
            bookings["creation_day"] <= bookings["entry_day"] - pd.Timedelta(days=h)
        ].copy()

        final_by_segment = (
            bookings
            .groupby(["entry_day", "duration_bucket", "entry_weekday_name", "entry_month"])["bookingUuid"]
            .nunique()
            .rename("final_bookings")
            .reset_index()
        )

        seen_by_segment = (
            seen
            .groupby(["entry_day", "duration_bucket", "entry_weekday_name", "entry_month"])["bookingUuid"]
            .nunique()
            .rename("seen_bookings")
            .reset_index()
        )

        compare = final_by_segment.merge(
            seen_by_segment,
            on=["entry_day", "duration_bucket", "entry_weekday_name", "entry_month"],
            how="left"
        )

        compare["seen_bookings"] = compare["seen_bookings"].fillna(0)
        compare["pct_visible"] = safe_divide(
            compare["seen_bookings"],
            compare["final_bookings"]
        ) * 100

        compare["horizon"] = horizon_label
        rows.append(compare)

    detail = pd.concat(rows, ignore_index=True)

    by_duration = (
        detail
        .groupby(["horizon", "duration_bucket"])
        .agg(
            avg_pct_visible=("pct_visible", "mean"),
            median_pct_visible=("pct_visible", "median"),
            days=("entry_day", "nunique"),
            final_bookings=("final_bookings", "sum"),
            seen_bookings=("seen_bookings", "sum"),
        )
        .reset_index()
    )

    by_weekday = (
        detail
        .groupby(["horizon", "entry_weekday_name"])
        .agg(
            avg_pct_visible=("pct_visible", "mean"),
            median_pct_visible=("pct_visible", "median"),
            days=("entry_day", "nunique"),
            final_bookings=("final_bookings", "sum"),
            seen_bookings=("seen_bookings", "sum"),
        )
        .reset_index()
    )

    by_month = (
        detail
        .groupby(["horizon", "entry_month"])
        .agg(
            avg_pct_visible=("pct_visible", "mean"),
            median_pct_visible=("pct_visible", "median"),
            days=("entry_day", "nunique"),
            final_bookings=("final_bookings", "sum"),
            seen_bookings=("seen_bookings", "sum"),
        )
        .reset_index()
    )

    return {
        "segmented_visibility_detail": detail,
        "visibility_by_duration": by_duration,
        "visibility_by_weekday": by_weekday,
        "visibility_by_month": by_month,
    }

def build_visibility_by_horizon_month(bookings_df, horizons=DEFAULT_HORIZONS):
    """
    Shows how visibility changes by horizon and month.

    Example output:

              2026-04  2026-05  2026-06
    D1          97.3     96.8     98.1
    D2          94.1     93.4     92.7
    ...
    D14         71.2     63.5     55.8
    ...
    """

    bookings = bookings_df.copy()

    bookings = bookings.loc[
        (~bookings["invalid_entry_lead"]) &
        (~bookings["invalid_exit_lead"])
    ].copy()

    final_bookings = (
        bookings
        .groupby(["entry_day", "entry_month"])["bookingUuid"]
        .nunique()
        .rename("final_bookings")
        .reset_index()
    )

    rows = []

    for h in horizons:

        seen = bookings.loc[
            bookings["creation_day"] <= bookings["entry_day"] - pd.Timedelta(days=h)
        ].copy()

        seen_bookings = (
            seen
            .groupby(["entry_day", "entry_month"])["bookingUuid"]
            .nunique()
            .rename("seen_bookings")
            .reset_index()
        )

        compare = final_bookings.merge(
            seen_bookings,
            on=["entry_day", "entry_month"],
            how="left"
        )

        compare["seen_bookings"] = compare["seen_bookings"].fillna(0)

        compare["pct_visible"] = np.where(
            compare["final_bookings"] > 0,
            compare["seen_bookings"] / compare["final_bookings"] * 100,
            np.nan
        )

        monthly = (
            compare
            .groupby("entry_month")["pct_visible"]
            .median()
        )

        for month, value in monthly.items():

            rows.append({
                "horizon": f"D{h}",
                "month": month,
                "pct_visible": value
            })

    result = pd.DataFrame(rows)

    pivot = result.pivot(
        index="horizon",
        columns="month",
        values="pct_visible"
    )

    horizon_sort = (
        pivot.index
        .str.replace("D", "", regex=False)
        .astype(int)
    )

    pivot["sort"] = horizon_sort
    pivot = pivot.sort_values("sort")
    pivot = pivot.drop(columns="sort")

    pivot = pivot.round(2)

    return pivot.reset_index()

# ---------------------------------------------------------------------
# Lead time analysis
# ---------------------------------------------------------------------
def build_leadtime_distribution(bookings_df):
    """
    Distribution of booking lead times.

    Used to validate booking behaviour against Tableau.
    """

    df = bookings_df.copy()

    leadtime_distribution = (
        df.groupby("lead_to_entry_calendar_days")
        .agg(
            bookings=("bookingUuid", "nunique")
        )
        .reset_index()
        .sort_values("lead_to_entry_calendar_days")
    )

    leadtime_distribution["booking_pct"] = (
        leadtime_distribution["bookings"]
        / leadtime_distribution["bookings"].sum()
        * 100
    )

    leadtime_distribution["cumulative_pct"] = (
        leadtime_distribution["booking_pct"]
        .cumsum()
    )

    return leadtime_distribution

# ---------------------------------------------------------------------
# Penetration analysis
# ---------------------------------------------------------------------
def build_daily_penetration(actual_entries_df, actual_exits_df, flight_df):
    """
    Calculate daily entry and exit penetration:
    - entry penetration = FastPark entries / departure pax
    - exit penetration = FastPark exits / arrival pax
    """
    departures = flight_df.loc[
        flight_df["ArrDeptureCode"] == "D"
    ].copy()

    arrivals = flight_df.loc[
        flight_df["ArrDeptureCode"] == "A"
    ].copy()

    dep_pax = (
        departures
        .groupby("service_day")["Pax_MostConfident"]
        .sum()
        .rename("departure_pax")
    )

    arr_pax = (
        arrivals
        .groupby("service_day")["Pax_MostConfident"]
        .sum()
        .rename("arrival_pax")
    )

    entry_counts = (
        actual_entries_df
        .groupby("entry_day")["BookingReference"]
        .nunique()
        .rename("actual_entries")
    )

    exit_counts = (
        actual_exits_df
        .groupby("exit_day")["BookingReference"]
        .nunique()
        .rename("actual_exits")
    )

    daily = pd.concat(
        [
            dep_pax,
            arr_pax,
            entry_counts,
            exit_counts,
        ],
        axis=1
    ).fillna(0)

    daily.index.name = "service_day"
    daily = daily.reset_index()
    daily["service_day"] = pd.to_datetime(daily["service_day"])
    daily["weekday"] = daily["service_day"].dt.weekday
    daily["weekday_name"] = daily["weekday"].map(lambda x: WEEKDAY_NAMES[x])
    daily["week_start"] = daily["service_day"] - pd.to_timedelta(daily["weekday"], unit="D")
    daily["month"] = daily["service_day"].dt.to_period("M").astype(str)

    daily["entry_penetration_pct"] = safe_divide(
        daily["actual_entries"],
        daily["departure_pax"]
    ) * 100

    daily["exit_penetration_pct"] = safe_divide(
        daily["actual_exits"],
        daily["arrival_pax"]
    ) * 100

    return daily


def build_penetration_period_tables(daily_penetration_df):
    """
    Build weekly, rolling fortnightly, rolling 4-week and monthly penetration tables.
    """
    daily = daily_penetration_df.copy()

    weekly = (
        daily
        .groupby("week_start")
        .agg(
            departure_pax=("departure_pax", "sum"),
            arrival_pax=("arrival_pax", "sum"),
            actual_entries=("actual_entries", "sum"),
            actual_exits=("actual_exits", "sum"),
        )
        .reset_index()
        .sort_values("week_start")
    )

    weekly["entry_penetration_pct"] = safe_divide(
        weekly["actual_entries"],
        weekly["departure_pax"]
    ) * 100

    weekly["exit_penetration_pct"] = safe_divide(
        weekly["actual_exits"],
        weekly["arrival_pax"]
    ) * 100

    weekly["entry_penetration_2w_rolling_pct"] = (
        weekly["actual_entries"].rolling(2).sum() /
        weekly["departure_pax"].rolling(2).sum()
    ) * 100

    weekly["exit_penetration_2w_rolling_pct"] = (
        weekly["actual_exits"].rolling(2).sum() /
        weekly["arrival_pax"].rolling(2).sum()
    ) * 100

    weekly["entry_penetration_4w_rolling_pct"] = (
        weekly["actual_entries"].rolling(4).sum() /
        weekly["departure_pax"].rolling(4).sum()
    ) * 100

    weekly["exit_penetration_4w_rolling_pct"] = (
        weekly["actual_exits"].rolling(4).sum() /
        weekly["arrival_pax"].rolling(4).sum()
    ) * 100

    monthly = (
        daily
        .groupby("month")
        .agg(
            departure_pax=("departure_pax", "sum"),
            arrival_pax=("arrival_pax", "sum"),
            actual_entries=("actual_entries", "sum"),
            actual_exits=("actual_exits", "sum"),
        )
        .reset_index()
        .sort_values("month")
    )

    monthly["entry_penetration_pct"] = safe_divide(
        monthly["actual_entries"],
        monthly["departure_pax"]
    ) * 100

    monthly["exit_penetration_pct"] = safe_divide(
        monthly["actual_exits"],
        monthly["arrival_pax"]
    ) * 100

    weekday = (
        daily
        .groupby("weekday_name")
        .agg(
            departure_pax=("departure_pax", "sum"),
            arrival_pax=("arrival_pax", "sum"),
            actual_entries=("actual_entries", "sum"),
            actual_exits=("actual_exits", "sum"),
        )
        .reindex(WEEKDAY_NAMES)
        .rename_axis("weekday_name")
        .reset_index()
    )
    weekday["entry_penetration_pct"] = safe_divide(
        weekday["actual_entries"],
        weekday["departure_pax"]
    ) * 100

    weekday["exit_penetration_pct"] = safe_divide(
        weekday["actual_exits"],
        weekday["arrival_pax"]
    ) * 100

    return {
        "weekly_penetration": weekly,
        "monthly_penetration": monthly,
        "weekday_penetration": weekday,
    }


def build_penetration_volatility_table(weekly_penetration):
    """
    Measure volatility between weekly, 2-week rolling and 4-week rolling penetration.
    This helps decide whether 2 weeks is too noisy.
    """
    weekly = weekly_penetration.copy()

    rows = []

    for metric in [
        "entry_penetration_pct",
        "exit_penetration_pct",
        "entry_penetration_2w_rolling_pct",
        "exit_penetration_2w_rolling_pct",
        "entry_penetration_4w_rolling_pct",
        "exit_penetration_4w_rolling_pct",
    ]:
        s = weekly[metric].dropna()

        rows.append({
            "metric": metric,
            "periods": len(s),
            "mean": s.mean(),
            "median": s.median(),
            "std_dev": s.std(),
            "min": s.min(),
            "max": s.max(),
            "range": s.max() - s.min(),
            "coefficient_of_variation": s.std() / s.mean() if s.mean() != 0 else np.nan,
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------
# Backtest current uplift vs recommended uplift
# ---------------------------------------------------------------------
def build_uplift_backtest(
    bookings_df,
    actual_entries_df,
    actual_exits_df,
    entry_uplift_array=None,
    exit_uplift_array=None,
    horizons=range(1, 15)
):
    """
    Backtest hard-coded uplift arrays against historical actuals.

    This answers:
    - If we use the current uplift at D1-D14, how close are we?
    - What bias do we have?
    - Do we over-forecast far out and under/over correct near term?
    """
    bookings = bookings_df.copy()

    bookings = bookings.loc[
        (~bookings["invalid_entry_lead"]) &
        (~bookings["invalid_exit_lead"])
    ].copy()

    if entry_uplift_array is None:
        entry_uplift_array = [5, 10, 15, 20, 25, 30, 35, 38, 41, 44, 47, 50, 53, 56]

    if exit_uplift_array is None:
        exit_uplift_array = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

    actual_entry_counts = (
        actual_entries_df
        .groupby("entry_day")["BookingReference"]
        .nunique()
        .rename("actual_entries")
    )

    actual_exit_counts = (
        actual_exits_df
        .groupby("exit_day")["BookingReference"]
        .nunique()
        .rename("actual_exits")
    )

    rows = []

    for i, h in enumerate(horizons):
        horizon_label = f"D{h}"

        entry_uplift = entry_uplift_array[i] if i < len(entry_uplift_array) else entry_uplift_array[-1]
        exit_uplift = exit_uplift_array[i] if i < len(exit_uplift_array) else exit_uplift_array[-1]

        seen_entry = bookings.loc[
            bookings["creation_day"] <= bookings["entry_day"] - pd.Timedelta(days=h)
        ].copy()

        seen_entry_counts = (
            seen_entry
            .groupby("entry_day")["bookingUuid"]
            .nunique()
            .rename("seen_entries")
        )

        entry_compare = pd.concat(
            [
                actual_entry_counts,
                seen_entry_counts,
            ],
            axis=1
        ).fillna(0)

        entry_compare["forecast_entries_current_uplift"] = (
            entry_compare["seen_entries"] * (1 + entry_uplift / 100)
        )

        entry_compare["entry_error"] = (
            entry_compare["forecast_entries_current_uplift"] -
            entry_compare["actual_entries"]
        )

        entry_compare["entry_abs_error"] = entry_compare["entry_error"].abs()

        entry_compare["entry_ape"] = safe_divide(
            entry_compare["entry_abs_error"],
            entry_compare["actual_entries"]
        ) * 100

        seen_exit = bookings.loc[
            bookings["creation_day"] <= bookings["exit_day"] - pd.Timedelta(days=h)
        ].copy()

        seen_exit_counts = (
            seen_exit
            .groupby("exit_day")["bookingUuid"]
            .nunique()
            .rename("seen_exits")
        )

        exit_compare = pd.concat(
            [
                actual_exit_counts,
                seen_exit_counts,
            ],
            axis=1
        ).fillna(0)

        exit_compare["forecast_exits_current_uplift"] = (
            exit_compare["seen_exits"] * (1 + exit_uplift / 100)
        )

        exit_compare["exit_error"] = (
            exit_compare["forecast_exits_current_uplift"] -
            exit_compare["actual_exits"]
        )

        exit_compare["exit_abs_error"] = exit_compare["exit_error"].abs()

        exit_compare["exit_ape"] = safe_divide(
            exit_compare["exit_abs_error"],
            exit_compare["actual_exits"]
        ) * 100

        rows.append({
            "horizon": horizon_label,
            "entry_current_uplift_pct": entry_uplift,
            "entry_mean_error": entry_compare["entry_error"].mean(),
            "entry_median_error": entry_compare["entry_error"].median(),
            "entry_mae": entry_compare["entry_abs_error"].mean(),
            "entry_mape": entry_compare["entry_ape"].replace([np.inf, -np.inf], np.nan).mean(),

            "exit_current_uplift_pct": exit_uplift,
            "exit_mean_error": exit_compare["exit_error"].mean(),
            "exit_median_error": exit_compare["exit_error"].median(),
            "exit_mae": exit_compare["exit_abs_error"].mean(),
            "exit_mape": exit_compare["exit_ape"].replace([np.inf, -np.inf], np.nan).mean(),
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------
# Forecast transition/drop analysis
# ---------------------------------------------------------------------
def build_forecast_transition_diagnostics(bookings_df, actual_entries_df, horizons=(28, 21, 14, 7, 3, 1)):
    """
    Shows how booking-based forecasts for the same service day change as the day gets closer.

    This is useful for your roster issue:
    - D21 may imply high demand
    - D7 may drop if bookings have not materialised
    """
    bookings = bookings_df.copy()

    bookings = bookings.loc[
        (~bookings["invalid_entry_lead"])
    ].copy()

    actual_entry_counts = (
        actual_entries_df
        .groupby("entry_day")["BookingReference"]
        .nunique()
        .rename("actual_entries")
    )

    final_booked_entries = (
        bookings
        .groupby("entry_day")["bookingUuid"]
        .nunique()
        .rename("final_booked_entries")
    )

    base = pd.concat([actual_entry_counts, final_booked_entries], axis=1).fillna(0)

    for h in horizons:
        seen = bookings.loc[
            bookings["creation_day"] <= bookings["entry_day"] - pd.Timedelta(days=h)
        ].copy()

        seen_counts = (
            seen
            .groupby("entry_day")["bookingUuid"]
            .nunique()
            .rename(f"seen_entries_D{h}")
        )

        base = base.join(seen_counts, how="left")
        base[f"seen_entries_D{h}"] = base[f"seen_entries_D{h}"].fillna(0)

    for h in horizons:
        base[f"pct_actual_visible_D{h}"] = safe_divide(
            base[f"seen_entries_D{h}"],
            base["actual_entries"]
        ) * 100

    if 21 in horizons and 7 in horizons:
        base["drop_in_seen_entries_D21_to_D7"] = base["seen_entries_D7"] - base["seen_entries_D21"]

    if 28 in horizons and 7 in horizons:
        base["drop_in_seen_entries_D28_to_D7"] = base["seen_entries_D7"] - base["seen_entries_D28"]

    base = base.reset_index().rename(columns={"entry_day": "service_day"})

    return base


# ---------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------
def run_fastpark_forecast_diagnostics(
    start="2026-04-01",
    end="2026-06-30",
    deduplicate=True
):
    """
    Complete diagnostic runner.

    Recommended initial use:
        run over the latest 3 completed months.

    Example:
        python scripts/fastpark_forecast_diagnostics.py
    """
    print_section("FASTPARK FORECAST DIAGNOSTICS")
    print(f"Analysis window: {start} to {end}")
    print(f"Output folder: {OUTPUT_DIR}")

    # ---------------------------------------------------------
    # Load data
    # ---------------------------------------------------------
    print_section("LOADING DATA")

    bookings_raw = load_airportx_fastpark_bookings(
        start=start,
        end=end,
        deduplicate=deduplicate
    )

    bookings = prepare_bookings(bookings_raw)

    actual_entries, actual_exits = load_fastpark_actuals(
        start=start,
        end=end
    )

    flight_df = load_flight_performance(
        start=start,
        end=end
    )

    print(f"Bookings loaded: {len(bookings):,.0f}")
    print(f"Actual entries loaded: {len(actual_entries):,.0f}")
    print(f"Actual exits loaded: {len(actual_exits):,.0f}")
    print(f"Flight rows loaded: {len(flight_df):,.0f}")

    if bookings.empty:
        raise ValueError("No bookings returned. Check date range/query.")
    if actual_entries.empty:
        raise ValueError("No actual entries returned. Check date range/query.")
    if actual_exits.empty:
        raise ValueError("No actual exits returned. Check date range/query.")
    if flight_df.empty:
        raise ValueError("No flight performance rows returned. Check date range/query.")
    
    workbook_sheets = {}

    # ---------------------------------------------------------
    # Lead time validation
    # ---------------------------------------------------------
    print_section("LEAD TIME DISTRIBUTION")

    leadtime_distribution = build_leadtime_distribution(bookings)

    print("\nLead time distribution:")
    print(leadtime_distribution.to_string(index=False))

    workbook_sheets["03_Leadtime_Distribution"] = leadtime_distribution

    # ---------------------------------------------------------
    # Lead-time visibility
    # ---------------------------------------------------------
    print_section("BOOKING VISIBILITY AND RECOMMENDED UPLIFTS")

    visibility_outputs = build_booking_visibility_curves(
        bookings_df=bookings,
        actual_entries_df=actual_entries,
        actual_exits_df=actual_exits,
        horizons=DEFAULT_HORIZONS
    )

    visibility_summary = visibility_outputs["visibility_summary"].round(2)

    print("\nVisibility summary:")
    print(visibility_summary.to_string(index=False))

    save_csv(visibility_summary, "visibility_summary_D1_D28.csv")
    save_csv(visibility_outputs["entry_daily_detail"], "entry_daily_visibility_detail.csv")
    save_csv(visibility_outputs["exit_daily_detail"], "exit_daily_visibility_detail.csv")
    save_csv(visibility_outputs["implied_exit_daily_detail"], "implied_exit_daily_visibility_detail.csv")
    
    workbook_sheets["09_Visibility_D1_D28"] = visibility_summary
    workbook_sheets["14_Entry_Detail"] = visibility_outputs["entry_daily_detail"]
    workbook_sheets["15_Exit_Detail"] = visibility_outputs["exit_daily_detail"]
    workbook_sheets["16_Implied_Exit_Detail"] = visibility_outputs["implied_exit_daily_detail"]

    # Ready-to-use arrays
    d1_d14 = visibility_summary[visibility_summary["horizon"].isin([f"D{i}" for i in range(1, 15)])].copy()

    entry_uplift_median_array = d1_d14["entry_uplift_median"].round(2).tolist()
    exit_uplift_median_array = d1_d14["exit_uplift_median"].round(2).tolist()
    implied_exit_uplift_median_array = d1_d14["implied_exit_uplift_median"].round(2).tolist()

    print("\nRecommended entry uplift array, D1-D14, median:")
    print(entry_uplift_median_array)

    print("\nRecommended exit uplift array, D1-D14, median:")
    print(exit_uplift_median_array)

    print("\nRecommended implied-exit uplift array, D1-D14, median:")
    print(implied_exit_uplift_median_array)

    # ---------------------------------------------------------
    # Segmented visibility
    # ---------------------------------------------------------
    print_section("SEGMENTED VISIBILITY")

    segmented_outputs = build_segmented_visibility(
        bookings_df=bookings,
        horizons=DEFAULT_HORIZONS
    )

    visibility_by_horizon_month = build_visibility_by_horizon_month(
    bookings_df=bookings,
    horizons=DEFAULT_HORIZONS
    )

    print("\nVisibility by duration:")
    print(segmented_outputs["visibility_by_duration"].round(2).to_string(index=False))

    print("\nVisibility by weekday:")
    print(segmented_outputs["visibility_by_weekday"].round(2).to_string(index=False))

    print("\nVisibility by month:")
    print(segmented_outputs["visibility_by_month"].round(2).to_string(index=False))

    print("\nVisibility by horizon and month:")
    print(visibility_by_horizon_month.to_string(index=False))

    save_csv(segmented_outputs["segmented_visibility_detail"], "segmented_visibility_detail.csv")
    save_csv(segmented_outputs["visibility_by_duration"], "visibility_by_duration.csv")
    save_csv(segmented_outputs["visibility_by_weekday"], "visibility_by_weekday.csv")
    save_csv(segmented_outputs["visibility_by_month"], "visibility_by_month.csv")

    workbook_sheets["10_Visibility_By_Weekday"] = segmented_outputs["visibility_by_weekday"]
    workbook_sheets["11_Visibility_By_Duration"] = segmented_outputs["visibility_by_duration"]
    workbook_sheets["12_Visibility_By_Month"] = segmented_outputs["visibility_by_month"]
    workbook_sheets["18_Visibility_Horizon_Month"] = visibility_by_horizon_month
    workbook_sheets["17_Segmented_Detail"] = segmented_outputs["segmented_visibility_detail"]

    # ---------------------------------------------------------
    # Penetration analysis
    # ---------------------------------------------------------
    print_section("PENETRATION RATE ANALYSIS")

    daily_penetration = build_daily_penetration(
        actual_entries_df=actual_entries,
        actual_exits_df=actual_exits,
        flight_df=flight_df
    )

    penetration_outputs = build_penetration_period_tables(daily_penetration)

    weekly_penetration = penetration_outputs["weekly_penetration"].round(3)
    monthly_penetration = penetration_outputs["monthly_penetration"].round(3)
    weekday_penetration = penetration_outputs["weekday_penetration"].round(3)

    print("\nDaily penetration:")
    print(daily_penetration.round(3).tail(20).to_string(index=False))

    print("\nWeekly penetration:")
    print(weekly_penetration.to_string(index=False))

    print("\nMonthly penetration:")
    print(monthly_penetration.to_string(index=False))

    print("\nWeekday penetration:")
    print(weekday_penetration.to_string(index=False))

    volatility = build_penetration_volatility_table(
        penetration_outputs["weekly_penetration"]
    ).round(4)

    print("\nPenetration volatility comparison:")
    print(volatility.to_string(index=False))

    save_csv(daily_penetration, "daily_penetration.csv")
    save_csv(penetration_outputs["weekly_penetration"], "weekly_penetration.csv")
    save_csv(penetration_outputs["monthly_penetration"], "monthly_penetration.csv")
    save_csv(penetration_outputs["weekday_penetration"], "weekday_penetration.csv")
    save_csv(volatility, "penetration_volatility_comparison.csv")

    workbook_sheets["04_Penetration_Daily"] = daily_penetration
    workbook_sheets["05_Penetration_Weekly"] = penetration_outputs["weekly_penetration"]
    workbook_sheets["06_Penetration_Monthly"] = penetration_outputs["monthly_penetration"]
    workbook_sheets["07_Penetration_Weekday"] = penetration_outputs["weekday_penetration"]
    workbook_sheets["08_Penetration_Volatility"] = volatility

    # ---------------------------------------------------------
    # Backtest current uplift
    # ---------------------------------------------------------
    print_section("BACKTEST CURRENT UPLIFT ARRAYS")

    uplift_backtest = build_uplift_backtest(
        bookings_df=bookings,
        actual_entries_df=actual_entries,
        actual_exits_df=actual_exits,
        horizons=range(1, 15)
    ).round(3)

    print("\nCurrent uplift backtest:")
    print(uplift_backtest.to_string(index=False))

    save_csv(uplift_backtest, "current_uplift_backtest_D1_D14.csv")

    workbook_sheets["03_Current_Uplift_Backtest"] = uplift_backtest

    # ---------------------------------------------------------
    # Forecast transition/drop diagnostics
    # ---------------------------------------------------------
    print_section("FORECAST TRANSITION / DROP DIAGNOSTICS")

    transition_diagnostics = build_forecast_transition_diagnostics(
        bookings_df=bookings,
        actual_entries_df=actual_entries,
        horizons=(28, 21, 14, 7, 3, 1)
    ).round(3)

    print("\nTransition diagnostics sample:")
    print(transition_diagnostics.tail(30).to_string(index=False))

    save_csv(transition_diagnostics, "forecast_transition_diagnostics.csv")

    workbook_sheets["13_Transition_Diagnostics"] = transition_diagnostics

    # ---------------------------------------------------------
    # Headline recommendation text
    # ---------------------------------------------------------
    print_section("HEADLINE INTERPRETATION HELPERS")

    latest_weekly = penetration_outputs["weekly_penetration"].dropna().tail(6)

    print("\nUse these checks:")
    print("1. If 2-week rolling penetration has much higher volatility than 4-week rolling, the current two-week penetration method is probably too reactive.")
    print("2. If current uplift backtest has positive mean error at D14/D21, it is over-forecasting far out.")
    print("3. If current uplift backtest has negative mean error at D7/D3, it is under-forecasting close-in.")
    print("4. If entry visibility varies strongly by duration bucket, use duration-specific uplift rather than one flat array.")
    print("5. If weekday visibility differs materially, use weekday-specific uplift for entries and exits.")

    print("\nSuggested next model design:")
    print("For D1-D14: use booking-based forecast with data-derived median uplift.")
    print("For D15-D28: use a blend of booking-based visibility and 4-week/weekday penetration.")
    print("For D29+: use penetration-based forecast, preferably 4-week or monthly, depending on the volatility table.")

    # ---------------------------------------------------------
    # Write all outputs to one Excel workbook
    # ---------------------------------------------------------
    print_section("WRITING EXCEL WORKBOOK")

    workbook_sheets = add_summary_sheet(
        workbook_sheets=workbook_sheets,
        start=start,
        end=end,
        bookings=bookings,
        actual_entries=actual_entries,
        actual_exits=actual_exits,
        flight_df=flight_df,
        visibility_summary=visibility_summary,
        volatility=volatility,
        uplift_backtest=uplift_backtest
    )

    workbook_sheets = add_recommended_arrays_sheet(
        workbook_sheets=workbook_sheets,
        visibility_summary=visibility_summary
    )

    # Re-order the sheets so Summary and Recommended Uplifts come first
    preferred_order = [
        "01_Summary",
        "02_Recommended_Uplifts",
        "03_Leadtime_Distribution",
        "03_Current_Uplift_Backtest",
        "04_Penetration_Daily",
        "05_Penetration_Weekly",
        "06_Penetration_Monthly",
        "07_Penetration_Weekday",
        "08_Penetration_Volatility",
        "09_Visibility_D1_D28",
        "10_Visibility_By_Weekday",
        "11_Visibility_By_Duration",
        "12_Visibility_By_Month",
        "13_Transition_Diagnostics",
        "14_Entry_Detail",
        "15_Exit_Detail",
        "16_Implied_Exit_Detail",
        "17_Segmented_Detail",
        "18_Visibility_Horizon_Month"
    ]

    ordered_workbook_sheets = {}

    for sheet in preferred_order:
        if sheet in workbook_sheets:
            ordered_workbook_sheets[sheet] = workbook_sheets[sheet]

    for sheet, df in workbook_sheets.items():
        if sheet not in ordered_workbook_sheets:
            ordered_workbook_sheets[sheet] = df

    write_diagnostics_workbook(
        workbook_sheets=ordered_workbook_sheets,
        output_path=EXCEL_OUTPUT_PATH
    )

    print(f"Saved to: {EXCEL_OUTPUT_PATH}")

    return {
        "bookings": bookings,
        "actual_entries": actual_entries,
        "actual_exits": actual_exits,
        "flight_df": flight_df,
        "visibility_summary": visibility_summary,
        "segmented_outputs": segmented_outputs,
        "daily_penetration": daily_penetration,
        "penetration_outputs": penetration_outputs,
        "volatility": volatility,
        "uplift_backtest": uplift_backtest,
        "transition_diagnostics": transition_diagnostics,
    }


# ---------------------------------------------------------------------
# Script entrypoint
# ---------------------------------------------------------------------
if __name__ == "__main__":
    outputs = run_fastpark_forecast_diagnostics(
        start="2026-07-21",
        end="2026-08-10",
        deduplicate=True
    )