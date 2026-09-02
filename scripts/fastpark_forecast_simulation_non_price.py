"""
FASTPARK FORECAST SIMULATION - NON-PRICE REFINEMENTS / GENERATION 2.1
=====================================================================

Purpose
-------
This file tests non-price improvements against the corrected Generation 2
rolling-origin demand forecast. It does not calculate FTE and does not use
pricing.

New tests
---------
1. Booking pace derived from the known-booking position at successive horizons.
2. Alternative historical same-weekday windows: 1, 2, 4, 6 and 8 weeks.
3. Explicit calendar categories supplied in MANUAL_CALENDAR_PERIODS.
4. Booking-position demand regimes.
5. Weekday-specific demand-level trend.
6. As-of cancellation-pattern analysis using cancellations known at each
 

Dependencies
------------
Place these files in the same scripts folder:

    fastpark_forecast_simulation.py
    fastpark_forecast_simulation_refined.py
    fastpark_forecast_simulation_non_price.py

The refined file is imported as a module. Its __main__ block will not run.
Generation 1 is imported through the refined configuration.

Manual inputs
-------------
1. Check the three module names in get_non_price_config().
2. Add holiday and school-holiday date ranges to MANUAL_CALENDAR_PERIODS.
3. Confirm the configured booking column names match the cleaned Generation 1
   booking dataframe.

"""

# =============================================================================
# 0. IMPORTS
# =============================================================================

import importlib
import pathlib
import sys
import time
from pathlib import Path

import numpy as np
import pandas as pd

try:
    from modules.utils.progress import step
except ImportError:
    def step(prev, msg):
        now = time.perf_counter()
        print(f"    OK {msg}   [{now - prev:0.2f}s]")
        return now


SCRIPT_DIR = pathlib.Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

for candidate in (SCRIPT_DIR, PROJECT_ROOT):
    if str(candidate) not in sys.path:
        sys.path.append(str(candidate))


# =============================================================================
# 1. MANUAL CALENDAR INPUTS
# =============================================================================

# Add only periods that were known before the forecast target date.
# Dates are inclusive. Leave a category as [] if it is not yet available.
#
# Example:
# "school_holiday": [
#     ("2025-06-30", "2025-08-15"),
#     ("2025-10-13", "2025-10-20"),
# ],
#
MANUAL_CALENDAR_PERIODS = {
    "public_holiday": [
        ("2025-05-03", "2025-05-05"),
        ("2025-05-24", "2025-05-26"),
        ("2025-08-02", "2025-08-04"),
        ("2025-11-29", "2025-12-01"),      
        ("2026-03-18", "2026-03-20"),
        ("2026-05-02", "2026-05-04"),
        ("2026-05-23", "2026-05-25"),
        ("2026-08-01", "2026-08-03"),
        ("2026-11-28", "2026-11-30"),
        ("2026-09-19", "2026-09-21"),
    ],
    "school_holiday": [
        ("2024-09-13", "2024-09-16"),
        ("2024-10-11", "2024-10-21"),
        ("2024-12-20", "2025-01-06"),
        ("2025-02-07", "2025-02-16"),
        ("2025-04-04", "2025-04-21"),
        ("2025-05-02", "2025-05-06"),
        ("2025-05-16", "2025-05-19"),
        ("2025-06-26", "2025-08-11"),
        ("2025-09-12", "2025-09-14"),
        ("2025-10-10", "2025-10-19"),
        ("2025-12-19", "2026-01-04"),
        ("2026-02-13", "2026-02-22"),
        ("2026-04-02", "2026-04-20"),
        ("2026-05-01", "2026-05-05"),
        ("2026-05-15", "2026-05-18"),
        ("2026-06-26", "2026-08-10"),


    ],
    "easter_period": [
        ("2025-03-17", "2025-03-21"),
        ("2026-03-02", "2026-03-06")
    ],
    "christmas_ny_period": [
        ("2025-12-19", "2026-01-03"),
    ],
    "special_event": [
        ("2025-06-19", "2025-06-22"), #Royal highland show 2025
        ("2026-06-18", "2026-06-21"), #Royal highland show 2026
        ("2025-08-22", "2025-08-26"), #summer sessions 2025
        ("2026-08-21", "2026-08-29"), #summer sessions 2026
        ("2025-08-01", "2025-08-25"), #Edinburgh festival 2025
        ("2026-08-07", "2026-08-30"), #Edinburgh festival 2026
    ],
}


# =============================================================================
# 2. CONFIGURATION
# =============================================================================

def get_non_price_config():
    return {
        # File/module names without .py
        "generation_1_module": "fastpark_forecast_simulation",
        "refined_module": "fastpark_forecast_simulation_refined",

        # Same historical and simulation dates as Generation 2.
        "history_start": "2024-01-01",
        "history_end": "2026-07-31",
        "simulation_start": "2025-07-01",
        "simulation_end": "2026-07-31",
        "forecast_horizons_days": [
            0, 1, 2, 3, 4, 5, 6, 7,
            14, 21, 28, 35, 42, 49, 56,
        ],

        # Rolling-origin setup.
        "minimum_training_months": 6,
        "validation_months": 1,
        "maximum_weight_training_months": 12,

        # Use 10% weights for initial feature screening. Once a feature is
        # proven useful, refine only the winning experiment at 5% or 2.5%.
        "weight_step": 0.10,
        "weight_stability_penalty": 0.0001,
        "booking_shrinkage_strengths": [0, 25, 50, 100],

        # Feature settings.
        "same_weekday_n": 4,
        "weekday_month_max_n": 8,
        "minimum_weekday_month_observations": 3,
        "trend_recent_days": 28,
        "trend_comparison_days": 84,
        "same_weekday_windows": [1, 2, 4, 6, 8],
        "pace_lookback_days": 7,
        "pace_bins": [-np.inf, -0.50, -0.15, 0.15, 0.50, np.inf],
        "pace_labels": [
            "very_low", "low", "normal", "high", "very_high"
        ],
        "regime_quantiles": 5,
        "minimum_segment_observations": 10,
        "calendar_adjustment_min": 0.70,
        "calendar_adjustment_max": 1.30,
        "pace_adjustment_min": 0.70,
        "pace_adjustment_max": 1.30,
        "regime_adjustment_min": 0.70,
        "regime_adjustment_max": 1.30,
        "weekday_trend_recent_occurrences": 4,
        "weekday_trend_comparison_occurrences": 12,
        "trend_factor_min": 0.70,
        "trend_factor_max": 1.30,

        # -----------------------------------------------------------------
        # Cancellation-pattern experiment
        # -----------------------------------------------------------------
        #
        # Generation 1 already excludes cancellations known at the historical
        # forecast cutoff from the active booking population.
        #
        # Generation 2.1 tests whether an unusually high or low recent
        # cancellation rate contains additional predictive information beyond
        # the standard booking curve.
        #
        # No cancellation after the historical cutoff is used.
        # -----------------------------------------------------------------

        "enable_cancellation_test": True,

        # Cleaned Generation 1 booking columns.
        "booking_id_col": "bookingId",
        "booking_created_timestamp_col": "createdAt",
        "cancellation_timestamp_col": "cancelledAt",
        "booking_entry_date_col": "entryDate",
        "booking_exit_date_col": "exitDate",

        # Calculate 1-day, 3-day and 7-day cancellation diagnostics.
        "cancellation_windows_days": [
            1,
            3,
            7,
        ],

        # Use the seven-day rate in the first cancellation experiment.
        "primary_cancellation_window_days": 7,

        # Compare each target date with the normal cancellation rate for the
        # same horizon and weekday in the training period.
        "cancellation_rate_bins": [
            -np.inf,
            -0.50,
            -0.15,
            0.15,
            0.50,
            np.inf,
        ],

        "cancellation_rate_labels": [
            "very_low",
            "low",
            "normal",
            "high",
            "very_high",
        ],

        # Minimum training observations needed before a segmented cancellation
        # adjustment can be applied.
        "minimum_cancellation_segment_observations": 10,

        # Prevent an unstable cancellation segment from making an extreme
        # forecast adjustment.
        "cancellation_adjustment_min": 0.70,
        "cancellation_adjustment_max": 1.30,

        # Runtime controls.
        "smoke_test": False,
        "smoke_test_target_count": 24,
        "smoke_test_horizons": [0, 7, 28, 56],
        "fail_fast": True,

        # Keep this False for the full run. Candidate grids can exceed Excel's
        # row limit. Selected weights and test predictions are still exported.
        "export_all_weight_tests": False,

        "output_path": (
            PROJECT_ROOT
            / "outputs"
            / "fastpark_forecast_simulation_non_price.xlsx"
        ),
        "checkpoint_dir": PROJECT_ROOT / "outputs" / "non_price_checkpoints",
    }


# =============================================================================
# 3. MODULE AND CONFIG HELPERS
# =============================================================================

def import_modules(config):
    refined = importlib.import_module(config["refined_module"])
    base = importlib.import_module(config["generation_1_module"])
    return base, refined


def build_base_config(base, config):
    base_config = base.get_simulation_config()
    base_config.update(
        {
            "history_start": config["history_start"],
            "history_end": config["history_end"],
            "simulation_start": config["simulation_start"],
            "simulation_end": config["simulation_end"],
            "forecast_horizons_days": config["forecast_horizons_days"],
            "test_days_of_month": list(range(1, 32)),
            "smoke_test": config["smoke_test"],
            "smoke_test_target_count": config["smoke_test_target_count"],
            "smoke_test_horizons": config["smoke_test_horizons"],
            "fail_fast": config["fail_fast"],
            "enable_hourly_analysis": False,
        }
    )
    return base_config


def create_test_dates(config):
    """
    Create all target dates for the full run.

    In smoke-test mode, select three dates from each of the first eight
    simulation months. This provides enough distinct months to create at least
    one six-month training, one-month validation, and one-month test fold.
    """

    dates = pd.date_range(
        start=config["simulation_start"],
        end=config["simulation_end"],
        freq="D",
    )

    if not config["smoke_test"]:
        return pd.DataFrame(
            {
                "target_date": dates.normalize()
            }
        )

    smoke_dates = []

    month_starts = pd.date_range(
        start=pd.Timestamp(config["simulation_start"])
        .to_period("M")
        .to_timestamp(),
        end=pd.Timestamp(config["simulation_end"])
        .to_period("M")
        .to_timestamp(),
        freq="MS",
    )

    for month_start in month_starts[:8]:
        month_end = month_start + pd.offsets.MonthEnd(0)

        for day_number in (1, 8, 15):
            candidate = month_start + pd.Timedelta(
                days=day_number - 1
            )

            if (
                candidate <= month_end
                and candidate >= pd.Timestamp(
                    config["simulation_start"]
                )
                and candidate <= pd.Timestamp(
                    config["simulation_end"]
                )
            ):
                smoke_dates.append(candidate.normalize())

    return pd.DataFrame(
        {
            "target_date": sorted(
                set(smoke_dates)
            )
        }
    )

# =============================================================================
# 4. GENERAL HELPERS
# =============================================================================

def safe_ratio(numerator, denominator, default=np.nan):
    if pd.isna(numerator) or pd.isna(denominator) or denominator == 0:
        return default
    return float(numerator) / float(denominator)


def clipped(value, lower, upper, default=1.0):
    if value is None or pd.isna(value) or not np.isfinite(value):
        return default
    return float(np.clip(value, lower, upper))


def add_error_columns(frame):
    output = frame.copy()
    output["error"] = output["forecast_value"] - output["actual_value"]
    output["absolute_error"] = output["error"].abs()
    output["squared_error"] = output["error"].pow(2)
    return output


def score_frame(refined, frame, groups):
    return refined.score_forecast_frame(
        dataframe=frame,
        group_columns=groups,
        forecast_col="forecast_value",
        actual_col="actual_value",
    )


def nearest_earlier_horizon(horizon, available_horizons, minimum_gap=1):
    candidates = [
        h for h in available_horizons
        if h > horizon and h - horizon >= minimum_gap
    ]
    return min(candidates) if candidates else None


# =============================================================================
# 5. CALENDAR FEATURES
# =============================================================================

def build_calendar_table(start_date, end_date, periods):
    """
    Create independent calendar flags.

    Start and end dates are inclusive. Overlapping calendar periods are
    retained rather than overwriting one another.
    """

    calendar = pd.DataFrame(
        {
            "target_date": pd.date_range(
                start=start_date,
                end=end_date,
                freq="D",
            )
        }
    )

    calendar["target_date"] = (
        pd.to_datetime(calendar["target_date"])
        .dt.normalize()
    )

    calendar_flags = []

    for category, ranges in periods.items():
        flag_column = f"is_{category}"
        calendar[flag_column] = False
        calendar_flags.append(flag_column)

        for start, end in ranges:
            start = pd.Timestamp(start).normalize()
            end = pd.Timestamp(end).normalize()

            if end < start:
                raise ValueError(
                    f"Calendar range end is before start: "
                    f"{category}, {start.date()} to {end.date()}"
                )

            mask = calendar["target_date"].between(
                start,
                end,
                inclusive="both",
            )

            calendar.loc[mask, flag_column] = True

    calendar["is_any_special_period"] = (
        calendar[calendar_flags].any(axis=1)
        if calendar_flags
        else False
    )

    special_dates = set(
        calendar.loc[
            calendar["is_any_special_period"],
            "target_date",
        ]
    )

    calendar["is_day_before_special"] = (
        calendar["target_date"].map(
            lambda date: (
                date + pd.Timedelta(days=1)
            ) in special_dates
        )
    )

    calendar["is_day_after_special"] = (
        calendar["target_date"].map(
            lambda date: (
                date - pd.Timedelta(days=1)
            ) in special_dates
        )
    )

    def combine_calendar_types(row):
        active_categories = []

        for category in periods:
            if row[f"is_{category}"]:
                active_categories.append(category)

        if active_categories:
            return "|".join(active_categories)

        if row["is_day_before_special"]:
            return "day_before_special"

        if row["is_day_after_special"]:
            return "day_after_special"

        return "normal"

    calendar["calendar_type"] = calendar.apply(
        combine_calendar_types,
        axis=1,
    )

    return calendar


# =============================================================================
# 6. ALTERNATIVE SAME-WEEKDAY COMPONENTS
# =============================================================================

def calculate_same_weekday_forecasts(
    base,
    data,
    component_table,
    windows,
):
    output = component_table.copy()

    for window in windows:
        output[f"entry_weekday_{window}"] = np.nan
        output[f"exit_weekday_{window}"] = np.nan

    total = len(output)
    for position, row in enumerate(output.itertuples(), start=1):
        if position == 1 or position % 250 == 0 or position == total:
            print(
                f"\rAlternative weekday forecasts {position:,}/{total:,}",
                end="",
            )

        for window in windows:
            entry_value = base.forecast_same_weekday(
                daily_actuals=data["daily_actuals"],
                target_date=row.target_date,
                cutoff_timestamp=row.cutoff_timestamp,
                target_col="entries",
                n=window,
                data=data,
            )
            exit_value = base.forecast_same_weekday(
                daily_actuals=data["daily_actuals"],
                target_date=row.target_date,
                cutoff_timestamp=row.cutoff_timestamp,
                target_col="exits",
                n=window,
                data=data,
            )
            output.at[row.Index, f"entry_weekday_{window}"] = entry_value
            output.at[row.Index, f"exit_weekday_{window}"] = exit_value

    print()
    return output


def choose_historical_window(training, validation, demand_type, windows, refined):
    actual_col = "actual_entries" if demand_type == "entry" else "actual_exits"
    rows = []
    for window in windows:
        column = f"{demand_type}_weekday_{window}"
        validation_wape = refined.wape(validation[actual_col], validation[column])
        training_wape = refined.wape(training[actual_col], training[column])
        rows.append(
            {
                "window": window,
                "column": column,
                "training_wape_pct": training_wape * 100,
                "validation_wape_pct": validation_wape * 100,
            }
        )
    diagnostics = pd.DataFrame(rows).sort_values(
        ["validation_wape_pct", "training_wape_pct", "window"]
    )
    return diagnostics.iloc[0].to_dict(), diagnostics


# =============================================================================
# 7. BOOKING PACE FEATURES
# =============================================================================

def add_booking_pace_features(component_table, config):
    output = component_table.copy()
    horizons = sorted(output["horizon_days"].unique())

    lookup = output.set_index(["target_date", "horizon_days"])

    for flow in ("entry", "exit"):
        known_col = f"{flow}_known_bookings"
        output[f"{flow}_pace_per_day"] = np.nan
        output[f"{flow}_pace_relative"] = np.nan
        output[f"{flow}_pace_source_horizon"] = np.nan

        for index, row in output.iterrows():
            earlier_horizon = nearest_earlier_horizon(
                int(row["horizon_days"]),
                horizons,
                minimum_gap=config["pace_lookback_days"],
            )
            if earlier_horizon is None:
                continue
            key = (row["target_date"], earlier_horizon)
            if key not in lookup.index:
                continue
            earlier_known = lookup.loc[key, known_col]
            current_known = row[known_col]
            gap = earlier_horizon - row["horizon_days"]
            pace = safe_ratio(current_known - earlier_known, gap)
            output.at[index, f"{flow}_pace_per_day"] = pace
            output.at[index, f"{flow}_pace_source_horizon"] = earlier_horizon

        # Relative pace and pace bands are fitted inside each rolling fold
        # using training data only. This avoids test-period leakage.

    return output


def apply_training_pace_bands(training, validation, test, flow, config):
    pace_col = f"{flow}_pace_per_day"
    relative_col = f"{flow}_pace_relative"
    band_col = f"{flow}_pace_band"

    medians = (
        training.groupby(["horizon_days", "weekday"], dropna=False)[pace_col]
        .median()
        .rename("training_pace_median")
        .reset_index()
    )

    outputs = []
    for frame in (training, validation, test):
        item = frame.merge(
            medians,
            on=["horizon_days", "weekday"],
            how="left",
            validate="many_to_one",
        )
        denominator = item["training_pace_median"].abs().clip(lower=1.0)
        item[relative_col] = (
            item[pace_col] - item["training_pace_median"]
        ) / denominator
        item[band_col] = pd.cut(
            item[relative_col],
            bins=config["pace_bins"],
            labels=config["pace_labels"],
            include_lowest=True,
        ).astype("object").fillna("unknown")
        item = item.drop(columns=["training_pace_median"])
        outputs.append(item)

    return tuple(outputs)


def fit_segment_adjustments(
    training,
    base_forecast_col,
    actual_col,
    segment_cols,
    minimum_observations,
    lower,
    upper,
):
    work = training[
        training[base_forecast_col].notna()
        & training[actual_col].notna()
        & training[base_forecast_col].gt(0)
    ].copy()
    work["ratio"] = work[actual_col] / work[base_forecast_col]

    grouped = work.groupby(segment_cols, dropna=False).agg(
        adjustment=("ratio", "median"),
        observations=("ratio", "size"),
    ).reset_index()
    grouped = grouped[grouped["observations"].ge(minimum_observations)].copy()
    grouped["adjustment"] = grouped["adjustment"].clip(lower, upper)
    return grouped


def apply_segment_adjustment(
    frame,
    base_forecast_col,
    adjustment_table,
    segment_cols,
    output_col,
):
    output = frame.copy()
    if adjustment_table.empty:
        output[output_col] = output[base_forecast_col]
        return output

    output = output.merge(adjustment_table, on=segment_cols, how="left")
    output["adjustment"] = output["adjustment"].fillna(1.0)
    output[output_col] = output[base_forecast_col] * output["adjustment"]
    output = output.drop(columns=["adjustment", "observations"], errors="ignore")
    return output


# =============================================================================
# 8. BOOKING REGIME FEATURES
# =============================================================================

def fit_regime_edges(training, flow, quantiles):
    known_col = f"{flow}_known_bookings"
    values = training[known_col].dropna().astype(float)
    if values.nunique() < 2:
        return None
    edges = np.unique(
        values.quantile(np.linspace(0, 1, quantiles + 1)).to_numpy()
    )
    if len(edges) < 3:
        return None
    edges[0] = -np.inf
    edges[-1] = np.inf
    return edges


def apply_regime_band(frame, flow, edges):
    output = frame.copy()
    col = f"{flow}_regime_band"
    if edges is None:
        output[col] = "normal"
        return output
    labels = [f"Q{i + 1}" for i in range(len(edges) - 1)]
    output[col] = pd.cut(
        output[f"{flow}_known_bookings"],
        bins=edges,
        labels=labels,
        include_lowest=True,
        duplicates="drop",
    ).astype("object").fillna("unknown")
    return output


# =============================================================================
# 9. WEEKDAY-SPECIFIC TREND
# =============================================================================

def weekday_specific_trend_forecast(
    daily_actuals,
    target_date,
    cutoff_timestamp,
    target_col,
    seasonal_forecast,
    recent_occurrences,
    comparison_occurrences,
    lower,
    upper,
):
    if pd.isna(seasonal_forecast):
        return np.nan

    cutoff_date = pd.Timestamp(cutoff_timestamp).normalize()
    target_weekday = pd.Timestamp(target_date).weekday()
    history = daily_actuals.copy()
    history["date"] = pd.to_datetime(history["date"]).dt.normalize()
    history = history[
        history["date"].lt(cutoff_date)
        & history["date"].dt.weekday.eq(target_weekday)
    ].sort_values("date")

    needed = recent_occurrences + comparison_occurrences
    values = history[target_col].dropna().tail(needed)
    if len(values) < needed:
        return seasonal_forecast

    recent = values.tail(recent_occurrences).mean()
    comparison = values.head(comparison_occurrences).mean()
    factor = clipped(safe_ratio(recent, comparison, 1.0), lower, upper)
    return max(0.0, float(seasonal_forecast) * factor)


def add_weekday_trend_features(component_table, daily_actuals, config):
    output = component_table.copy()
    output["entry_weekday_trend"] = np.nan
    output["exit_weekday_trend"] = np.nan

    for row in output.itertuples():
        output.at[row.Index, "entry_weekday_trend"] = (
            weekday_specific_trend_forecast(
                daily_actuals=daily_actuals,
                target_date=row.target_date,
                cutoff_timestamp=row.cutoff_timestamp,
                target_col="entries",
                seasonal_forecast=row.entry_month,
                recent_occurrences=config["weekday_trend_recent_occurrences"],
                comparison_occurrences=config[
                    "weekday_trend_comparison_occurrences"
                ],
                lower=config["trend_factor_min"],
                upper=config["trend_factor_max"],
            )
        )
        output.at[row.Index, "exit_weekday_trend"] = (
            weekday_specific_trend_forecast(
                daily_actuals=daily_actuals,
                target_date=row.target_date,
                cutoff_timestamp=row.cutoff_timestamp,
                target_col="exits",
                seasonal_forecast=row.exit_month,
                recent_occurrences=config["weekday_trend_recent_occurrences"],
                comparison_occurrences=config[
                    "weekday_trend_comparison_occurrences"
                ],
                lower=config["trend_factor_min"],
                upper=config["trend_factor_max"],
            )
        )
    return output


# =============================================================================
# 10. OPTIONAL CANCELLATION FEATURE VALIDATION
# =============================================================================

def validate_cancellation_configuration(
    bookings,
    config,
):
    """
    Validate all cleaned booking columns required for the as-of
    cancellation-pattern experiment.
    """

    if not config["enable_cancellation_test"]:
        print(
            "Cancellation-pattern experiment disabled."
        )
        return

    configured_columns = {
        "booking_id_col":
            config["booking_id_col"],

        "booking_created_timestamp_col":
            config["booking_created_timestamp_col"],

        "cancellation_timestamp_col":
            config["cancellation_timestamp_col"],

        "booking_entry_date_col":
            config["booking_entry_date_col"],

        "booking_exit_date_col":
            config["booking_exit_date_col"],
    }

    missing = []

    for config_name, column_name in configured_columns.items():

        if not column_name:

            missing.append(
                f"{config_name}=None"
            )

        elif column_name not in bookings.columns:

            missing.append(
                f"{config_name}={column_name!r}"
            )

    if missing:

        raise KeyError(
            "Cancellation analysis cannot run because the following "
            "configured booking columns were not found:\n\n"
            +
            "\n".join(
                f"    {item}"
                for item in missing
            )
            +
            "\n\nAvailable cleaned booking columns:\n\n"
            +
            "\n".join(
                f"    {column}"
                for column in bookings.columns
            )
        )

    print()
    print("Cancellation configuration validated:")

    print(
        f"    Booking ID: "
        f"{config['booking_id_col']}"
    )

    print(
        f"    Created timestamp: "
        f"{config['booking_created_timestamp_col']}"
    )

    print(
        f"    Cancellation timestamp: "
        f"{config['cancellation_timestamp_col']}"
    )

    print(
        f"    Entry target: "
        f"{config['booking_entry_date_col']}"
    )

    print(
        f"    Exit target: "
        f"{config['booking_exit_date_col']}"
    )

# =============================================================================
# 10A. AS-OF CANCELLATION-PATTERN FEATURES
# =============================================================================

def prepare_cancellation_booking_data(
    bookings,
    config,
):
    """
    Prepare a cancellation-specific copy of the cleaned booking data.

    This does not alter the Generation 1 bookings dataframe.
    """

    output = bookings.copy()

    created_col = config[
        "booking_created_timestamp_col"
    ]

    cancelled_col = config[
        "cancellation_timestamp_col"
    ]

    entry_col = config[
        "booking_entry_date_col"
    ]

    exit_col = config[
        "booking_exit_date_col"
    ]

    output[created_col] = pd.to_datetime(
        output[created_col],
        errors="coerce",
    )

    output[cancelled_col] = pd.to_datetime(
        output[cancelled_col],
        errors="coerce",
    )

    output[entry_col] = pd.to_datetime(
        output[entry_col],
        errors="coerce",
    ).dt.normalize()

    output[exit_col] = pd.to_datetime(
        output[exit_col],
        errors="coerce",
    ).dt.normalize()

    return output


def build_cancellation_features_for_flow(
    component_table,
    cancellation_bookings,
    flow,
    config,
):
    """
    Build as-of cancellation metrics for ENTRY or EXIT snapshots.

    A cancellation is counted only when:

        createdAt <= cutoff_timestamp
        cancelledAt <= cutoff_timestamp
        planned entry/exit date == target_date

    Therefore no future cancellation information is used.
    """

    created_col = config[
        "booking_created_timestamp_col"
    ]

    cancelled_col = config[
        "cancellation_timestamp_col"
    ]

    booking_id_col = config[
        "booking_id_col"
    ]

    if flow == "entry":

        target_col = config[
            "booking_entry_date_col"
        ]

    else:

        target_col = config[
            "booking_exit_date_col"
        ]

    snapshots = component_table[
        [
            "target_date",
            "horizon_days",
            "cutoff_timestamp",
        ]
    ].drop_duplicates().copy()

    snapshots["target_date"] = pd.to_datetime(
        snapshots["target_date"],
        errors="coerce",
    ).dt.normalize()

    snapshots["cutoff_timestamp"] = pd.to_datetime(
        snapshots["cutoff_timestamp"],
        errors="coerce",
    )

    booking_subset = cancellation_bookings[
        [
            booking_id_col,
            created_col,
            cancelled_col,
            target_col,
        ]
    ].copy()

    booking_subset = booking_subset.rename(
        columns={
            target_col: "target_date",
        }
    )

    booking_subset["target_date"] = pd.to_datetime(
        booking_subset["target_date"],
        errors="coerce",
    ).dt.normalize()

    # Match each target-date booking with every forecast horizon tested
    # for that same target date.
    expanded = snapshots.merge(
        booking_subset,
        on="target_date",
        how="left",
    )

    # The booking must have existed by the historical cutoff.
    expanded["created_by_cutoff"] = (
        expanded[created_col].notna()
        &
        expanded[created_col].le(
            expanded["cutoff_timestamp"]
        )
    )

    # The cancellation must also have been known by the cutoff.
    expanded["cancelled_by_cutoff"] = (
        expanded["created_by_cutoff"]
        &
        expanded[cancelled_col].notna()
        &
        expanded[cancelled_col].le(
            expanded["cutoff_timestamp"]
        )
    )

    for window_days in config[
        "cancellation_windows_days"
    ]:

        window_start = (
            expanded["cutoff_timestamp"]
            -
            pd.Timedelta(
                days=window_days
            )
        )

        expanded[
            f"cancelled_last_{window_days}d"
        ] = (
            expanded["cancelled_by_cutoff"]
            &
            expanded[cancelled_col].gt(
                window_start
            )
        )

    aggregation = {
        "created_by_cutoff": "sum",
        "cancelled_by_cutoff": "sum",
    }

    for window_days in config[
        "cancellation_windows_days"
    ]:

        aggregation[
            f"cancelled_last_{window_days}d"
        ] = "sum"

    summary = (
        expanded
        .groupby(
            [
                "target_date",
                "horizon_days",
            ],
            as_index=False,
        )
        .agg(aggregation)
    )

    summary = summary.rename(
        columns={
            "created_by_cutoff":
                f"{flow}_gross_created_by_cutoff",

            "cancelled_by_cutoff":
                f"{flow}_cancelled_by_cutoff",
        }
    )

    gross_col = (
        f"{flow}_gross_created_by_cutoff"
    )

    cumulative_cancelled_col = (
        f"{flow}_cancelled_by_cutoff"
    )

    summary[
        f"{flow}_cumulative_cancellation_rate"
    ] = (
        summary[cumulative_cancelled_col]
        /
        summary[gross_col].replace(
            0,
            np.nan,
        )
    ).fillna(0.0)

    for window_days in config[
        "cancellation_windows_days"
    ]:

        source_col = (
            f"cancelled_last_{window_days}d"
        )

        count_col = (
            f"{flow}_cancellations_last_"
            f"{window_days}d"
        )

        rate_col = (
            f"{flow}_cancellation_rate_last_"
            f"{window_days}d"
        )

        summary = summary.rename(
            columns={
                source_col: count_col,
            }
        )

        summary[rate_col] = (
            summary[count_col]
            /
            summary[gross_col].replace(
                0,
                np.nan,
            )
        ).fillna(0.0)

    return summary


def add_cancellation_features(
    component_table,
    bookings,
    config,
):
    """
    Add ENTRY and EXIT as-of cancellation features.

    These are contextual features. They do not replace Generation 1's
    established active-booking logic.
    """

    if not config["enable_cancellation_test"]:

        return component_table.copy()

    print()
    print(
        "Building as-of cancellation-pattern features..."
    )

    start_time = time.perf_counter()

    cancellation_bookings = (
        prepare_cancellation_booking_data(
            bookings=bookings,
            config=config,
        )
    )

    entry_features = (
        build_cancellation_features_for_flow(
            component_table=component_table,
            cancellation_bookings=(
                cancellation_bookings
            ),
            flow="entry",
            config=config,
        )
    )

    exit_features = (
        build_cancellation_features_for_flow(
            component_table=component_table,
            cancellation_bookings=(
                cancellation_bookings
            ),
            flow="exit",
            config=config,
        )
    )

    output = component_table.merge(
        entry_features,
        on=[
            "target_date",
            "horizon_days",
        ],
        how="left",
        validate="one_to_one",
    )

    output = output.merge(
        exit_features,
        on=[
            "target_date",
            "horizon_days",
        ],
        how="left",
        validate="one_to_one",
    )

    cancellation_columns = [
        column
        for column in output.columns
        if (
            "cancellation" in column
            or "cancelled_by_cutoff" in column
            or "gross_created_by_cutoff" in column
        )
    ]

    output[cancellation_columns] = (
        output[cancellation_columns]
        .fillna(0.0)
    )

    print(
        "Completed as-of cancellation features "
        f"[{time.perf_counter() - start_time:.2f}s]"
    )

    return output


def apply_training_cancellation_bands(
    training,
    validation,
    test,
    flow,
    config,
):
    """
    Classify cancellation activity relative to training-period behaviour.

    Training-period medians are used for training, validation, and test.
    Validation and test dates do not contribute to their own band definitions.
    """

    window_days = config[
        "primary_cancellation_window_days"
    ]

    cancellation_rate_col = (
        f"{flow}_cancellation_rate_last_"
        f"{window_days}d"
    )

    relative_col = (
        f"{flow}_cancellation_rate_relative"
    )

    band_col = (
        f"{flow}_cancellation_band"
    )

    training_medians = (
        training
        .groupby(
            [
                "horizon_days",
                "weekday",
            ],
            dropna=False,
        )[cancellation_rate_col]
        .median()
        .rename(
            "training_cancellation_median"
        )
        .reset_index()
    )

    outputs = []

    for frame in (
        training,
        validation,
        test,
    ):

        item = frame.merge(
            training_medians,
            on=[
                "horizon_days",
                "weekday",
            ],
            how="left",
            validate="many_to_one",
        )

        denominator = item[
            "training_cancellation_median"
        ].abs().clip(
            lower=0.0001
        )

        item[relative_col] = (
            item[cancellation_rate_col]
            -
            item[
                "training_cancellation_median"
            ]
        ) / denominator

        item[band_col] = pd.cut(
            item[relative_col],
            bins=config[
                "cancellation_rate_bins"
            ],
            labels=config[
                "cancellation_rate_labels"
            ],
            include_lowest=True,
        ).astype(
            "object"
        ).fillna(
            "unknown"
        )

        item = item.drop(
            columns=[
                "training_cancellation_median"
            ]
        )

        outputs.append(item)

    return tuple(outputs)


# =============================================================================
# 11. EXPERIMENT DEFINITIONS
# =============================================================================

def get_experiments():
    return {
        "BASELINE": {
            "entry": [
                "entry_booking_shrunk",
                "entry_weekday_selected",
                "entry_month",
                "entry_trend_month",
            ],
            "exit": [
                "exit_booking_shrunk",
                "exit_duration",
                "exit_weekday_selected",
                "exit_trend_month",
            ],
        },
        "ADD_BOOKING_PACE": {
            "entry": [
                "entry_booking_shrunk",
                "entry_booking_pace_adjusted",
                "entry_weekday_selected",
                "entry_month",
            ],
            "exit": [
                "exit_booking_shrunk",
                "exit_booking_pace_adjusted",
                "exit_duration",
                "exit_weekday_selected",
            ],
        },
        "ADD_CANCELLATION": {
            "entry": [
                "entry_booking_shrunk",
                "entry_booking_cancellation_adjusted",
                "entry_weekday_selected",
                "entry_month",
            ],

            "exit": [
                "exit_booking_shrunk",
                "exit_booking_cancellation_adjusted",
                "exit_duration",
                "exit_weekday_selected",
            ],
        },
        "ADD_CALENDAR": {
            "entry": [
                "entry_booking_shrunk",
                "entry_weekday_selected",
                "entry_month",
                "entry_calendar_adjusted",
            ],
            "exit": [
                "exit_booking_shrunk",
                "exit_duration",
                "exit_weekday_selected",
                "exit_calendar_adjusted",
            ],
        },
        "ADD_BOOKING_REGIME": {
            "entry": [
                "entry_booking_shrunk",
                "entry_booking_regime_adjusted",
                "entry_weekday_selected",
                "entry_month",
            ],
            "exit": [
                "exit_booking_shrunk",
                "exit_booking_regime_adjusted",
                "exit_duration",
                "exit_weekday_selected",
            ],
        },
        "ADD_WEEKDAY_TREND": {
            "entry": [
                "entry_booking_shrunk",
                "entry_weekday_selected",
                "entry_month",
                "entry_weekday_trend",
            ],
            "exit": [
                "exit_booking_shrunk",
                "exit_duration",
                "exit_weekday_selected",
                "exit_weekday_trend",
            ],
        },
        "ALL_NON_PRICE": {
            "entry": [
                "entry_booking_shrunk",
                "entry_booking_pace_adjusted",
                "entry_booking_regime_adjusted",
                "entry_calendar_adjusted",
                "entry_weekday_selected",
            ],
            "exit": [
                "exit_booking_shrunk",
                "exit_duration",
                "exit_booking_pace_adjusted",
                "exit_booking_regime_adjusted",
                "exit_calendar_adjusted",
            ],
        },
    }


# =============================================================================
# 12. PREPARE FOLD-SPECIFIC CANDIDATES
# =============================================================================

def prepare_fold_candidates(
    training,
    validation,
    test,
    flow,
    historical_selection,
    config,
):
    actual_col = "actual_entries" if flow == "entry" else "actual_exits"
    booking_col = f"{flow}_booking"
    selected_weekday_col = historical_selection["column"]

    outputs = []
    for frame in (training, validation, test):
        item = frame.copy()
        item[f"{flow}_weekday_selected"] = item[selected_weekday_col]
        outputs.append(item)
    training, validation, test = outputs

    training, validation, test = apply_training_pace_bands(
        training=training,
        validation=validation,
        test=test,
        flow=flow,
        config=config,
    )

    if config["enable_cancellation_test"]:

        training, validation, test = (
            apply_training_cancellation_bands(
                training=training,
                validation=validation,
                test=test,
                flow=flow,
                config=config,
            )
        )

    # Pace adjustment learned only from training data.
    pace_adjustments = fit_segment_adjustments(
        training=training,
        base_forecast_col=booking_col,
        actual_col=actual_col,
        segment_cols=["horizon_days", "weekday", f"{flow}_pace_band"],
        minimum_observations=config["minimum_segment_observations"],
        lower=config["pace_adjustment_min"],
        upper=config["pace_adjustment_max"],
    )
    prepared = []
    for frame in (training, validation, test):
        prepared.append(
            apply_segment_adjustment(
                frame=frame,
                base_forecast_col=booking_col,
                adjustment_table=pace_adjustments,
                segment_cols=[
                    "horizon_days", "weekday", f"{flow}_pace_band"
                ],
                output_col=f"{flow}_booking_pace_adjusted",
            )
        )

    training, validation, test = prepared

    # -----------------------------------------------------------------
    # Cancellation-pattern adjustment
    # -----------------------------------------------------------------

    if config["enable_cancellation_test"]:

        cancellation_band_col = (
            f"{flow}_cancellation_band"
        )

        cancellation_adjustments = (
            fit_segment_adjustments(
                training=training,
                base_forecast_col=booking_col,
                actual_col=actual_col,
                segment_cols=[
                    "horizon_days",
                    "weekday",
                    cancellation_band_col,
                ],
                minimum_observations=config[
                    "minimum_cancellation_segment_observations"
                ],
                lower=config[
                    "cancellation_adjustment_min"
                ],
                upper=config[
                    "cancellation_adjustment_max"
                ],
            )
        )

        prepared = []

        for frame in (
            training,
            validation,
            test,
        ):

            prepared.append(
                apply_segment_adjustment(
                    frame=frame,
                    base_forecast_col=booking_col,
                    adjustment_table=(
                        cancellation_adjustments
                    ),
                    segment_cols=[
                        "horizon_days",
                        "weekday",
                        cancellation_band_col,
                    ],
                    output_col=(
                        f"{flow}_booking_"
                        f"cancellation_adjusted"
                    ),
                )
            )

        training, validation, test = prepared

    else:

        prepared = []

        for frame in (
            training,
            validation,
            test,
        ):

            item = frame.copy()

            item[
                f"{flow}_booking_"
                f"cancellation_adjusted"
            ] = item[booking_col]

            prepared.append(item)

        training, validation, test = prepared




    # Calendar adjustment learned only from training data.
    calendar_base = f"{flow}_month"
    calendar_adjustments = fit_segment_adjustments(
        training=training,
        base_forecast_col=calendar_base,
        actual_col=actual_col,
        segment_cols=["horizon_days", "calendar_type"],
        minimum_observations=config["minimum_segment_observations"],
        lower=config["calendar_adjustment_min"],
        upper=config["calendar_adjustment_max"],
    )
    prepared = []
    for frame in (training, validation, test):
        prepared.append(
            apply_segment_adjustment(
                frame=frame,
                base_forecast_col=calendar_base,
                adjustment_table=calendar_adjustments,
                segment_cols=["horizon_days", "calendar_type"],
                output_col=f"{flow}_calendar_adjusted",
            )
        )
    training, validation, test = prepared

    # Booking regime adjustment learned only from training data.
    edges = fit_regime_edges(training, flow, config["regime_quantiles"])
    training = apply_regime_band(training, flow, edges)
    validation = apply_regime_band(validation, flow, edges)
    test = apply_regime_band(test, flow, edges)

    regime_adjustments = fit_segment_adjustments(
        training=training,
        base_forecast_col=booking_col,
        actual_col=actual_col,
        segment_cols=["horizon_days", f"{flow}_regime_band"],
        minimum_observations=config["minimum_segment_observations"],
        lower=config["regime_adjustment_min"],
        upper=config["regime_adjustment_max"],
    )
    prepared = []
    for frame in (training, validation, test):
        prepared.append(
            apply_segment_adjustment(
                frame=frame,
                base_forecast_col=booking_col,
                adjustment_table=regime_adjustments,
                segment_cols=["horizon_days", f"{flow}_regime_band"],
                output_col=f"{flow}_booking_regime_adjusted",
            )
        )
    return tuple(prepared)


# =============================================================================
# 13. RUN FEATURE ABLATION TOURNAMENT
# =============================================================================

def run_experiments(component_table, refined, config):
    folds = refined.create_rolling_month_folds(
        component_table=component_table,
        minimum_training_months=config["minimum_training_months"],
        validation_months=config["validation_months"],
        maximum_training_months=config["maximum_weight_training_months"],
    )
    experiments = get_experiments()
    horizons = sorted(component_table["horizon_days"].unique())

    predictions = []
    selected_rows = []
    all_tests = []
    historical_diagnostics = []
    prior_weight_lookup = {}

    tasks = len(folds) * 2 * len(horizons) * len(experiments)
    counter = 0

    for fold in folds:
        for flow in ("entry", "exit"):
            actual_col = "actual_entries" if flow == "entry" else "actual_exits"

            for horizon in horizons:
                horizon_frame = component_table[
                    component_table["horizon_days"].eq(horizon)
                ].copy()
                training = horizon_frame[
                    horizon_frame["month_start"].isin(fold["training_months"])
                ].copy()
                validation = horizon_frame[
                    horizon_frame["month_start"].isin(fold["validation_months"])
                ].copy()
                test = horizon_frame[
                    horizon_frame["month_start"].eq(fold["test_month"])
                ].copy()

                if training.empty or validation.empty or test.empty:
                    continue

                historical_selection, history_diag = choose_historical_window(
                    training=training,
                    validation=validation,
                    demand_type=flow,
                    windows=config["same_weekday_windows"],
                    refined=refined,
                )
                history_diag["fold_id"] = fold["fold_id"]
                history_diag["test_month"] = fold["test_month"]
                history_diag["flow"] = flow
                history_diag["horizon_days"] = horizon
                historical_diagnostics.append(history_diag)

                training, validation, test = prepare_fold_candidates(
                    training=training,
                    validation=validation,
                    test=test,
                    flow=flow,
                    historical_selection=historical_selection,
                    config=config,
                )

                for experiment_name, experiment in experiments.items():
                    counter += 1
                    print(
                        f"\rExperiment {counter:,}/{tasks:,} "
                        f"| {fold['test_month']:%Y-%m} | {flow.upper()} "
                        f"| T-{horizon} | {experiment_name}",
                        end="",
                    )

                    components = experiment[flow]
                    weight_grid = refined.generate_weight_grid(
                        component_names=components,
                        step=config["weight_step"],
                    )
                    prior_key = (experiment_name, flow, int(horizon))

                    best, candidates = refined.select_best_weights(
                        training_frame=training,
                        validation_frame=validation,
                        demand_type=flow,
                        component_names=components,
                        actual_col=actual_col,
                        weight_grid=weight_grid,
                        shrinkage_strengths=config[
                            "booking_shrinkage_strengths"
                        ],
                        prior_weights=prior_weight_lookup.get(prior_key),
                        stability_penalty=config["weight_stability_penalty"],
                    )
                    if best is None:
                        continue

                    prior_weight_lookup[prior_key] = {
                        name: best[name] for name in components
                    }

                    if config["export_all_weight_tests"]:
                        candidates["experiment"] = experiment_name
                        candidates["fold_id"] = fold["fold_id"]
                        candidates["test_month"] = fold["test_month"]
                        candidates["flow"] = flow
                        candidates["horizon_days"] = horizon
                        all_tests.append(candidates)

                    selected_rows.append(
                        {
                            "experiment": experiment_name,
                            "fold_id": fold["fold_id"],
                            "test_month": fold["test_month"],
                            "flow": flow,
                            "horizon_days": horizon,
                            "selected_historical_window": historical_selection[
                                "window"
                            ],
                            "shrinkage_strength": best["shrinkage_strength"],
                            "training_wape_pct": best["training_wape_pct"],
                            "validation_wape_pct": best[
                                "validation_wape_pct"
                            ],
                            **{name: best[name] for name in components},
                        }
                    )

                    test_candidate = refined.add_shrunk_booking_component(
                        frame=test,
                        demand_type=flow,
                        shrinkage_strength=best["shrinkage_strength"],
                    )
                    test_candidate["forecast_value"] = (
                        refined.forecast_from_weights(
                            frame=test_candidate,
                            component_names=components,
                            weight_row=best,
                        )
                    )
                    test_candidate["actual_value"] = test_candidate[actual_col]
                    test_candidate["experiment"] = experiment_name
                    test_candidate["flow"] = flow
                    test_candidate["fold_id"] = fold["fold_id"]
                    test_candidate["test_month"] = fold["test_month"]
                    test_candidate["selected_historical_window"] = (
                        historical_selection["window"]
                    )
                    predictions.append(add_error_columns(test_candidate))

    print()
    return (
        pd.concat(predictions, ignore_index=True) if predictions else pd.DataFrame(),
        pd.DataFrame(selected_rows),
        pd.concat(all_tests, ignore_index=True) if all_tests else pd.DataFrame(),
        pd.concat(historical_diagnostics, ignore_index=True)
        if historical_diagnostics else pd.DataFrame(),
        pd.DataFrame(folds),
    )


# =============================================================================
# 14. PERFORMANCE AND INCREMENTAL BENEFIT
# =============================================================================

def build_incremental_benefit(performance):
    if performance.empty:
        return pd.DataFrame()
    baseline = performance[
        performance["experiment"].eq("BASELINE")
    ][["flow", "horizon_days", "wape_pct", "mae", "rmse", "bias"]].rename(
        columns={
            "wape_pct": "baseline_wape_pct",
            "mae": "baseline_mae",
            "rmse": "baseline_rmse",
            "bias": "baseline_bias",
        }
    )
    result = performance.merge(
        baseline, on=["flow", "horizon_days"], how="left"
    )
    result["wape_improvement_points"] = (
        result["baseline_wape_pct"] - result["wape_pct"]
    )
    result["relative_wape_reduction_pct"] = np.where(
        result["baseline_wape_pct"].gt(0),
        result["wape_improvement_points"]
        / result["baseline_wape_pct"]
        * 100,
        np.nan,
    )
    result["mae_improvement"] = result["baseline_mae"] - result["mae"]
    result["rmse_improvement"] = result["baseline_rmse"] - result["rmse"]
    result["absolute_bias_change"] = (
        result["baseline_bias"].abs() - result["bias"].abs()
    )
    return result.sort_values(
        ["flow", "horizon_days", "wape_pct"]
    ).reset_index(drop=True)


def build_fold_win_rate(predictions):
    if predictions.empty:
        return pd.DataFrame()
    fold_perf_rows = []
    for keys, group in predictions.groupby(
        ["flow", "horizon_days", "fold_id", "test_month", "experiment"]
    ):
        flow, horizon, fold_id, test_month, experiment = keys
        actual = group["actual_value"]
        forecast = group["forecast_value"]
        fold_perf_rows.append(
            {
                "flow": flow,
                "horizon_days": horizon,
                "fold_id": fold_id,
                "test_month": test_month,
                "experiment": experiment,
                "wape_pct": (
                    (forecast - actual).abs().sum() / actual.abs().sum() * 100
                ),
            }
        )
    fold_perf = pd.DataFrame(fold_perf_rows)
    baseline = fold_perf[fold_perf["experiment"].eq("BASELINE")].rename(
        columns={"wape_pct": "baseline_wape_pct"}
    )[["flow", "horizon_days", "fold_id", "baseline_wape_pct"]]
    comparison = fold_perf.merge(
        baseline,
        on=["flow", "horizon_days", "fold_id"],
        how="left",
    )
    comparison["beat_baseline"] = (
        comparison["wape_pct"] < comparison["baseline_wape_pct"]
    )
    return comparison.groupby(
        ["flow", "horizon_days", "experiment"], as_index=False
    ).agg(
        folds=("fold_id", "nunique"),
        folds_beating_baseline=("beat_baseline", "sum"),
        win_rate_pct=("beat_baseline", lambda s: s.mean() * 100),
        average_test_wape_pct=("wape_pct", "mean"),
        average_baseline_wape_pct=("baseline_wape_pct", "mean"),
    )


# =============================================================================
# 15. CHECKPOINTS AND EXPORT
# =============================================================================

def checkpoint(frame, path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_pickle(path)


def format_workbook(workbook):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for cells in worksheet.columns:
            width = min(
                max(max((len(str(c.value)) if c.value is not None else 0) for c in cells) + 2, 10),
                60,
            )
            worksheet.column_dimensions[
                get_column_letter(cells[0].column)
            ].width = width


def export_results(
    config,
    performance,
    incremental,
    fold_win_rate,
    selected_weights,
    predictions,
    component_table,
    historical_diagnostics,
    folds,
    all_weight_tests,
):
    path = Path(config["output_path"])
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        performance.to_excel(excel_writer = writer, sheet_name= "Experiment Performance", index=False)
        incremental.to_excel(excel_writer = writer, sheet_name= "Incremental Benefit", index=False)
        fold_win_rate.to_excel(excel_writer = writer, sheet_name= "Fold Win Rate", index=False)
        selected_weights.to_excel(excel_writer = writer, sheet_name= "Selected Weights", index=False)
        historical_diagnostics.to_excel(
            excel_writer = writer, sheet_name= "Historical Window Tests", index=False
        )
        folds.to_excel(excel_writer = writer, sheet_name= "Rolling Folds", index=False)
        predictions.to_excel(excel_writer = writer, sheet_name= "Out of Sample Forecasts", index=False)
        component_table.to_excel(excel_writer = writer, sheet_name= "Extended Components", index=False)
        if not all_weight_tests.empty:
            all_weight_tests.to_excel(excel_writer = writer, sheet_name= "All Weight Tests", index=False)
        format_workbook(writer.book)
    return path


# =============================================================================
# 16. MAIN
# =============================================================================

if __name__ == "__main__":
    timer = time.perf_counter()
    config = get_non_price_config()
    base, refined = import_modules(config)
    base_config = build_base_config(base, config)
    test_dates = create_test_dates(config)
    horizons = (
        config["smoke_test_horizons"]
        if config["smoke_test"]
        else config["forecast_horizons_days"]
    )

    print("=" * 90)
    print("FASTPARK GENERATION 2.1 - NON-PRICE COMPONENT TOURNAMENT")
    print("=" * 90)
    print(f"Target dates: {len(test_dates):,}")
    print(f"Horizons: {len(horizons):,}")

    raw_data = base.load_simulation_data(base_config)
    base.validate_loaded_data(raw_data, base_config)
    timer = step(timer, "Loaded and validated source data")

    model_data = refined.prepare_refined_model_data(
        base=base,
        raw_data=raw_data,
        base_config=base_config,
        refined_config=config,
        test_dates=test_dates,
        horizons=horizons,
    )
    validate_cancellation_configuration(model_data["bookings"], config)
    timer = step(timer, "Prepared Generation 1 data and all-date caches")

    component_table = refined.build_component_table(
        base=base,
        data=model_data,
        test_dates=test_dates,
        horizons=horizons,
        refined_config=config,
    )
    timer = step(timer, f"Built base component table ({len(component_table):,} rows)")

    checkpoint_dir = Path(config["checkpoint_dir"])
    checkpoint(component_table, checkpoint_dir / "01_base_components.pkl")

    component_table = calculate_same_weekday_forecasts(
        base=base,
        data=model_data,
        component_table=component_table,
        windows=config["same_weekday_windows"],
    )
    timer = step(timer, "Added alternative same-weekday windows")

    component_table = add_booking_pace_features(component_table, config)
    timer = step(timer, "Added booking-pace features")

    if config["enable_cancellation_test"]:

        component_table = add_cancellation_features(
            component_table=component_table,
            bookings=model_data["bookings"],
            config=config,
        )

        timer = step(
            timer,
            "Added as-of cancellation-pattern features",
        )

    calendar = build_calendar_table(
        config["simulation_start"],
        config["simulation_end"],
        MANUAL_CALENDAR_PERIODS,
    )
    component_table = component_table.merge(
        calendar,
        on="target_date",
        how="left",
        validate="many_to_one",
    )
    timer = step(timer, "Added manual calendar features")

    component_table = add_weekday_trend_features(
        component_table=component_table,
        daily_actuals=model_data["daily_actuals"],
        config=config,
    )
    timer = step(timer, "Added weekday-specific trend forecasts")

    checkpoint(component_table, checkpoint_dir / "02_extended_components.pkl")

    predictions, selected_weights, all_weight_tests, history_diag, folds = (
        run_experiments(
            component_table=component_table,
            refined=refined,
            config=config,
        )
    )
    timer = step(
        timer,
        f"Completed non-price tournament ({len(predictions):,} forecasts)",
    )

    # Critical outputs are checkpointed before post-processing/export.
    checkpoint(predictions, checkpoint_dir / "03_predictions.pkl")
    checkpoint(selected_weights, checkpoint_dir / "04_selected_weights.pkl")
    checkpoint(all_weight_tests, checkpoint_dir / "05_all_weight_tests.pkl")

    performance = score_frame(
        refined,
        predictions,
        ["flow", "experiment", "horizon_days"],
    ).sort_values(["flow", "horizon_days", "wape_pct"])

    incremental = build_incremental_benefit(performance)
    fold_win_rate = build_fold_win_rate(predictions)

    output_path = export_results(
        config=config,
        performance=performance,
        incremental=incremental,
        fold_win_rate=fold_win_rate,
        selected_weights=selected_weights,
        predictions=predictions,
        component_table=component_table,
        historical_diagnostics=history_diag,
        folds=folds,
        all_weight_tests=all_weight_tests,
    )
    timer = step(timer, f"Exported {output_path.name}")

    print("\nTop experiment by flow and horizon:")
    print(
        performance.groupby(["flow", "horizon_days"], as_index=False)
        .first()[["flow", "horizon_days", "experiment", "wape_pct", "bias"]]
        .to_string(index=False)
    )
    print(f"\nOutput workbook: {output_path}")
