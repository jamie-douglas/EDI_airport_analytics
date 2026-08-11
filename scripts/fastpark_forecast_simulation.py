

"""
FASTPARK FORECAST SIMULATION / HISTORICAL BACKTESTING
======================================================

Purpose
-------
This script is designed to answer:

    "If we had stood on a historical day at T-56, T-28, T-14,
     T-7, T-1, T-0 etc., what forecast would we have made using
     ONLY information that was available at that point?"

We then compare that forecast against the actual entries/exits that
eventually occurred.

This is therefore a HISTORICAL FORECAST BACKTEST / MODEL TOURNAMENT.

It is not a Monte Carlo simulation.

----------------------------------------------------------------------
TEST PERIOD
----------------------------------------------------------------------

Target dates:

    1st - 7th
    14th - 20th

of each month from:

    July 2025 -> March 2026

This gives us 126 target dates:

    9 months x 14 days = 126 days

----------------------------------------------------------------------
FORECAST HORIZONS
----------------------------------------------------------------------

For every target date we test:

    T-0
    T-1
    T-2
    T-3
    T-4
    T-5
    T-6
    T-7
    T-14
    T-21
    T-28
    T-35
    T-42
    T-49
    T-56

----------------------------------------------------------------------
CORE PRINCIPLE: NO LOOK-AHEAD BIAS
----------------------------------------------------------------------

At T-28 for a target date, the model may use:

    * bookings created by T-28
    * cancellations that had happened by T-28
    * historical actuals before T-28
    * historical passenger patterns available before T-28
    * historical model relationships learned before T-28

It MUST NOT use:

    * future bookings
    * future cancellations
    * the target day's actual entries
    * the target day's actual exits
    * the target day's actual passenger count

This distinction is particularly important for booking curves.

Example:

    Booking created: 1 August
    Target entry:    1 September
    Cancellation:    25 August

At T-28 (4 August), the booking WAS known and WAS active.

Therefore it must be included in the T-28 book of business.

We cannot simply filter on today's final booking status because that would
remove the booking retrospectively and introduce look-ahead bias.

----------------------------------------------------------------------
DATA SOURCES
----------------------------------------------------------------------

The existing FastPark analysis already establishes these sources:

    AirportX.v_Bookings
    FastPark.v_EntryAndExits
    EAL.FlightPerformance

We deliberately reuse the existing analysis extraction/cleaning functions
rather than recreating them here.

That keeps the forecasting framework aligned with the existing analysis.
"""

# =============================================================================
# 0. IMPORTS
# =============================================================================

import sys
import pathlib
import time

from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd


# -----------------------------------------------------------------------------
# Project path
# -----------------------------------------------------------------------------

PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))


# -----------------------------------------------------------------------------
# Existing project utilities
# -----------------------------------------------------------------------------

from modules.utils.progress import step
from modules.utils.db import get_engine


# -----------------------------------------------------------------------------
# IMPORTANT
# -----------------------------------------------------------------------------
# Re-use the existing extraction and cleaning functions from
# fastpark_forecast_analysis.py.
#
# The existing analysis already defines:
#
#     AirportX.v_Bookings
#     FastPark.v_EntryAndExits
#     EAL.FlightPerformance
#
# and the important field definitions.
#
# This means we don't want the simulation creating a second, subtly
# different version of the data.
# -----------------------------------------------------------------------------

from fastpark_forecast_analysis import (
    get_fastpark_bookings,
    get_fastpark_entry_exits,
    get_historical_flight_performance,
    clean_bookings,
    clean_operations,
    clean_flights,
    reconcile_bookings_to_operations,
)


# =============================================================================
# 1. CONFIGURATION
# =============================================================================

def get_simulation_config():
    """
    Central configuration for the forecasting experiment.

    Keeping this in one place is intentional.

    The aim is that someone can later change:

        * test period
        * test dates
        * forecast horizons
        * historical look-back windows
        * output location

    without having to search through the modelling code.
    """

    config = {

        # ---------------------------------------------------------------------
        # PRODUCT
        # ---------------------------------------------------------------------

        "asset_name": "FastPark",

        # ---------------------------------------------------------------------
        # SETTINGS REQUIRED BY REUSED ANALYSIS CLEANING FUNCTIONS
        # ---------------------------------------------------------------------

        # Booking status definitions
        "valid_booking_status": "B",
        "cancelled_status": "CX",
        "unknown_statuses": ["F"],

        # Duration bands used by clean_bookings()
        "duration_bins_days": [
            0,
            1,
            3,
            7,
            10,
            14,
            21,
            999,
        ],

        "duration_labels": [
            "0-1 days",
            "2-3 days",
            "4-7 days",
            "8-10 days",
            "11-14 days",
            "15-21 days",
            "22+ days",
        ],

        # Operational timestamp definitions used by clean_operations()
        "actual_entry_timestamp_col": "CheckInEnded",
        "actual_exit_timestamp_col": "ActualCheckedOutDate",

        # Passenger field used by clean_flights()
        "historical_pax_col": "Passengers",



        # ---------------------------------------------------------------------
        # HISTORICAL DATA EXTRACTION WINDOW
        # ---------------------------------------------------------------------
        #
        # We need history BEFORE July 2025 because the first forecast point
        # we test is T-56.
        #
        # We use 2024 onwards to provide plenty of historical observations
        # for same-weekday, seasonal and booking-curve scenarios.
        # ---------------------------------------------------------------------

        "history_start": "2024-01-01",
        "history_end": "2026-07-31",


        # ---------------------------------------------------------------------
        # ACTUAL TEST PERIOD
        # ---------------------------------------------------------------------
        #
        # The requested test:
        #
        #       1st -> 7th
        #       14th -> 20th
        #
        # for every month July 2025 through July 2026.
        # ---------------------------------------------------------------------

        "simulation_start": "2025-07-01",
        "simulation_end": "2026-07-31",

        "test_days_of_month": (
            list(range(1, 8))
            +
            list(range(14, 21))
        ),


        # ---------------------------------------------------------------------
        # FORECAST HORIZONS
        # ---------------------------------------------------------------------
        #
        # These are deliberately kept exactly as requested.
        # ---------------------------------------------------------------------

        "forecast_horizons_days": [
            0,
            1,
            2,
            3,
            4,
            5,
            6,
            7,
            14,
            21,
            28,
            35,
            42,
            49,
            56,
        ],


        # ---------------------------------------------------------------------
        # SAME-WEEKDAY HISTORY
        # ---------------------------------------------------------------------
        #
        # The existing analysis tested same-weekday tendency windows.
        #
        # We retain these because the existing results suggested that
        # recent same-weekday methods are strong baselines.
        # ---------------------------------------------------------------------

        "same_weekday_windows": [
            1,
            2,
            4,
            6,
            8,
            10,
            12,
        ],


        # ---------------------------------------------------------------------
        # BOOKING CURVE
        # ---------------------------------------------------------------------

        # Minimum number of historical comparable dates we would ideally
        # like before trusting a segmented estimate.
        "minimum_curve_observations": 10,

        # ---------------------------------------------------------------------
        # WEIGHT CALIBRATION
        # ---------------------------------------------------------------------
        #
        # Hybrid weights are calibrated separately from the scenario catalogue.
        # They are not hard-coded as individual E/X scenarios.
        #
        # The grid below tests booking weights from 0% to 100% in 5% increments.
        # The complementary component receives:
        #
        #     1 - booking_weight
        #
        # Examples:
        #
        #     booking_weight = 0.80
        #     comparison_weight = 0.20
        #
        # Weight calibration is performed independently for every forecast
        # horizon to avoid assuming the same blend is optimal at T-0 and T-56.
        # ---------------------------------------------------------------------

        "weight_grid": [
            round(x, 2)
            for x in np.arange(
                0.00,
                1.001,
                0.05,
            )
        ],

        # Minimum number of observations needed before a calibrated weight
        # can be treated as valid.
        "minimum_weight_calibration_observations": 30,

        # ---------------------------------------------------------------------
        # EXECUTION / ERROR HANDLING
        # ---------------------------------------------------------------------

        # True:
        #     stop immediately if any scenario raises an exception.
        #
        # False:
        #     record the exception in the simulation output and continue.
        #
        # Keep this True while developing the revised script.
        "fail_fast": True,

        # ---------------------------------------------------------------------
        # HOURLY ANALYSIS
        # ---------------------------------------------------------------------

        # Run hourly profile validation after the daily tournament.
        "enable_hourly_analysis": True,

        # Test alternative same-weekday hourly profile windows rather than
        # assuming that six weeks is optimal.
        "hourly_profile_windows": [
            4,
            6,
            8,
            12,
        ],

        "hourly_output_path": (
            PROJECT_ROOT
            / "outputs"
            / "fastpark_hourly_forecast_simulation.xlsx"
        ),

        # ---------------------------------------------------------------------
        # WEEKDAY-MONTH SEASONALITY
        # ---------------------------------------------------------------------

        # Test whether same-weekday-same-month history outperforms same-weekday-any-month history.

        "use_weekday_month_seasonality": True,

        #Minimum observations required for weekday-month method. 

        "minimum_weekday_month_observations": 3,

        # ---------------------------------------------------------------------
        # SMOKE TEST
        # ---------------------------------------------------------------------

        "smoke_test": False,

        "smoke_test_target_count": 2,

        "smoke_test_horizons": [
            0,
            7,
            28,
            56,
        ],

        "output_path": (
            PROJECT_ROOT
            / "outputs"
            / "fastpark_forecast_simulation_v2.xlsx"
        ),

    }

    return config


# =============================================================================
# 2. DATA LOADING
# =============================================================================

def load_simulation_data(config):
    """
    Load all data required for the forecasting simulation.

    We intentionally call the existing analysis functions.

    This keeps the simulation consistent with the analysis around:

        * booking status
        * date fields
        * operational entry
        * operational exit
        * passenger field
        * flight filtering
        * booking/operations reconciliation
    """

    print("=" * 80)
    print("FASTPARK FORECAST SIMULATION")
    print("=" * 80)

    print("\nLoading database engine...")

    engine = get_engine()


    # -------------------------------------------------------------------------
    # BOOKINGS
    # -------------------------------------------------------------------------

    print("\n[1/5] Loading FastPark bookings...")

    bookings = get_fastpark_bookings(
        start=config["history_start"],
        end=config["history_end"],
        statuses=[
            "B",
            "CX",
            "F",
        ],
        asset_name=config["asset_name"],
        engine=engine,
    )


    # -------------------------------------------------------------------------
    # OPERATIONAL ENTRIES / EXITS
    # -------------------------------------------------------------------------

    print("[2/5] Loading FastPark operational entries/exits...")

    operations = get_fastpark_entry_exits(
        start=config["history_start"],
        end=config["history_end"],
        engine=engine,
    )


    # -------------------------------------------------------------------------
    # FLIGHT / PASSENGER DATA
    # -------------------------------------------------------------------------

    print("[3/5] Loading airport flight/passenger history...")

    flights = get_historical_flight_performance(
        start=config["history_start"],
        end=config["history_end"],
        engine=engine,
    )


    # -------------------------------------------------------------------------
    # CLEAN
    # -------------------------------------------------------------------------

    print("[4/5] Cleaning datasets...")

    bookings_clean = clean_bookings(
        bookings,
        config,
    )

    operations_clean = clean_operations(
        operations,
        config,
    )

    flights_clean = clean_flights(
        flights,
        config,
    )


    # -------------------------------------------------------------------------
    # RECONCILE BOOKINGS TO OPERATIONS
    # -------------------------------------------------------------------------

    print("[5/5] Reconciling bookings to operational records...")

    master = reconcile_bookings_to_operations(
        bookings_clean,
        operations_clean,
    )


    return {
        "bookings": bookings_clean,
        "operations": operations_clean,
        "flights": flights_clean,
        "master": master,
    }


# =============================================================================
# 3. DAILY ACTUALS
# =============================================================================

def create_daily_actuals(master):
    """
    Create the actual daily FastPark demand used as the validation target.

    Entry definition
    ----------------
    Actual operational entry = CheckInEnded.

    Exit definition
    ---------------
    Actual operational exit = ActualCheckedOutDate.

    These definitions follow the existing analysis.

    IMPORTANT
    ---------
    This table represents the FINAL ACTUALS.

    It is ONLY used as the thing we are trying to predict.

    The forecasting scenarios themselves must not use the target day's
    actual values.
    """

    df = master.copy()


    # -------------------------------------------------------------------------
    # ENTRIES
    # -------------------------------------------------------------------------

    entries = (
        df
        .dropna(subset=["actual_entry_ts"])
        .groupby("actual_entry_date")
        .agg(
            entries=("bookingId", "nunique")
        )
        .reset_index()
        .rename(
            columns={
                "actual_entry_date": "date"
            }
        )
    )


    # -------------------------------------------------------------------------
    # EXITS
    # -------------------------------------------------------------------------

    exits = (
        df
        .dropna(subset=["actual_exit_ts"])
        .groupby("actual_exit_date")
        .agg(
            exits=("bookingId", "nunique")
        )
        .reset_index()
        .rename(
            columns={
                "actual_exit_date": "date"
            }
        )
    )


    # -------------------------------------------------------------------------
    # COMBINE
    # -------------------------------------------------------------------------

    daily = entries.merge(
        exits,
        on="date",
        how="outer",
    )

    daily["entries"] = (
        daily["entries"]
        .fillna(0)
        .astype(int)
    )

    daily["exits"] = (
        daily["exits"]
        .fillna(0)
        .astype(int)
    )

    daily["date"] = pd.to_datetime(
        daily["date"],
        errors="coerce",
    )

    daily["weekday"] = (
        daily["date"]
        .dt.day_name()
    )

    daily["weekday_num"] = (
        daily["date"]
        .dt.weekday
    )

    daily["month"] = (
        daily["date"]
        .dt.month
    )

    daily["year"] = (
        daily["date"]
        .dt.year
    )

    daily["movements"] = (
        daily["entries"]
        +
        daily["exits"]
    )

    daily["net_flow"] = (
        daily["entries"]
        -
        daily["exits"]
    )

    return (
        daily
        .sort_values("date")
        .reset_index(drop=True)
    )


# =============================================================================
# 4. HOURLY ACTUALS
# =============================================================================

def create_hourly_actuals(master):
    """
    Create actual hourly entries and exits on a complete hourly date grid.
    """

    df = master.copy()

    df["actual_entry_ts"] = pd.to_datetime(
        df["actual_entry_ts"],
        errors="coerce",
    )

    df["actual_exit_ts"] = pd.to_datetime(
        df["actual_exit_ts"],
        errors="coerce",
    )

    entries = (
        df
        .dropna(
            subset=[
                "actual_entry_ts"
            ]
        )
        .assign(
            datetime=lambda x:
                x["actual_entry_ts"]
                .dt.floor("h")
        )
        .groupby(
            "datetime",
            as_index=False,
        )
        .agg(
            entries=(
                "bookingId",
                "nunique",
            )
        )
    )

    exits = (
        df
        .dropna(
            subset=[
                "actual_exit_ts"
            ]
        )
        .assign(
            datetime=lambda x:
                x["actual_exit_ts"]
                .dt.floor("h")
        )
        .groupby(
            "datetime",
            as_index=False,
        )
        .agg(
            exits=(
                "bookingId",
                "nunique",
            )
        )
    )

    minimum_datetime = min(
        entries["datetime"].min(),
        exits["datetime"].min(),
    ).floor("D")

    maximum_datetime = max(
        entries["datetime"].max(),
        exits["datetime"].max(),
    ).ceil("D") - pd.Timedelta(hours=1)

    complete_grid = pd.DataFrame(
        {
            "datetime": pd.date_range(
                start=minimum_datetime,
                end=maximum_datetime,
                freq="h",
            )
        }
    )

    hourly = (
        complete_grid
        .merge(
            entries,
            on="datetime",
            how="left",
        )
        .merge(
            exits,
            on="datetime",
            how="left",
        )
    )

    hourly["entries"] = (
        hourly["entries"]
        .fillna(0)
        .astype(int)
    )

    hourly["exits"] = (
        hourly["exits"]
        .fillna(0)
        .astype(int)
    )

    hourly["date"] = (
        hourly["datetime"]
        .dt.normalize()
    )

    hourly["hour"] = (
        hourly["datetime"]
        .dt.hour
    )

    hourly["weekday"] = (
        hourly["datetime"]
        .dt.day_name()
    )

    hourly["weekday_num"] = (
        hourly["datetime"]
        .dt.weekday
    )

    hourly["month"] = (
        hourly["datetime"]
        .dt.month
    )

    return hourly


# =============================================================================
# 5. DAILY PASSENGER CONTEXT
# =============================================================================

def create_daily_passenger_context(flights):
    """
    Create historical daily passenger context.

    Entry demand is linked primarily to:

        departing passengers

    Exit demand is linked primarily to:

        arriving passengers

    We also preserve:

        domestic passenger volume
        international passenger volume
        international passenger share

    because the existing analysis explicitly investigated passenger mix.

    IMPORTANT
    ---------
    This is historical context used to learn relationships.

    We must NOT feed the target day's actual passenger count into a
    production-like forecast.

    Instead, the forecasting scenarios will estimate target-day passenger
    demand from historical information.
    """

    df = flights.copy()

    df["ScheduledDateTime_Local"] = pd.to_datetime(
        df["ScheduledDateTime_Local"],
        errors="coerce",
    )

    df["date"] = (
        df["ScheduledDateTime_Local"]
        .dt.normalize()
    )

    # -------------------------------------------------------------------------
    # Departure / arrival flags
    # -------------------------------------------------------------------------

    df["is_departure"] = (
        df["ArrDeptureCode"]
        .eq("D")
    )

    df["is_arrival"] = (
        df["ArrDeptureCode"]
        .eq("A")
    )

    df["is_domestic"] = (
        df["Domestic_International"]
        .eq("Domestic")
    )

    df["is_international"] = (
        df["Domestic_International"]
        .eq("International")
    )


    # -------------------------------------------------------------------------
    # Build separate components first.
    #
    # This is deliberately clearer than doing everything inside a single
    # groupby lambda.
    # -------------------------------------------------------------------------

    departure = df[df["is_departure"]].copy()

    arrival = df[df["is_arrival"]].copy()


    # -------------------------------------------------------------------------
    # DEPARTURES
    # -------------------------------------------------------------------------

    departing = (
        departure
        .groupby("date")
        .agg(
            departing_pax=(
                "analysis_pax",
                "sum",
            ),
            domestic_departing_pax=(
                "analysis_pax",
                lambda x: x[
                    departure.loc[
                        x.index,
                        "is_domestic"
                    ]
                ].sum()
            ),
            international_departing_pax=(
                "analysis_pax",
                lambda x: x[
                    departure.loc[
                        x.index,
                        "is_international"
                    ]
                ].sum()
            ),
            departing_flights=(
                "FlightID",
                "nunique",
            ),
        )
        .reset_index()
    )


    # -------------------------------------------------------------------------
    # ARRIVALS
    # -------------------------------------------------------------------------

    arriving = (
        arrival
        .groupby("date")
        .agg(
            arriving_pax=(
                "analysis_pax",
                "sum",
            ),
            domestic_arriving_pax=(
                "analysis_pax",
                lambda x: x[
                    arrival.loc[
                        x.index,
                        "is_domestic"
                    ]
                ].sum()
            ),
            international_arriving_pax=(
                "analysis_pax",
                lambda x: x[
                    arrival.loc[
                        x.index,
                        "is_international"
                    ]
                ].sum()
            ),
            arriving_flights=(
                "FlightID",
                "nunique",
            ),
        )
        .reset_index()
    )


    # -------------------------------------------------------------------------
    # COMBINE
    # -------------------------------------------------------------------------

    daily = departing.merge(
        arriving,
        on="date",
        how="outer",
    )

    numeric_cols = [
        "departing_pax",
        "domestic_departing_pax",
        "international_departing_pax",
        "departing_flights",
        "arriving_pax",
        "domestic_arriving_pax",
        "international_arriving_pax",
        "arriving_flights",
    ]

    for col in numeric_cols:
        if col not in daily.columns:
            daily[col] = 0

        daily[col] = (
            daily[col]
            .fillna(0)
        )


    # -------------------------------------------------------------------------
    # PASSENGER MIX
    # -------------------------------------------------------------------------

    daily["international_departing_share"] = (
        daily["international_departing_pax"]
        /
        daily["departing_pax"].replace(
            0,
            np.nan,
        )
    )

    daily["international_arriving_share"] = (
        daily["international_arriving_pax"]
        /
        daily["arriving_pax"].replace(
            0,
            np.nan,
        )
    )


    # -------------------------------------------------------------------------
    # CALENDAR
    # -------------------------------------------------------------------------

    daily["weekday"] = (
        daily["date"]
        .dt.day_name()
    )

    daily["weekday_num"] = (
        daily["date"]
        .dt.weekday
    )

    daily["month"] = (
        daily["date"]
        .dt.month
    )

    daily["year"] = (
        daily["date"]
        .dt.year
    )

    return (
        daily
        .sort_values("date")
        .reset_index(drop=True)
    )


# =============================================================================
# 6. TEST DATE GENERATOR
# =============================================================================

def create_test_dates(config):
    """
    Generate the exact target dates for the requested simulation.

    July 2025 -> March 2026

    Days:
        1-7
        14-20

    The output is one row per target date.
    """

    dates = pd.date_range(
        start=config["simulation_start"],
        end=config["simulation_end"],
        freq="D",
    )

    selected_dates = [
        date
        for date in dates
        if date.day
        in config["test_days_of_month"]
    ]

    return pd.DataFrame(
        {
            "target_date": selected_dates
        }
    )


# =============================================================================
# 7. AS-OF BOOKING LOGIC
# =============================================================================

def get_bookings_active_as_of(
    bookings,
    cutoff_timestamp,
):
    """
    Return bookings active at the historical forecast cutoff.

    A booking is active when:
        createdAt <= cutoff
        AND (
            cancelledAt is null
            OR cancelledAt > cutoff
        )
        AND status != "F"

    The booking dataframe is expected to have already been cleaned, so
    datetime conversion is not repeated here.
    """

    cutoff_timestamp = pd.Timestamp(
        cutoff_timestamp
    )

    active_mask = (
        bookings["createdAt"].le(
            cutoff_timestamp
        )
        &
        (
            bookings["cancelledAt"].isna()
            |
            bookings["cancelledAt"].gt(
                cutoff_timestamp
            )
        )
        &
        bookings["status"].ne("F")
    )

    return bookings.loc[
        active_mask
    ]


# =============================================================================
# 8. TARGET-DATE BOOK OF BUSINESS
# =============================================================================

def get_known_target_bookings(
    bookings,
    target_date,
    cutoff_timestamp,
    demand_type,
):
    """
    Return the bookings visible at the forecast cut-off for the target date.

    For ENTRY:
        entryDate == target_date

    For EXIT:
        exitDate == target_date

    This function is deliberately based on the AS-OF booking population,
    rather than final booking status.
    """

    active = get_bookings_active_as_of(
        bookings=bookings,
        cutoff_timestamp=cutoff_timestamp,
    )

    if demand_type == "entry":

        entry_dates = pd.to_datetime(
            active["entryDate"],
            errors="coerce",
        ).dt.normalize()

        return active.loc[
            entry_dates.eq(
                target_date.normalize()
            )
        ].copy()


    if demand_type == "exit":

        exit_dates = pd.to_datetime(
            active["exitDate"],
            errors="coerce",
        ).dt.normalize()

        return active.loc[
            exit_dates.eq(
                target_date.normalize()
            )
        ].copy()


    raise ValueError(
        f"Unknown demand_type: {demand_type}"
    )


# =============================================================================
# 9. HISTORICAL LOOK-BACK HELPERS
# =============================================================================

def previous_same_weekdays(
    dataframe,
    target_date,
    cutoff_timestamp,
    n,
):
    """
    Return the previous n observations with the same weekday as the target.

    Example:

        Target = Saturday 20 September

        n = 4

    returns the four previous Saturdays, provided those dates were already
    historical by the forecast cut-off.

    This is the foundation for several of the baseline scenarios.
    """

    df = dataframe.copy()

    df["date"] = pd.to_datetime(
        df["date"],
        errors="coerce",
    )

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    result = df[
        (df["date"] < target_date)
        &
        (df["date"] < cutoff_date)
        &
        (
            df["date"].dt.weekday
            ==
            target_date.weekday()
        )
    ].sort_values(
        "date"
    )

    return result.tail(n)


def previous_days(
    dataframe,
    cutoff_timestamp,
    n,
):
    """
    Return the previous n historical observations available at the
    forecast cut-off.
    """

    df = dataframe.copy()

    df["date"] = pd.to_datetime(
        df["date"],
        errors="coerce",
    )

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    return (
        df[
            df["date"]
            < cutoff_date
        ]
        .sort_values("date")
        .tail(n)
    )


def safe_mean(values):
    """
    Mean which safely ignores NaN / infinite values.
    """

    values = (
        pd.Series(values)
        .replace(
            [np.inf, -np.inf],
            np.nan,
        )
        .dropna()
    )

    if values.empty:
        return np.nan

    return float(
        values.mean()
    )


def safe_weighted_mean(
    values,
    weights,
):
    """
    Weighted mean which safely handles missing values.
    """

    values = pd.Series(
        values,
        dtype=float,
    )

    weights = pd.Series(
        weights,
        dtype=float,
    )

    mask = (
        values.notna()
        &
        weights.notna()
    )

    if not mask.any():
        return np.nan

    return float(
        np.average(
            values.loc[mask],
            weights=weights.loc[mask],
        )
    )

# =============================================================================
# 9A. PRE-COMPUTATION CACHE
# =============================================================================
#
# These functions build lookup tables ONCE before the simulation runs.
#
# This eliminates redundant filtering of large dataframes during the
# 54,810 forecast iterations.
# =============================================================================

def precompute_known_booking_counts(
    bookings,
    test_dates,
    horizons,
):
    """
    Pre-compute visible entry and exit booking counts using one vectorised
    booking-table calculation per horizon.

    Historical as-of rules:
        createdAt <= cutoff
        cancelledAt is null or cancelledAt > cutoff
        status != F
    """

    print("  Pre-computing known booking counts...")

    start_time = time.perf_counter()

    booking_df = bookings[
        [
            "bookingId",
            "createdAt",
            "cancelledAt",
            "status",
            "entryDate",
            "exitDate",
        ]
    ].copy()

    for column in [
        "createdAt",
        "cancelledAt",
        "entryDate",
        "exitDate",
    ]:
        booking_df[column] = pd.to_datetime(
            booking_df[column],
            errors="coerce",
        )

    booking_df = booking_df[
        booking_df["bookingId"].notna()
        &
        booking_df["createdAt"].notna()
        &
        booking_df["status"].ne("F")
    ].copy()

    booking_df["entry_target_date"] = (
        booking_df["entryDate"]
        .dt.normalize()
    )

    booking_df["exit_target_date"] = (
        booking_df["exitDate"]
        .dt.normalize()
    )

    test_date_index = pd.DatetimeIndex(
        pd.to_datetime(
            test_dates["target_date"],
            errors="coerce",
        ).dt.normalize()
    )

    result_frames = []

    for demand_type, target_column in [
        ("entry", "entry_target_date"),
        ("exit", "exit_target_date"),
    ]:

        demand_bookings = booking_df[
            booking_df[target_column].notna()
            &
            booking_df[target_column].isin(
                test_date_index
            )
        ].copy()

        for horizon_days in horizons:

            cutoff_by_booking = (
                demand_bookings[target_column]
                - pd.Timedelta(days=horizon_days)
                + pd.Timedelta(hours=7)
            )

            active_mask = (
                demand_bookings["createdAt"].le(
                    cutoff_by_booking
                )
                &
                (
                    demand_bookings["cancelledAt"].isna()
                    |
                    demand_bookings["cancelledAt"].gt(
                        cutoff_by_booking
                    )
                )
            )

            counts = (
                demand_bookings.loc[
                    active_mask,
                    [
                        target_column,
                        "bookingId",
                    ],
                ]
                .groupby(
                    target_column,
                    as_index=False,
                )
                .agg(
                    known_bookings=(
                        "bookingId",
                        "nunique",
                    )
                )
                .rename(
                    columns={
                        target_column: "target_date"
                    }
                )
            )

            complete_index = pd.DataFrame(
                {
                    "target_date": test_date_index
                }
            )

            counts = complete_index.merge(
                counts,
                on="target_date",
                how="left",
            )

            counts["known_bookings"] = (
                counts["known_bookings"]
                .fillna(0)
                .astype(int)
            )

            counts["horizon_days"] = horizon_days
            counts["demand_type"] = demand_type

            result_frames.append(counts)

    counts_df = pd.concat(
        result_frames,
        ignore_index=True,
    )

    cache = {
        (
            pd.Timestamp(row.target_date).normalize(),
            int(row.horizon_days),
            row.demand_type,
        ): int(row.known_bookings)
        for row in counts_df.itertuples(
            index=False
        )
    }

    print(
        "  Completed known booking counts "
        f"[{time.perf_counter() - start_time:.2f}s]"
    )

    return cache

def precompute_booking_curve_factors(
    entry_curve,
    exit_curve,
    test_dates,
    horizons,
):
    """
    Pre-compute booking curve completion factors for each target date
    and horizon.
    """

    print("  Pre-computing booking curve factors...")

    start_time = time.perf_counter()

    results = {}

    target_dates_list = pd.to_datetime(
        test_dates["target_date"]
    ).dt.normalize().tolist()

    for target_date in target_dates_list:

        for horizon_days in horizons:

            cutoff_timestamp = (
                target_date
                - pd.Timedelta(days=horizon_days)
                + pd.Timedelta(hours=7)
            )

            # Entry factor.
            entry_factor = get_booking_curve_factor(
                curve=entry_curve,
                target_date=target_date,
                cutoff_timestamp=cutoff_timestamp,
                cutoff_days=horizon_days,
                n=20,
            )

            # Exit factor.
            exit_factor = get_booking_curve_factor(
                curve=exit_curve,
                target_date=target_date,
                cutoff_timestamp=cutoff_timestamp,
                cutoff_days=horizon_days,
                n=20,
            )

            results[(target_date, horizon_days, "entry")] = entry_factor
            results[(target_date, horizon_days, "exit")] = exit_factor

    print(
        f"  Completed booking curve factors "
        f"[{time.perf_counter() - start_time:.2f}s]"
    )

    return results


def precompute_same_weekday_history(
    daily_actuals,
    test_dates,
    horizons,
    max_n=12,
):
    """
    Pre-compute same-weekday historical values for each target date.

    Returns a dictionary keyed by (target_date, horizon_days) containing
    the previous N same-weekday observations for entries and exits.
    """

    print("  Pre-computing same-weekday history...")

    start_time = time.perf_counter()

    actuals = daily_actuals.copy()
    actuals["date"] = pd.to_datetime(
        actuals["date"],
        errors="coerce",
    ).dt.normalize()

    results = {}

    target_dates_list = pd.to_datetime(
        test_dates["target_date"]
    ).dt.normalize().tolist()

    for target_date in target_dates_list:

        weekday = target_date.weekday()

        for horizon_days in horizons:

            cutoff_date = (
                target_date
                - pd.Timedelta(days=horizon_days)
            ).normalize()

            # Get same-weekday history before both target and cutoff.
            history = actuals[
                (actuals["date"] < target_date)
                &
                (actuals["date"] < cutoff_date)
                &
                (actuals["date"].dt.weekday == weekday)
            ].sort_values("date").tail(max_n)

            results[(target_date, horizon_days)] = {
                "entries": history["entries"].tolist(),
                "exits": history["exits"].tolist(),
                "dates": history["date"].tolist(),
            }

    print(
        f"  Completed same-weekday history "
        f"[{time.perf_counter() - start_time:.2f}s]"
    )

    return results


def precompute_weekday_month_history(
    daily_actuals,
    test_dates,
    horizons,
    max_n=6,
):
    """
    Pre-compute same-weekday-same-month historical values.

    This tests whether Saturday-in-July history is more predictive of
    Saturday-in-July demand than Saturday-in-any-month history.
    """

    print("  Pre-computing weekday-month history...")

    start_time = time.perf_counter()

    actuals = daily_actuals.copy()
    actuals["date"] = pd.to_datetime(
        actuals["date"],
        errors="coerce",
    ).dt.normalize()

    actuals["month"] = actuals["date"].dt.month

    results = {}

    target_dates_list = pd.to_datetime(
        test_dates["target_date"]
    ).dt.normalize().tolist()

    for target_date in target_dates_list:

        weekday = target_date.weekday()
        month = target_date.month

        for horizon_days in horizons:

            cutoff_date = (
                target_date
                - pd.Timedelta(days=horizon_days)
            ).normalize()

            # Same weekday AND same month.
            history = actuals[
                (actuals["date"] < target_date)
                &
                (actuals["date"] < cutoff_date)
                &
                (actuals["date"].dt.weekday == weekday)
                &
                (actuals["month"] == month)
            ].sort_values("date").tail(max_n)

            results[(target_date, horizon_days)] = {
                "entries": history["entries"].tolist(),
                "exits": history["exits"].tolist(),
                "dates": history["date"].tolist(),
                "count": len(history),
            }

    print(
        f"  Completed weekday-month history "
        f"[{time.perf_counter() - start_time:.2f}s]"
    )

    return results


# def precompute_passenger_estimates(
#     daily_passengers,
#     test_dates,
#     horizons,
#     n=6,
# ):
#     """
#     Pre-compute estimated passenger demand for each target date.
#     """

#     print("  Pre-computing passenger estimates...")

#     start_time = time.perf_counter()

#     pax = daily_passengers.copy()
#     pax["date"] = pd.to_datetime(
#         pax["date"],
#         errors="coerce",
#     ).dt.normalize()

#     results = {}

#     target_dates_list = pd.to_datetime(
#         test_dates["target_date"]
#     ).dt.normalize().tolist()

#     for target_date in target_dates_list:

#         weekday = target_date.weekday()

#         for horizon_days in horizons:

#             cutoff_date = (
#                 target_date
#                 - pd.Timedelta(days=horizon_days)
#             ).normalize()

#             history = pax[
#                 (pax["date"] < target_date)
#                 &
#                 (pax["date"] < cutoff_date)
#                 &
#                 (pax["date"].dt.weekday == weekday)
#             ].sort_values("date").tail(n)

#             if len(history) >= n:
#                 est_departing = history["departing_pax"].mean()
#                 est_arriving = history["arriving_pax"].mean()
#             else:
#                 est_departing = np.nan
#                 est_arriving = np.nan

#             results[(target_date, horizon_days)] = {
#                 "departing_pax": est_departing,
#                 "arriving_pax": est_arriving,
#             }

#     print(
#         f"  Completed passenger estimates "
#         f"[{time.perf_counter() - start_time:.2f}s]"
#     )

#     return results

def precompute_passenger_forecasts(
    daily_actuals,
    daily_passengers,
    test_dates,
    horizons,
):
    """
    Pre-compute the four passenger forecast components used by the
    E9, E10, E13, X9, X10 and X15 scenarios.

    The underlying historical filtering remains as-of aware.
    """

    print("  Pre-computing passenger forecasts...")

    start_time = time.perf_counter()

    cache = {}

    target_dates = pd.to_datetime(
        test_dates["target_date"],
        errors="coerce",
    ).dt.normalize()

    specifications = [
        (
            "entry",
            "entries",
            "departing_pax",
            "ratio_of_sums",
        ),
        (
            "entry",
            "entries",
            "departing_pax",
            "average_ratio",
        ),
        (
            "exit",
            "exits",
            "arriving_pax",
            "ratio_of_sums",
        ),
        (
            "exit",
            "exits",
            "arriving_pax",
            "average_ratio",
        ),
    ]

    for target_date in target_dates:

        for horizon_days in horizons:

            cutoff_timestamp = (
                target_date
                - pd.Timedelta(days=horizon_days)
                + pd.Timedelta(hours=7)
            )

            for (
                demand_type,
                target_col,
                passenger_col,
                method,
            ) in specifications:

                value = forecast_passenger_model(
                    daily_actuals=daily_actuals,
                    daily_passengers=daily_passengers,
                    target_date=target_date,
                    cutoff_timestamp=cutoff_timestamp,
                    target_col=target_col,
                    passenger_col=passenger_col,
                    n=6,
                    penetration_method=method,
                    data=None,
                )

                cache[
                    (
                        pd.Timestamp(
                            target_date
                        ).normalize(),
                        int(horizon_days),
                        demand_type,
                        method,
                    )
                ] = value

    print(
        "  Completed passenger forecasts "
        f"[{time.perf_counter() - start_time:.2f}s]"
    )

    return cache
# =============================================================================
# 10. HISTORICAL PASSENGER ESTIMATE
# =============================================================================

def estimate_target_passengers(
    target_date,
    cutoff_timestamp,
    daily_passengers,
    demand_type,
    n=6,
):
    """
    Estimate the target day's airport passenger demand using ONLY historical
    information.

    This is important.

    We could make the forecast look artificially good by doing:

        forecast entries =
            target day's ACTUAL departing passengers
            x historical penetration

    But that would be cheating.

    Instead:

        forecast departing passengers =
            historical same-weekday passenger average

    and similarly for arrivals.

    This gives us a genuine historical forecast input.
    """

    pax_col = (
        "departing_pax"
        if demand_type == "entry"
        else "arriving_pax"
    )

    history = previous_same_weekdays(
        dataframe=daily_passengers,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        n=n,
    )

    if len(history) < n:
        return np.nan

    return safe_mean(
        history[pax_col]
    )


# =============================================================================
# SECTION 2
# FORECASTING SCENARIOS
# =============================================================================
#
# This section contains the actual forecasting "tournament".
#
# The key principle is:
#
#     Every scenario receives only information that would have been
#     available at the historical forecast cut-off.
#
# For example:
#
#     Target date = 20 September
#     Horizon     = T-14
#     Cut-off     = 6 September
#
# The model can use:
#
#     * bookings visible by 6 September
#     * cancellations known by 6 September
#     * historical demand before 6 September
#     * historical passenger relationships
#     * historical duration / return behaviour
#
# It CANNOT use:
#
#     * bookings created after 6 September
#     * cancellations after 6 September
#     * actual passenger numbers on 20 September
#     * actual FastPark demand on 20 September
#
# =============================================================================


# =============================================================================
# 11. SCENARIO DEFINITIONS
# =============================================================================
#
# The scenario catalogue is intentionally explicit.
#
# We want the final Excel output to answer:
#
#     "Which method works best?"
#
# rather than simply:
#
#     "Which statistical calculation did we run?"
#
# -----------------------------------------------------------------------------


ENTRY_SCENARIOS = {

    "E1": {
        "name": "Last Same Weekday",
        "family": "Historical Tendency",
        "description": (
            "Uses the most recent historical occurrence of the same "
            "weekday as the target date."
        ),
        "evidence": (
            "The existing analysis explicitly tests same-weekday methods. "
            "This is the simplest operational benchmark."
        ),
    },

    "E2": {
        "name": "Average Last 2 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual entries from the previous two occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests whether a small same-weekday sample is more stable "
            "than relying on a single prior observation."
        ),
    },

    "E3": {
        "name": "Average Last 4 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual entries from the previous four occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Balances recent trading conditions against individual-day noise."
        ),
    },

    "E4": {
        "name": "Average Last 6 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual entries from the previous six occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests a longer same-weekday history while retaining weekday "
            "seasonality."
        ),
    },

    "E5": {
        "name": "Average Last 8 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual entries from the previous eight occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests whether additional historical observations improve "
            "stability at the expense of recency."
        ),
    },

    "E6": {
        "name": "Weighted Last 4 Same Weekdays",
        "family": "Weighted Historical Tendency",
        "description": (
            "Weighted average of the previous four same-weekday entry days, "
            "giving more importance to recent observations."
        ),
        "evidence": (
            "The existing analysis explicitly tests weighted same-weekday "
            "rolling penetration because recent trading conditions may "
            "be more representative."
        ),
    },

    "E7": {
        "name": "Weighted Last 6 Same Weekdays",
        "family": "Weighted Historical Tendency",
        "description": (
            "Weighted average of the previous six same-weekday entry days."
        ),
        "evidence": (
            "Tests a compromise between recent conditions and a larger "
            "historical sample."
        ),
    },

    "E8": {
        "name": "Weighted Last 8 Same Weekdays",
        "family": "Weighted Historical Tendency",
        "description": (
            "Weighted average of the previous eight same-weekday entry days."
        ),
        "evidence": (
            "Tests whether a longer weighted history is more robust than "
            "a short rolling window."
        ),
    },

    "E9": {
        "name": "Same Weekday Passenger Penetration",
        "family": "Passenger Driver",
        "description": (
            "Forecasts entries as estimated departing passengers multiplied "
            "by historical same-weekday FastPark entry penetration."
        ),
        "evidence": (
            "The existing analysis explicitly relates entries to departing "
            "passengers and calculates entry penetration."
        ),
    },

    "E10": {
        "name": "Average Passenger Penetration",
        "family": "Passenger Driver",
        "description": (
            "Forecasts entries using estimated departing passengers multiplied "
            "by the average historical entry penetration."
        ),
        "evidence": (
            "Tests whether airport passenger demand explains FastPark demand "
            "better than a pure historical entry count."
        ),
    },

    "E11": {
        "name": "Booking Visibility Curve",
        "family": "Booking Curve",
        "description": (
            "Takes the number of bookings currently visible for the target "
            "entry date and scales them using the historical booking "
            "visibility/completion curve for the current lead time."
        ),
        "evidence": (
            "The existing analysis explicitly constructs entry booking curves "
            "showing how much final entry demand is visible at each lead time."
        ),
    },

    "E12": {
        "name": "Booking Curve + Same Weekday",
        "family": "Booking Curve Hybrid",
        "description": (
            "Combines the current visible booking position with the historical "
            "same-weekday demand pattern."
        ),
        "evidence": (
            "Tests whether the live book of business plus weekday seasonality "
            "beats either signal independently."
        ),
    },

    "E13": {
        "name": "Booking + Passenger Hybrid",
        "family": "Hybrid",
        "description": (
            "Combines current visible bookings with the passenger-based "
            "forecast, allowing the model to use both realised demand "
            "visibility and the size of the airport demand pool."
        ),
        "evidence": (
            "Tests the two strongest conceptual demand signals together: "
            "current FastPark bookings and airport passenger volume."
        ),
    },

    "E14": {
        "name": "Weekday-Month Historical",
        "family": "Seasonal Historical",
        "description": (
            "Uses historical same-weekday-same-month demand rather than "
            "same-weekday-any-month. Tests whether July Saturdays predict "
            "July Saturdays better than all Saturdays."
        ),
        "evidence": (
            "Airport demand has strong monthly seasonality. Combining "
            "weekday and month may capture this better."
        ),
    },


    "E15": {
        "name": "Weekday-Month + Booking Curve",
        "family": "Seasonal Hybrid",
        "description": (
            "Combines booking curve visibility with weekday-month "
            "seasonal patterns."
        ),
        "evidence": (
            "Tests whether month-specific seasonality improves the hybrid."
        ),
    },
}


EXIT_SCENARIOS = {

    "X1": {
        "name": "Last Same Weekday",
        "family": "Historical Tendency",
        "description": (
            "Uses the most recent historical occurrence of the same weekday "
            "as the target exit date."
        ),
        "evidence": (
            "Provides the simplest operational baseline for exit demand."
        ),
    },

    "X2": {
        "name": "Average Last 2 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual exits from the previous two occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests whether two observations smooth daily noise without "
            "losing too much recency."
        ),
    },

    "X3": {
        "name": "Average Last 4 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual exits from the previous four occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests a medium-term weekday baseline."
        ),
    },

    "X4": {
        "name": "Average Last 6 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual exits from the previous six occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests whether a longer history improves stability."
        ),
    },

    "X5": {
        "name": "Average Last 8 Same Weekdays",
        "family": "Historical Tendency",
        "description": (
            "Average actual exits from the previous eight occurrences "
            "of the target weekday."
        ),
        "evidence": (
            "Tests longer-term weekday seasonality."
        ),
    },

    "X6": {
        "name": "Weighted Last 4 Same Weekdays",
        "family": "Weighted Historical Tendency",
        "description": (
            "Weighted average of the previous four same-weekday exit days."
        ),
        "evidence": (
            "Tests whether recent exit behaviour deserves greater weight."
        ),
    },

    "X7": {
        "name": "Weighted Last 6 Same Weekdays",
        "family": "Weighted Historical Tendency",
        "description": (
            "Weighted average of the previous six same-weekday exit days."
        ),
        "evidence": (
            "Tests a larger weighted history."
        ),
    },

    "X8": {
        "name": "Weighted Last 8 Same Weekdays",
        "family": "Weighted Historical Tendency",
        "description": (
            "Weighted average of the previous eight same-weekday exit days."
        ),
        "evidence": (
            "Tests whether longer weighted history is more robust."
        ),
    },

    "X9": {
        "name": "Same Weekday Passenger Penetration",
        "family": "Passenger Driver",
        "description": (
            "Forecasts exits from estimated arriving passengers multiplied "
            "by historical same-weekday exit penetration."
        ),
        "evidence": (
            "The existing analysis explicitly relates exits to arriving "
            "passenger volume."
        ),
    },

    "X10": {
        "name": "Average Passenger Penetration",
        "family": "Passenger Driver",
        "description": (
            "Forecasts exits using estimated arriving passengers multiplied "
            "by average historical exit penetration."
        ),
        "evidence": (
            "Tests the general relationship between airport arrival volume "
            "and FastPark exit demand."
        ),
    },

    "X11": {
        "name": "Exit Booking Visibility Curve",
        "family": "Booking Curve",
        "description": (
            "Uses the number of target-date exit bookings currently visible "
            "and scales them using the historical exit booking visibility curve."
        ),
        "evidence": (
            "The existing analysis explicitly creates exit booking curves "
            "and notes that exit visibility is important for roster planning."
        ),
    },

    "X12": {
        "name": "Exit Curve + Duration",
        "family": "Booking Curve + Duration",
        "description": (
            "Adjusts the visible exit booking forecast according to the "
            "planned duration mix of the currently visible bookings."
        ),
        "evidence": (
            "The existing analysis explicitly segments exit booking curves "
            "by planned duration and investigates whether short-duration "
            "bookings have different visibility."
        ),
    },

    "X13": {
        "name": "Expected Return Cohort",
        "family": "Return Behaviour",
        "description": (
            "Forecasts exits from bookings whose expected operational return "
            "date falls on the target date."
        ),
        "evidence": (
            "FastPark operational data contains ExpectedReturnDate, and the "
            "existing analysis specifically investigates advised/expected "
            "return dates."
        ),
    },

    "X14": {
        "name": "Expected Return + Historical Deviation",
        "family": "Return Behaviour",
        "description": (
            "Starts with the expected-return cohort and adjusts it using "
            "historical early/late return behaviour."
        ),
        "evidence": (
            "The existing analysis calculates actual deviation from both "
            "booking exit date and expected return date, including early "
            "and late return flags."
        ),
    },

    "X15": {
        "name": "Exit Booking + Passenger Hybrid",
        "family": "Hybrid",
        "description": (
            "Combines the visible exit book of business with the "
            "arriving-passenger-based forecast."
        ),
        "evidence": (
            "Tests whether current FastPark booking visibility plus the "
            "airport arrival demand pool is stronger than either signal alone."
        ),
    },

    "X16": {
        "name": "Entry Cohort + Duration",
        "family": "Cohort / Duration",
        "description": (
            "Forecasts future exits from cars already entering FastPark, "
            "using historical stay-duration behaviour to estimate when "
            "those entry cohorts will return."
        ),
        "evidence": (
            "The existing analysis explicitly identifies entry cohorts, "
            "planned duration, actual duration and return deviation as "
            "candidate explanations for exit demand."
        ),
    },

    "X17": {
        "name": "Weekday-Month Historical",
        "family": "Seasonal Historical",
        "description": (
            "Uses historical same-weekday-same-month exit demand."
        ),
        "evidence": (
            "Tests monthly seasonality for exit forecasting."
        ),
    },

}


# =============================================================================
# 12. GENERAL FORECAST HELPERS
# =============================================================================

def clip_forecast(value):
    """
    Forecasts cannot be negative.

    We deliberately do NOT round here.

    Keeping decimals allows the evaluation layer to distinguish between:

        23.2
        23.8

    before operational rounding is applied.
    """

    if value is None:
        return np.nan

    if pd.isna(value):
        return np.nan

    return max(
        0.0,
        float(value),
    )


def weighted_average_recent(
    values,
):
    """
    Weighted average where the most recent observation receives the largest
    weight.

    Example:

        values = [oldest, ..., newest]

    gives:

        oldest  -> weight 1
        ...
        newest  -> weight N

    This mirrors the concept of the weighted same-weekday methods tested in
    the original analysis.
    """

    values = pd.Series(
        values,
        dtype=float,
    ).replace(
        [np.inf, -np.inf],
        np.nan,
    ).dropna()

    if values.empty:
        return np.nan

    weights = np.arange(
        1,
        len(values) + 1,
        dtype=float,
    )

    return float(
        np.average(
            values.values,
            weights=weights,
        )
    )


def get_same_weekday_history(
    daily_actuals,
    target_date,
    cutoff_timestamp,
    n,
    target_col,
):
    """
    Retrieve historical actual demand for previous occurrences of the same
    weekday.

    IMPORTANT:

        The target date itself is excluded.

        Any date on/after the forecast cut-off is excluded.

    This prevents look-ahead bias.
    """

    history = previous_same_weekdays(
        dataframe=daily_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        n=n,
    )

    if history.empty:
        return pd.Series(
            dtype=float
        )

    return (
        history[target_col]
        .astype(float)
    )


def forecast_same_weekday(
    daily_actuals,
    target_date,
    cutoff_timestamp,
    target_col,
    n,
    data=None,
):
    """
    Simple average of previous same-weekday observations.

    OPTIMISATION: If pre-computed cache is available, use it.
    """

    # -------------------------------------------------------------------------
    # Try to use pre-computed cache.
    # -------------------------------------------------------------------------

    if data is not None:

        weekday_cache = data.get("weekday_cache")

        if weekday_cache is not None:

            target_date_norm = pd.Timestamp(target_date).normalize()
            cutoff_norm = pd.Timestamp(cutoff_timestamp).normalize()

            lead_days = int(
                (target_date_norm - cutoff_norm).days
            )

            cache_key = (target_date_norm, lead_days)
            cached = weekday_cache.get(cache_key)

            if cached is not None:

                values = cached.get(
                    "entries" if target_col == "entries" else "exits",
                    []
                )

                # Take the last n values.
                values = values[-n:] if len(values) >= n else []

                if len(values) >= n:
                    return clip_forecast(
                        safe_mean(values)
                    )

    # -------------------------------------------------------------------------
    # Fall back to original calculation.
    # -------------------------------------------------------------------------

    values = get_same_weekday_history(
        daily_actuals=daily_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        n=n,
        target_col=target_col,
    )

    if len(values) < n:
        return np.nan

    return clip_forecast(
        values.mean()
    )


def forecast_weighted_same_weekday(
    daily_actuals,
    target_date,
    cutoff_timestamp,
    target_col,
    n,
    data=None,
):
    """
    Weighted average of previous same-weekday observations.

    OPTIMISATION: If pre-computed cache is available, use it.
    """

    # -------------------------------------------------------------------------
    # Try to use pre-computed cache.
    # -------------------------------------------------------------------------

    if data is not None:

        weekday_cache = data.get("weekday_cache")

        if weekday_cache is not None:

            target_date_norm = pd.Timestamp(target_date).normalize()
            cutoff_norm = pd.Timestamp(cutoff_timestamp).normalize()

            lead_days = int(
                (target_date_norm - cutoff_norm).days
            )

            cache_key = (target_date_norm, lead_days)
            cached = weekday_cache.get(cache_key)

            if cached is not None:

                values = cached.get(
                    "entries" if target_col == "entries" else "exits",
                    []
                )

                # Take the last n values.
                values = values[-n:] if len(values) >= n else []

                if len(values) >= n:
                    return clip_forecast(
                        weighted_average_recent(values)
                    )

    # -------------------------------------------------------------------------
    # Fall back to original calculation.
    # -------------------------------------------------------------------------

    values = get_same_weekday_history(
        daily_actuals=daily_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        n=n,
        target_col=target_col,
    )

    if len(values) < n:
        return np.nan

    return clip_forecast(
        weighted_average_recent(
            values
        )
    )


# =============================================================================
# 13. PASSENGER PENETRATION MODELS
# =============================================================================

def calculate_historical_penetration(
    daily_actuals,
    daily_passengers,
    target_date,
    cutoff_timestamp,
    target_col,
    passenger_col,
    n,
    method="ratio_of_sums",
):
    """
    Calculate historical FastPark penetration against airport passenger demand.

    Entry:
        entries / departing passengers

    Exit:
        exits / arriving passengers

    Two versions are tested:

        ratio_of_sums:
            sum(FastPark demand) / sum(passengers)

        average_ratio:
            average of each day's individual penetration

    The distinction matters because the ratio-of-sums method gives more
    weight to high-volume days naturally.
    """

    actuals = daily_actuals.copy()
    pax = daily_passengers.copy()

    actuals["date"] = pd.to_datetime(
        actuals["date"],
        errors="coerce",
    )

    pax["date"] = pd.to_datetime(
        pax["date"],
        errors="coerce",
    )

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    weekday = target_date.weekday()

    historical = actuals[
        (actuals["date"] < target_date)
        &
        (actuals["date"] < cutoff_date)
        &
        (actuals["date"].dt.weekday == weekday)
    ].merge(
        pax[
            [
                "date",
                passenger_col,
            ]
        ],
        on="date",
        how="inner",
    )

    historical = (
        historical
        .replace(
            [np.inf, -np.inf],
            np.nan,
        )
        .dropna(
            subset=[
                target_col,
                passenger_col,
            ]
        )
    )

    historical = historical[
        historical[passenger_col] > 0
    ]

    if len(historical) < n:
        return np.nan

    historical = historical.tail(n)

    if method == "ratio_of_sums":

        denominator = (
            historical[passenger_col]
            .sum()
        )

        if denominator <= 0:
            return np.nan

        return float(
            historical[target_col].sum()
            /
            denominator
        )

    if method == "average_ratio":

        penetration = (
            historical[target_col]
            /
            historical[passenger_col]
        )

        return safe_mean(
            penetration
        )

    raise ValueError(
        f"Unknown penetration method: {method}"
    )


def forecast_passenger_model(
    daily_actuals,
    daily_passengers,
    target_date,
    cutoff_timestamp,
    target_col,
    passenger_col,
    n,
    penetration_method,
    data=None,
):
    """
    Passenger-based forecast using historical passenger estimates and
    historical FastPark penetration.

    If a pre-computed passenger forecast is available, return it directly.
    """

    target_date_norm = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    horizon_days = int(
        (
            target_date_norm
            - cutoff_date
        ).days
    )

    demand_type = (
        "entry"
        if target_col == "entries"
        else "exit"
    )

    if data is not None:

        passenger_forecast_cache = data.get(
            "passenger_forecast_cache"
        )

        if passenger_forecast_cache is not None:

            cache_key = (
                target_date_norm,
                horizon_days,
                demand_type,
                penetration_method,
            )

            cached_value = (
                passenger_forecast_cache
                .get(cache_key)
            )

            if cached_value is not None:
                return cached_value

    estimated_pax = estimate_target_passengers(
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        daily_passengers=daily_passengers,
        demand_type=demand_type,
        n=n,
    )

    if pd.isna(estimated_pax):
        return np.nan

    penetration = calculate_historical_penetration(
        daily_actuals=daily_actuals,
        daily_passengers=daily_passengers,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col=target_col,
        passenger_col=passenger_col,
        n=n,
        method=penetration_method,
    )

    if pd.isna(penetration):
        return np.nan

    return clip_forecast(
        estimated_pax
        *
        penetration
    )


# =============================================================================
# 14. BOOKING CURVE HELPERS
# =============================================================================

def build_historical_booking_curve(
    bookings,
    master,
    target_field,
    actual_field,
    cutoff_days,
    target_date_col=None,
):
    """
    Build the historical booking visibility curve efficiently.

    For each forecast horizon, this function evaluates all historical
    target dates together rather than scanning the full booking table
    separately for every target date.

    Historical as-of logic:
        createdAt <= cutoff
        AND (
            cancelledAt is null
            OR cancelledAt > cutoff
        )
        AND status != "F"

    This means a booking whose final status is CX remains included at a
    historical cutoff if its cancellation happened after that cutoff.
    """

    # Use only the booking columns needed by this calculation.
    booking_df = bookings[
        [
            "bookingId",
            "createdAt",
            "cancelledAt",
            "status",
            target_field,
        ]
    ].copy()

    booking_df["createdAt"] = pd.to_datetime(
        booking_df["createdAt"],
        errors="coerce",
    )

    booking_df["cancelledAt"] = pd.to_datetime(
        booking_df["cancelledAt"],
        errors="coerce",
    )

    booking_df["booking_target_date"] = pd.to_datetime(
        booking_df[target_field],
        errors="coerce",
    ).dt.normalize()

    booking_df = booking_df[
        booking_df["bookingId"].notna()
        &
        booking_df["createdAt"].notna()
        &
        booking_df["booking_target_date"].notna()
        &
        booking_df["status"].ne("F")
    ].copy()

    # -------------------------------------------------------------
    # Final actual operational demand by actual movement date.
    #
    # Entry curve:
    #     actual_entry_ts
    #
    # Exit curve:
    #     actual_exit_ts
    # -------------------------------------------------------------

    actual_df = master[
        [
            "bookingId",
            actual_field,
        ]
    ].copy()

    actual_df[actual_field] = pd.to_datetime(
        actual_df[actual_field],
        errors="coerce",
    )

    actual_df["target_date"] = (
        actual_df[actual_field]
        .dt.normalize()
    )

    actual_demand = (
        actual_df
        .dropna(
            subset=[
                "bookingId",
                "target_date",
            ]
        )
        .groupby(
            "target_date",
            as_index=False,
        )
        .agg(
            actual_demand=(
                "bookingId",
                "nunique",
            )
        )
    )

    if actual_demand.empty:
        return pd.DataFrame(
            columns=[
                "target_date",
                "lead_days",
                "bookings_known",
                "actual_demand",
                "completion_ratio",
            ]
        )

    historical_target_dates = set(
        actual_demand["target_date"]
    )

    # There is no reason to retain bookings for target dates with no
    # actual operational validation target.
    booking_df = booking_df[
        booking_df["booking_target_date"].isin(
            historical_target_dates
        )
    ].copy()

    rows = []

    # -------------------------------------------------------------
    # Only 15 full-table operations are now required.
    #
    # Previously, the script performed one full-table operation for
    # every target date and horizon combination.
    # -------------------------------------------------------------

    for lead_days in cutoff_days:

        cutoff_by_booking = (
            booking_df["booking_target_date"]
            - pd.Timedelta(days=lead_days)
            + pd.Timedelta(hours=7)
        )

        active_mask = (
            booking_df["createdAt"].le(
                cutoff_by_booking
            )
            &
            (
                booking_df["cancelledAt"].isna()
                |
                booking_df["cancelledAt"].gt(
                    cutoff_by_booking
                )
            )
        )

        known_counts = (
            booking_df.loc[
                active_mask,
                [
                    "booking_target_date",
                    "bookingId",
                ],
            ]
            .groupby(
                "booking_target_date",
                as_index=False,
            )
            .agg(
                bookings_known=(
                    "bookingId",
                    "nunique",
                )
            )
            .rename(
                columns={
                    "booking_target_date":
                        "target_date",
                }
            )
        )

        horizon_result = actual_demand.merge(
            known_counts,
            on="target_date",
            how="left",
        )

        horizon_result["bookings_known"] = (
            horizon_result["bookings_known"]
            .fillna(0)
            .astype(int)
        )

        horizon_result["lead_days"] = (
            lead_days
        )

        rows.append(
            horizon_result
        )

    curve = pd.concat(
        rows,
        ignore_index=True,
    )

    curve["completion_ratio"] = (
        curve["actual_demand"]
        /
        curve["bookings_known"].replace(
            0,
            np.nan,
        )
    )

    return (
        curve[
            [
                "target_date",
                "lead_days",
                "bookings_known",
                "actual_demand",
                "completion_ratio",
            ]
        ]
        .sort_values(
            [
                "target_date",
                "lead_days",
            ]
        )
        .reset_index(drop=True)
    )

def get_booking_curve_factor(
    curve,
    target_date,
    cutoff_timestamp,
    cutoff_days,
    n=20,
):
    """
    Estimate the historical booking-curve completion factor using ONLY
    information that would have been available at the forecast cut-off.

    IMPORTANT:
    ----------
    This function is deliberately AS-OF aware.

    Example:

        Target date:       20 September 2025
        Forecast horizon:  T-28
        Cut-off date:      23 August 2025

    A historical booking-curve observation for:

        10 August 2025

    is allowed because its final actual demand was already known by
    23 August.

    A historical observation for:

        10 September 2025

    is NOT allowed because, at the 23 August forecast cut-off, that
    historical outcome had not happened yet.

    This prevents future booking behaviour from leaking into the
    historical forecast.

    We also exclude the target date itself.

    Parameters
    ----------
    curve:
        Historical booking-curve dataset created by
        build_historical_booking_curve().

    target_date:
        Date we are trying to forecast.

    cutoff_timestamp:
        The date/time at which we pretend the forecast was made.

    cutoff_days:
        Forecast lead time, e.g. 28 for T-28.

    n:
        Maximum number of recent historical observations to use.

    Returns
    -------
    float
        Median historical completion factor.
    """

    if curve is None or curve.empty:
        return np.nan

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    history = curve.copy()

    history["target_date"] = pd.to_datetime(
        history["target_date"],
        errors="coerce",
    ).dt.normalize()

    # -------------------------------------------------------------------------
    # CRITICAL AS-OF FILTER
    # -------------------------------------------------------------------------
    #
    # We may ONLY learn from historical target dates whose final outcome
    # would already have been known at the forecast cut-off.
    #
    # Therefore:
    #
    #     historical target date < forecast cut-off
    #
    # NOT merely:
    #
    #     historical target date < target date
    #
    # The latter would allow future information into the backtest.
    # -------------------------------------------------------------------------

    history = history[
        history["target_date"].notna()
        &
        history["target_date"].lt(cutoff_date)
        &
        history["target_date"].lt(target_date)
    ].copy()

    # Use the same forecast horizon being tested.
    history = history[
        history["lead_days"].eq(cutoff_days)
    ].copy()

    # We cannot estimate a meaningful completion factor where there
    # were no visible bookings.
    history = history[
        history["bookings_known"].gt(0)
    ].copy()

    history = history[
        history["completion_ratio"].notna()
    ].copy()

    if history.empty:
        return np.nan

    # Prefer the most recent historical observations that were genuinely
    # available at the forecast cut-off.
    history = (
        history
        .sort_values("target_date")
        .tail(n)
    )

    # Median is deliberately used because booking curves can contain
    # extreme ratios when only one or two bookings were visible at
    # long lead times.
    return float(
        history["completion_ratio"].median()
    )


def forecast_booking_curve(
    bookings,
    curve,
    target_date,
    cutoff_timestamp,
    demand_type,
    data=None,
):
    """
    Forecast target demand using the currently visible booking population
    multiplied by the historical completion factor.

    This is the central "how much do we know already?" model.

    OPTIMISATION: If pre-computed caches are available in `data`, use them
    instead of recalculating.
    """

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_timestamp = pd.Timestamp(
        cutoff_timestamp
    )

    lead_days = int(
        np.floor(
            (
                target_date
                -
                cutoff_timestamp.normalize()
            ).total_seconds()
            /
            86400
        )
    )

    # -------------------------------------------------------------------------
    # Try to use pre-computed caches.
    # -------------------------------------------------------------------------

    if data is not None:

        # Known booking count from cache.
        known_booking_cache = data.get("known_booking_cache")

        if known_booking_cache is not None:
            cache_key = (target_date, lead_days, demand_type)
            known_count = known_booking_cache.get(cache_key)

            if known_count is not None:

                if known_count == 0:
                    return 0.0

                # Curve factor from cache.
                curve_factor_cache = data.get("curve_factor_cache")

                if curve_factor_cache is not None:
                    factor = curve_factor_cache.get(cache_key)

                    if factor is not None and not pd.isna(factor):
                        return clip_forecast(known_count * factor)

    # -------------------------------------------------------------------------
    # Fall back to original calculation if caches unavailable.
    # -------------------------------------------------------------------------

    target_field = (
        "entryDate"
        if demand_type == "entry"
        else "exitDate"
    )

    known_bookings = get_known_target_bookings(
        bookings=bookings,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type=demand_type,
    )

    known_count = (
        known_bookings["bookingId"]
        .nunique()
    )

    if known_count == 0:
        return 0.0

    factor = get_booking_curve_factor(
        curve=curve,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        cutoff_days=lead_days,
    )

    if pd.isna(factor):
        return np.nan

    return clip_forecast(
        known_count
        *
        factor
    )


# =============================================================================
# 15. DURATION-SEGMENTED BOOKING CURVE
# =============================================================================

def get_duration_adjusted_booking_curve(
    bookings,
    curve,
    target_date,
    cutoff_timestamp,
    demand_type,
    historical_duration_factors,
):
    """
    Booking curve adjusted for the duration mix currently visible.

    Why do this?

    The original analysis explicitly investigates booking visibility by
    planned duration. It is therefore reasonable to test whether:

        20 visible bookings

    means the same thing when they are:

        20 x 1-day stays

    versus:

        20 x 14-day stays.

    The duration mix can alter the relationship between bookings visible
    today and eventual operational demand.
    """

    known = get_known_target_bookings(
        bookings=bookings,
        target_date=pd.Timestamp(target_date).normalize(),
        cutoff_timestamp=cutoff_timestamp,
        demand_type=demand_type,
    )

    if known.empty:
        return 0.0

    if (
        historical_duration_factors is None
        or historical_duration_factors.empty
    ):
        return forecast_booking_curve(
            bookings=bookings,
            curve=curve,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            demand_type=demand_type,
        )

    duration_col = "planned_duration_days_calc"

    if duration_col not in known.columns:
        return forecast_booking_curve(
            bookings=bookings,
            curve=curve,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            demand_type=demand_type,
        )

    known["duration_band"] = pd.cut(
        known[duration_col],
        bins=[
            -np.inf,
            1,
            3,
            7,
            14,
            28,
            np.inf,
        ],
        labels=[
            "0-1",
            "2-3",
            "4-7",
            "8-14",
            "15-28",
            "29+",
        ],
    )

    forecasts = []

    for band, group in known.groupby(
        "duration_band",
        observed=False,
    ):

        visible = (
            group["bookingId"]
            .nunique()
        )

        factor_row = historical_duration_factors[
            historical_duration_factors[
                "duration_band"
            ].eq(band)
        ]

        if factor_row.empty:
            return forecast_booking_curve(
                bookings=bookings,
                curve=curve,
                target_date=target_date,
                cutoff_timestamp=cutoff_timestamp,
                demand_type=demand_type,
            )

        factor = float(
            factor_row[
                "completion_ratio"
            ].iloc[0]
        )

        if not np.isfinite(factor) or factor <= 0:
            return forecast_booking_curve(
                bookings=bookings,
                curve=curve,
                target_date=target_date,
                cutoff_timestamp=cutoff_timestamp,
                demand_type=demand_type,
            )

        forecasts.append(
            visible * factor
        )

    return clip_forecast(
        sum(forecasts)
    )


# =============================================================================
# 16. EXIT RETURN-BEHAVIOUR HELPERS
# =============================================================================

def create_return_behaviour_factors(
    master,
):
    """
    Estimate historical early/late return behaviour.

    The original analysis explicitly calculates:

        actual exit - booking exit date
        actual exit - expected return date

    and flags early/late returns.

    For the forecasting simulation we turn that information into a simple
    historical adjustment factor.

    The important restriction is that this function should only be supplied
    with historical records available before the forecast cut-off.
    """

    df = master.copy()

    df["actual_exit_ts"] = pd.to_datetime(
        df["actual_exit_ts"],
        errors="coerce",
    )

    df["ExpectedReturnDate"] = pd.to_datetime(
        df["ExpectedReturnDate"],
        errors="coerce",
    )

    df["exitDate"] = pd.to_datetime(
        df["exitDate"],
        errors="coerce",
    )

    valid = df.dropna(
        subset=[
            "actual_exit_ts",
            "ExpectedReturnDate",
        ]
    ).copy()

    if valid.empty:
        return {
            "expected_return_hit_rate": np.nan,
            "booking_exit_hit_rate": np.nan,
            "early_rate": np.nan,
            "late_rate": np.nan,
            "median_expected_return_deviation_hours": np.nan,
        }

    expected_deviation_hours = (
        (
            valid["actual_exit_ts"]
            -
            valid["ExpectedReturnDate"]
        )
        .dt.total_seconds()
        /
        3600
    )

    booking_deviation_hours = (
        (
            valid["actual_exit_ts"]
            -
            valid["exitDate"]
        )
        .dt.total_seconds()
        /
        3600
    )

    return {
        "expected_return_hit_rate": (
            expected_deviation_hours.abs() <= 24
        ).mean(),

        "booking_exit_hit_rate": (
            booking_deviation_hours.abs() <= 24
        ).mean(),

        "early_rate": (
            expected_deviation_hours < -24
        ).mean(),

        "late_rate": (
            expected_deviation_hours > 24
        ).mean(),

        "median_expected_return_deviation_hours": (
            expected_deviation_hours.median()
        ),
    }


def forecast_expected_return_cohort(
    bookings,
    operations,
    target_date,
    cutoff_timestamp,
):
    """
    Count bookings whose expected operational return date is the target date.

    This is an "information-rich" exit forecast because it attempts to use
    operational return information rather than simply booking exitDate.
    """

    cutoff_timestamp = pd.Timestamp(
        cutoff_timestamp
    )

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    ops = operations.copy()

    ops["ExpectedReturnDate"] = pd.to_datetime(
        ops["ExpectedReturnDate"],
        errors="coerce",
    )

    ops["RecordUpdatedDateTime"] = pd.to_datetime(
        ops["RecordUpdatedDateTime"],
        errors="coerce",
    )

    # Only use operational records that existed by the cut-off.
    #
    # If RecordUpdatedDateTime is unavailable, we retain the record because
    # the historical database may not have reliable update timestamps.
    if "RecordUpdatedDateTime" in ops.columns:
        known = ops[
            ops["RecordUpdatedDateTime"].isna()
            |
            ops["RecordUpdatedDateTime"].le(
                cutoff_timestamp
            )
        ].copy()
    else:
        known = ops.copy()

    known = known[
        known["ExpectedReturnDate"]
        .dt.normalize()
        .eq(target_date)
    ]

    return float(
        known["BookingReference"]
        .nunique()
    )


# =============================================================================
# 17. EXIT ENTRY-COHORT MODEL
# =============================================================================

def estimate_duration_probability_table(
    master,
    cutoff_timestamp,
):
    """
    Learn historical stay-duration probabilities from information available
    before the cut-off.

    Duration is measured in calendar days between actual operational entry
    and actual operational exit.

    This supports X16.

    Example:

        Cars entering 7 days before target date

        Historical probability of exiting after 7 days = 35%

    Therefore:

        known entry cohort x 35%

    contributes to the target day's exit forecast.
    """

    cutoff_timestamp = pd.Timestamp(
        cutoff_timestamp
    )

    df = master.copy()

    df["actual_entry_ts"] = pd.to_datetime(
        df["actual_entry_ts"],
        errors="coerce",
    )

    df["actual_exit_ts"] = pd.to_datetime(
        df["actual_exit_ts"],
        errors="coerce",
    )

    df = df[
        df["actual_entry_ts"].notna()
        &
        df["actual_exit_ts"].notna()
        &
        df["actual_exit_ts"].lt(
            cutoff_timestamp
        )
    ].copy()

    if df.empty:
        return pd.DataFrame()

    df["actual_duration_days"] = (
        (
            df["actual_exit_ts"]
            -
            df["actual_entry_ts"]
        )
        .dt.total_seconds()
        /
        86400
    )

    # Round to nearest day because this is a daily forecasting model.
    df["duration_days"] = (
        df["actual_duration_days"]
        .round()
        .clip(lower=0)
        .astype(int)
    )

    distribution = (
        df.groupby("duration_days")
        .size()
        .rename("observations")
        .reset_index()
    )

    distribution["probability"] = (
        distribution["observations"]
        /
        distribution["observations"].sum()
    )

    return distribution


def forecast_entry_cohort_exits(
    master,
    target_date,
    cutoff_timestamp,
):
    """
    Forecast target-day exits generated by historical/known entry cohorts.

    We only use actual entries that occurred before the forecast cut-off.

    For each historical entry cohort:

        expected exits on target date
        =
        known entries
        x
        historical probability of staying
        exactly the required number of days.

    This is particularly important for exits because exit demand is not
    necessarily generated by today's passenger volume alone.

    It is generated by customers who entered FastPark previously.
    """

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    df = master.copy()

    df["actual_entry_ts"] = pd.to_datetime(
        df["actual_entry_ts"],
        errors="coerce",
    )

    df["actual_exit_ts"] = pd.to_datetime(
        df["actual_exit_ts"],
        errors="coerce",
    )

    historical = df[
        df["actual_entry_ts"].notna()
        &
        df["actual_entry_ts"].lt(
            pd.Timestamp(cutoff_timestamp)
        )
        &
        (
            df["actual_exit_ts"].isna()
            |
            df["actual_exit_ts"].ge(
                pd.Timestamp(cutoff_timestamp)
            )
        )
    ].copy()

    if historical.empty:
        return np.nan

    duration_distribution = estimate_duration_probability_table(
        master=master,
        cutoff_timestamp=cutoff_timestamp,
    )

    if duration_distribution.empty:
        return np.nan

    probability_lookup = dict(
        zip(
            duration_distribution["duration_days"],
            duration_distribution["probability"],
        )
    )

    # -------------------------------------------------------------------------
    # Find the historical entry cohorts currently "in the system".
    #
    # We aggregate by entry date because the model is forecasting daily exits.
    # -------------------------------------------------------------------------

    cohort_counts = (
        historical
        .groupby(
            historical["actual_entry_ts"].dt.normalize()
        )
        .agg(
            entries=("bookingId", "nunique")
        )
        .reset_index()
        .rename(
            columns={
                "actual_entry_ts": "entry_date"
            }
        )
    )

    forecast = 0.0

    for _, row in cohort_counts.iterrows():

        entry_date = pd.Timestamp(
            row["entry_date"]
        ).normalize()

        duration = (
            target_date
            -
            entry_date
        ).days

        if duration < 0:
            continue

        probability = probability_lookup.get(
            duration,
            0.0,
        )

        forecast += (
            float(row["entries"])
            *
            float(probability)
        )

    return clip_forecast(
        forecast
    )


# =============================================================================
# 18. BUILD HISTORICAL DURATION FACTORS
# =============================================================================

def build_historical_duration_exit_curve(
    bookings,
    master,
    cutoff_days,
):
    """
    Build duration-specific exit booking-curve observations efficiently.

    The booking table is evaluated once per forecast horizon rather than
    separately for every historical target date and horizon.

    Visible bookings are grouped by:
        planned exit date
        planned duration band
        forecast horizon

    Final actual exits are grouped by:
        actual operational exit date
        planned duration band
    """

    duration_bins = [
        -np.inf,
        1,
        3,
        7,
        14,
        28,
        np.inf,
    ]

    duration_labels = [
        "0-1",
        "2-3",
        "4-7",
        "8-14",
        "15-28",
        "29+",
    ]

    # -------------------------------------------------------------
    # Booking population
    # -------------------------------------------------------------

    booking_df = bookings[
        [
            "bookingId",
            "createdAt",
            "cancelledAt",
            "status",
            "entryDate",
            "exitDate",
        ]
    ].copy()

    for column in [
        "createdAt",
        "cancelledAt",
        "entryDate",
        "exitDate",
    ]:
        booking_df[column] = pd.to_datetime(
            booking_df[column],
            errors="coerce",
        )

    booking_df["booking_target_date"] = (
        booking_df["exitDate"]
        .dt.normalize()
    )

    booking_df["planned_duration_days_calc"] = (
        (
            booking_df["exitDate"]
            - booking_df["entryDate"]
        )
        .dt.total_seconds()
        / 86400
    )

    booking_df["duration_band"] = pd.cut(
        booking_df[
            "planned_duration_days_calc"
        ],
        bins=duration_bins,
        labels=duration_labels,
    )

    booking_df = booking_df[
        booking_df["bookingId"].notna()
        &
        booking_df["createdAt"].notna()
        &
        booking_df["booking_target_date"].notna()
        &
        booking_df["duration_band"].notna()
        &
        booking_df["status"].ne("F")
    ].copy()

    # Converting to string avoids categorical merge/grouping issues.
    booking_df["duration_band"] = (
        booking_df["duration_band"]
        .astype(str)
    )

    # -------------------------------------------------------------
    # Final actual exits
    # -------------------------------------------------------------

    actual_df = master[
        [
            "bookingId",
            "entryDate",
            "exitDate",
            "actual_exit_ts",
        ]
    ].copy()

    for column in [
        "entryDate",
        "exitDate",
        "actual_exit_ts",
    ]:
        actual_df[column] = pd.to_datetime(
            actual_df[column],
            errors="coerce",
        )

    actual_df["target_date"] = (
        actual_df["actual_exit_ts"]
        .dt.normalize()
    )

    actual_df["planned_duration_days_calc"] = (
        (
            actual_df["exitDate"]
            - actual_df["entryDate"]
        )
        .dt.total_seconds()
        / 86400
    )

    actual_df["duration_band"] = pd.cut(
        actual_df[
            "planned_duration_days_calc"
        ],
        bins=duration_bins,
        labels=duration_labels,
    )

    actual_df = actual_df.dropna(
        subset=[
            "bookingId",
            "target_date",
            "duration_band",
        ]
    ).copy()

    actual_df["duration_band"] = (
        actual_df["duration_band"]
        .astype(str)
    )

    actuals = (
        actual_df
        .groupby(
            [
                "target_date",
                "duration_band",
            ],
            as_index=False,
        )
        .agg(
            actual_exits=(
                "bookingId",
                "nunique",
            )
        )
    )

    if actuals.empty:
        return pd.DataFrame(
            columns=[
                "target_date",
                "lead_days",
                "duration_band",
                "visible_bookings",
                "actual_exits",
                "completion_ratio",
            ]
        )

    historical_target_dates = set(
        actuals["target_date"]
    )

    booking_df = booking_df[
        booking_df["booking_target_date"].isin(
            historical_target_dates
        )
    ].copy()

    rows = []

    # -------------------------------------------------------------
    # One vectorised calculation per horizon.
    # -------------------------------------------------------------

    for lead_days in cutoff_days:

        cutoff_by_booking = (
            booking_df["booking_target_date"]
            - pd.Timedelta(days=lead_days)
            + pd.Timedelta(hours=7)
        )

        active_mask = (
            booking_df["createdAt"].le(
                cutoff_by_booking
            )
            &
            (
                booking_df["cancelledAt"].isna()
                |
                booking_df["cancelledAt"].gt(
                    cutoff_by_booking
                )
            )
        )

        visible_counts = (
            booking_df.loc[
                active_mask,
                [
                    "booking_target_date",
                    "duration_band",
                    "bookingId",
                ],
            ]
            .groupby(
                [
                    "booking_target_date",
                    "duration_band",
                ],
                as_index=False,
            )
            .agg(
                visible_bookings=(
                    "bookingId",
                    "nunique",
                )
            )
            .rename(
                columns={
                    "booking_target_date":
                        "target_date",
                }
            )
        )

        horizon_result = actuals.merge(
            visible_counts,
            on=[
                "target_date",
                "duration_band",
            ],
            how="left",
        )

        horizon_result["visible_bookings"] = (
            horizon_result["visible_bookings"]
            .fillna(0)
            .astype(int)
        )

        horizon_result["lead_days"] = (
            lead_days
        )

        horizon_result["completion_ratio"] = (
            horizon_result["actual_exits"]
            /
            horizon_result[
                "visible_bookings"
            ].replace(
                0,
                np.nan,
            )
        )

        rows.append(
            horizon_result
        )

    return (
        pd.concat(
            rows,
            ignore_index=True,
        )
        [
            [
                "target_date",
                "lead_days",
                "duration_band",
                "visible_bookings",
                "actual_exits",
                "completion_ratio",
            ]
        ]
        .sort_values(
            [
                "target_date",
                "lead_days",
                "duration_band",
            ]
        )
        .reset_index(drop=True)
    )

def get_duration_exit_factors_as_of(
    duration_curve,
    target_date,
    cutoff_timestamp,
    lead_days,
    minimum_observations=10,
    maximum_observations=20,
):
    """
    Return duration-specific exit completion factors learned only from
    historical target dates completed before the forecast cutoff.
    """

    if duration_curve is None or duration_curve.empty:
        return pd.DataFrame()

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    history = duration_curve.copy()

    history["target_date"] = pd.to_datetime(
        history["target_date"],
        errors="coerce",
    ).dt.normalize()

    history = history[
        history["target_date"].lt(cutoff_date)
        &
        history["target_date"].lt(target_date)
        &
        history["lead_days"].eq(lead_days)
        &
        history["completion_ratio"].notna()
        &
        history["visible_bookings"].gt(0)
    ].copy()

    rows = []

    for duration_band, group in history.groupby(
        "duration_band",
        observed=False,
    ):

        group = (
            group
            .sort_values("target_date")
            .tail(maximum_observations)
        )

        if len(group) < minimum_observations:
            continue

        rows.append(
            {
                "duration_band": duration_band,
                "completion_ratio": float(
                    group["completion_ratio"].median()
                ),
                "observations": len(group),
            }
        )

    return pd.DataFrame(rows)


# =============================================================================
# 19. SCENARIO E1-E15
# =============================================================================

def forecast_entry_scenario(
    scenario_id,
    target_date,
    cutoff_timestamp,
    data,
):
    """
    Execute exactly one entry scenario.

    Returns:
        forecast_value
        supporting diagnostics
    """

    daily_actuals = data["daily_actuals"]
    daily_passengers = data["daily_passengers"]
    bookings = data["bookings"]
    components = data.get(
        "current_snapshot_components"
        )

    # -------------------------------------------------------------------------
    # E1 - Last same weekday
    # -------------------------------------------------------------------------

    if scenario_id == "E1":

        value = components["entry_weekday"][1]

        return value, {
            "method_detail": "previous_same_weekday",
            "history_n": 1,
        }


    # -------------------------------------------------------------------------
    # E2 - Average last 2 same weekdays
    # -------------------------------------------------------------------------

    if scenario_id == "E2":

        value = components["entry_weekday"][2]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 2,
        }


    # -------------------------------------------------------------------------
    # E3 - Average last 4
    # -------------------------------------------------------------------------

    if scenario_id == "E3":

        value = components["entry_weekday"][4]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 4,
        }


    # -------------------------------------------------------------------------
    # E4 - Average last 6
    # -------------------------------------------------------------------------

    if scenario_id == "E4":

        value = components["entry_weekday"][6]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # E5 - Average last 8
    # -------------------------------------------------------------------------

    if scenario_id == "E5":

        value = components["entry_weekday"][8]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 8,
        }


    # -------------------------------------------------------------------------
    # E6 - Weighted last 4
    # -------------------------------------------------------------------------

    if scenario_id == "E6":

        value = components["entry_weighted"][4]

        return value, {
            "method_detail": "weighted_same_weekday",
            "history_n": 4,
        }


    # -------------------------------------------------------------------------
    # E7 - Weighted last 6
    # -------------------------------------------------------------------------

    if scenario_id == "E7":

        value = components["entry_weighted"][6]

        return value, {
            "method_detail": "weighted_same_weekday",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # E8 - Weighted last 8
    # -------------------------------------------------------------------------

    if scenario_id == "E8":

        value = components["entry_weighted"][8]

        return value, {
            "method_detail": "weighted_same_weekday",
            "history_n": 8,
        }


    # -------------------------------------------------------------------------
    # E9 - Same weekday passenger penetration
    # -------------------------------------------------------------------------

    if scenario_id == "E9":

        value = components["entry_passenger_ratio"]

        return value, {
            "method_detail": "departing_pax_x_same_weekday_penetration",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # E10 - Average passenger penetration
    # -------------------------------------------------------------------------

    if scenario_id == "E10":

        value = components["entry_passenger_average"]

        return value, {
            "method_detail": "departing_pax_x_average_penetration",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # E11 - Booking visibility curve
    # -------------------------------------------------------------------------

    if scenario_id == "E11":

        return components["entry_booking"], {
            "method_detail":
                "visible_bookings_x_historical_completion",

            "known_bookings":
                components["entry_known_bookings"],
        }


    # -------------------------------------------------------------------------
    # E12 - Booking curve + same weekday
    # -------------------------------------------------------------------------

    if scenario_id == "E12":

        booking_forecast = components[
            "entry_booking"
        ]

        weekday_forecast = components[
            "entry_weekday"
        ][4]

        values = [
            value
            for value in [
                booking_forecast,
                weekday_forecast,
            ]
            if pd.notna(value)
        ]

        if not values:
            return np.nan, {
                "method_detail": "no_valid_components"
            }

        return clip_forecast(
            np.mean(values)
        ), {
            "method_detail":
                "50pct_booking_curve_50pct_same_weekday",

            "booking_forecast":
                booking_forecast,

            "weekday_forecast":
                weekday_forecast,
        }


    # -------------------------------------------------------------------------
    # E13 - Booking + passenger hybrid
    # -------------------------------------------------------------------------

    if scenario_id == "E13":

        booking_forecast = components[
            "entry_booking"
        ]

        passenger_forecast = components[
            "entry_passenger_ratio"
        ]

        values = [
            value
            for value in [
                booking_forecast,
                passenger_forecast,
            ]
            if pd.notna(value)
        ]

        if not values:
            return np.nan, {
                "method_detail": "no_valid_components"
            }

        return clip_forecast(
            np.mean(values)
        ), {
            "method_detail":
                "booking_curve_plus_passenger",

            "booking_forecast":
                booking_forecast,

            "passenger_forecast":
                passenger_forecast,
        }


    # -------------------------------------------------------------------------
    # E14 - Weekday-Month Historical
    # -------------------------------------------------------------------------

    if scenario_id == "E14":

        return clip_forecast(
            components["entry_weekday_month"]
        ), {
            "method_detail":
                "weekday_month_average",

            "history_n":
                components[
                    "weekday_month_observations"
                ],

            "used_general_weekday_fallback": (
                components[
                    "weekday_month_observations"
                ]
                == 0
            ),
        }


    # -------------------------------------------------------------------------
    # E15 - Weekday-Month + Booking Curve
    # -------------------------------------------------------------------------

    if scenario_id == "E15":

        booking_forecast = components[
            "entry_booking"
        ]

        weekday_month_forecast = components[
            "entry_weekday_month"
        ]

        values = [
            value
            for value in [
                booking_forecast,
                weekday_month_forecast,
            ]
            if pd.notna(value)
        ]

        if not values:
            return np.nan, {
                "method_detail": "no_valid_components"
            }

        return clip_forecast(
            np.mean(values)
        ), {
            "method_detail":
                "booking_plus_weekday_month",

            "booking_forecast":
                booking_forecast,

            "weekday_month_forecast":
                weekday_month_forecast,

            "weekday_month_observations":
                components[
                    "weekday_month_observations"
                ],
        }

    raise ValueError(
        f"Unknown entry scenario: {scenario_id}"
    )


# =============================================================================
# 20. SCENARIO X1-X17
# =============================================================================

def forecast_exit_scenario(
    scenario_id,
    target_date,
    cutoff_timestamp,
    data,
):
    """
    Execute exactly one exit scenario.

    Returns:
        forecast_value
        supporting diagnostics
    """

    daily_actuals = data["daily_actuals"]
    daily_passengers = data["daily_passengers"]
    bookings = data["bookings"]
    operations = data["operations"]
    master = data["master"]
    components = data.get(
        "current_snapshot_components"
    )


    # -------------------------------------------------------------------------
    # X1 - Last same weekday
    # -------------------------------------------------------------------------

    if scenario_id == "X1":

        value = components["exit_weekday"][1]

        return value, {
            "method_detail": "previous_same_weekday",
            "history_n": 1,
        }


    # -------------------------------------------------------------------------
    # X2 - Average last 2
    # -------------------------------------------------------------------------

    if scenario_id == "X2":

        value = components["exit_weekday"][2]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 2,
        }


    # -------------------------------------------------------------------------
    # X3 - Average last 4
    # -------------------------------------------------------------------------

    if scenario_id == "X3":

        value = components["exit_weekday"][4]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 4,
        }


    # -------------------------------------------------------------------------
    # X4 - Average last 6
    # -------------------------------------------------------------------------

    if scenario_id == "X4":

        value = components["exit_weekday"][6]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # X5 - Average last 8
    # -------------------------------------------------------------------------

    if scenario_id == "X5":

        value = components["exit_weekday"][8]

        return value, {
            "method_detail": "same_weekday_average",
            "history_n": 8,
        }


    # -------------------------------------------------------------------------
    # X6 - Weighted last 4
    # -------------------------------------------------------------------------

    if scenario_id == "X6":

        value = components["exit_weighted"][4]

        return value, {
            "method_detail": "weighted_same_weekday",
            "history_n": 4,
        }


    # -------------------------------------------------------------------------
    # X7 - Weighted last 6
    # -------------------------------------------------------------------------

    if scenario_id == "X7":

        value = components["exit_weighted"][6]

        return value, {
            "method_detail": "weighted_same_weekday",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # X8 - Weighted last 8
    # -------------------------------------------------------------------------

    if scenario_id == "X8":

        value = components["exit_weighted"][8]

        return value, {
            "method_detail": "weighted_same_weekday",
            "history_n": 8,
        }


    # -------------------------------------------------------------------------
    # X9 - Same weekday passenger penetration
    # -------------------------------------------------------------------------

    if scenario_id == "X9":

        value = components["exit_passenger_ratio"]

        return value, {
            "method_detail": "arriving_pax_x_same_weekday_penetration",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # X10 - Average passenger penetration
    # -------------------------------------------------------------------------

    if scenario_id == "X10":

        value = components["exit_passenger_average"]

        return value, {
            "method_detail": "arriving_pax_x_average_penetration",
            "history_n": 6,
        }


    # -------------------------------------------------------------------------
    # X11 - Exit booking visibility curve
    # -------------------------------------------------------------------------

    if scenario_id == "X11":

        return components["exit_booking"], {
            "method_detail":
                "visible_exit_bookings_x_completion",

            "known_bookings":
                components["exit_known_bookings"],
        }

    # -------------------------------------------------------------------------
    # X12 - Exit curve + duration
    # -------------------------------------------------------------------------

    if scenario_id == "X12":

        lead_days = (
                pd.Timestamp(target_date).normalize()
            - pd.Timestamp(cutoff_timestamp).normalize()
        ).days

        historical_duration_factors = (
            get_duration_exit_factors_as_of(
                duration_curve=data[
                    "duration_exit_booking_curve"
                ],
                target_date=target_date,
                cutoff_timestamp=cutoff_timestamp,
                lead_days=lead_days,
                minimum_observations=10,
                maximum_observations=20,
            )
        )   

        value = get_duration_adjusted_booking_curve(
            bookings=bookings,
            curve=data["exit_booking_curve"],
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            demand_type="exit",
            historical_duration_factors=historical_duration_factors
        )

        return value, {
            "method_detail": "duration_adjusted_exit_booking_curve",
            "duration_factors_as_of": pd.Timestamp(cutoff_timestamp).normalize(),
        }


    # -------------------------------------------------------------------------
    # X13 - Expected return cohort
    # -------------------------------------------------------------------------

    if scenario_id == "X13":

        value = forecast_expected_return_cohort(
            bookings=bookings,
            operations=operations,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
        )

        return value, {
            "method_detail": "expected_return_date_cohort",
        }


    # -------------------------------------------------------------------------
    # X14 - Expected return + historical deviation
    # -------------------------------------------------------------------------

    if scenario_id == "X14":

        base = forecast_expected_return_cohort(
            bookings=bookings,
            operations=operations,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
        )

        if pd.isna(base):
            return np.nan, {
                "method_detail": "no_expected_return_cohort",
            }

        historical_master = master.copy()

        cutoff_date = pd.Timestamp(
            cutoff_timestamp
        ).normalize()

        historical_master["actual_exit_ts"] = pd.to_datetime(
            historical_master["actual_exit_ts"],
            errors="coerce",
        )

        historical_master = historical_master[
            historical_master["actual_exit_ts"].notna()
            &
            historical_master["actual_exit_ts"].dt.normalize().lt(
                cutoff_date
            )
        ]

        factors = create_return_behaviour_factors(
            historical_master
        )

        deviation = factors[
            "median_expected_return_deviation_hours"
        ]

        if pd.isna(deviation):
            return base, {
                "method_detail": "expected_return_no_adjustment",
            }

        # Positive median deviation means customers historically returned
        # later than the expected return time.
        #
        # For a DAILY forecast, we do not shift the entire cohort by a
        # fractional day. Instead we use historical return-date accuracy
        # to adjust the cohort conservatively.
        #
        # The adjustment is deliberately modest because this scenario is
        # testing whether the signal exists, not attempting to overfit it.
        if deviation > 24:
            adjustment = 1.10
        elif deviation < -24:
            adjustment = 0.90
        else:
            adjustment = 1.00

        value = (
            base
            *
            adjustment
        )

        return clip_forecast(value), {
            "method_detail": "expected_return_plus_historical_deviation",
            "median_deviation_hours": deviation,
            "adjustment_factor": adjustment,
        }


    # -------------------------------------------------------------------------
    # X15 - Exit booking + passenger hybrid
    # -------------------------------------------------------------------------

    if scenario_id == "X15":

        booking_forecast = components[
            "exit_booking"
        ]

        passenger_forecast = components[
            "exit_passenger_ratio"
        ]

        values = [
            value
            for value in [
                booking_forecast,
                passenger_forecast,
            ]
            if pd.notna(value)
        ]

        if not values:
            return np.nan, {
                "method_detail": "no_valid_components"
            }

        return clip_forecast(
            np.mean(values)
        ), {
            "method_detail":
                "exit_booking_plus_arriving_pax",

            "booking_forecast":
                booking_forecast,

            "passenger_forecast":
                passenger_forecast,
        }


    # -------------------------------------------------------------------------
    # X16 - Entry cohort + duration
    # -------------------------------------------------------------------------

    if scenario_id == "X16":

        value = forecast_entry_cohort_exits(
            master=master,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
        )

        return value, {
            "method_detail": "known_entry_cohort_x_duration_probability",
        }

    # -------------------------------------------------------------------------
    # X17 - Weekday-Month Historical
    # -------------------------------------------------------------------------

    if scenario_id == "X17":

        return clip_forecast(
            components["exit_weekday_month"]
        ), {
            "method_detail":
                "weekday_month_average",

            "history_n":
                components[
                    "weekday_month_observations"
                ],

            "used_general_weekday_fallback": (
                components[
                    "weekday_month_observations"
                ]
                == 0
            ),
        }



# =============================================================================
# 21. SCENARIO VALIDATION
# =============================================================================

def validate_scenario_catalogue():
    """
    Fail early if the scenario catalogue differs from the approved
    forecasting tournament.

    Approved entry scenarios:
        E1-E15

    Approved exit scenarios:
        X1-X17
    """

    expected_entries = [
        f"E{i}"
        for i in range(1, 16)
    ]

    expected_exits = [
        f"X{i}"
        for i in range(1, 18)
    ]

    actual_entries = list(
        ENTRY_SCENARIOS.keys()
    )

    actual_exits = list(
        EXIT_SCENARIOS.keys()
    )

    if actual_entries != expected_entries:
        raise RuntimeError(
            "Entry scenario catalogue is not E1-E15.\n"
            f"Expected: {expected_entries}\n"
            f"Found:    {actual_entries}"
        )

    if actual_exits != expected_exits:
        raise RuntimeError(
            "Exit scenario catalogue is not X1-X17.\n"
            f"Expected: {expected_exits}\n"
            f"Found:    {actual_exits}"
        )

    print(
        "Scenario catalogue validated: "
        f"{len(actual_entries)} entry models + "
        f"{len(actual_exits)} exit models."
    )


# =============================================================================
# 22. PREPARE MODEL DATA
# =============================================================================

def prepare_forecast_model_data(
    data,
    config,
):
    """
    Prepare the historical datasets needed by the scenario engine.

    This function creates:

        daily actuals
        hourly actuals
        daily passenger demand
        entry booking curve
        exit booking curve
        duration completion factors
        PRE-COMPUTED CACHES FOR PERFORMANCE

    The resulting dictionary is passed into every scenario.

    Keeping this preparation separate from scenario execution makes the
    backtest much easier to debug.
    """

    master = data["master"]
    bookings = data["bookings"]
    flights = data["flights"]

    print("\nPreparing forecasting model datasets...")

    preparation_start = time.perf_counter()

    # -------------------------------------------------------------
    # Actual demand
    # -------------------------------------------------------------

    stage_start = time.perf_counter()

    daily_actuals = create_daily_actuals(
        master
    )

    hourly_actuals = create_hourly_actuals(
        master
    )

    print(
        "  Created daily and hourly actuals "
        f"[{time.perf_counter() - stage_start:.2f}s]"
    )

    # -------------------------------------------------------------
    # Passenger context
    # -------------------------------------------------------------

    stage_start = time.perf_counter()

    daily_passengers = (
        create_daily_passenger_context(
            flights
        )
    )

    print(
        "  Created daily passenger context "
        f"[{time.perf_counter() - stage_start:.2f}s]"
    )

    curve_horizons = config[
        "forecast_horizons_days"
    ]

    # -------------------------------------------------------------
    # Entry booking curve
    # -------------------------------------------------------------

    stage_start = time.perf_counter()

    print(
        "  Building entry booking visibility curve..."
    )

    entry_booking_curve = (
        build_historical_booking_curve(
            bookings=bookings,
            master=master,
            target_field="entryDate",
            actual_field="actual_entry_ts",
            cutoff_days=curve_horizons,
            target_date_col="actual_entry_date",
        )
    )

    print(
        "  Completed entry booking visibility curve "
        f"[{time.perf_counter() - stage_start:.2f}s]"
    )

    # -------------------------------------------------------------
    # Exit booking curve
    # -------------------------------------------------------------

    stage_start = time.perf_counter()

    print(
        "  Building exit booking visibility curve..."
    )

    exit_booking_curve = (
        build_historical_booking_curve(
            bookings=bookings,
            master=master,
            target_field="exitDate",
            actual_field="actual_exit_ts",
            cutoff_days=curve_horizons,
            target_date_col="actual_exit_date",
        )
    )

    print(
        "  Completed exit booking visibility curve "
        f"[{time.perf_counter() - stage_start:.2f}s]"
    )

    # -------------------------------------------------------------
    # Duration-segmented exit curve
    # -------------------------------------------------------------

    stage_start = time.perf_counter()

    print(
        "  Building duration-segmented exit booking curve..."
    )

    duration_exit_booking_curve = (
        build_historical_duration_exit_curve(
            bookings=bookings,
            master=master,
            cutoff_days=curve_horizons,
        )
    )

    print(
        "  Completed duration-segmented exit booking curve "
        f"[{time.perf_counter() - stage_start:.2f}s]"
    )

    # =================================================================
    # PRE-COMPUTATION CACHES
    # =================================================================
    #
    # These caches eliminate redundant calculations during the main
    # simulation loop, reducing runtime from ~7.6 hours to ~30-60 minutes.
    # =================================================================

    print("\n  Building pre-computation caches...")

    # Get test dates for cache building.
    test_dates = create_test_dates(
        config
    )

    cache_horizons = curve_horizons

    if config.get(
        "smoke_test",
        False,
    ):

        test_dates = test_dates.head(
            config.get(
                "smoke_test_target_count",
                2,
            )
        )

        cache_horizons = config.get(
            "smoke_test_horizons",
            curve_horizons,
        )

    # -------------------------------------------------------------
    # Known booking counts cache
    # -------------------------------------------------------------
    known_booking_cache = (
        precompute_known_booking_counts(
            bookings=bookings,
            test_dates=test_dates,
            horizons=cache_horizons,
        )
    )

    # -------------------------------------------------------------
    # Booking curve factors cache
    # -------------------------------------------------------------

    curve_factor_cache = precompute_booking_curve_factors(
        entry_curve=entry_booking_curve,
        exit_curve=exit_booking_curve,
        test_dates=test_dates,
        horizons=cache_horizons,
    )

    # -------------------------------------------------------------
    # Same-weekday history cache
    # -------------------------------------------------------------

    weekday_cache = precompute_same_weekday_history(
        daily_actuals=daily_actuals,
        test_dates=test_dates,
        horizons=cache_horizons,
        max_n=12,
    )

    # -------------------------------------------------------------
    # Weekday-month history cache
    # -------------------------------------------------------------

    weekday_month_cache = precompute_weekday_month_history(
        daily_actuals=daily_actuals,
        test_dates=test_dates,
        horizons=cache_horizons,
        max_n=6,
    )

    # -------------------------------------------------------------
    # Passenger estimates cache
    # -------------------------------------------------------------

    passenger_forecast_cache = (
        precompute_passenger_forecasts(
            daily_actuals=daily_actuals,
            daily_passengers=daily_passengers,
            test_dates=test_dates,
            horizons=cache_horizons,
        )
    )

    print(
        "  Forecast model preparation complete "
        f"[{time.perf_counter() - preparation_start:.2f}s]"
    )

    return {
        **data,
        "config": config,

        "daily_actuals": daily_actuals,
        "hourly_actuals": hourly_actuals,
        "daily_passengers": daily_passengers,

        "entry_booking_curve": entry_booking_curve,
        "exit_booking_curve": exit_booking_curve,
        "duration_exit_booking_curve": duration_exit_booking_curve,

        # Pre-computation caches.
        "known_booking_cache": known_booking_cache,
        "curve_factor_cache": curve_factor_cache,
        "weekday_cache": weekday_cache,
        "weekday_month_cache": weekday_month_cache,
        "passenger_forecast_cache": passenger_forecast_cache,
    }


# =============================================================================
# 23. CREATE ONE FORECAST SNAPSHOT
# =============================================================================

def build_snapshot_components(
    target_date,
    horizon_days,
    data,
):
    """
    Calculate reusable forecast components once for one target-date/horizon
    snapshot.

    Scenario functions then reuse these components instead of repeatedly
    recalculating the same forecasts.
    """
    snapshot_start = time.perf_counter()

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_timestamp = (
        target_date
        - pd.Timedelta(days=horizon_days)
        + pd.Timedelta(hours=7)
    )

    daily_actuals = data["daily_actuals"]
    daily_passengers = data["daily_passengers"]
    bookings = data["bookings"]

    known_booking_cache = data[
        "known_booking_cache"
    ]

    entry_booking = forecast_booking_curve(
        bookings=bookings,
        curve=data["entry_booking_curve"],
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="entry",
        data=data,
    )

    exit_booking = forecast_booking_curve(
        bookings=bookings,
        curve=data["exit_booking_curve"],
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="exit",
        data=data,
    )

    entry_weekday = {
        n: forecast_same_weekday(
            daily_actuals=daily_actuals,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            target_col="entries",
            n=n,
            data=data,
        )
        for n in [
            1,
            2,
            4,
            6,
            8,
        ]
    }

    exit_weekday = {
        n: forecast_same_weekday(
            daily_actuals=daily_actuals,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            target_col="exits",
            n=n,
            data=data,
        )
        for n in [
            1,
            2,
            4,
            6,
            8,
        ]
    }

    entry_weighted = {
        n: forecast_weighted_same_weekday(
            daily_actuals=daily_actuals,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            target_col="entries",
            n=n,
            data=data,
        )
        for n in [
            4,
            6,
            8,
        ]
    }

    exit_weighted = {
        n: forecast_weighted_same_weekday(
            daily_actuals=daily_actuals,
            target_date=target_date,
            cutoff_timestamp=cutoff_timestamp,
            target_col="exits",
            n=n,
            data=data,
        )
        for n in [
            4,
            6,
            8,
        ]
    }

    entry_passenger_ratio = forecast_passenger_model(
        daily_actuals=daily_actuals,
        daily_passengers=daily_passengers,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col="entries",
        passenger_col="departing_pax",
        n=6,
        penetration_method="ratio_of_sums",
        data=data,
    )

    entry_passenger_average = forecast_passenger_model(
        daily_actuals=daily_actuals,
        daily_passengers=daily_passengers,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col="entries",
        passenger_col="departing_pax",
        n=6,
        penetration_method="average_ratio",
        data=data,
    )

    exit_passenger_ratio = forecast_passenger_model(
        daily_actuals=daily_actuals,
        daily_passengers=daily_passengers,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col="exits",
        passenger_col="arriving_pax",
        n=6,
        penetration_method="ratio_of_sums",
        data=data,
    )

    exit_passenger_average = forecast_passenger_model(
        daily_actuals=daily_actuals,
        daily_passengers=daily_passengers,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        target_col="exits",
        passenger_col="arriving_pax",
        n=6,
        penetration_method="average_ratio",
        data=data,
    )

    weekday_month = data.get(
        "weekday_month_cache",
        {},
    ).get(
        (
            target_date,
            horizon_days,
        )
    )

    if (
        weekday_month is not None
        and weekday_month.get("count", 0)
        >= data["config"][
            "minimum_weekday_month_observations"
        ]
    ):
        entry_weekday_month = safe_mean(
            weekday_month["entries"]
        )
        exit_weekday_month = safe_mean(
            weekday_month["exits"]
        )
        weekday_month_observations = (
            weekday_month["count"]
        )
    else:
        entry_weekday_month = (
            entry_weekday[4]
        )
        exit_weekday_month = (
            exit_weekday[4]
        )
        weekday_month_observations = 0

    if horizon_days == 56:
        print(
            f"\nSnapshot components built in "
            f"{time.perf_counter()-snapshot_start:.2f}s"
        )

    return {
        "target_date": target_date,
        "cutoff_timestamp": cutoff_timestamp,
        "horizon_days": horizon_days,

        "entry_known_bookings":
            known_booking_cache.get(
                (
                    target_date,
                    horizon_days,
                    "entry",
                ),
                0,
            ),

        "exit_known_bookings":
            known_booking_cache.get(
                (
                    target_date,
                    horizon_days,
                    "exit",
                ),
                0,
            ),

        "entry_booking": entry_booking,
        "exit_booking": exit_booking,

        "entry_weekday": entry_weekday,
        "exit_weekday": exit_weekday,

        "entry_weighted": entry_weighted,
        "exit_weighted": exit_weighted,

        "entry_passenger_ratio":
            entry_passenger_ratio,

        "entry_passenger_average":
            entry_passenger_average,

        "exit_passenger_ratio":
            exit_passenger_ratio,

        "exit_passenger_average":
            exit_passenger_average,

        "entry_weekday_month":
            entry_weekday_month,

        "exit_weekday_month":
            exit_weekday_month,

        "weekday_month_observations":
            weekday_month_observations,
    }

def run_daily_forecast_snapshot(
    target_date,
    horizon_days,
    data,
):
    """
    Run all entry and exit scenarios for one target date and horizon.

    Common forecast components are calculated once and reused by all
    scenarios in this snapshot.
    """

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_timestamp = (
        target_date
        - pd.Timedelta(days=horizon_days)
        + pd.Timedelta(hours=7)
    )

    components = build_snapshot_components(
        target_date=target_date,
        horizon_days=horizon_days,
        data=data,
    )

    data["current_snapshot_components"] = (
        components
    )

    fail_fast = data["config"].get(
        "fail_fast",
        True,
    )

    rows = []

    for scenario_id in ENTRY_SCENARIOS:

        try:
            forecast, diagnostics = (
                forecast_entry_scenario(
                    scenario_id=scenario_id,
                    target_date=target_date,
                    cutoff_timestamp=cutoff_timestamp,
                    data=data,
                )
            )

        except Exception as exc:

            if fail_fast:
                raise RuntimeError(
                    "Entry scenario failed: "
                    f"{scenario_id}, "
                    f"target={target_date.date()}, "
                    f"horizon=T-{horizon_days}"
                ) from exc

            forecast = np.nan

            diagnostics = {
                "method_detail": "ERROR",
                "error_message": str(exc),
            }

        rows.append(
            {
                "target_date": target_date,
                "cutoff_timestamp":
                    cutoff_timestamp,

                "horizon_days":
                    horizon_days,

                "demand_type":
                    "entry",

                "scenario_id":
                    scenario_id,

                "scenario_name":
                    ENTRY_SCENARIOS[
                        scenario_id
                    ]["name"],

                "scenario_family":
                    ENTRY_SCENARIOS[
                        scenario_id
                    ]["family"],

                "scenario_description":
                    ENTRY_SCENARIOS[
                        scenario_id
                    ]["description"],

                "forecast_value":
                    forecast,

                **diagnostics,
            }
        )

    for scenario_id in EXIT_SCENARIOS:

        try:
            forecast, diagnostics = (
                forecast_exit_scenario(
                    scenario_id=scenario_id,
                    target_date=target_date,
                    cutoff_timestamp=cutoff_timestamp,
                    data=data,
                )
            )

        except Exception as exc:

            if fail_fast:
                raise RuntimeError(
                    "Exit scenario failed: "
                    f"{scenario_id}, "
                    f"target={target_date.date()}, "
                    f"horizon=T-{horizon_days}"
                ) from exc

            forecast = np.nan

            diagnostics = {
                "method_detail": "ERROR",
                "error_message": str(exc),
            }

        rows.append(
            {
                "target_date": target_date,
                "cutoff_timestamp":
                    cutoff_timestamp,

                "horizon_days":
                    horizon_days,

                "demand_type":
                    "exit",

                "scenario_id":
                    scenario_id,

                "scenario_name":
                    EXIT_SCENARIOS[
                        scenario_id
                    ]["name"],

                "scenario_family":
                    EXIT_SCENARIOS[
                        scenario_id
                    ]["family"],

                "scenario_description":
                    EXIT_SCENARIOS[
                        scenario_id
                    ]["description"],

                "forecast_value":
                    forecast,

                **diagnostics,
            }
        )

    data.pop(
        "current_snapshot_components",
        None,
    )

    return pd.DataFrame(rows)


# =============================================================================
# 24. BUILD THE COMPLETE DAILY SIMULATION
# =============================================================================

def run_daily_simulation(
    data,
    config,
):
    """
    Run the full historical daily forecasting tournament.

    Requested target dates:

        1-7
        14-20

    July 2025 -> March 2026

    Forecast horizons:

        T-0
        T-1
        T-2
        T-3
        T-4
        T-5
        T-6
        T-7
        T-14
        T-21
        T-28
        T-35
        T-42
        T-49
        T-56

    With:

        126 target dates
        x 15 horizons
        x 13 entry scenarios
        x 16 exit scenarios

    the model tournament generates:

        126 x 15 x 32
        =
        60,480

    daily forecast observations.

    That is exactly what we want.

    We are not trying to decide the best model by intuition.

    We are allowing historical performance to decide it.
    """

    test_dates = create_test_dates(
        config
    )

    horizons = config[
        "forecast_horizons_days"
    ]

    if config.get(
        "smoke_test",
        False,
    ):

        test_dates = test_dates.head(
            config.get(
                "smoke_test_target_count",
                2,
            )
        )

        horizons = config.get(
            "smoke_test_horizons",
            horizons,
        )

    all_results = []

    total = (
        len(test_dates)
        *
        len(horizons)
    )

    counter = 0

    print("\n" + "=" * 80)
    print("RUNNING DAILY FORECAST SIMULATION")
    print("=" * 80)

    print(
        f"\nTarget dates: {len(test_dates)}"
    )

    print(
        f"Forecast horizons: {len(horizons)}"
    )

    print(
        f"Entry scenarios: {len(ENTRY_SCENARIOS)}"
    )

    print(
        f"Exit scenarios: {len(EXIT_SCENARIOS)}"
    )

    print(
        f"Total daily scenario forecasts: "
        f"{total * (len(ENTRY_SCENARIOS) + len(EXIT_SCENARIOS)):,}"
    )


    for _, row in test_dates.iterrows():

        target_date = pd.Timestamp(
            row["target_date"]
        ).normalize()

        for horizon_days in horizons:

            counter += 1

            print(
                f"\rRunning "
                f"{counter:,}/{total:,} "
                f"| Target {target_date.date()} "
                f"| T-{horizon_days}",
                end="",
            )

            result = run_daily_forecast_snapshot(
                target_date=target_date,
                horizon_days=horizon_days,
                data=data,
            )

            all_results.append(
                result
            )


    print("\n")

    if not all_results:
        return pd.DataFrame()

    forecasts = pd.concat(
        all_results,
        ignore_index=True,
    )


    # -------------------------------------------------------------------------
    # Attach actual validation target
    # -------------------------------------------------------------------------

    actuals = data[
        "daily_actuals"
    ].copy()

    actuals["date"] = pd.to_datetime(
        actuals["date"]
    )

    actuals = actuals.rename(
        columns={
            "date": "target_date",
        }
    )

    forecasts = forecasts.merge(
        actuals[
            [
                "target_date",
                "entries",
                "exits",
            ]
        ],
        on="target_date",
        how="left",
    )


    forecasts["actual_value"] = np.where(
        forecasts["demand_type"].eq("entry"),
        forecasts["entries"],
        forecasts["exits"],
    )


    # -------------------------------------------------------------------------
    # Errors
    # -------------------------------------------------------------------------

    forecasts["error"] = (
        forecasts["forecast_value"]
        -
        forecasts["actual_value"]
    )

    forecasts["absolute_error"] = (
        forecasts["error"]
        .abs()
    )

    forecasts["squared_error"] = (
        forecasts["error"]
        ** 2
    )

    forecasts["absolute_percentage_error"] = (
        forecasts["absolute_error"]
        /
        forecasts["actual_value"].replace(
            0,
            np.nan,
        )
    )

    forecasts["weekday"] = (
        forecasts["target_date"]
        .dt.day_name()
    )

    forecasts["weekday_num"] = (
        forecasts["target_date"]
        .dt.weekday
    )

    forecasts["month"] = (
        forecasts["target_date"]
        .dt.month
    )

    forecasts["year"] = (
        forecasts["target_date"]
        .dt.year
    )

    forecasts["month_year"] = (
        forecasts["target_date"]
        .dt.to_period("M")
        .astype(str)
    )


    return forecasts


# =============================================================================
# 25. DAILY SIMULATION QUICK CHECK
# =============================================================================

def print_simulation_quick_check(
    daily_results,
):
    """
    Print a simple sanity check before we move to scoring.

    This is deliberately useful for debugging.

    We want to see:

        * number of target dates
        * number of horizons
        * number of scenarios
        * number of forecasts with actuals
        * forecast range
        * actual range
    """

    print("\n" + "=" * 80)
    print("DAILY SIMULATION QUICK CHECK")
    print("=" * 80)

    if daily_results.empty:
        print("No simulation results were generated.")
        return

    print(
        f"Rows generated: "
        f"{len(daily_results):,}"
    )

    print(
        f"Target dates: "
        f"{daily_results['target_date'].nunique():,}"
    )

    print(
        f"Horizons: "
        f"{daily_results['horizon_days'].nunique():,}"
    )

    print(
        f"Scenarios: "
        f"{daily_results['scenario_id'].nunique():,}"
    )

    print(
        "\nBy demand type:"
    )

    print(
        daily_results[
            "demand_type"
        ]
        .value_counts()
        .to_string()
    )

    print(
        "\nMissing forecast values:"
    )

    print(
        daily_results[
            "forecast_value"
        ]
        .isna()
        .sum()
    )

    print(
        "\nMissing actual validation values:"
    )

    print(
        daily_results[
            "actual_value"
        ]
        .isna()
        .sum()
    )

def validate_no_scenario_errors(
    daily_results,
):
    """
    Fail the run if any scenario exception was converted into an ERROR row.
    """

    if "method_detail" not in daily_results.columns:
        return

    errors = daily_results[
        daily_results["method_detail"]
        .eq("ERROR")
    ].copy()

    if errors.empty:
        print(
            "No scenario execution errors found."
        )
        return

    columns = [
        "target_date",
        "horizon_days",
        "demand_type",
        "scenario_id",
        "error_message",
    ]

    available_columns = [
        column
        for column in columns
        if column in errors.columns
    ]

    raise RuntimeError(
        "Scenario execution errors were found:\n"
        +
        errors[
            available_columns
        ]
        .head(20)
        .to_string(index=False)
    )    


# =============================================================================
# SECTION 3 — RUN, VALIDATE, SCORE AND EXPORT THE FORECAST SIMULATION
# =============================================================================
#
# PURPOSE
# -------
#
# This section is the final orchestration layer.
#
# Sections 1-2 have already defined:
#
#   1. How the data is loaded and cleaned
#   2. The E1-E13 and X1-X16 forecasting scenarios
#   3. The historical backtesting engine
#
# Section 3 now:
#
#   - loads the data
#   - cleans and reconciles it
#   - builds the model datasets
#   - validates that the test sample is correct
#   - runs the complete 29-scenario tournament
#   - scores the forecasts against actual outcomes
#   - ranks the scenarios
#   - identifies the best model at each T-minus horizon
#   - exports everything to Excel
#

# =============================================================================
# 3.1 — FINAL SIMULATION CONFIGURATION
# =============================================================================
#
# We use the configuration already defined in Section 1.
#
# This function simply adds a few reporting / validation settings rather than
# creating a second competing configuration.
# =============================================================================


def get_final_simulation_config():
    """
    Get the central simulation configuration and add validation settings.
    """

    config = get_simulation_config()

    config["expected_target_dates"] = 182

    config["expected_entry_scenarios"] = [
        f"E{i}"
        for i in range(1, 16)
    ]

    config["expected_exit_scenarios"] = [
        f"X{i}"
        for i in range(1, 18)
    ]

    config["expected_horizons"] = [
        0,
        1,
        2,
        3,
        4,
        5,
        6,
        7,
        14,
        21,
        28,
        35,
        42,
        49,
        56,
    ]

    config["expected_scenario_count"] = (
        len(config["expected_entry_scenarios"])
        +
        len(config["expected_exit_scenarios"])
    )

    config["expected_simulation_rows"] = (
        config["expected_target_dates"]
        *
        len(config["expected_horizons"])
        *
        config["expected_scenario_count"]
    )

    return config


# =============================================================================
# 3.2 — VALIDATE THE INPUT DATASETS
# =============================================================================
#
# Before running 54,810 forecast observations, we want to make sure the
# underlying data actually looks sensible.
#
# This is particularly important because the simulation is only as good as:
#
#       bookings
#       operations
#       flights
#       reconciliation
#
# The analysis script already uses:
#
#       B  = valid booking
#       CX = cancelled
#       F  = unknown
#
# We retain those definitions rather than inventing a different cleaning
# process here.
# =============================================================================


def validate_loaded_data(data, config):
    """
    Validate that the loaded and cleaned datasets contain the fields and
    records required by the forecasting scenarios.
    """

    required_datasets = [
        "bookings",
        "operations",
        "flights",
        "master",
    ]

    for dataset_name in required_datasets:

        if dataset_name not in data:
            raise RuntimeError(
                f"Required dataset '{dataset_name}' is missing."
            )

        if data[dataset_name] is None:
            raise RuntimeError(
                f"Required dataset '{dataset_name}' is None."
            )

    bookings = data["bookings"]
    operations = data["operations"]
    flights = data["flights"]
    master = data["master"]

    required_booking_columns = [
        "bookingId",
        "createdAt",
        "cancelledAt",
        "entryDate",
        "exitDate",
        "status",
    ]

    required_operation_columns = [
        "bookingId",
        "actual_entry_ts",
        "actual_exit_ts",
    ]

    missing_booking_columns = [
        c
        for c in required_booking_columns
        if c not in bookings.columns
    ]

    missing_operation_columns = [
        c
        for c in required_operation_columns
        if c not in master.columns
    ]

    if missing_booking_columns:
        raise RuntimeError(
            "Missing booking columns: "
            + ", ".join(missing_booking_columns)
        )

    if missing_operation_columns:
        raise RuntimeError(
            "Missing operational/master columns: "
            + ", ".join(missing_operation_columns)
        )

    if bookings.empty:
        raise RuntimeError(
            "Cleaned bookings dataframe is empty."
        )

    if master.empty:
        raise RuntimeError(
            "Reconciled master dataframe is empty."
        )

    # -------------------------------------------------------------------------
    # Date conversion sanity checks
    # -------------------------------------------------------------------------

    for column in [
        "createdAt",
        "cancelledAt",
        "entryDate",
        "exitDate",
    ]:

        bookings[column] = pd.to_datetime(
            bookings[column],
            errors="coerce",
        )

    for column in [
        "actual_entry_ts",
        "actual_exit_ts",
    ]:

        master[column] = pd.to_datetime(
            master[column],
            errors="coerce",
        )

    # -------------------------------------------------------------------------
    # Status distribution
    # -------------------------------------------------------------------------

    status_counts = (
        bookings["status"]
        .value_counts(dropna=False)
        .rename_axis("status")
        .reset_index(name="bookings")
    )

    print()
    print("Booking status distribution:")
    print(status_counts.to_string(index=False))

    # -------------------------------------------------------------------------
    # Basic actual availability
    # -------------------------------------------------------------------------

    entry_actuals = master[
        "actual_entry_ts"
    ].notna().sum()

    exit_actuals = master[
        "actual_exit_ts"
    ].notna().sum()

    if entry_actuals == 0:
        raise RuntimeError(
            "No actual entry timestamps were found."
        )

    if exit_actuals == 0:
        raise RuntimeError(
            "No actual exit timestamps were found."
        )

    print()
    print(
        f"Operational actual entries available: "
        f"{entry_actuals:,}"
    )

    print(
        f"Operational actual exits available: "
        f"{exit_actuals:,}"
    )

    return {
        "status_counts": status_counts,
        "entry_actual_rows": entry_actuals,
        "exit_actual_rows": exit_actuals,
        "bookings_rows": len(bookings),
        "operations_rows": len(operations),
        "flights_rows": len(flights),
        "master_rows": len(master),
    }


# =============================================================================
# 3.3 — VALIDATE TEST DATES
# =============================================================================


def validate_test_dates(config):
    """
    Validate that the requested historical test sample is exactly correct.

    Expected:
        1st-7th
        14th-20th

    for:
        July 2025 -> March 2026

    Expected total:
        126 target dates
    """

    test_dates = create_test_dates(
        config
    )

    if test_dates.empty:
        raise RuntimeError(
            "No simulation target dates were generated."
        )

    test_dates["target_date"] = pd.to_datetime(
        test_dates["target_date"],
        errors="coerce",
    ).dt.normalize()

    expected_dates = pd.date_range(
        start=config["simulation_start"],
        end=config["simulation_end"],
        freq="D",
    )

    expected_dates = expected_dates[
        expected_dates.day.isin(
            config["test_days_of_month"]
        )
    ].normalize()

    actual_set = set(
        test_dates["target_date"]
    )

    expected_set = set(
        expected_dates
    )

    missing_dates = (
        expected_set - actual_set
    )

    unexpected_dates = (
        actual_set - expected_set
    )

    if missing_dates:
        raise RuntimeError(
            "Simulation is missing expected target dates: "
            + ", ".join(
                str(x.date())
                for x in sorted(missing_dates)
            )
        )

    if unexpected_dates:
        raise RuntimeError(
            "Simulation contains unexpected target dates: "
            + ", ".join(
                str(x.date())
                for x in sorted(unexpected_dates)
            )
        )

    if len(test_dates) != config["expected_target_dates"]:
        raise RuntimeError(
            "Unexpected number of target dates. "
            f"Expected {config['expected_target_dates']}, "
            f"found {len(test_dates)}."
        )

    print()
    print(
        f"Validated target dates: "
        f"{len(test_dates):,}"
    )

    print(
        f"Target range: "
        f"{test_dates['target_date'].min().date()} "
        f"-> "
        f"{test_dates['target_date'].max().date()}"
    )

    return test_dates


# =============================================================================
# 3.4 — VALIDATE SCENARIO CATALOGUE
# =============================================================================
#
# We explicitly require:
#
#       E1-E13
#       X1-X16
#
# This prevents a scenario accidentally disappearing from the tournament.
# =============================================================================


def validate_final_scenario_catalogue(config):
    """
    Confirm that the final scenario catalogue contains exactly E1-E15
    and X1-X17.
    """

    validate_scenario_catalogue()

    actual_entries = list(
        ENTRY_SCENARIOS.keys()
    )

    actual_exits = list(
        EXIT_SCENARIOS.keys()
    )

    expected_entries = config[
        "expected_entry_scenarios"
    ]

    expected_exits = config[
        "expected_exit_scenarios"
    ]

    if actual_entries != expected_entries:
        raise RuntimeError(
            "Entry scenario catalogue does not match E1-E15.\n"
            f"Expected: {expected_entries}\n"
            f"Found:    {actual_entries}"
        )

    if actual_exits != expected_exits:
        raise RuntimeError(
            "Exit scenario catalogue does not match X1-X17.\n"
            f"Expected: {expected_exits}\n"
            f"Found:    {actual_exits}"
        )

    print()
    print("Scenario catalogue validated:")
    print(
        f"  Entries: {len(actual_entries)} "
        "(E1-E15)"
    )
    print(
        f"  Exits:   {len(actual_exits)} "
        "(X1-X17)"
    )
    print(
        f"  Total:   "
        f"{len(actual_entries) + len(actual_exits)} scenarios"
    )

# =============================================================================
# 3.5 — VALIDATE FORECAST HORIZONS
# =============================================================================


def validate_horizons(config):
    """
    Confirm that the simulation uses exactly the requested horizons.
    """

    actual = list(
        config["forecast_horizons_days"]
    )

    expected = list(
        config["expected_horizons"]
    )

    if actual != expected:
        raise RuntimeError(
            "Forecast horizon configuration does not match expected values.\n"
            f"Expected: {expected}\n"
            f"Found:    {actual}"
        )

    print()
    print(
        f"Validated forecast horizons: "
        f"{len(actual)}"
    )

    print(
        "  "
        + ", ".join(
            f"T-{x}"
            for x in actual
        )
    )


# =============================================================================
# 3.6 — VALIDATE DAILY ACTUAL VALIDATION TARGET
# =============================================================================
#
# The actual validation target must come from operational reality:
#
#       Entry = actual_entry_ts / CheckInEnded
#       Exit  = actual_exit_ts / ActualCheckedOutDate
#
# as defined in Section 1.
#
# We do NOT use bookings as the actual validation target.
# =============================================================================


def validate_daily_actuals(
    model_data,
    test_dates,
):
    """
    Validate that the target dates have usable actual entry and exit demand.
    """

    daily_actuals = model_data[
        "daily_actuals"
    ].copy()

    daily_actuals["date"] = pd.to_datetime(
        daily_actuals["date"],
        errors="coerce",
    ).dt.normalize()

    target_dates = test_dates[
        "target_date"
    ].copy()

    target_dates = pd.to_datetime(
        target_dates,
        errors="coerce",
    ).dt.normalize()

    validation = target_dates.to_frame(
        name="target_date"
    ).merge(
        daily_actuals[
            [
                "date",
                "entries",
                "exits",
            ]
        ].rename(
            columns={
                "date": "target_date"
            }
        ),
        on="target_date",
        how="left",
    )

    missing_entries = validation[
        validation["entries"].isna()
    ]

    missing_exits = validation[
        validation["exits"].isna()
    ]

    print()
    print(
        f"Target dates with entry actuals: "
        f"{validation['entries'].notna().sum():,} / "
        f"{len(validation):,}"
    )

    print(
        f"Target dates with exit actuals: "
        f"{validation['exits'].notna().sum():,} / "
        f"{len(validation):,}"
    )

    if not missing_entries.empty:
        print(
            "\nWARNING: target dates with no entry actual:"
        )

        print(
            missing_entries[
                ["target_date"]
            ].to_string(index=False)
        )

    if not missing_exits.empty:
        print(
            "\nWARNING: target dates with no exit actual:"
        )

        print(
            missing_exits[
                ["target_date"]
            ].to_string(index=False)
        )

    return validation


# =============================================================================
# 3.7 — SCORE THE NEW SECTION 2 SIMULATION RESULTS
# =============================================================================
#
# The new run_daily_simulation() already calculates:
#
#       forecast_value
#       actual_value
#       error
#       absolute_error
#       squared_error
#       absolute_percentage_error
#
# We therefore score that exact output rather than rerunning another
# simulation engine.
# =============================================================================


def score_daily_simulation_results(
    daily_results,
):
    """
    Create scenario/horizon performance summaries from the actual Section 2
    simulation results.
    """

    if daily_results.empty:
        return pd.DataFrame()

    df = daily_results.copy()

    df["target_date"] = pd.to_datetime(
        df["target_date"],
        errors="coerce",
    ).dt.normalize()

    # -------------------------------------------------------------------------
    # Use only observations with both forecast and actual.
    # -------------------------------------------------------------------------

    valid = df[
        df["forecast_value"].notna()
        &
        df["actual_value"].notna()
    ].copy()

    if valid.empty:
        return pd.DataFrame()

    rows = []

    for (
        demand_type,
        scenario_id,
        scenario_name,
        scenario_family,
        horizon_days,
    ), group in valid.groupby(
        [
            "demand_type",
            "scenario_id",
            "scenario_name",
            "scenario_family",
            "horizon_days",
        ]
    ):

        actual = group[
            "actual_value"
        ]

        forecast = group[
            "forecast_value"
        ]

        error = (
            forecast
            - actual
        )

        absolute_error = (
            error.abs()
        )

        squared_error = (
            error ** 2
        )

        actual_total = (
            actual.abs().sum()
        )

        if actual_total > 0:

            wape = (
                absolute_error.sum()
                / actual_total
            )

        else:

            wape = np.nan

        mape_values = (
            absolute_error
            /
            actual.abs().replace(
                0,
                np.nan,
            )
        )

        rows.append(
            {
                "demand_type": demand_type,

                "flow_type": (
                    "ENTRY"
                    if demand_type == "entry"
                    else "EXIT"
                ),

                "scenario_id": scenario_id,

                "scenario_name": scenario_name,

                "scenario_family": scenario_family,

                "horizon_days": horizon_days,

                "horizon_label": f"T-{horizon_days}",

                "observations": len(group),

                "forecast_mean": forecast.mean(),

                "actual_mean": actual.mean(),

                "forecast_total": forecast.sum(),

                "actual_total": actual.sum(),

                "mae": absolute_error.mean(),

                "rmse": np.sqrt(
                    squared_error.mean()
                ),

                "bias": error.mean(),

                "median_absolute_error":
                    absolute_error.median(),

                "mape_pct":
                    mape_values.mean() * 100,

                "wape_pct":
                    wape * 100
                    if pd.notna(wape)
                    else np.nan,

                "mean_error":
                    error.mean(),

                "min_error":
                    error.min(),

                "max_error":
                    error.max(),
            }
        )

    return (
        pd.DataFrame(rows)
        .sort_values(
            [
                "flow_type",
                "horizon_days",
                "wape_pct",
            ]
        )
        .reset_index(drop=True)
    )


# =============================================================================
# 3.8 — SCENARIO COVERAGE
# =============================================================================
#
#
# A scenario that only forecasts 40% of cases cannot fairly beat a scenario
# that forecasts 100% of cases.
#
# We therefore report:
#
#       expected observations
#       forecast available
#       actual available
#       forecast coverage
#       actual coverage
#
# =============================================================================


def calculate_final_coverage(
    daily_results,
    config,
):
    """
    Calculate forecast and actual coverage by scenario and horizon.
    """

    if daily_results.empty:
        return pd.DataFrame()

    df = daily_results.copy()

    grouped = (
        df
        .groupby(
            [
                "demand_type",
                "scenario_id",
                "scenario_name",
                "horizon_days",
            ]
        )
        .agg(
            expected_observations=(
                "target_date",
                "count",
            ),

            forecasts_available=(
                "forecast_value",
                lambda x: x.notna().sum(),
            ),

            actuals_available=(
                "actual_value",
                lambda x: x.notna().sum(),
            ),
        )
        .reset_index()
    )

    grouped["forecast_coverage_pct"] = (
        grouped["forecasts_available"]
        /
        grouped["expected_observations"]
        * 100
    )

    grouped["actual_coverage_pct"] = (
        grouped["actuals_available"]
        /
        grouped["expected_observations"]
        * 100
    )

    grouped["flow_type"] = np.where(
        grouped["demand_type"].eq("entry"),
        "ENTRY",
        "EXIT",
    )

    return grouped.sort_values(
        [
            "flow_type",
            "horizon_days",
            "scenario_id",
        ]
    ).reset_index(drop=True)


# =============================================================================
# 3.9 — RANK MODELS BY HORIZON
# =============================================================================
#
# WAPE is the primary model-selection metric.
#
# But we retain MAE, RMSE and bias because:
#
#       low WAPE + systematic over-forecast
#
# may not be operationally preferable to:
#
#       slightly higher WAPE + very balanced forecast.
#
# =============================================================================


def rank_final_models(
    performance,
):
    """
    Rank models separately for Entries/Exits and each T-minus horizon.
    """

    if performance.empty:
        return pd.DataFrame()

    ranked = performance.copy()

    ranked["wape_rank"] = (
        ranked
        .groupby(
            [
                "flow_type",
                "horizon_days",
            ]
        )["wape_pct"]
        .rank(
            method="min",
            ascending=True,
        )
    )

    ranked["mae_rank"] = (
        ranked
        .groupby(
            [
                "flow_type",
                "horizon_days",
            ]
        )["mae"]
        .rank(
            method="min",
            ascending=True,
        )
    )

    ranked["rmse_rank"] = (
        ranked
        .groupby(
            [
                "flow_type",
                "horizon_days",
            ]
        )["rmse"]
        .rank(
            method="min",
            ascending=True,
        )
    )

    return ranked.sort_values(
        [
            "flow_type",
            "horizon_days",
            "wape_rank",
            "scenario_id",
        ]
    ).reset_index(drop=True)


# =============================================================================
# 3.10 — BEST MODEL BY HORIZON
# =============================================================================


def create_best_by_horizon(
    ranked_performance,
    coverage,
):
    """
    Identify the best scenario at each T-minus point.

    We require a forecast coverage threshold so that a model cannot win merely
    because it successfully forecasts a small subset of the validation sample.
    """

    if ranked_performance.empty:
        return pd.DataFrame()

    merged = ranked_performance.merge(
        coverage[
            [
                "flow_type",
                "scenario_id",
                "horizon_days",
                "forecast_coverage_pct",
                "actual_coverage_pct",
            ]
        ],
        on=[
            "flow_type",
            "scenario_id",
            "horizon_days",
        ],
        how="left",
    )

    # -------------------------------------------------------------------------
    # For the primary winner table, use models with at least 80% forecast
    # coverage.
    #
    # This threshold is deliberately not treated as a final production rule.
    # It simply stops a low-coverage scenario from winning unfairly.
    # -------------------------------------------------------------------------

    eligible = merged[
        merged["forecast_coverage_pct"] >= 80
    ].copy()

    if eligible.empty:
        eligible = merged.copy()

    best = (
        eligible
        .sort_values(
            [
                "flow_type",
                "horizon_days",
                "wape_pct",
                "mae",
            ]
        )
        .groupby(
            [
                "flow_type",
                "horizon_days",
            ],
            as_index=False,
        )
        .first()
    )

    return best.sort_values(
        [
            "flow_type",
            "horizon_days",
        ]
    ).reset_index(drop=True)


# =============================================================================
# 3.11 — OVERALL MODEL RANKING
# =============================================================================
#
# This is deliberately secondary to the horizon-by-horizon result.
#
# We ultimately care about:
#
#       "What should we use at T-56?"
#       "What should we use at T-28?"
#       "What should we use at T-7?"
#
# rather than assuming one model must be best at every point.
# =============================================================================


def create_overall_model_ranking(
    daily_results,
):
    """
    Rank each scenario across all horizons for Entries and Exits.
    """

    if daily_results.empty:
        return pd.DataFrame()

    valid = daily_results[
        daily_results["forecast_value"].notna()
        &
        daily_results["actual_value"].notna()
    ].copy()

    rows = []

    for (
        demand_type,
        scenario_id,
        scenario_name,
        scenario_family,
    ), group in valid.groupby(
        [
            "demand_type",
            "scenario_id",
            "scenario_name",
            "scenario_family",
        ]
    ):

        error = (
            group["forecast_value"]
            - group["actual_value"]
        )

        absolute_error = (
            error.abs()
        )

        actual_total = (
            group["actual_value"]
            .abs()
            .sum()
        )

        if actual_total > 0:

            wape = (
                absolute_error.sum()
                /
                actual_total
            )

        else:

            wape = np.nan

        rows.append(
            {
                "flow_type": (
                    "ENTRY"
                    if demand_type == "entry"
                    else "EXIT"
                ),

                "scenario_id": scenario_id,

                "scenario_name": scenario_name,

                "scenario_family": scenario_family,

                "observations": len(group),

                "horizons_tested":
                    group["horizon_days"].nunique(),

                "mae":
                    absolute_error.mean(),

                "rmse":
                    np.sqrt(
                        error.pow(2).mean()
                    ),

                "bias":
                    error.mean(),

                "wape_pct":
                    wape * 100
                    if pd.notna(wape)
                    else np.nan,

                "forecast_coverage_pct":
                    group["forecast_value"].notna().mean()
                    * 100,
            }
        )

    ranking = pd.DataFrame(rows)

    ranking["overall_rank"] = (
        ranking
        .groupby("flow_type")["wape_pct"]
        .rank(
            method="min",
            ascending=True,
        )
    )

    return ranking.sort_values(
        [
            "flow_type",
            "overall_rank",
            "scenario_id",
        ]
    ).reset_index(drop=True)


# =============================================================================
# 3.12 — MODEL STABILITY ACROSS HORIZONS
# =============================================================================
#
# A useful model is not necessarily the model that wins once.
#
# We therefore measure:
#
#       average WAPE rank
#       median WAPE rank
#       number of horizon wins
#       win rate
#
# This will help us distinguish:
#
#       "best one-off model"
#
# from:
#
#       "robust forecasting methodology".
# =============================================================================


def create_model_stability(
    ranked_performance,
):
    """
    Summarise how consistently each scenario performs across horizons.
    """

    if ranked_performance.empty:
        return pd.DataFrame()

    rows = []

    for (
        flow_type,
        scenario_id,
        scenario_name,
        scenario_family,
    ), group in ranked_performance.groupby(
        [
            "flow_type",
            "scenario_id",
            "scenario_name",
            "scenario_family",
        ]
    ):

        rows.append(
            {
                "flow_type": flow_type,

                "scenario_id": scenario_id,

                "scenario_name": scenario_name,

                "scenario_family": scenario_family,

                "horizons_tested":
                    group["horizon_days"].nunique(),

                "average_wape_rank":
                    group["wape_rank"].mean(),

                "median_wape_rank":
                    group["wape_rank"].median(),

                "horizon_wins":
                    (
                        group["wape_rank"] == 1
                    ).sum(),

                "average_wape_pct":
                    group["wape_pct"].mean(),

                "average_mae":
                    group["mae"].mean(),

                "average_bias":
                    group["bias"].mean(),
            }
        )

    output = pd.DataFrame(rows)

    output["win_rate_pct"] = (
        output["horizon_wins"]
        /
        output["horizons_tested"]
        * 100
    )

    return output.sort_values(
        [
            "flow_type",
            "average_wape_rank",
        ]
    ).reset_index(drop=True)

def calculate_weighted_forecast(
    primary_forecast,
    comparison_forecast,
    primary_weight,
):
    """
    Blend two forecasts using a tested primary-component weight.
    """

    if (
        pd.isna(primary_forecast)
        and pd.isna(comparison_forecast)
    ):
        return np.nan

    if pd.isna(primary_forecast):
        return comparison_forecast

    if pd.isna(comparison_forecast):
        return primary_forecast

    return clip_forecast(
        primary_weight
        * primary_forecast
        +
        (1.0 - primary_weight)
        * comparison_forecast
    )


def calibrate_two_component_weights(
    daily_results,
    primary_scenario,
    comparison_scenario,
    flow_type,
    blend_name,
    weight_grid,
    minimum_observations=30,
):
    """
    Test all configured weights for a two-component blend independently
    at each forecast horizon.

    This is an analytical calibration. It does not add fixed-weight
    scenarios to the main tournament.
    """

    demand_type = (
        "entry"
        if flow_type == "ENTRY"
        else "exit"
    )

    required = daily_results[
        daily_results["demand_type"].eq(
            demand_type
        )
        &
        daily_results["scenario_id"].isin(
            [
                primary_scenario,
                comparison_scenario,
            ]
        )
    ][
        [
            "target_date",
            "horizon_days",
            "scenario_id",
            "forecast_value",
            "actual_value",
        ]
    ].copy()

    pivot = required.pivot_table(
        index=[
            "target_date",
            "horizon_days",
            "actual_value",
        ],
        columns="scenario_id",
        values="forecast_value",
        aggfunc="first",
    ).reset_index()

    if (
        primary_scenario not in pivot.columns
        or comparison_scenario not in pivot.columns
    ):
        return pd.DataFrame()

    rows = []

    for horizon_days, group in pivot.groupby(
        "horizon_days"
    ):

        group = group[
            group["actual_value"].notna()
        ].copy()

        if len(group) < minimum_observations:
            continue

        for primary_weight in weight_grid:

            blended = [
                calculate_weighted_forecast(
                    primary_forecast=primary,
                    comparison_forecast=comparison,
                    primary_weight=primary_weight,
                )
                for primary, comparison in zip(
                    group[primary_scenario],
                    group[comparison_scenario],
                )
            ]

            blended = pd.Series(
                blended,
                index=group.index,
                dtype=float,
            )

            valid_mask = (
                blended.notna()
                &
                group["actual_value"].notna()
            )

            actual = group.loc[
                valid_mask,
                "actual_value",
            ].astype(float)

            forecast = blended.loc[
                valid_mask
            ]

            if actual.empty:
                continue

            error = forecast - actual

            denominator = (
                actual.abs().sum()
            )

            wape_pct = (
                error.abs().sum()
                /
                denominator
                * 100
                if denominator > 0
                else np.nan
            )

            rows.append(
                {
                    "flow_type": flow_type,
                    "blend_name": blend_name,
                    "primary_scenario":
                        primary_scenario,
                    "comparison_scenario":
                        comparison_scenario,
                    "horizon_days":
                        horizon_days,
                    "horizon_label":
                        f"T-{horizon_days}",
                    "primary_weight":
                        primary_weight,
                    "comparison_weight":
                        1.0 - primary_weight,
                    "observations":
                        int(valid_mask.sum()),
                    "wape_pct":
                        wape_pct,
                    "mae":
                        error.abs().mean(),
                    "rmse":
                        np.sqrt(
                            error.pow(2).mean()
                        ),
                    "bias":
                        error.mean(),
                }
            )

    return pd.DataFrame(rows)


def create_weight_calibration_outputs(
    daily_results,
    config,
):
    """
    Calibrate the main booking-based blends by horizon.
    """

    calibration_frames = [
        calibrate_two_component_weights(
            daily_results=daily_results,
            primary_scenario="E11",
            comparison_scenario="E3",
            flow_type="ENTRY",
            blend_name="Entry Booking + Weekday",
            weight_grid=config["weight_grid"],
            minimum_observations=config[
                "minimum_weight_calibration_observations"
            ],
        ),

        calibrate_two_component_weights(
            daily_results=daily_results,
            primary_scenario="E11",
            comparison_scenario="E9",
            flow_type="ENTRY",
            blend_name="Entry Booking + Passenger",
            weight_grid=config["weight_grid"],
            minimum_observations=config[
                "minimum_weight_calibration_observations"
            ],
        ),

        calibrate_two_component_weights(
            daily_results=daily_results,
            primary_scenario="E11",
            comparison_scenario="E14",
            flow_type="ENTRY",
            blend_name="Entry Booking + Weekday-Month",
            weight_grid=config["weight_grid"],
            minimum_observations=config[
                "minimum_weight_calibration_observations"
            ],
        ),

        calibrate_two_component_weights(
            daily_results=daily_results,
            primary_scenario="X11",
            comparison_scenario="X3",
            flow_type="EXIT",
            blend_name="Exit Booking + Weekday",
            weight_grid=config["weight_grid"],
            minimum_observations=config[
                "minimum_weight_calibration_observations"
            ],
        ),

        calibrate_two_component_weights(
            daily_results=daily_results,
            primary_scenario="X11",
            comparison_scenario="X9",
            flow_type="EXIT",
            blend_name="Exit Booking + Passenger",
            weight_grid=config["weight_grid"],
            minimum_observations=config[
                "minimum_weight_calibration_observations"
            ],
        ),
    ]

    calibration_frames = [
        frame
        for frame in calibration_frames
        if frame is not None
        and not frame.empty
    ]

    if not calibration_frames:
        return (
            pd.DataFrame(),
            pd.DataFrame(),
        )

    all_weights = pd.concat(
        calibration_frames,
        ignore_index=True,
    )

    best_weights = (
        all_weights
        .sort_values(
            [
                "flow_type",
                "blend_name",
                "horizon_days",
                "wape_pct",
                "mae",
                "primary_weight",
            ]
        )
        .groupby(
            [
                "flow_type",
                "blend_name",
                "horizon_days",
            ],
            as_index=False,
        )
        .first()
        .sort_values(
            [
                "flow_type",
                "blend_name",
                "horizon_days",
            ]
        )
        .reset_index(drop=True)
    )

    return all_weights, best_weights

def summarise_performance_by_columns(
    daily_results,
    group_columns,
):
    """
    Calculate WAPE, MAE, RMSE and bias for arbitrary diagnostic groupings.
    """

    valid = daily_results[
        daily_results["forecast_value"].notna()
        &
        daily_results["actual_value"].notna()
    ].copy()

    rows = []

    full_group_columns = [
        "demand_type",
        "scenario_id",
        "scenario_name",
        *group_columns,
    ]

    for group_key, group in valid.groupby(
        full_group_columns,
        dropna=False,
    ):

        if not isinstance(group_key, tuple):
            group_key = (group_key,)

        key_values = dict(
            zip(
                full_group_columns,
                group_key,
            )
        )

        actual = group[
            "actual_value"
        ].astype(float)

        forecast = group[
            "forecast_value"
        ].astype(float)

        error = forecast - actual

        denominator = actual.abs().sum()

        rows.append(
            {
                **key_values,

                "flow_type": (
                    "ENTRY"
                    if key_values[
                        "demand_type"
                    ] == "entry"
                    else "EXIT"
                ),

                "observations": len(group),

                "wape_pct": (
                    error.abs().sum()
                    /
                    denominator
                    * 100
                    if denominator > 0
                    else np.nan
                ),

                "mae":
                    error.abs().mean(),

                "rmse":
                    np.sqrt(
                        error.pow(2).mean()
                    ),

                "bias":
                    error.mean(),

                "actual_mean":
                    actual.mean(),

                "forecast_mean":
                    forecast.mean(),
            }
        )

    return pd.DataFrame(rows)


def create_weekday_analysis(
    daily_results,
):
    return summarise_performance_by_columns(
        daily_results,
        [
            "weekday_num",
            "weekday",
            "horizon_days",
        ],
    )


def create_month_analysis(
    daily_results,
):
    return summarise_performance_by_columns(
        daily_results,
        [
            "month",
            "horizon_days",
        ],
    )


def add_demand_bands(
    daily_results,
):
    """
    Create within-flow demand quartiles from actual demand.
    """

    df = daily_results.copy()

    df["demand_band"] = pd.NA

    labels = [
        "Low",
        "Medium",
        "High",
        "Peak",
    ]

    for demand_type in [
        "entry",
        "exit",
    ]:

        mask = (
            df["demand_type"]
            .eq(demand_type)
            &
            df["actual_value"]
            .notna()
        )

        unique_actuals = (
            df.loc[
                mask,
                [
                    "target_date",
                    "actual_value",
                ],
            ]
            .drop_duplicates()
        )

        try:
            unique_actuals[
                "demand_band"
            ] = pd.qcut(
                unique_actuals[
                    "actual_value"
                ],
                q=4,
                labels=labels,
                duplicates="drop",
            )

        except ValueError:
            unique_actuals[
                "demand_band"
            ] = "Unclassified"

        mapping = dict(
            zip(
                unique_actuals[
                    "target_date"
                ],
                unique_actuals[
                    "demand_band"
                ].astype(str),
            )
        )

        df.loc[
            mask,
            "demand_band",
        ] = (
            df.loc[
                mask,
                "target_date",
            ]
            .map(mapping)
        )

    return df


def create_demand_band_analysis(
    daily_results,
):
    banded = add_demand_bands(
        daily_results
    )

    return summarise_performance_by_columns(
        banded,
        [
            "demand_band",
            "horizon_days",
        ],
    )


def create_visibility_analysis(
    daily_results,
):
    """
    Analyse known booking visibility for booking-driven scenarios.
    """

    scenario_columns = [
        "target_date",
        "horizon_days",
        "demand_type",
        "scenario_id",
        "scenario_name",
        "forecast_value",
        "actual_value",
        "error",
        "absolute_error",
        "known_bookings",
    ]

    available_columns = [
        column
        for column in scenario_columns
        if column in daily_results.columns
    ]

    visibility = daily_results[
        daily_results["scenario_id"].isin(
            [
                "E11",
                "X11",
            ]
        )
    ][
        available_columns
    ].copy()

    if (
        visibility.empty
        or "known_bookings"
        not in visibility.columns
    ):
        return pd.DataFrame()

    visibility["visibility_pct"] = (
        visibility["known_bookings"]
        /
        visibility["actual_value"]
        .replace(0, np.nan)
        * 100
    )

    visibility["actual_completion_factor"] = (
        visibility["actual_value"]
        /
        visibility["known_bookings"]
        .replace(0, np.nan)
    )

    visibility["estimated_completion_factor"] = (
        visibility["forecast_value"]
        /
        visibility["known_bookings"]
        .replace(0, np.nan)
    )

    visibility["completion_factor_error"] = (
        visibility[
            "estimated_completion_factor"
        ]
        -
        visibility[
            "actual_completion_factor"
        ]
    )

    visibility["visibility_band"] = pd.cut(
        visibility["visibility_pct"],
        bins=[
            -np.inf,
            20,
            40,
            60,
            80,
            100,
            np.inf,
        ],
        labels=[
            "0-20%",
            "20-40%",
            "40-60%",
            "60-80%",
            "80-100%",
            "100%+",
        ],
    )

    return visibility.sort_values(
        [
            "demand_type",
            "target_date",
            "horizon_days",
        ]
    ).reset_index(drop=True)


def create_visibility_summary(
    visibility_analysis,
):
    if visibility_analysis.empty:
        return pd.DataFrame()

    rows = []

    for (
        demand_type,
        horizon_days,
        visibility_band,
    ), group in visibility_analysis.groupby(
        [
            "demand_type",
            "horizon_days",
            "visibility_band",
        ],
        observed=True,
    ):

        denominator = (
            group["actual_value"]
            .abs()
            .sum()
        )

        rows.append(
            {
                "flow_type": (
                    "ENTRY"
                    if demand_type == "entry"
                    else "EXIT"
                ),

                "horizon_days":
                    horizon_days,

                "visibility_band":
                    visibility_band,

                "observations":
                    len(group),

                "average_visibility_pct":
                    group[
                        "visibility_pct"
                    ].mean(),

                "average_known_bookings":
                    group[
                        "known_bookings"
                    ].mean(),

                "wape_pct": (
                    group[
                        "absolute_error"
                    ].sum()
                    /
                    denominator
                    * 100
                    if denominator > 0
                    else np.nan
                ),

                "bias":
                    group[
                        "error"
                    ].mean(),

                "average_actual_completion":
                    group[
                        "actual_completion_factor"
                    ].mean(),

                "average_estimated_completion":
                    group[
                        "estimated_completion_factor"
                    ].mean(),
            }
        )

    return pd.DataFrame(rows)


def create_worst_forecast_days(
    daily_results,
    top_n=100,
):
    """
    Return the largest absolute forecast errors.
    """

    return (
        daily_results[
            daily_results["forecast_value"].notna()
            &
            daily_results["actual_value"].notna()
        ]
        .sort_values(
            "absolute_error",
            ascending=False,
        )
        .head(top_n)
        .reset_index(drop=True)
    )


def create_family_comparison(
    daily_results,
):
    return summarise_performance_by_columns(
        daily_results,
        [
            "scenario_family",
            "horizon_days",
        ],
    )


def create_incremental_benefit_analysis(
    performance,
):
    """
    Compare component models against the principal booking benchmarks.
    """

    comparisons = [
        (
            "ENTRY",
            "E11",
            "E12",
            "Weekday added to booking curve",
        ),
        (
            "ENTRY",
            "E11",
            "E13",
            "Passenger context added to booking curve",
        ),
        (
            "ENTRY",
            "E11",
            "E15",
            "Weekday-month added to booking curve",
        ),
        (
            "EXIT",
            "X11",
            "X12",
            "Duration added to exit booking curve",
        ),
        (
            "EXIT",
            "X11",
            "X15",
            "Passenger context added to exit booking curve",
        ),
    ]

    rows = []

    for (
        flow_type,
        baseline_id,
        challenger_id,
        comparison_name,
    ) in comparisons:

        baseline = performance[
            performance["flow_type"].eq(
                flow_type
            )
            &
            performance["scenario_id"].eq(
                baseline_id
            )
        ][
            [
                "horizon_days",
                "wape_pct",
                "mae",
                "rmse",
                "bias",
            ]
        ].rename(
            columns={
                "wape_pct":
                    "baseline_wape_pct",
                "mae":
                    "baseline_mae",
                "rmse":
                    "baseline_rmse",
                "bias":
                    "baseline_bias",
            }
        )

        challenger = performance[
            performance["flow_type"].eq(
                flow_type
            )
            &
            performance["scenario_id"].eq(
                challenger_id
            )
        ][
            [
                "horizon_days",
                "wape_pct",
                "mae",
                "rmse",
                "bias",
            ]
        ].rename(
            columns={
                "wape_pct":
                    "challenger_wape_pct",
                "mae":
                    "challenger_mae",
                "rmse":
                    "challenger_rmse",
                "bias":
                    "challenger_bias",
            }
        )

        merged = baseline.merge(
            challenger,
            on="horizon_days",
            how="inner",
        )

        if merged.empty:
            continue

        merged["flow_type"] = flow_type
        merged["comparison"] = comparison_name
        merged["baseline_scenario"] = baseline_id
        merged["challenger_scenario"] = challenger_id

        merged["wape_improvement_points"] = (
            merged["baseline_wape_pct"]
            -
            merged["challenger_wape_pct"]
        )

        merged["mae_improvement"] = (
            merged["baseline_mae"]
            -
            merged["challenger_mae"]
        )

        rows.append(merged)

    if not rows:
        return pd.DataFrame()

    return pd.concat(
        rows,
        ignore_index=True,
    )

# =============================================================================
# 3.13 — WAPE MATRIX
# =============================================================================
#
# This gives us a very useful visual/table representation:
#
#                       T-56   T-49   T-42 ... T-0
#
#       E1
#       E2
#       E3
#       ...
#
# and separately:
#
#       X1
#       X2
#       ...
#
# It will make it much easier to see whether the optimal forecasting approach
# changes through the booking curve.
# =============================================================================


def create_wape_matrix(
    performance,
    flow_type,
):
    """
    Create scenario x horizon WAPE matrix.
    """

    if performance.empty:
        return pd.DataFrame()

    df = performance[
        performance["flow_type"] == flow_type
    ].copy()

    if df.empty:
        return pd.DataFrame()

    df["scenario_label"] = (
        df["scenario_id"]
        + " — "
        + df["scenario_name"]
    )

    matrix = df.pivot_table(
        index="scenario_label",
        columns="horizon_label",
        values="wape_pct",
        aggfunc="first",
    )

    horizon_order = sorted(
        matrix.columns,
        key=lambda x: int(
            str(x).replace(
                "T-",
                "",
            )
        ),
        reverse=True,
    )

    matrix = matrix.reindex(
        columns=horizon_order
    )

    return matrix.reset_index()


# =============================================================================
# 3.14 — SCENARIO DOCUMENTATION TABLE
# =============================================================================
#
# Pull the descriptions directly from Section 2.
#
# This is much safer than duplicating scenario names/descriptions manually in
# Section 4 and then having them drift apart.
# =============================================================================


def create_scenario_documentation():
    """
    Build a single documentation table directly from Section 2.
    """

    rows = []

    for scenario_id, details in ENTRY_SCENARIOS.items():

        rows.append(
            {
                "flow_type": "ENTRY",
                "scenario_id": scenario_id,
                "scenario_name": details.get(
                    "name",
                    "",
                ),
                "scenario_family": details.get(
                    "family",
                    "",
                ),
                "description": details.get(
                    "description",
                    "",
                ),
                "evidence": details.get(
                    "evidence",
                    "",
                ),
            }
        )

    for scenario_id, details in EXIT_SCENARIOS.items():

        rows.append(
            {
                "flow_type": "EXIT",
                "scenario_id": scenario_id,
                "scenario_name": details.get(
                    "name",
                    "",
                ),
                "scenario_family": details.get(
                    "family",
                    "",
                ),
                "description": details.get(
                    "description",
                    "",
                ),
                "evidence": details.get(
                    "evidence",
                    "",
                ),
            }
        )

    return pd.DataFrame(rows)


# =============================================================================
# 3.15 — DATA QUALITY SUMMARY
# =============================================================================


def create_data_quality_summary(
    data,
    model_data,
    test_validation,
):
    """
    Create a compact data-quality summary for the Excel workbook.
    """

    bookings = data["bookings"]
    operations = data["operations"]
    flights = data["flights"]
    master = data["master"]

    rows = [
        {
            "metric": "Clean bookings rows",
            "value": len(bookings),
            "notes": "Cleaned FastPark booking population.",
        },
        {
            "metric": "Operational rows",
            "value": len(operations),
            "notes": "FastPark operational entry/exit records.",
        },
        {
            "metric": "Flight rows",
            "value": len(flights),
            "notes": "Cleaned historical flight/passenger records.",
        },
        {
            "metric": "Reconciled master rows",
            "value": len(master),
            "notes": "Booking/operations reconciliation output.",
        },
        {
            "metric": "Daily actual dates",
            "value": model_data["daily_actuals"]["date"].nunique(),
            "notes": "Dates with actual entry/exit activity.",
        },
        {
            "metric": "Hourly actual rows",
            "value": len(model_data["hourly_actuals"]),
            "notes": "Hourly operational actuals; not yet a separate hourly model tournament.",
        },
        {
            "metric": "Target validation dates",
            "value": len(test_validation),
            "notes": "Expected 126 target dates.",
        },
        {
            "metric": "Target dates with entry actual",
            "value": test_validation["entries"].notna().sum(),
            "notes": "Validation coverage.",
        },
        {
            "metric": "Target dates with exit actual",
            "value": test_validation["exits"].notna().sum(),
            "notes": "Validation coverage.",
        },
    ]

    return pd.DataFrame(rows)


# =============================================================================
# 3.16 — EXCEL FORMATTING
# =============================================================================


def format_output_workbook(
    workbook,
):
    """
    Apply simple functional formatting to every output sheet.
    """

    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

        # Header formatting.
        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

        # Column widths.
        for column_cells in worksheet.columns:

            max_length = 0

            for cell in column_cells:

                if cell.value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

            column_letter = get_column_letter(
                column_cells[0].column
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(
                    max_length + 2,
                    10,
                ),
                60,
            )

        # Percentage formatting based on header names.
        percentage_headers = {
            "wape_pct",
            "mape_pct",
            "forecast_coverage_pct",
            "actual_coverage_pct",
            "win_rate_pct",
        }

        header_lookup = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        for header in percentage_headers:

            if header not in header_lookup:
                continue

            column_number = (
                header_lookup[header]
            )

            for row in worksheet.iter_rows(
                min_row=2,
                min_col=column_number,
                max_col=column_number,
            ):

                row[0].number_format = (
                    "0.0"
                )


# =============================================================================
# 3.17 — EXPORT RESULTS TO EXCEL
# =============================================================================


def export_simulation_results(
    output_path,
    data,
    model_data,
    test_dates,
    test_validation,
    daily_results,
    performance,
    ranked_performance,
    coverage,
    best_by_horizon,
    overall_ranking,
    stability,
    all_weight_results,
    best_weight_results,
    weekday_analysis,
    month_analysis,
    demand_band_analysis,
    visibility_analysis,
    visibility_summary,
    worst_forecast_days,
    family_comparison,
    incremental_benefit,
    
):
    """
    Export the full simulation results.

    The workbook deliberately includes both:
        - management-level summary
        - detailed model-level diagnostics
    """

    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wape_entry = create_wape_matrix(
        performance,
        "ENTRY",
    )

    wape_exit = create_wape_matrix(
        performance,
        "EXIT",
    )

    scenario_documentation = (
        create_scenario_documentation()
    )

    data_quality = (
        create_data_quality_summary(
            data,
            model_data,
            test_validation,
        )
    )

    # -------------------------------------------------------------------------
    # Winner table
    # -------------------------------------------------------------------------

    winner_table = best_by_horizon.copy()

    if not winner_table.empty:

        winner_table = winner_table[
            [
                "flow_type",
                "horizon_days",
                "horizon_label",
                "scenario_id",
                "scenario_name",
                "scenario_family",
                "wape_pct",
                "mae",
                "rmse",
                "bias",
                "forecast_coverage_pct",
            ]
        ]

    # -------------------------------------------------------------------------
    # Workbook
    # -------------------------------------------------------------------------

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        # =====================================================================
        # 1 — EXECUTIVE WINNERS
        # =====================================================================

        winner_table.to_excel(
            writer,
            sheet_name="Best By Horizon",
            index=False,
        )

        # =====================================================================
        # 2 — ENTRY PERFORMANCE
        # =====================================================================

        performance[
            performance["flow_type"] == "ENTRY"
        ].to_excel(
            writer,
            sheet_name="Entry Performance",
            index=False,
        )

        # =====================================================================
        # 3 — EXIT PERFORMANCE
        # =====================================================================

        performance[
            performance["flow_type"] == "EXIT"
        ].to_excel(
            writer,
            sheet_name="Exit Performance",
            index=False,
        )

        # =====================================================================
        # 4 — ENTRY WAPE MATRIX
        # =====================================================================

        wape_entry.to_excel(
            writer,
            sheet_name="Entry WAPE Matrix",
            index=False,
        )

        # =====================================================================
        # 5 — EXIT WAPE MATRIX
        # =====================================================================

        wape_exit.to_excel(
            writer,
            sheet_name="Exit WAPE Matrix",
            index=False,
        )

        # =====================================================================
        # 6 — OVERALL RANKING
        # =====================================================================

        overall_ranking.to_excel(
            writer,
            sheet_name="Overall Ranking",
            index=False,
        )

        # =====================================================================
        # 7 — MODEL STABILITY
        # =====================================================================

        stability.to_excel(
            writer,
            sheet_name="Model Stability",
            index=False,
        )

        # =====================================================================
        # 8 — COVERAGE
        # =====================================================================

        coverage.to_excel(
            writer,
            sheet_name="Forecast Coverage",
            index=False,
        )

        # =====================================================================
        # 9 — SCENARIO DEFINITIONS
        # =====================================================================

        scenario_documentation.to_excel(
            writer,
            sheet_name="Scenario Definitions",
            index=False,
        )

        # =====================================================================
        # 10 — TEST DATES
        # =====================================================================

        test_dates.to_excel(
            writer,
            sheet_name="Test Dates",
            index=False,
        )

        # =====================================================================
        # 11 — DATA QUALITY
        # =====================================================================

        data_quality.to_excel(
            writer,
            sheet_name="Data Quality",
            index=False,
        )

        # =====================================================================
        # 12 — VALIDATION TARGETS
        # =====================================================================

        test_validation.to_excel(
            writer,
            sheet_name="Validation Targets",
            index=False,
        )

        # =====================================================================
        # 13 — DAILY SIMULATION
        # =====================================================================

        daily_results.to_excel(
            writer,
            sheet_name="Daily Simulation",
            index=False,
        )

        # =====================================================================
        # 14 — HOURLY ACTUALS
        # =====================================================================
        #
        # We include hourly actuals now because they are important for the
        # eventual operational forecasting model.
        #
        # BUT:
        #
        # We do NOT pretend that the current E/X scenario tournament is an
        # hourly forecasting tournament.
        #
        # That should be a separate next stage where each scenario is tested
        # against hourly demand curves.
        # =====================================================================

        model_data[
            "hourly_actuals"
        ].to_excel(
            writer,
            sheet_name="Hourly Actuals",
            index=False,
        )

        # =====================================================================
        # 15 — DAILY ACTUALS
        # =====================================================================

        model_data[
            "daily_actuals"
        ].to_excel(
            writer,
            sheet_name="Daily Actuals",
            index=False,
        )

        # =====================================================================
        # 16 — DAILY PASSENGER CONTEXT
        # =====================================================================

        model_data[
            "daily_passengers"
        ].to_excel(
            writer,
            sheet_name="Passenger Context",
            index=False,
        )

        # =====================================================================
        # 17 — WEIGHT CALIBRATION AND ANALYSIS
        # =====================================================================
        best_weight_results.to_excel(
            writer,
            sheet_name="Best Hybrid Weights",
            index=False,
        )

        all_weight_results.to_excel(
            writer,
            sheet_name="All Weight Tests",
            index=False,
        )

        weekday_analysis.to_excel(
            writer,
            sheet_name="WAPE by Weekday",
            index=False,
        )

        month_analysis.to_excel(
            writer,
            sheet_name="WAPE by Month",
            index=False,
        )

        demand_band_analysis.to_excel(
            writer,
            sheet_name="Demand Band Analysis",
            index=False,
        )

        visibility_summary.to_excel(
            writer,
            sheet_name="Visibility Summary",
            index=False,
        )

        visibility_analysis.to_excel(
            writer,
            sheet_name="Visibility Detail",
            index=False,
        )

        worst_forecast_days.to_excel(
            writer,
            sheet_name="Worst Forecast Days",
            index=False,
        )

        family_comparison.to_excel(
            writer,
            sheet_name="Family Comparison",
            index=False,
        )

        incremental_benefit.to_excel(
            writer,
            sheet_name="Incremental Benefit",
            index=False,
        )

        # =====================================================================
        # FORMAT
        # =====================================================================

        format_output_workbook(
            writer.book
        )

    return output_path


# =============================================================================
# 3.18 — FINAL CONSOLE SUMMARY
# =============================================================================


def print_final_results_summary(
    performance,
    best_by_horizon,
    overall_ranking,
    coverage,
):
    """
    Print a concise summary after the simulation completes.
    """

    print()
    print("=" * 90)
    print("FASTPARK FORECAST SIMULATION — FINAL RESULTS")
    print("=" * 90)

    if performance.empty:

        print(
            "No valid performance results were generated."
        )

        return

    # -------------------------------------------------------------------------
    # Overall winners
    # -------------------------------------------------------------------------

    print()
    print("-" * 90)
    print("OVERALL BEST MODELS")
    print("-" * 90)

    for flow_type in [
        "ENTRY",
        "EXIT",
    ]:

        flow = overall_ranking[
            overall_ranking["flow_type"]
            == flow_type
        ].copy()

        if flow.empty:
            continue

        print()
        print(
            f"{flow_type}:"
        )

        print(
            flow[
                [
                    "scenario_id",
                    "scenario_name",
                    "wape_pct",
                    "mae",
                    "rmse",
                    "bias",
                    "overall_rank",
                ]
            ]
            .head(5)
            .to_string(
                index=False
            )
        )

    # -------------------------------------------------------------------------
    # Best by horizon
    # -------------------------------------------------------------------------

    print()
    print("-" * 90)
    print("BEST MODEL BY FORECAST HORIZON")
    print("-" * 90)

    if not best_by_horizon.empty:

        print(
            best_by_horizon[
                [
                    "flow_type",
                    "horizon_label",
                    "scenario_id",
                    "scenario_name",
                    "wape_pct",
                    "forecast_coverage_pct",
                ]
            ].to_string(
                index=False
            )
        )

    # -------------------------------------------------------------------------
    # Coverage warnings
    # -------------------------------------------------------------------------

    low_coverage = coverage[
        coverage["forecast_coverage_pct"] < 80
    ].copy()

    if not low_coverage.empty:

        print()
        print("-" * 90)
        print("LOW FORECAST COVERAGE — REVIEW BEFORE USING MODEL")
        print("-" * 90)

        print(
            low_coverage[
                [
                    "flow_type",
                    "scenario_id",
                    "scenario_name",
                    "horizon_days",
                    "forecast_coverage_pct",
                ]
            ]
            .sort_values(
                "forecast_coverage_pct"
            )
            .head(20)
            .to_string(
                index=False
            )
        )

    print()
    print("=" * 90)
    print("SIMULATION COMPLETE")
    print("=" * 90)

# =============================================================================
# SECTION 4 — HOURLY FORECASTING FRAMEWORK
# =============================================================================
#
# PURPOSE
# -------
#
# The daily forecasting tournament identifies the best daily model.
#
# But operational staffing requires HOURLY forecasts:
#
#     "How many cars will arrive at 08:00?"
#     "When is the peak exit hour?"
#
# This section distributes daily forecasts across hours using historical
# hourly profiles.
#
# =============================================================================


def build_hourly_profile(
    hourly_actuals,
    target_date,
    cutoff_timestamp,
    demand_type,
    n_weeks,
):
    """
    Build a same-weekday hourly demand profile using only fully completed
    historical dates before the forecast cutoff.
    """

    target_date = pd.Timestamp(
        target_date
    ).normalize()

    cutoff_date = pd.Timestamp(
        cutoff_timestamp
    ).normalize()

    weekday = target_date.weekday()

    df = hourly_actuals

    historical = df[
        df["date"].lt(target_date)
        &
        df["date"].lt(cutoff_date)
        &
        df["weekday_num"].eq(weekday)
    ].copy()

    if historical.empty:
        return None

    selected_dates = (
        historical["date"]
        .drop_duplicates()
        .sort_values()
        .tail(n_weeks)
    )

    if len(selected_dates) < n_weeks:
        return None

    historical = historical[
        historical["date"].isin(
            selected_dates
        )
    ]

    demand_col = (
        "entries"
        if demand_type == "entry"
        else "exits"
    )

    hourly_totals = (
        historical
        .groupby("hour")[demand_col]
        .sum()
        .reindex(
            range(24),
            fill_value=0,
        )
        .astype(float)
    )

    total_demand = hourly_totals.sum()

    if total_demand <= 0:
        return None

    profile = (
        hourly_totals
        /
        total_demand
    )

    if not np.isclose(
        profile.sum(),
        1.0,
        atol=1e-10,
    ):
        raise RuntimeError(
            "Hourly profile does not sum to 1."
        )

    return profile.to_dict()


def distribute_daily_to_hourly(
    daily_forecast,
    hourly_profile,
):
    """
    Distribute a daily forecast across hours using the hourly profile.

    Parameters
    ----------
    daily_forecast : float
        Total daily demand forecast.

    hourly_profile : dict
        Mapping of hour (0-23) to proportion of daily demand.

    Returns
    -------
    dict
        Mapping of hour (0-23) to hourly forecast.
    """

    if hourly_profile is None:
        return None

    if pd.isna(daily_forecast):
        return None

    hourly_forecast = {}

    for hour in range(24):
        proportion = hourly_profile.get(hour, 0.0)
        hourly_forecast[hour] = daily_forecast * proportion

    return hourly_forecast


def create_hourly_forecast_for_date(
    target_date,
    horizon_days,
    daily_forecast_entry,
    daily_forecast_exit,
    hourly_actuals,
):
    """
    Create complete hourly forecasts for entries and exits on a target date.
    """

    target_date = pd.Timestamp(target_date).normalize()

    cutoff_timestamp = (
        target_date
        - pd.Timedelta(days=horizon_days)
        + pd.Timedelta(hours=7)
    )

    # Build hourly profiles.
    entry_profile = build_hourly_profile(
        hourly_actuals=hourly_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="entry",
        n_weeks=6,
    )

    exit_profile = build_hourly_profile(
        hourly_actuals=hourly_actuals,
        target_date=target_date,
        cutoff_timestamp=cutoff_timestamp,
        demand_type="exit",
        n_weeks=6,
    )

    # Distribute daily forecasts.
    hourly_entries = distribute_daily_to_hourly(
        daily_forecast=daily_forecast_entry,
        hourly_profile=entry_profile,
    )

    hourly_exits = distribute_daily_to_hourly(
        daily_forecast=daily_forecast_exit,
        hourly_profile=exit_profile,
    )

    # Build output rows.
    rows = []

    for hour in range(24):

        hour_datetime = target_date + pd.Timedelta(hours=hour)

        rows.append({
            "target_date": target_date,
            "hour": hour,
            "datetime": hour_datetime,
            "horizon_days": horizon_days,
            "forecast_entries": (
                hourly_entries.get(hour)
                if hourly_entries
                else np.nan
            ),
            "forecast_exits": (
                hourly_exits.get(hour)
                if hourly_exits
                else np.nan
            ),
            "entry_profile_pct": (
                entry_profile.get(hour, 0) * 100
                if entry_profile
                else np.nan
            ),
            "exit_profile_pct": (
                exit_profile.get(hour, 0) * 100
                if exit_profile
                else np.nan
            ),
        })

    return pd.DataFrame(rows)


def run_hourly_forecast_simulation(
    daily_results,
    model_data,
    config,
    entry_model_by_horizon,
    exit_model_by_horizon,
):
    """
    Run the hourly profile-window tournament using the winning daily model
    for each forecast horizon.
    """

    print("\n" + "=" * 80)
    print("RUNNING HOURLY FORECAST SIMULATION")
    print("=" * 80)

    hourly_actuals = model_data[
        "hourly_actuals"
    ]

    daily_lookup = (
        daily_results[
            [
                "target_date",
                "horizon_days",
                "demand_type",
                "scenario_id",
                "forecast_value",
            ]
        ]
        .set_index(
            [
                "target_date",
                "horizon_days",
                "demand_type",
                "scenario_id",
            ]
        )["forecast_value"]
        .to_dict()
    )

    test_dates = pd.to_datetime(
        create_test_dates(
            config
        )["target_date"]
    ).dt.normalize()

    horizons = config[
        "forecast_horizons_days"
    ]

    profile_windows = config[
        "hourly_profile_windows"
    ]

    all_rows = []

    for target_date in test_dates:

        for horizon_days in horizons:

            cutoff_timestamp = (
                target_date
                - pd.Timedelta(
                    days=horizon_days
                )
                + pd.Timedelta(hours=7)
            )

            entry_model = (
                entry_model_by_horizon.get(
                    horizon_days
                )
            )

            exit_model = (
                exit_model_by_horizon.get(
                    horizon_days
                )
            )

            if (
                entry_model is None
                or exit_model is None
            ):
                continue

            daily_entry = daily_lookup.get(
                (
                    target_date,
                    horizon_days,
                    "entry",
                    entry_model,
                ),
                np.nan,
            )

            daily_exit = daily_lookup.get(
                (
                    target_date,
                    horizon_days,
                    "exit",
                    exit_model,
                ),
                np.nan,
            )

            for profile_weeks in profile_windows:

                entry_profile = build_hourly_profile(
                    hourly_actuals=hourly_actuals,
                    target_date=target_date,
                    cutoff_timestamp=cutoff_timestamp,
                    demand_type="entry",
                    n_weeks=profile_weeks,
                )

                exit_profile = build_hourly_profile(
                    hourly_actuals=hourly_actuals,
                    target_date=target_date,
                    cutoff_timestamp=cutoff_timestamp,
                    demand_type="exit",
                    n_weeks=profile_weeks,
                )

                hourly_entries = (
                    distribute_daily_to_hourly(
                        daily_forecast=daily_entry,
                        hourly_profile=entry_profile,
                    )
                )

                hourly_exits = (
                    distribute_daily_to_hourly(
                        daily_forecast=daily_exit,
                        hourly_profile=exit_profile,
                    )
                )

                for hour in range(24):

                    all_rows.append(
                        {
                            "target_date":
                                target_date,

                            "datetime": (
                                target_date
                                +
                                pd.Timedelta(
                                    hours=hour
                                )
                            ),

                            "hour":
                                hour,

                            "horizon_days":
                                horizon_days,

                            "profile_weeks":
                                profile_weeks,

                            "entry_model":
                                entry_model,

                            "exit_model":
                                exit_model,

                            "forecast_entries": (
                                hourly_entries.get(
                                    hour
                                )
                                if hourly_entries
                                is not None
                                else np.nan
                            ),

                            "forecast_exits": (
                                hourly_exits.get(
                                    hour
                                )
                                if hourly_exits
                                is not None
                                else np.nan
                            ),

                            "entry_profile_pct": (
                                entry_profile.get(
                                    hour,
                                    0.0,
                                )
                                * 100
                                if entry_profile
                                is not None
                                else np.nan
                            ),

                            "exit_profile_pct": (
                                exit_profile.get(
                                    hour,
                                    0.0,
                                )
                                * 100
                                if exit_profile
                                is not None
                                else np.nan
                            ),
                        }
                    )

    hourly_forecasts = pd.DataFrame(
        all_rows
    )

    actuals = hourly_actuals[
        [
            "datetime",
            "entries",
            "exits",
        ]
    ].rename(
        columns={
            "entries":
                "actual_entries",

            "exits":
                "actual_exits",
        }
    )

    hourly_forecasts = (
        hourly_forecasts
        .merge(
            actuals,
            on="datetime",
            how="left",
        )
    )

    for column in [
        "actual_entries",
        "actual_exits",
    ]:
        hourly_forecasts[column] = (
            hourly_forecasts[column]
            .fillna(0)
        )

    hourly_forecasts["entry_error"] = (
        hourly_forecasts[
            "forecast_entries"
        ]
        -
        hourly_forecasts[
            "actual_entries"
        ]
    )

    hourly_forecasts["exit_error"] = (
        hourly_forecasts[
            "forecast_exits"
        ]
        -
        hourly_forecasts[
            "actual_exits"
        ]
    )

    hourly_forecasts["entry_abs_error"] = (
        hourly_forecasts[
            "entry_error"
        ].abs()
    )

    hourly_forecasts["exit_abs_error"] = (
        hourly_forecasts[
            "exit_error"
        ].abs()
    )

    return hourly_forecasts


def score_hourly_forecasts(
    hourly_forecasts,
):
    """
    Score entry and exit hourly forecasts independently by horizon,
    profile window and hour of day.
    """

    if hourly_forecasts.empty:
        return pd.DataFrame()

    rows = []

    specifications = [
        (
            "ENTRY",
            "forecast_entries",
            "actual_entries",
        ),
        (
            "EXIT",
            "forecast_exits",
            "actual_exits",
        ),
    ]

    for (
        flow_type,
        forecast_col,
        actual_col,
    ) in specifications:

        valid = hourly_forecasts[
            hourly_forecasts[
                forecast_col
            ].notna()
            &
            hourly_forecasts[
                actual_col
            ].notna()
        ].copy()

        if valid.empty:
            continue

        for (
            horizon_days,
            profile_weeks,
            hour,
        ), group in valid.groupby(
            [
                "horizon_days",
                "profile_weeks",
                "hour",
            ]
        ):

            actual = group[
                actual_col
            ].astype(float)

            forecast = group[
                forecast_col
            ].astype(float)

            error = forecast - actual

            denominator = actual.abs().sum()

            rows.append(
                {
                    "flow_type":
                        flow_type,

                    "horizon_days":
                        horizon_days,

                    "horizon_label":
                        f"T-{horizon_days}",

                    "profile_weeks":
                        profile_weeks,

                    "hour":
                        hour,

                    "observations":
                        len(group),

                    "wape_pct": (
                        error.abs().sum()
                        /
                        denominator
                        * 100
                        if denominator > 0
                        else np.nan
                    ),

                    "mae":
                        error.abs().mean(),

                    "rmse":
                        np.sqrt(
                            error.pow(2).mean()
                        ),

                    "bias":
                        error.mean(),

                    "average_actual":
                        actual.mean(),

                    "average_forecast":
                        forecast.mean(),
                }
            )

    return pd.DataFrame(rows)

def create_best_hourly_profile_windows(
    hourly_scores,
):
    """
    Choose the lowest-WAPE hourly profile window for each flow and horizon,
    aggregating errors across all hours.
    """

    if hourly_scores.empty:
        return pd.DataFrame()

    summary = (
        hourly_scores
        .groupby(
            [
                "flow_type",
                "horizon_days",
                "horizon_label",
                "profile_weeks",
            ],
            as_index=False,
        )
        .agg(
            average_hourly_wape_pct=(
                "wape_pct",
                "mean",
            ),
            average_hourly_mae=(
                "mae",
                "mean",
            ),
            average_hourly_bias=(
                "bias",
                "mean",
            ),
        )
    )

    return (
        summary
        .sort_values(
            [
                "flow_type",
                "horizon_days",
                "average_hourly_wape_pct",
                "average_hourly_mae",
            ]
        )
        .groupby(
            [
                "flow_type",
                "horizon_days",
            ],
            as_index=False,
        )
        .first()
        .sort_values(
            [
                "flow_type",
                "horizon_days",
            ]
        )
        .reset_index(drop=True)
    )

def get_recommended_models_by_horizon(
    best_by_horizon,
):
    """
    Extract the recommended model for each horizon from the daily results.

    Returns two dictionaries for use in hourly forecasting.
    """

    entry_models = {}
    exit_models = {}

    for _, row in best_by_horizon.iterrows():

        horizon = row["horizon_days"]
        scenario = row["scenario_id"]
        flow_type = row["flow_type"]

        if flow_type == "ENTRY":
            entry_models[horizon] = scenario
        else:
            exit_models[horizon] = scenario

    return entry_models, exit_models


def export_hourly_results(
    hourly_forecasts,
    hourly_scores,
    best_profile_windows,
    output_path,
):
    """
    Export hourly profile-window validation results.
    """

    output_path = Path(
        output_path
    )

    peak_analysis = (
        hourly_forecasts
        .sort_values(
            [
                "target_date",
                "horizon_days",
                "profile_weeks",
                "hour",
            ]
        )
        .groupby(
            [
                "target_date",
                "horizon_days",
                "profile_weeks",
            ]
        )
        .apply(
            lambda group: pd.Series(
                {
                    "peak_entry_hour": (
                        group.loc[
                            group[
                                "forecast_entries"
                            ].idxmax(),
                            "hour",
                        ]
                        if group[
                            "forecast_entries"
                        ].notna().any()
                        else np.nan
                    ),

                    "peak_exit_hour": (
                        group.loc[
                            group[
                                "forecast_exits"
                            ].idxmax(),
                            "hour",
                        ]
                        if group[
                            "forecast_exits"
                        ].notna().any()
                        else np.nan
                    ),
                }
            ),
            include_groups=False,
        )
        .reset_index()
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        best_profile_windows.to_excel(
            writer,
            sheet_name="Best Profile Windows",
            index=False,
        )

        hourly_scores.to_excel(
            writer,
            sheet_name="Hourly Performance",
            index=False,
        )

        peak_analysis.to_excel(
            writer,
            sheet_name="Peak Hour Analysis",
            index=False,
        )

        hourly_forecasts.to_excel(
            writer,
            sheet_name="Hourly Forecasts",
            index=False,
        )

        format_output_workbook(
            writer.book
        )

    print(
        f"Exported hourly results to: "
        f"{output_path}"
    )

    return output_path

# =============================================================================
# 3.19 — MAIN EXECUTION
# =============================================================================
#
# THIS IS THE ONLY PART THAT SHOULD ACTUALLY RUN THE PIPELINE.
#
# Everything above defines validation, scoring and export functions.
#
# =============================================================================


if __name__ == "__main__":

    # -------------------------------------------------------------------------
    # Start timer
    # -------------------------------------------------------------------------

    t0 = time.perf_counter()

    print()
    print("=" * 90)
    print("FASTPARK FORECAST SIMULATION")
    print("HISTORICAL BACKTEST / MODEL TOURNAMENT")
    print("=" * 90)

    # -------------------------------------------------------------------------
    # STEP 1 — CONFIG
    # -------------------------------------------------------------------------

    config = get_final_simulation_config()

    t1 = step(
        t0,
        "Loaded simulation configuration",
    )

    # -------------------------------------------------------------------------
    # STEP 2 — SCENARIO CATALOGUE
    # -------------------------------------------------------------------------

    validate_final_scenario_catalogue(
        config
    )

    t2 = step(
        t1,
        "Validated E1-E15 and X1-X17 scenario catalogue",
    )

    # -------------------------------------------------------------------------
    # STEP 3 — HORIZONS
    # -------------------------------------------------------------------------

    validate_horizons(
        config
    )

    t3 = step(
        t2,
        "Validated T-minus forecast horizons",
    )

    # -------------------------------------------------------------------------
    # STEP 4 — TEST DATES
    # -------------------------------------------------------------------------

    test_dates = validate_test_dates(
        config
    )

    t4 = step(
        t3,
        "Validated 126 historical target dates",
    )

    # -------------------------------------------------------------------------
    # STEP 5 — LOAD RAW DATA
    # -------------------------------------------------------------------------
    #
    # This calls the same extraction functions used by the FastPark analysis.
    #
    # We deliberately DO NOT recreate the SQL here.
    #
    # This is important because the analysis already established:
    #
    #     AirportX.v_Bookings
    #     FastPark.v_EntryAndExits
    #     EAL.FlightPerformance
    #
    # and the corresponding field definitions.
    # -------------------------------------------------------------------------

    print()
    print("=" * 90)
    print("STEP 5 — LOADING AND CLEANING HISTORICAL DATA")
    print("=" * 90)

    raw_data = load_simulation_data(
        config
    )

    t5 = step(
        t4,
        (
            "Loaded and cleaned bookings, operations "
            "and flight/passenger data"
        ),
    )

    # -------------------------------------------------------------------------
    # STEP 6 — DATA QUALITY VALIDATION
    # -------------------------------------------------------------------------

    validate_loaded_data(
        raw_data,
        config,
    )

    t6 = step(
        t5,
        "Validated cleaned source datasets",
    )

    # -------------------------------------------------------------------------
    # STEP 7 — PREPARE FORECASTING MODEL DATA
    # -------------------------------------------------------------------------
    #
    # This creates:
    #
    #     daily_actuals
    #     hourly_actuals
    #     daily_passengers
    #     entry_booking_curve
    #     exit_booking_curve
    #     duration_completion_factors
    #
    # These are the inputs used by the E1-E13 / X1-X16 scenario engine.
    # -------------------------------------------------------------------------

    print()
    print("=" * 90)
    print("STEP 7 — PREPARING FORECAST MODEL DATA")
    print("=" * 90)

    model_data = prepare_forecast_model_data(
        data=raw_data,
        config=config,
    )

    t7 = step(
        t6,
        (
            "Prepared daily actuals, hourly actuals, "
            "passenger context, booking curves and duration factors"
        ),
    )

    # -------------------------------------------------------------------------
    # STEP 8 — VALIDATE ACTUAL TARGETS
    # -------------------------------------------------------------------------

    test_validation = validate_daily_actuals(
        model_data=model_data,
        test_dates=test_dates,
    )

    t8 = step(
        t7,
        "Validated historical actual demand targets",
    )

    # -------------------------------------------------------------------------
    # STEP 9 — RUN THE ACTUAL 32-SCENARIO TOURNAMENT
    # -------------------------------------------------------------------------
    #
    # IMPORTANT:
    #
    # Use run_daily_simulation().
    #
    # Do NOT use run_and_analyse_simulation().
    #
    # run_daily_simulation() is connected to:
    #
    #     forecast_entry_scenario()
    #     forecast_exit_scenario()
    #
    # which are the functions containing the actual E1-E15 and X1-X17 logic.
    #

    # -------------------------------------------------------------------------

    print()
    print("=" * 90)
    print("STEP 9 — RUNNING 32-SCENARIO HISTORICAL BACKTEST")
    print("=" * 90)

    daily_results = run_daily_simulation(
        data=model_data,
        config=config,
    )

    t9 = step(
        t8,
        (
            f"Completed daily simulation "
            f"({len(daily_results):,} rows)"
        ),
    )

    # -------------------------------------------------------------------------
    # STEP 10 — QUICK CHECK
    # -------------------------------------------------------------------------

    print_simulation_quick_check(
        daily_results
    )

    validate_no_scenario_errors(
        daily_results
    )
    
    t10 = step(
        t9,
        "Completed simulation sanity check",
    )

    # -------------------------------------------------------------------------
    # STEP 11 — STRICT RESULT VALIDATION
    # -------------------------------------------------------------------------

    if daily_results.empty:

        raise RuntimeError(
            "The simulation returned no results."
        )

    if config.get(
        "smoke_test",
        False,
    ):

        expected_rows = (
            config["smoke_test_target_count"]
            *
            len(config["smoke_test_horizons"])
            *
            config["expected_scenario_count"]
        )

    else:

        expected_rows = config[
            "expected_simulation_rows"
        ]

    print()
    print(
        f"Expected theoretical rows: "
        f"{expected_rows:,}"
    )

    print(
        f"Actual rows generated: "
        f"{len(daily_results):,}"
    )

    # The simulation should generate exactly:
    #
    #     126 x 15 x 29
    #
    # = 54,810 rows.
    #
    # If not, something has gone wrong with the scenario catalogue, test
    # dates or horizons.

    if len(daily_results) != expected_rows:

        raise RuntimeError(
            "Unexpected simulation row count.\n"
            f"Expected: {expected_rows:,}\n"
            f"Found:    {len(daily_results):,}"
        )

    actual_entry_scenarios = set(
        daily_results.loc[
            daily_results["demand_type"].eq("entry"),
            "scenario_id",
        ].unique()
    )

    actual_exit_scenarios = set(
        daily_results.loc[
            daily_results["demand_type"].eq("exit"),
            "scenario_id",
        ].unique()
    )

    if actual_entry_scenarios != set(
        config["expected_entry_scenarios"]
    ):

        raise RuntimeError(
            "Simulation did not produce exactly E1-E13.\n"
            f"Found: {sorted(actual_entry_scenarios)}"
        )

    if actual_exit_scenarios != set(
        config["expected_exit_scenarios"]
    ):

        raise RuntimeError(
            "Simulation did not produce exactly X1-X16.\n"
            f"Found: {sorted(actual_exit_scenarios)}"
        )

    t11 = step(
        t10,
        (
            f"Validated {expected_rows:,} simulation rows "
            f"and {config['expected_scenario_count']} scenarios"
        )
    )

    # -------------------------------------------------------------------------
    # STEP 12 — SCORE
    # -------------------------------------------------------------------------

    print()
    print("=" * 90)
    print("STEP 12 — SCORING FORECASTS AGAINST ACTUALS")
    print("=" * 90)

    performance = score_daily_simulation_results(
        daily_results
    )

    t12 = step(
        t11,
        "Calculated WAPE, MAE, RMSE and bias",
    )

    # -------------------------------------------------------------------------
    # STEP 13 — COVERAGE
    # -------------------------------------------------------------------------

    coverage = calculate_final_coverage(
        daily_results,
        config,
    )

    t13 = step(
        t12,
        "Calculated scenario forecast coverage",
    )

    # -------------------------------------------------------------------------
    # STEP 14 — RANK MODELS
    # -------------------------------------------------------------------------

    ranked_performance = rank_final_models(
        performance
    )

    best_by_horizon = create_best_by_horizon(
        ranked_performance,
        coverage,
    )

    overall_ranking = (
        create_overall_model_ranking(
            daily_results
        )
    )

    stability = create_model_stability(
        ranked_performance
    )

    all_weight_results, best_weight_results = (
        create_weight_calibration_outputs(
            daily_results=daily_results,
            config=config,
        )
    )

    weekday_analysis = create_weekday_analysis(
        daily_results
    )

    month_analysis = create_month_analysis(
        daily_results
    )

    demand_band_analysis = (
        create_demand_band_analysis(
            daily_results
        )
    )

    visibility_analysis = (
        create_visibility_analysis(
            daily_results
        )
    )

    visibility_summary = (
        create_visibility_summary(
            visibility_analysis
        )
    )

    worst_forecast_days = (
        create_worst_forecast_days(
            daily_results,
            top_n=100,
        )
    )

    family_comparison = (
        create_family_comparison(
            daily_results
        )
    )

    incremental_benefit = (
        create_incremental_benefit_analysis(
            performance
        )
    )    

    t14 = step(
        t13,
        "Ranked models and identified best scenario by horizon",
    )

    # -------------------------------------------------------------------------
    # STEP 15 — EXPORT
    # -------------------------------------------------------------------------

    print()
    print("=" * 90)
    print("STEP 15 — EXPORTING EXCEL RESULTS")
    print("=" * 90)

    output_path = export_simulation_results(
        output_path=config["output_path"],
        data=raw_data,
        model_data=model_data,
        test_dates=test_dates,
        test_validation=test_validation,
        daily_results=daily_results,
        performance=performance,
        ranked_performance=ranked_performance,
        coverage=coverage,
        best_by_horizon=best_by_horizon,
        overall_ranking=overall_ranking,
        stability=stability,
        all_weight_results=all_weight_results,
        best_weight_results=best_weight_results,
        weekday_analysis=weekday_analysis,
        month_analysis=month_analysis,
        demand_band_analysis=demand_band_analysis,
        visibility_analysis=visibility_analysis,
        visibility_summary=visibility_summary,
        worst_forecast_days=worst_forecast_days,
        family_comparison=family_comparison,
        incremental_benefit=incremental_benefit,
    )

    t15 = step(
        t14,
        f"Exported simulation workbook: {output_path}",
    )

    # -------------------------------------------------------------------------
    # STEP 16 — OPTIONAL HOURLY PROFILE VALIDATION
    # -------------------------------------------------------------------------

    if config.get(
        "enable_hourly_analysis",
        False,
    ):

        print()
        print("=" * 90)
        print(
            "STEP 16 — RUNNING HOURLY PROFILE VALIDATION"
        )
        print("=" * 90)

        (
            entry_models_by_horizon,
            exit_models_by_horizon,
        ) = get_recommended_models_by_horizon(
            best_by_horizon
        )

        hourly_forecasts = (
            run_hourly_forecast_simulation(
                daily_results=daily_results,
                model_data=model_data,
                config=config,
                entry_model_by_horizon=
                    entry_models_by_horizon,
                exit_model_by_horizon=
                    exit_models_by_horizon,
            )
        )

        hourly_scores = score_hourly_forecasts(
            hourly_forecasts
        )

        best_hourly_profiles = (
            create_best_hourly_profile_windows(
                hourly_scores
            )
        )

        hourly_output_path = (
            export_hourly_results(
                hourly_forecasts=
                    hourly_forecasts,

                hourly_scores=
                    hourly_scores,

                best_profile_windows=
                    best_hourly_profiles,

                output_path=config[
                    "hourly_output_path"
                ],
            )
        )

        t16 = step(
            t15,
            (
                "Completed hourly profile "
                f"validation: {hourly_output_path}"
            ),
        )

    else:

        t16 = t15

    # -------------------------------------------------------------------------
    # FINAL REPORT
    # -------------------------------------------------------------------------

    print_final_results_summary(
        performance=performance,
        best_by_horizon=best_by_horizon,
        overall_ranking=overall_ranking,
        coverage=coverage,
    )

    step(
        t16,
        "FastPark forecast simulation complete",
    )


